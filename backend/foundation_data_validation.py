from __future__ import annotations

import json
import random
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(Path(__file__).resolve().parent) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parent))

from database import SessionLocal
from email_import_service import EmailImportService
from main import _build_candidate_entries_from_record, _create_candidate_record, parse_kb_excel_import


OUT_DIR = ROOT / "测试数据" / "基础数据抽样验证_20260420"
SEED = 20260420


SOURCES = [
    {
        "key": "dev_email",
        "label": "销售开发类邮件（正文）",
        "path": ROOT / "docs" / "开发邮件出已标识意图.xlsx",
        "kind": "email",
        "source_type": "email_excel",
    },
    {
        "key": "biz_email",
        "label": "除开发类，日常业务沟通邮件（正文）",
        "path": ROOT / "docs" / "除开发类销售沟通邮件出已标识意图.xlsx",
        "kind": "email",
        "source_type": "email_excel",
    },
    {
        "key": "business_knowledge",
        "label": "业务知识、业务流程、话术、案例",
        "path": ROOT / "docs" / "业务知识、业务流程、话术、案例.xlsx",
        "kind": "knowledge_xlsx",
        "source_type": "case_extract",
    },
    {
        "key": "translation_quote_doc",
        "label": "笔译报价规范",
        "path": ROOT / "docs" / "笔译报价.doc",
        "kind": "doc_quote",
        "source_type": "case_extract",
    },
]


def truncate(text: object, n: int = 500) -> str:
    value = str(text or "").replace("\r", "").strip()
    return value[:n] + ("..." if len(value) > n else "")


def load_email_records(path: Path) -> list[dict]:
    return EmailImportService.parse_file(path.read_bytes(), path.name)


def load_business_records(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    records = []
    for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        title = str(row[0] or "").strip() if len(row) > 0 else ""
        content = str(row[1] or "").strip() if len(row) > 1 else ""
        if not title or not content:
            continue
        records.append({"row": row_no, "title": title, "subject": title, "content": content})
    return records


def inspect_doc_package(path: Path) -> dict:
    import zipfile

    if not zipfile.is_zipfile(path):
        return {
            "extractable": False,
            "reason": "不是 zip/docx 包，当前环境无 Word/soffice/antiword 可转换 .doc",
        }
    with zipfile.ZipFile(path) as package:
        names = package.namelist()
    has_document = any(name in {"word/document.xml", "word/document2.xml"} for name in names)
    return {
        "extractable": has_document,
        "reason": "" if has_document else "未找到 word/document.xml，包内只有主题/样式资源，未发现正文",
        "package_files": names[:20],
    }


def write_sample_json(key: str, samples: list[dict]) -> str:
    path = OUT_DIR / f"{key}_sample5.json"
    path.write_text(json.dumps(samples, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return str(path.relative_to(ROOT))


def write_business_standard_excel(samples: list[dict]) -> Path:
    path = OUT_DIR / "business_knowledge_sample5_standard_import.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "kb_import"
    headers = [
        "标题",
        "内容",
        "知识类型",
        "切片类型",
        "业务线",
        "子服务",
        "语种",
        "服务范围",
        "地区",
        "客户层级",
        "优先级",
        "风险等级",
        "生效日期",
        "失效日期",
        "单位",
        "币种",
        "价格",
        "最高价格",
        "最低收费",
        "加急倍率",
        "税费",
        "标签",
    ]
    sheet.append(headers)
    for item in samples:
        text = f"{item.get('title', '')}\n{item.get('content', '')}"
        if any(word in text for word in ["报价", "价格", "收费", "费用", "最低收费", "加急", "税点", "折扣", "元/千字", "元每千字"]):
            knowledge_type, chunk_type, risk_level = "pricing", "rule", "high"
        elif "流程" in text:
            knowledge_type, chunk_type, risk_level = "process", "rule", "medium"
        elif any(word in text for word in ["话术", "怎么回复", "客户问"]):
            knowledge_type, chunk_type, risk_level = "faq", "template", "medium"
        else:
            knowledge_type, chunk_type, risk_level = "faq", "faq", "medium"

        if any(word in text for word in ["印刷", "画册", "纸张", "装订"]):
            business_line = "printing"
        elif any(word in text for word in ["展台", "展会", "搭建", "撤展"]):
            business_line = "exhibition"
        elif any(word in text for word in ["翻译", "笔译", "口译", "同传", "语种"]):
            business_line = "translation"
        else:
            business_line = "general"

        sheet.append([
            item["title"],
            item["content"],
            knowledge_type,
            chunk_type,
            business_line,
            "",
            "",
            "general",
            "",
            "",
            70,
            risk_level,
            "2026-04-20",
            "2030-12-31",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "foundation_sample",
        ])
    workbook.save(path)
    return path


def validate_candidates(source: dict, samples: list[dict]) -> dict:
    db = SessionLocal()
    results = []
    created_total = 0
    try:
        for index, record in enumerate(samples, start=1):
            row_result = {
                "sample_index": index,
                "source_row": record.get("row"),
                "title_or_subject": truncate(record.get("title") or record.get("subject"), 120),
                "content_preview": truncate(record.get("content"), 240),
                "status": "pending",
                "created_count": 0,
                "candidates": [],
            }
            try:
                entries = _build_candidate_entries_from_record(record, source["source_type"])
                for offset, entry in enumerate(entries, start=1):
                    candidate = _create_candidate_record(
                        db,
                        title=entry["title"],
                        content=entry["content"],
                        knowledge_type=entry["knowledge_type"],
                        chunk_type=entry["chunk_type"],
                        business_line=entry["business_line"],
                        sub_service=entry.get("sub_service"),
                        language_pair=entry.get("language_pair"),
                        service_scope=entry.get("service_scope"),
                        region=entry.get("region"),
                        customer_tier=entry.get("customer_tier"),
                        priority=int(entry.get("priority") or 60),
                        risk_level=entry.get("risk_level") or "medium",
                        effective_from=entry.get("effective_from"),
                        effective_to=entry.get("effective_to"),
                        pricing_rule=entry.get("pricing_rule"),
                        source_type=source["source_type"],
                        source_ref=f"foundation_validation_20260420_{source['key']}_{record.get('row')}_{offset}",
                        source_snapshot=entry.get("source_snapshot") or {"sample_record": record},
                        owner="foundation_data_validation",
                        operator="codex_validation",
                        review_notes="基础数据随机抽样验证生成",
                    )
                    row_result["candidates"].append({
                        "candidate_id": str(candidate.candidate_id),
                        "title": candidate.title,
                        "knowledge_type": candidate.knowledge_type,
                        "business_line": candidate.business_line,
                        "risk_level": candidate.risk_level,
                    })
                row_result["created_count"] = len(row_result["candidates"])
                row_result["status"] = "success"
                created_total += row_result["created_count"]
                db.commit()
            except Exception as exc:
                db.rollback()
                row_result["status"] = "error"
                row_result["error"] = f"{type(exc).__name__}: {exc}"
            results.append(row_result)
    finally:
        db.close()
    return {"created_total": created_total, "items": results}


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    random.seed(SEED)
    report = {"run_at": datetime.now().isoformat(timespec="seconds"), "seed": SEED, "sources": []}

    for source in SOURCES:
        item = {
            "key": source["key"],
            "label": source["label"],
            "path": str(source["path"].relative_to(ROOT)),
            "kind": source["kind"],
        }
        if source["kind"] == "email":
            records = load_email_records(source["path"])
            samples = random.sample(records, min(5, len(records)))
            item["parsed_count"] = len(records)
            item["sample_file"] = write_sample_json(source["key"], samples)
            item["validation"] = validate_candidates(source, samples)
            item["usable_as_foundation"] = True
            item["ingest_recommendation"] = "适合作为候选知识来源；建议先进入 candidate 池，人工审核后再转正式知识。"
        elif source["kind"] == "knowledge_xlsx":
            records = load_business_records(source["path"])
            samples = random.sample(records, min(5, len(records)))
            item["parsed_count"] = len(records)
            item["sample_file"] = write_sample_json(source["key"], samples)
            standard_path = write_business_standard_excel(samples)
            item["standard_import_file"] = str(standard_path.relative_to(ROOT))
            parsed = parse_kb_excel_import(standard_path.read_bytes(), standard_path.name, "faq")
            item["standard_preview"] = {
                "valid_count": len(parsed["valid_rows"]),
                "failed_count": len(parsed["failed"]),
                "skipped_count": len(parsed["skipped"]),
            }
            item["validation"] = validate_candidates(source, samples)
            item["usable_as_foundation"] = True
            item["ingest_recommendation"] = "适合作为基础知识来源；原文件缺标准表头，建议先转换为标准模板后入库。"
        else:
            doc_info = inspect_doc_package(source["path"])
            item.update(doc_info)
            item["parsed_count"] = 0
            item["sample_file"] = None
            item["validation"] = {"created_total": 0, "items": []}
            item["usable_as_foundation"] = bool(doc_info.get("extractable"))
            item["ingest_recommendation"] = "当前文件未发现可抽取正文，不能作为入库基础数据；请重新提供包含正文表格/段落的 .docx/.xlsx/.pdf 或可解析 .doc。"
        report["sources"].append(item)

    report_path = OUT_DIR / "基础数据抽样验证报告.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    print(f"REPORT_PATH={report_path}")


if __name__ == "__main__":
    main()
