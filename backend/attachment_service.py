"""
Attachment parsing service.
Ported from other/KnowledgeBase/BackEnd/app/services/attachment_parser.py

Supported formats: txt/md, json, csv, xml, html, pdf, docx, doc, xlsx, xls,
                   zip (with nested extraction), jpg/png/etc (OCR if available)

Optional dependencies (gracefully degraded if missing):
  pip install pypdf python-docx openpyxl xlrd olefile Pillow pytesseract py7zr rarfile
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import string
import zipfile
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from uuid import uuid4
from xml.etree import ElementTree

logger = logging.getLogger(__name__)

UPLOADS_DIR = Path(__file__).parent / "uploads"
ARTIFACTS_DIR = UPLOADS_DIR / "parsed-attachments"

TEXT_EXTENSIONS = {".txt", ".md", ".log", ".ini", ".yaml", ".yml"}
STRUCTURED_EXTENSIONS = {".json", ".csv", ".tsv", ".xml", ".html", ".htm"}
ARCHIVE_EXTENSIONS = {".zip", ".7z", ".rar"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tif", ".tiff"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls"}
MAX_ATTACHMENT_SIZE = 25 * 1024 * 1024
MAX_ARCHIVE_ENTRY_SIZE = 10 * 1024 * 1024
MAX_ARCHIVE_DEPTH = 2
MAX_ARCHIVE_ENTRIES = 50
MAX_TEXT_LENGTH = 20000


# ── Optional imports ──────────────────────────────────────────────────────────

try:
    from pypdf import PdfReader as _PdfReader
except ImportError:
    _PdfReader = None

try:
    from docx import Document as _DocxDocument
except ImportError:
    _DocxDocument = None

try:
    from openpyxl import load_workbook as _load_workbook
except ImportError:
    _load_workbook = None

try:
    import xlrd as _xlrd
except ImportError:
    _xlrd = None

try:
    import olefile as _olefile
except ImportError:
    _olefile = None

try:
    from PIL import Image as _PILImage
except ImportError:
    _PILImage = None

try:
    import pytesseract as _pytesseract
    import os as _os
    _TESS_WIN = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if _os.name == "nt" and _os.path.exists(_TESS_WIN):
        _pytesseract.pytesseract.tesseract_cmd = _TESS_WIN
except ImportError:
    _pytesseract = None

try:
    import py7zr as _py7zr
except ImportError:
    _py7zr = None

try:
    import rarfile as _rarfile
except ImportError:
    _rarfile = None


# ── Text helpers ──────────────────────────────────────────────────────────────

class _HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        text = data.strip()
        if text:
            self.parts.append(text)

    def get_text(self) -> str:
        return "\n".join(self.parts)


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _decode_bytes(payload: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "gb18030", "latin-1"):
        try:
            return payload.decode(enc)
        except Exception:
            continue
    return payload.decode("utf-8", errors="ignore")


def _parse_html_text(text: str) -> str:
    p = _HTMLTextExtractor()
    p.feed(text)
    return p.get_text()


def _build_summary(text: str) -> str:
    return text[:300].strip()


def _extract_printable(payload: bytes) -> str:
    printable = set(string.printable)
    runs, current = [], []
    for byte in payload:
        ch = chr(byte)
        if ch in printable:
            current.append(ch)
        else:
            if len(current) >= 4:
                runs.append("".join(current))
            current = []
    if current:
        runs.append("".join(current))
    return "\n".join(runs)[:MAX_TEXT_LENGTH]


# ── Format parsers ────────────────────────────────────────────────────────────

def _parse_text(payload: bytes) -> str:
    return _decode_bytes(payload)[:MAX_TEXT_LENGTH]


def _parse_json(payload: bytes) -> str:
    try:
        obj = json.loads(_decode_bytes(payload))
        return json.dumps(obj, ensure_ascii=False, indent=2)[:MAX_TEXT_LENGTH]
    except Exception:
        return _decode_bytes(payload)[:MAX_TEXT_LENGTH]


def _parse_csv_bytes(payload: bytes) -> str:
    text = _decode_bytes(payload)
    lines = []
    for delim in (",", "\t", ";"):
        try:
            reader = csv.reader(io.StringIO(text), delimiter=delim)
            rows = list(reader)
            if rows and max(len(r) for r in rows) > 1:
                lines = ["\t".join(r) for r in rows[:200]]
                break
        except Exception:
            continue
    return "\n".join(lines)[:MAX_TEXT_LENGTH] or text[:MAX_TEXT_LENGTH]


def _parse_xml(payload: bytes) -> str:
    try:
        root = ElementTree.fromstring(payload)
        parts: list[str] = []
        for elem in root.iter():
            if elem.text and elem.text.strip():
                parts.append(f"{elem.tag}: {elem.text.strip()}")
        return "\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception:
        return _decode_bytes(payload)[:MAX_TEXT_LENGTH]


def _parse_html(payload: bytes) -> str:
    return _parse_html_text(_decode_bytes(payload))[:MAX_TEXT_LENGTH]


def _parse_pdf(payload: bytes) -> str:
    if _PdfReader is None:
        return "[pypdf 未安装，无法解析 PDF]"
    try:
        reader = _PdfReader(io.BytesIO(payload))
        parts = []
        for page in reader.pages[:50]:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[PDF 解析失败: {exc}]"


def _parse_docx(payload: bytes) -> str:
    if _DocxDocument is None:
        return "[python-docx 未安装，无法解析 DOCX]"
    try:
        doc = _DocxDocument(io.BytesIO(payload))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[DOCX 解析失败: {exc}]"


def _parse_doc(payload: bytes) -> str:
    if _olefile is None:
        return "[olefile 未安装，无法解析 DOC]"
    try:
        ole = _olefile.OleFileIO(io.BytesIO(payload))
        if ole.exists("WordDocument"):
            raw = ole.openstream("WordDocument").read()
            return _extract_printable(raw)[:MAX_TEXT_LENGTH]
        return "[无法提取 DOC 文本]"
    except Exception as exc:
        return f"[DOC 解析失败: {exc}]"


def _parse_xlsx(payload: bytes) -> str:
    if _load_workbook is None:
        return "[openpyxl 未安装，无法解析 XLSX]"
    try:
        wb = _load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        parts = []
        for ws in wb.worksheets[:5]:
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append("\t".join(cells))
        return "\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[XLSX 解析失败: {exc}]"


def _parse_xls(payload: bytes) -> str:
    if _xlrd is None:
        return "[xlrd 未安装，无法解析 XLS]"
    try:
        wb = _xlrd.open_workbook(file_contents=payload)
        parts = []
        for ws in wb.sheets()[:5]:
            for i in range(min(ws.nrows, 200)):
                cells = [str(ws.cell_value(i, j)) for j in range(ws.ncols)]
                if any(cells):
                    parts.append("\t".join(cells))
        return "\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[XLS 解析失败: {exc}]"


def _parse_image_ocr(payload: bytes) -> str:
    if _PILImage is None:
        return "[Pillow 未安装，无法解析图片]"
    if _pytesseract is None:
        return "[pytesseract 未安装，无法 OCR]"
    try:
        img = _PILImage.open(io.BytesIO(payload))
        return _pytesseract.image_to_string(img, lang="chi_sim+eng")[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[图片 OCR 失败: {exc}]"


def _parse_zip(payload: bytes, depth: int = 0) -> str:
    if depth >= MAX_ARCHIVE_DEPTH:
        return "[嵌套压缩层数超限]"
    try:
        parts = []
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            entries = [e for e in zf.infolist() if not e.is_dir()][:MAX_ARCHIVE_ENTRIES]
            for entry in entries:
                if entry.file_size > MAX_ARCHIVE_ENTRY_SIZE:
                    parts.append(f"[{entry.filename}: 文件过大，跳过]")
                    continue
                try:
                    data = zf.read(entry.filename)
                    ext = Path(entry.filename).suffix.lower()
                    text = parse_attachment_bytes(data, ext, depth=depth + 1)
                    if text:
                        parts.append(f"=== {entry.filename} ===\n{text}")
                except Exception as exc:
                    parts.append(f"[{entry.filename}: {exc}]")
        return "\n\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[ZIP 解析失败: {exc}]"


def _parse_7z(payload: bytes) -> str:
    if _py7zr is None:
        return "[py7zr 未安装，无法解析 7z]"
    try:
        parts = []
        with _py7zr.SevenZipFile(io.BytesIO(payload)) as sz:
            for fname, bio in (sz.read() or {}).items():
                if bio is None:
                    continue
                data = bio.read()
                ext = Path(fname).suffix.lower()
                text = parse_attachment_bytes(data, ext)
                if text:
                    parts.append(f"=== {fname} ===\n{text}")
        return "\n\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[7z 解析失败: {exc}]"


def _parse_rar(payload: bytes) -> str:
    if _rarfile is None:
        return "[rarfile 未安装，无法解析 RAR]"
    try:
        parts = []
        with _rarfile.RarFile(io.BytesIO(payload)) as rf:
            for fname in rf.namelist()[:MAX_ARCHIVE_ENTRIES]:
                try:
                    data = rf.read(fname)
                    ext = Path(fname).suffix.lower()
                    text = parse_attachment_bytes(data, ext)
                    if text:
                        parts.append(f"=== {fname} ===\n{text}")
                except Exception as exc:
                    parts.append(f"[{fname}: {exc}]")
        return "\n\n".join(parts)[:MAX_TEXT_LENGTH]
    except Exception as exc:
        return f"[RAR 解析失败: {exc}]"


# ── Main dispatcher ───────────────────────────────────────────────────────────

def parse_attachment_bytes(payload: bytes, file_ext: str, depth: int = 0) -> str:
    """Parse attachment bytes by extension, return extracted text."""
    if len(payload) > MAX_ATTACHMENT_SIZE:
        return f"[文件过大 ({len(payload) // 1024}KB)，跳过解析]"
    ext = file_ext.lower()
    try:
        if ext in TEXT_EXTENSIONS:
            return _parse_text(payload)
        if ext == ".json":
            return _parse_json(payload)
        if ext in {".csv", ".tsv"}:
            return _parse_csv_bytes(payload)
        if ext == ".xml":
            return _parse_xml(payload)
        if ext in {".html", ".htm"}:
            return _parse_html(payload)
        if ext == ".pdf":
            return _parse_pdf(payload)
        if ext == ".docx":
            return _parse_docx(payload)
        if ext == ".doc":
            return _parse_doc(payload)
        if ext == ".xlsx":
            return _parse_xlsx(payload)
        if ext == ".xls":
            return _parse_xls(payload)
        if ext == ".zip":
            return _parse_zip(payload, depth=depth)
        if ext == ".7z":
            return _parse_7z(payload)
        if ext == ".rar":
            return _parse_rar(payload)
        if ext in IMAGE_EXTENSIONS:
            return _parse_image_ocr(payload)
    except Exception as exc:
        logger.warning("[attachment] parse %s failed: %s", ext, exc)
        return f"[解析失败: {exc}]"
    return ""


def parse_and_store(mail_uid: str, attachment: dict[str, Any]) -> dict[str, Any]:
    """
    Parse an attachment dict (from mail parsing) and return result dict.
    attachment keys: filename, file_ext, storage_path, file_size, ...
    Returns dict with: parsed_text, summary, parse_status, parse_error, artifact_path
    """
    filename = attachment.get("filename", "")
    file_ext = attachment.get("file_ext", Path(filename).suffix.lower())
    storage_path = attachment.get("storage_path", "")
    result: dict[str, Any] = {
        "mail_uid": mail_uid, "filename": filename, "file_ext": file_ext,
        "parse_status": "skipped", "parsed_text": "", "summary": "",
        "artifact_path": "", "parse_error": None,
    }
    if not storage_path:
        result["parse_error"] = "no storage_path"
        return result
    path = Path(storage_path)
    if not path.exists():
        result["parse_error"] = f"file not found: {storage_path}"
        return result
    try:
        payload = path.read_bytes()
        text = parse_attachment_bytes(payload, file_ext)
        if text:
            ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)
            safe_stem = re.sub(r"[^\w\-]", "_", Path(filename).stem)[:48] or "attachment"
            artifact_path = ARTIFACTS_DIR / f"{safe_stem}_{uuid4().hex[:8]}.txt"
            artifact_path.write_text(text, encoding="utf-8")
            result.update({
                "parse_status": "completed", "parsed_text": text,
                "summary": _build_summary(text), "artifact_path": str(artifact_path),
            })
        else:
            result["parse_status"] = "empty"
    except Exception as exc:
        result["parse_status"] = "failed"
        result["parse_error"] = str(exc)
    return result


def get_available_parsers() -> dict[str, bool]:
    """Report which optional parsers are available."""
    return {
        "pdf (pypdf)": _PdfReader is not None,
        "docx (python-docx)": _DocxDocument is not None,
        "xlsx (openpyxl)": _load_workbook is not None,
        "xls (xlrd)": _xlrd is not None,
        "doc (olefile)": _olefile is not None,
        "image OCR (Pillow)": _PILImage is not None,
        "image OCR (pytesseract)": _pytesseract is not None,
        "7z (py7zr)": _py7zr is not None,
        "rar (rarfile)": _rarfile is not None,
    }
