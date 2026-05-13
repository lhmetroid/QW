from fastapi import FastAPI, Request, BackgroundTasks, Query, HTTPException, Depends, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse
from concurrent.futures import Future, ThreadPoolExecutor
import xml.etree.ElementTree as ET
import os
import logging
import json
import uuid
import re
import hashlib
import subprocess
import threading
import requests
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from time import perf_counter
from typing import Any
from urllib.parse import urlparse
from sqlalchemy import and_, or_, text, func
from sqlalchemy.orm import Session
from pydantic import BaseModel
from config import settings
from logging_config import setup_logging, sanitize_text
setup_logging()
from qywx_utils import QYWXUtils
from intent_engine import IntentEngine
from embedding_service import EmbeddingService
from email_import_service import EmailImportService
from business_csv_import import (
    DEFAULT_BUSINESS_CSV_FILENAME,
    DEFAULT_ROW_END as DEFAULT_BUSINESS_CSV_ROW_END,
    DEFAULT_ROW_START as DEFAULT_BUSINESS_CSV_ROW_START,
    DEFAULT_SOURCE_TYPE as BUSINESS_CSV_SOURCE_TYPE,
    run_business_csv_import,
)
from database import (
    SessionLocal, MessageLog, init_db, KnowledgeBase, IntentSummary, get_db,
    KnowledgeDocument, KnowledgeChunk, PricingRule, KnowledgeHitLog, KnowledgeCandidate,
    KnowledgeVersionSnapshot, JobTask, ThreadBusinessFact, EmailThreadAsset,
    EmailFragmentAsset, EmailEffectFeedback, ModelTrainingSample, ReplyChainSnapshot,
    ApiAssistInvocation, WecomTriggerRecord,
)
from worker import start_job
from knowledge_governance import (
    build_thread_business_fact,
    detect_mixed_knowledge,
    infer_function_fragment,
    infer_library_type,
    infer_scenario_intent,
    merge_tags,
    score_content_governance,
    validate_thread_state_consistency,
)

app = FastAPI(title="企微智能实时提醒后台")
logger = logging.getLogger(__name__)
_RUNTIME_ENV_CACHE: dict[str, Any] = {"mtime": None, "values": {}}
LLM_COMPARE_RUNTIME_STATUS: dict[str, dict[str, Any]] = {}
REPLY_CHAIN_EXECUTOR = ThreadPoolExecutor(max_workers=8, thread_name_prefix="reply-chain")
REPLY_CHAIN_ACTIVE_FUTURES: dict[str, Future] = {}
REPLY_CHAIN_ACTIVE_LOCK = threading.Lock()
SIDEBAR_ASSIST_RECENT_DEDUPE_TTL_SECONDS = 180
SIDEBAR_ASSIST_ACTIVE_FUTURES: dict[str, Future] = {}
SIDEBAR_ASSIST_ACTIVE_LOCK = threading.Lock()
SIDEBAR_ASSIST_RESULT_CACHE: dict[str, dict[str, Any]] = {}
SIDEBAR_ASSIST_CACHE_LOCK = threading.Lock()
POSITIVE_FEEDBACK_STATUSES = {"useful", "adopted", "won", "advanced"}
NEGATIVE_FEEDBACK_STATUSES = {"needs_fix", "rejected"}
TRAINING_SAMPLE_TYPES = {"embedding_corpus", "reply_fragment_sft", "thread_reply_sft", "retrieval_pair"}
LEGACY_SALES_KB_HOSTS = {"192.168.31.124"}

def _initial_cors_allow_origins() -> list[str]:
    origins = set()
    external_base_url = str(settings.EXTERNAL_API_BASE_URL or "").strip().rstrip("/")
    if external_base_url:
        origins.add(external_base_url)
    return sorted(origins)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_initial_cors_allow_origins(),
    allow_origin_regex=r"https?://(localhost|127\.0\.0\.1|\[::1\])(?::\d+)?",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def index_redirect():
    return RedirectResponse(url="/static/index.html")

@app.middleware("http")
async def request_observability_middleware(request: Request, call_next):
    request_id = request.headers.get("X-Request-ID") or uuid.uuid4().hex
    started = perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        elapsed_ms = round((perf_counter() - started) * 1000)
        logger.exception(
            "REQUEST_ERROR path=%s method=%s request_id=%s elapsed_ms=%s",
            request.url.path,
            request.method,
            request_id,
            elapsed_ms,
        )
        raise
    elapsed_ms = round((perf_counter() - started) * 1000)
    response.headers["X-Request-ID"] = request_id
    response.headers["X-Response-Time-Ms"] = str(elapsed_ms)
    if elapsed_ms >= settings.SLOW_REQUEST_MS:
        logger.warning(
            "SLOW_REQUEST path=%s method=%s status=%s request_id=%s elapsed_ms=%s",
            request.url.path,
            request.method,
            response.status_code,
            request_id,
            elapsed_ms,
        )
    return response


def _optional_embedding_for_storage(text: str, *, context: str) -> list[float] | None:
    try:
        embedding = EmbeddingService.embed(text)
    except Exception as exc:
        logger.warning("Embedding 调用失败，继续按无向量模式处理。 context=%s error=%s", context, sanitize_text(str(exc)))
        return None
    if not embedding:
        logger.warning("Embedding 未返回向量，继续按无向量模式处理。 context=%s", context)
        return None
    return embedding


def _embedding_metadata_fields(embedding: list[float] | None) -> dict[str, Any]:
    return {
        "embedding_provider": settings.EMBEDDING_PROVIDER if embedding else None,
        "embedding_model": settings.EMBEDDING_MODEL if embedding else None,
        "embedding_dim": len(embedding) if embedding else None,
    }

from callback import router as callback_router
app.include_router(callback_router)

from crm_profile import router as crm_profile_router
app.include_router(crm_profile_router)

# 挂载前端静态文件
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_path):
    app.mount("/static", StaticFiles(directory=frontend_path, html=True), name="static")

# 初始化数据库结构
def auto_patch_db():
    try:
        init_db() # 基础建表
        db = SessionLocal()
        # 强制补丁：为旧表增加 is_mock 列
        db.execute(text("ALTER TABLE message_logs ADD COLUMN IF NOT EXISTS is_mock BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE message_logs ADD COLUMN IF NOT EXISTS archive_msg_id VARCHAR(120);"))
        db.execute(text("ALTER TABLE message_logs ADD COLUMN IF NOT EXISTS archive_seq VARCHAR(40);"))
        db.execute(text("ALTER TABLE message_logs ALTER COLUMN user_id TYPE VARCHAR(120);"))
        db.execute(text("ALTER TABLE intent_summaries ALTER COLUMN user_id TYPE VARCHAR(120);"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS llm1_compare_summary JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS llm1_compare_prompt_trace JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS crm_info JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS crm_status VARCHAR(50);"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS thread_business_fact JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_log_id VARCHAR(120);"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_v2 JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_external_api JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_status VARCHAR(50);"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_confidence_score NUMERIC(8, 6);"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS knowledge_manual_review_required BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS sales_advice_compare_v2 TEXT;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS sales_advice_compare_prompt_trace_v2 JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS reply_style_results_v2 JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS reply_scores_v2 JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS assist_validation JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS assist_compare_validation JSON;"))
        db.execute(text("ALTER TABLE intent_summaries ADD COLUMN IF NOT EXISTS stage_status JSON;"))
        db.execute(text("ALTER TABLE knowledge_document ADD COLUMN IF NOT EXISTS library_type VARCHAR(20) DEFAULT 'reference';"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS library_type VARCHAR(20) DEFAULT 'reference';"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS allowed_for_generation BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS usable_for_reply BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS publishable BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS topic_clarity_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS completeness_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS reusability_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS evidence_reliability_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS useful_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS effect_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS feedback_count INTEGER DEFAULT 0;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS positive_feedback_count INTEGER DEFAULT 0;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS last_feedback_at TIMESTAMP;"))
        db.execute(text("ALTER TABLE knowledge_chunk ADD COLUMN IF NOT EXISTS quality_notes JSON;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS library_type VARCHAR(20) DEFAULT 'reference';"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS allowed_for_generation BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS usable_for_reply BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS publishable BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS topic_clarity_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS completeness_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS reusability_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS evidence_reliability_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS useful_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS effect_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS feedback_count INTEGER DEFAULT 0;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS positive_feedback_count INTEGER DEFAULT 0;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS last_feedback_at TIMESTAMP;"))
        db.execute(text("ALTER TABLE knowledge_candidate ADD COLUMN IF NOT EXISTS quality_notes JSON;"))
        db.execute(text("ALTER TABLE thread_business_fact ADD COLUMN IF NOT EXISTS effect_score NUMERIC(6, 4);"))
        db.execute(text("ALTER TABLE thread_business_fact ADD COLUMN IF NOT EXISTS outcome_feedback JSON;"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS retrieval_quality VARCHAR(50);"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS confidence_score NUMERIC(8, 6);"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS insufficient_info BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS manual_review_required BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS final_response TEXT;"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS manual_feedback JSON;"))
        db.execute(text("ALTER TABLE knowledge_hit_logs ADD COLUMN IF NOT EXISTS feedback_status VARCHAR(50);"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS visible_message_ids JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS latest_dialog_count INTEGER DEFAULT 0;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS fast_track JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS crm_info JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS crm_status VARCHAR(50);"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS thread_business_fact JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_log_id VARCHAR(120);"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_v2 JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_external_api JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_status VARCHAR(50);"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_confidence_score NUMERIC(8, 6);"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS knowledge_manual_review_required BOOLEAN DEFAULT FALSE;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS llm1_compare_summary JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS llm1_compare_prompt_trace JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS reply_style_results_v2 JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS reply_scores_v2 JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS assist_validation JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS assist_compare_validation JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS actual_sales_replies JSON;"))
        db.execute(text("ALTER TABLE reply_chain_snapshot ADD COLUMN IF NOT EXISTS stage_status JSON;"))
        db.execute(text(
            "CREATE TABLE IF NOT EXISTS api_assist_invocation ("
            "invocation_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),"
            "session_id VARCHAR(120) NOT NULL,"
            "requested_session_id VARCHAR(120),"
            "external_userid VARCHAR(120),"
            "sales_userid VARCHAR(120),"
            "anchor_message_id INTEGER,"
            "anchor_message_time TIMESTAMP,"
            "anchor_message_text TEXT,"
            "visible_message_ids JSON,"
            "latest_dialog_count INTEGER DEFAULT 0,"
            "trigger_source VARCHAR(30),"
            "trigger_kind VARCHAR(50),"
            "triggered_at TIMESTAMP DEFAULT now(),"
            "stage_status JSON,"
            "result_payload JSON NOT NULL,"
            "actual_sales_replies JSON,"
            "actual_sales_reply_text TEXT,"
            "actual_sales_reply_hash VARCHAR(64),"
            "quality_similarity JSON,"
            "quality_score NUMERIC(6, 2),"
            "quality_status VARCHAR(50),"
            "quality_scored_at TIMESTAMP,"
            "quality_annotations JSON,"
            "created_at TIMESTAMP DEFAULT now(),"
            "updated_at TIMESTAMP DEFAULT now()"
            ");"
        ))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS trigger_source VARCHAR(30);"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS trigger_kind VARCHAR(50);"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS kb1_eval_score NUMERIC(6, 2);"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS kb1_eval_reason TEXT;"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS kb2_eval_score NUMERIC(6, 2);"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS kb2_eval_reason TEXT;"))
        db.execute(text("ALTER TABLE api_assist_invocation ADD COLUMN IF NOT EXISTS quality_annotations JSON;"))
        db.execute(text("ALTER TABLE wecom_trigger_record ADD COLUMN IF NOT EXISTS kb1_eval_score NUMERIC(6, 2);"))
        db.execute(text("ALTER TABLE wecom_trigger_record ADD COLUMN IF NOT EXISTS kb1_eval_reason TEXT;"))
        db.execute(text("ALTER TABLE wecom_trigger_record ADD COLUMN IF NOT EXISTS kb2_eval_score NUMERIC(6, 2);"))
        db.execute(text("ALTER TABLE wecom_trigger_record ADD COLUMN IF NOT EXISTS kb2_eval_reason TEXT;"))
        db.execute(text("ALTER TABLE wecom_trigger_record ADD COLUMN IF NOT EXISTS quality_annotations JSON;"))
        db.execute(text(
            "CREATE TABLE IF NOT EXISTS wecom_trigger_record ("
            "record_id UUID PRIMARY KEY DEFAULT gen_random_uuid(),"
            "session_id VARCHAR(120) NOT NULL,"
            "snapshot_id VARCHAR(120),"
            "run_id VARCHAR(40),"
            "trigger_source VARCHAR(30) NOT NULL DEFAULT 'web_manual',"
            "trigger_kind VARCHAR(50) NOT NULL,"
            "requested_step INTEGER,"
            "requested_channel VARCHAR(50),"
            "request_status VARCHAR(30) NOT NULL DEFAULT 'queued',"
            "anchor_message_id INTEGER,"
            "anchor_message_time TIMESTAMP,"
            "anchor_message_text TEXT,"
            "visible_message_ids JSON,"
            "input_messages JSON,"
            "recent_customer_messages JSON,"
            "latest_dialog_count INTEGER DEFAULT 0,"
            "actual_sales_replies JSON,"
            "actual_sales_reply_text TEXT,"
            "stage_status JSON,"
            "result_payload JSON,"
            "request_payload JSON,"
            "error_message TEXT,"
            "quality_annotations JSON,"
            "triggered_at TIMESTAMP DEFAULT now(),"
            "finished_at TIMESTAMP,"
            "created_at TIMESTAMP DEFAULT now(),"
            "updated_at TIMESTAMP DEFAULT now()"
            ");"
        ))
        db.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_reply_chain_snapshot_anchor ON reply_chain_snapshot (session_id, anchor_message_id);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_reply_chain_snapshot_session_updated ON reply_chain_snapshot (session_id, updated_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_api_ai_session_triggered ON api_assist_invocation (session_id, triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_api_ai_anchor_triggered ON api_assist_invocation (session_id, anchor_message_id, triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_api_ai_quality_status ON api_assist_invocation (quality_status, triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_wtr_triggered ON wecom_trigger_record (triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_wtr_source_triggered ON wecom_trigger_record (trigger_source, triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_wtr_session_triggered ON wecom_trigger_record (session_id, triggered_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_wtr_run_id ON wecom_trigger_record (run_id);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_message_logs_archive_msg_id ON message_logs (archive_msg_id);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_message_logs_archive_seq ON message_logs (archive_seq);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kc_active_effective ON knowledge_chunk (status, business_line, knowledge_type, effective_from, effective_to);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kc_replyable ON knowledge_chunk (status, usable_for_reply, allowed_for_generation, useful_score DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kc_priority ON knowledge_chunk (priority DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_pr_active_scope ON pricing_rule (status, business_line, language_pair, service_scope, customer_tier);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kd_effective_to ON knowledge_document (status, effective_to);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kcand_status_created ON knowledge_candidate (status, created_at DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kcand_replyable ON knowledge_candidate (status, usable_for_reply, useful_score DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kc_effect_score ON knowledge_chunk (status, effect_score DESC, useful_score DESC);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_kcand_source ON knowledge_candidate (source_type, source_ref);"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_khl_status_created ON knowledge_hit_logs (status, created_at DESC);"))
        if settings.KB_FULLTEXT_INDEX_ENABLED:
            db.execute(text(
                "CREATE INDEX IF NOT EXISTS idx_kc_fulltext_simple ON knowledge_chunk "
                "USING GIN (to_tsvector('simple', coalesce(title,'') || ' ' || coalesce(keyword_text,'') || ' ' || coalesce(content,'')));"
            ))
        if settings.PGVECTOR_ENABLED:
            db.execute(text("CREATE EXTENSION IF NOT EXISTS vector;"))
        db.commit()
        db.close()
        logger.info("数据库 Schema 自动检查/修复完成")
    except Exception as e:
        logger.error(f"数据库补丁执行异常 (可能由于已存在): {e}")

auto_patch_db()

import asyncio
from archive_service import ArchiveService

@app.on_event("startup")
async def startup_event():
    archive_status = ArchiveService.config_status()
    if not settings.ENABLE_ARCHIVE_POLLING:
        logger.info("企微会话存档自动轮询未启用；知识库后台正常启动。")
        return
    if not archive_status["ready"]:
        logger.warning("企微会话存档自动轮询已请求但配置不完整，跳过启动: %s", archive_status["missing"])
        return

    from kb_evaluation_worker import start_kb_evaluation_thread
    start_kb_evaluation_thread()

    logger.info("启动企微会话存档 15 分钟异步轮询任务")
    async def background_poll_worker():
        while True:
            try:
                # 使用 to_thread 将底层 C-SDK 同步阻塞(10秒超时网络I/O) 推到独立线程池，绝不卡死 FastAPI 主服务
                result = await asyncio.to_thread(ArchiveService.sync_today_data)
                if result.get("status") != "success":
                    logger.warning("企微会话存档轮询未成功: %s", result.get("msg"))
            except Exception as e:
                logger.error(f"企微会话存档自动轮询异常: {e}")
            await asyncio.sleep(900)
    
    asyncio.create_task(background_poll_worker())

@app.get("/cb/qywx")
async def qywx_verify(
    msg_signature: str = Query(...),
    timestamp: str = Query(...),
    nonce: str = Query(...),
    echostr: str = Query(...)
):
    """企微回调开启时的 URL 验证接口"""
    logger.info(f"收到企微验证请求")
    res = QYWXUtils.decrypt_message(msg_signature, timestamp, nonce, echostr, is_verify=True)
    return res

@app.post("/cb/qywx")
async def qywx_receive_message(
    request: Request,
    background_tasks: BackgroundTasks,
    msg_signature: str = Query(...),
    timestamp: str = Query(...),
    nonce: str = Query(...)
):
    """企微回调消息接收接口"""
    body = await request.body()
    xml_content = QYWXUtils.decrypt_message(msg_signature, timestamp, nonce, body, is_verify=False)
    if not xml_content:
        raise HTTPException(status_code=400, detail="Decrypt failed")
    
    try:
        root = ET.fromstring(xml_content)
        from_user = (root.findtext("FromUserName") or "").strip()
        content = (
            root.findtext("Content")
            or root.findtext("Recognition")
            or root.findtext("Title")
            or ""
        ).strip()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"企微消息 XML 解析失败: {exc}") from exc

    if not from_user or not content:
        raise HTTPException(status_code=400, detail="企微回调缺少 FromUserName 或文本内容")
    
    # 1. 实时旁路扫描 (主线程)
    IntentEngine.fast_track_scan(from_user, content)
    
    # 2. 异步深度分析 (改用原生 BackgroundTasks 以适配 Windows)
    background_tasks.add_task(process_deep_analyze_task, from_user, content, "customer")
    
    return "success"

def process_deep_analyze_task(user_id, content, sender_type):
    """后台长对话分析任务"""
    db = SessionLocal()
    try:
        # 消息持久化
        log = MessageLog(user_id=user_id, content=content, sender_type=sender_type)
        db.add(log)
        db.commit()
        
        # 获取上下文
        context_logs = db.query(MessageLog).filter(MessageLog.user_id == user_id).order_by(MessageLog.id.desc()).limit(10).all()
        context = [{"content": l.content, "sender_type": l.sender_type} for l in reversed(context_logs)]
        
        # 执行 AI 分析链路
        IntentEngine.slow_track_analyze(user_id, context)
    finally:
        db.close()

# --- 知识库管理 API ---

@app.post("/api/knowledge")
async def add_knowledge(title: str, content: str, category: str = "通用"):
    embedding = IntentEngine.get_embedding(content)

    db = SessionLocal()
    try:
        new_item = KnowledgeBase(title=title, content=content, category=category, embedding=embedding)
        db.add(new_item)
        db.commit()
        return {"status": "success", "id": new_item.id}
    finally:
        db.close()

@app.get("/api/knowledge")
async def list_knowledge():
    db = SessionLocal()
    try:
        items = db.query(KnowledgeBase.id, KnowledgeBase.title, KnowledgeBase.category).all()
        return [{"id": i.id, "title": i.title, "category": i.category} for i in items]
    finally:
        db.close()

# --- 知识库 V2 治理 API：仅负责知识入库、审核状态和证据管理 ---

class KnowledgeManualCreate(BaseModel):
    title: str
    content: str
    knowledge_class: str | None = None
    knowledge_type: str = "faq"
    business_line: str = "general"
    sub_service: str | None = None
    chunk_type: str = "rule"
    language_pair: str | None = None
    service_scope: str | None = None
    region: str | None = None
    customer_tier: str | None = None
    source_type: str = "manual"
    source_ref: str | None = None
    owner: str | None = None
    priority: int = 50
    risk_level: str | None = None
    review_required: bool = True
    tags: dict | None = None
    effective_from: datetime | None = None
    effective_to: datetime | None = None
    pricing_rule: dict | None = None

class KnowledgeRetrieveRequest(BaseModel):
    query_text: str
    query_features: dict | None = None
    top_k: int = 5
    request_id: str | None = None
    session_id: str | None = None

class KnowledgeChunkCreate(BaseModel):
    title: str
    content: str
    knowledge_class: str | None = None
    knowledge_type: str = "faq"
    chunk_type: str = "rule"
    business_line: str = "general"
    sub_service: str | None = None
    language_pair: str | None = None
    service_scope: str | None = None
    region: str | None = None
    customer_tier: str | None = None
    priority: int = 50
    tags: dict | None = None
    pricing_rule: dict | None = None

class KnowledgeMultiChunkCreate(BaseModel):
    title: str
    knowledge_class: str | None = None
    knowledge_type: str = "faq"
    business_line: str = "general"
    sub_service: str | None = None
    source_type: str = "manual"
    source_ref: str | None = None
    owner: str | None = None
    risk_level: str | None = None
    review_required: bool = True
    tags: dict | None = None
    effective_from: datetime | None = None
    effective_to: datetime | None = None
    chunks: list[KnowledgeChunkCreate]

class KnowledgeDocumentUpdate(BaseModel):
    title: str | None = None
    knowledge_class: str | None = None
    knowledge_type: str | None = None
    business_line: str | None = None
    sub_service: str | None = None
    risk_level: str | None = None
    review_required: bool | None = None
    tags: dict | None = None
    effective_from: datetime | None = None
    effective_to: datetime | None = None

class KnowledgeChunkUpdate(BaseModel):
    title: str | None = None
    content: str | None = None
    knowledge_class: str | None = None
    chunk_type: str | None = None
    knowledge_type: str | None = None
    business_line: str | None = None
    sub_service: str | None = None
    language_pair: str | None = None
    service_scope: str | None = None
    region: str | None = None
    customer_tier: str | None = None
    priority: int | None = None
    structured_tags: dict | None = None

class KnowledgeAssistTextImport(BaseModel):
    title: str
    content: str
    business_line: str = "general"
    source_type: str = "manual_text"
    source_ref: str | None = None
    owner: str | None = None

class KnowledgeBulkAction(BaseModel):
    document_ids: list[str]
    force: bool = False
    force_reason: str | None = None
    operator: str | None = None

class KnowledgeHitFeedback(BaseModel):
    feedback_status: str
    manual_feedback: dict | None = None

class AssistFeedback(BaseModel):
    session_id: str
    feedback_status: str
    manual_feedback: dict | None = None
    final_response: str | None = None
    snapshot_id: str | None = None


class TriggerAnalyticsQualityAnnotationRequest(BaseModel):
    row_source: str
    row_id: str
    score_key: str
    labels: list[str] | None = None
    custom_text: str | None = None

class ExcellentReplyExtractRequest(BaseModel):
    session_id: str
    owner: str | None = None
    operator: str | None = None
    force: bool = False

class TrainingSamplePrepareRequest(BaseModel):
    sample_types: list[str] | None = None
    min_quality_score: float = 0.65
    min_effect_score: float = 0.55
    limit_per_source: int = 200
    operator: str | None = None

class TrainingSampleExportRequest(BaseModel):
    sample_types: list[str] | None = None
    max_samples: int = 1000
    operator: str | None = None


class TrainingExecutionRequest(BaseModel):
    sample_types: list[str] | None = None
    min_quality_score: float = 0.45
    min_effect_score: float = 0.45
    limit_per_source: int = 200
    max_samples: int = 1000
    execute_runner: bool = True
    runner_command: str | None = None
    runner_workdir: str | None = None
    runner_timeout_seconds: int | None = None
    operator: str | None = None


class AnalysisCompletionRequest(BaseModel):
    source_type: str = "email_excel"
    rebuild_candidate_governance: bool = True
    backfill_email_assets: bool = True
    backfill_thread_facts: bool = True
    bootstrap_positive_feedback_count: int = 3
    bootstrap_feedback_status: str = "adopted"
    extract_excellent_replies: bool = True
    prepare_training_samples: bool = True
    export_training_samples: bool = True
    sample_types: list[str] | None = None
    min_quality_score: float = 0.45
    min_effect_score: float = 0.45
    limit_per_source: int = 200
    max_samples: int = 500
    operator: str | None = None

class KnowledgeRegressionRunRequest(BaseModel):
    case_ids: list[str] | None = None
    groups: list[str] | None = None
    category: str | None = None
    risk_level: str | None = None
    business_line: str | None = None
    top_k: int = 5
    min_score: float | None = None
    cleanup_logs: bool = True
    include_hits: bool = False
    run_id: str | None = None

class JobRetryRequest(BaseModel):
    operator: str | None = None

class KnowledgePublishRequest(BaseModel):
    force: bool = False
    force_reason: str | None = None
    operator: str | None = None

class KnowledgeCandidateUpdate(BaseModel):
    title: str | None = None
    content: str | None = None
    knowledge_type: str | None = None
    chunk_type: str | None = None
    business_line: str | None = None
    sub_service: str | None = None
    language_pair: str | None = None
    service_scope: str | None = None
    region: str | None = None
    customer_tier: str | None = None
    priority: int | None = None
    risk_level: str | None = None
    effective_from: datetime | None = None
    effective_to: datetime | None = None
    pricing_rule: dict | None = None
    owner: str | None = None
    operator: str | None = None
    review_notes: str | None = None
    status: str | None = None

class KnowledgeCandidateFromFeedbackRequest(BaseModel):
    log_id: str
    owner: str | None = None
    operator: str | None = None
    note: str | None = None
    title: str | None = None

class KnowledgeCandidatePromoteRequest(BaseModel):
    owner: str | None = None
    operator: str | None = None
    title: str | None = None
    content: str | None = None
    knowledge_type: str | None = None
    chunk_type: str | None = None
    business_line: str | None = None
    sub_service: str | None = None
    language_pair: str | None = None
    service_scope: str | None = None
    region: str | None = None
    customer_tier: str | None = None
    priority: int | None = None
    risk_level: str | None = None
    effective_from: datetime | None = None
    effective_to: datetime | None = None
    pricing_rule: dict | None = None
    review_notes: str | None = None

class KnowledgeCandidateBatchPromoteRequest(BaseModel):
    candidate_ids: list[str]
    owner: str | None = None
    operator: str | None = None

class KnowledgeExtractFromMessagesRequest(BaseModel):
    session_id: str | None = None
    messages: list[str] | None = None
    title: str | None = None
    source_type: str = "message_extract"
    owner: str | None = None
    operator: str | None = None
    max_messages: int = 50
    max_candidates: int = 20

class PricingRulePayload(BaseModel):
    business_line: str | None = None
    sub_service: str | None = None
    language_pair: str | None = None
    service_scope: str | None = None
    unit: str = "per_1000_chars"
    currency: str = "CNY"
    price_min: float | str | None = None
    price_max: float | str | None = None
    urgent_multiplier: float | str | None = None
    tax_policy: str | None = None
    min_charge: float | str | None = None
    customer_tier: str | None = None
    region: str | None = None
    source_ref: str | None = None

KB_LABELS = {
    "business_line": {
        "translation": "翻译",
        "printing": "印刷",
        "interpretation": "同传",
        "multimedia": "多媒体译制",
        "exhibition": "展台搭建",
        "gifts": "礼品",
        "general": "通用",
    },
    "sub_service": {},
    "knowledge_type": {
        "capability": "能力知识",
        "pricing": "报价知识",
        "process": "流程知识",
        "faq": "常见问答",
    },
    "status": {
        "draft": "草稿",
        "review": "审核中",
        "active": "已发布",
        "archived": "已归档",
        "rejected": "已驳回",
    },
    "document_stage": {
        "draft": "草稿",
        "review": "审核中",
        "approved": "审核同意",
        "active": "已发布",
        "archived": "已归档",
        "rejected": "已驳回",
    },
    "candidate_status": {
        "candidate": "候选中",
        "promoted": "已转为草稿知识",
        "rejected": "已拒绝",
        "archived": "已归档",
    },
    "candidate_source_type": {
        "feedback": "销售反馈",
        "message_extract": "会话抽取",
        "case_extract": "历史案例",
        "email_excel": "邮件整理表",
        "excellent_reply": "优秀回复抽取",
    },
    "review_status": {
        "pending": "待审核",
        "in_review": "审核中",
        "approved": "审核通过",
        "rejected": "审核驳回",
        "rolled_back": "已回滚",
        "auto_ready": "自动就绪",
    },
    "risk_level": {
        "low": "低",
        "medium": "中",
        "high": "高",
    },
    "language_pair": {
        "en->fr": "英译法",
        "en->zh": "英译中",
        "zh->en": "中译英",
        "en->ru": "英译俄",
        "ja->zh": "日译中",
        "ko->zh": "韩译中",
        "zh->ja": "中译日",
        "fr->zh": "法译中",
        "zh->fr": "中译法",
        "de->zh": "德译中",
        "zh->de": "中译德",
        "ru->zh": "俄译中",
        "zh->ru": "中译俄",
        "zh->ko": "中译韩",
        "it->zh": "意译中",
        "zh->it": "中译意",
        "es->zh": "西译中",
        "zh->es": "中译西",
        "ar->zh": "阿译中",
        "zh->ar": "中译阿",
        "da->zh": "丹麦语译中",
        "zh->da": "中译丹麦语",
        "pt->zh": "葡译中",
        "zh->pt": "中译葡",
        "nl->zh": "荷译中",
        "zh->nl": "中译荷",
        "sv->zh": "瑞典语译中",
        "zh->sv": "中译瑞典语",
        "no->zh": "挪威语译中",
        "zh->no": "中译挪威语",
        "el->zh": "希腊语译中",
        "zh->el": "中译希腊语",
        "tr->zh": "土耳其语译中",
        "zh->tr": "中译土耳其语",
        "fr->en": "法译英",
        "de->en": "德译英",
        "en->de": "英译德",
        "en->ja": "英译日",
        "ja->en": "日译英",
        "en->en": "英文润稿",
    },
    "service_scope": {
        "general": "普通资料",
        "medical": "医学资料",
        "legal": "法律资料",
        "technical": "技术资料",
        "financial": "财务/金融资料",
        "marketing": "市场宣传资料",
        "certified": "认证/盖章文件",
        "confidential": "保密资料",
        "literary_marketing": "菜单/品牌/文学类资料",
        "presentation": "幻灯片资料",
        "native_polishing": "外籍母语润稿",
        "native_translation": "外籍母语翻译",
        "rare_language": "欧洲稀有语种资料",
        "bilingual_foreign": "双外语翻译",
        "formatting": "输入或排版服务",
    },
    "customer_tier": {
        "common": "普通客户",
        "key": "重点客户",
        "strategic": "战略客户",
        "vip": "VIP客户",
    },
    "chunk_type": {
        "rule": "规则",
        "faq": "问答",
        "example": "案例",
        "template": "话术模板",
        "constraint": "限制条件",
        "definition": "定义",
    },
    "knowledge_class": {
        "pricing_constraint": "报价限制条件",
        "capability": "能力知识",
        "process": "流程规则",
        "faq": "FAQ常见问答",
        "example": "案例",
        "script": "话术模板",
        "email_template": "邮件模板",
        "definition": "名词定义",
    },
    "intent_type": {
        "pricing": "报价咨询",
        "capability": "能力咨询",
        "process": "流程咨询",
        "faq": "常见问题",
    },
}

KB_PRICING_UNIT_LABELS = {
    "per_1000_chars": "每千中文字符",
    "per_project": "每个项目",
    "per_hour": "每小时",
    "per_day": "每天",
    "per_slide": "每页幻灯片",
    "per_english_word": "每个英文单词",
    "per_source_word": "每个原文单词",
    "per_japanese_char": "每个日文字符",
    "per_a4_original": "每页 A4 原稿",
}

def label_for(dict_type: str, code: str | None) -> str | None:
    if code is None:
        return None
    return KB_LABELS.get(dict_type, {}).get(code, code)

def document_stage_code(doc: "KnowledgeDocument") -> str:
    if doc.status == "archived":
        return "archived"
    if doc.status == "rejected" or doc.review_status == "rejected":
        return "rejected"
    if doc.status == "active":
        return "active"
    if doc.status == "review" and doc.review_status == "approved":
        return "approved"
    if doc.status == "review":
        return "review"
    return "draft"

def document_stage_label(doc: "KnowledgeDocument") -> str:
    return label_for("document_stage", document_stage_code(doc)) or "-"

def document_allowed_actions(doc: "KnowledgeDocument") -> list[str]:
    stage = document_stage_code(doc)
    if stage == "draft":
        return ["edit", "submit_review", "archive", "reject"]
    if stage == "review":
        return ["edit", "approve", "reject", "archive"]
    if stage == "approved":
        return ["edit", "publish", "reject", "archive"]
    if stage == "active":
        return ["archive"]
    if stage == "rejected":
        return ["archive"]
    return []

KB_ACTION_LABELS = {
    "edit": "编辑文档",
    "submit_review": "提交审核",
    "approve": "审核同意",
    "publish": "发布",
    "archive": "归档",
    "reject": "驳回",
}

def _ensure_document_action_allowed(doc: "KnowledgeDocument", action: str) -> None:
    if action in document_allowed_actions(doc):
        return
    stage_label = document_stage_label(doc)
    action_label = KB_ACTION_LABELS.get(action, action)
    raise HTTPException(
        status_code=409,
        detail={
            "message": f"当前状态为{stage_label}，不允许执行{action_label}",
            "document_id": str(doc.document_id),
            "document_stage": document_stage_code(doc),
            "allowed_actions": document_allowed_actions(doc),
        },
    )

def _ensure_documents_action_allowed(docs: list["KnowledgeDocument"], action: str) -> None:
    for doc in docs:
        _ensure_document_action_allowed(doc, action)

def normalize_code(dict_type: str, value):
    if value is None:
        return None, None
    code = str(value).strip()
    if not code:
        return None, None
    if code in KB_LABELS.get(dict_type, {}):
        return code, None
    for item_code, label in KB_LABELS.get(dict_type, {}).items():
        if code == label:
            return item_code, None
    return None, code

KB_KNOWLEDGE_CLASS_CONFIG = {
    "pricing_constraint": {"knowledge_type": "pricing", "chunk_type": "constraint", "risk_level": "high"},
    "capability": {"knowledge_type": "capability", "chunk_type": "rule", "risk_level": "medium"},
    "process": {"knowledge_type": "process", "chunk_type": "rule", "risk_level": "medium"},
    "faq": {"knowledge_type": "faq", "chunk_type": "faq", "risk_level": "medium"},
    "example": {"knowledge_type": "faq", "chunk_type": "example", "risk_level": "medium"},
    "script": {"knowledge_type": "faq", "chunk_type": "template", "risk_level": "medium"},
    "email_template": {"knowledge_type": "faq", "chunk_type": "template", "risk_level": "medium"},
    "definition": {"knowledge_type": "faq", "chunk_type": "definition", "risk_level": "low"},
}

def normalize_knowledge_class(value):
    return normalize_code("knowledge_class", value)

def normalize_knowledge_class_or_raise(value: str | None, field_name: str = "knowledge_class") -> str | None:
    code, raw_value = normalize_knowledge_class(value)
    if raw_value:
        options = "、".join(KB_LABELS["knowledge_class"].values())
        raise HTTPException(status_code=400, detail=f"{field_name} 不合法，必须是以下之一：{options}")
    return code

def knowledge_class_from_pair(knowledge_type: str | None, chunk_type: str | None) -> str | None:
    for class_code, config in KB_KNOWLEDGE_CLASS_CONFIG.items():
        if config["knowledge_type"] == knowledge_type and config["chunk_type"] == chunk_type:
            return class_code
    if knowledge_type == "pricing":
        return "pricing_constraint" if chunk_type == "constraint" else None
    if knowledge_type == "capability":
        return "capability"
    if knowledge_type == "process":
        return "process"
    if chunk_type == "example":
        return "example"
    if chunk_type == "template":
        return "email_template"
    if chunk_type == "definition":
        return "definition"
    if knowledge_type == "faq":
        return "faq"
    return None

def knowledge_class_label(knowledge_type: str | None, chunk_type: str | None, tags: dict | None = None) -> str | None:
    class_code = (tags or {}).get("knowledge_class") or knowledge_class_from_pair(knowledge_type, chunk_type)
    return label_for("knowledge_class", class_code)

def merge_knowledge_class_tags(tags: dict | None, knowledge_class: str | None) -> dict | None:
    merged = dict(tags or {})
    if knowledge_class:
        merged["knowledge_class"] = knowledge_class
    return merged or None

def resolve_knowledge_class_fields(
    *,
    knowledge_class: str | None,
    knowledge_type: str | None,
    chunk_type: str | None,
    risk_level: str | None = None,
):
    class_code = knowledge_class or knowledge_class_from_pair(knowledge_type, chunk_type) or "faq"
    config = KB_KNOWLEDGE_CLASS_CONFIG.get(class_code, KB_KNOWLEDGE_CLASS_CONFIG["faq"])
    resolved_type = config["knowledge_type"] if knowledge_class else (knowledge_type or config["knowledge_type"])
    resolved_chunk_type = config["chunk_type"] if knowledge_class else (chunk_type or config["chunk_type"])
    resolved_risk = risk_level or config["risk_level"]
    return class_code, resolved_type, resolved_chunk_type, resolved_risk

def _decimal_score(value: float | None) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(round(float(value), 4)))

def _governance_metadata(
    *,
    business_line: str | None = None,
    language_pair: str | None = None,
    service_scope: str | None = None,
    customer_tier: str | None = None,
) -> dict:
    return {
        "business_line": business_line,
        "language_pair": language_pair,
        "service_scope": service_scope,
        "customer_tier": customer_tier,
    }

def _governance_tags(
    tags: dict | None,
    *,
    source_type: str | None,
    title: str | None,
    content: str | None,
) -> dict | None:
    tags = dict(tags or {})
    fragment = infer_function_fragment(title=title, content=content, source_type=source_type, tags=tags)
    scenario_label, intent_label, language_style = infer_scenario_intent(title=title, content=content, tags=tags)
    return merge_tags(
        tags,
        function_fragment=fragment,
        scenario_label=scenario_label,
        intent_label=intent_label,
        language_style=language_style,
        thread_id=tags.get("thread_id") or tags.get("session_id"),
    )

def _chunk_quality_payload(
    *,
    title: str,
    content: str,
    knowledge_type: str,
    chunk_type: str,
    source_type: str | None,
    tags: dict | None,
    pricing_rule: dict | None,
    source_ref: str | None = None,
    metadata: dict | None = None,
) -> dict:
    merged_tags = _governance_tags(tags, source_type=source_type, title=title, content=content)
    return score_content_governance(
        title=title,
        content=content,
        knowledge_type=knowledge_type,
        chunk_type=chunk_type,
        source_type=source_type,
        tags=merged_tags,
        pricing_rule=pricing_rule,
        has_source_ref=bool(source_ref),
        metadata=metadata,
    ) | {"structured_tags": merged_tags}

def _apply_chunk_governance(
    chunk: KnowledgeChunk,
    *,
    source_type: str | None,
    pricing_rule: dict | None = None,
    source_ref: str | None = None,
) -> dict:
    quality = _chunk_quality_payload(
        title=chunk.title,
        content=chunk.content,
        knowledge_type=chunk.knowledge_type,
        chunk_type=chunk.chunk_type,
        source_type=source_type,
        tags=chunk.structured_tags,
        pricing_rule=pricing_rule,
        source_ref=source_ref,
        metadata=_governance_metadata(
            business_line=chunk.business_line,
            language_pair=chunk.language_pair,
            service_scope=chunk.service_scope,
            customer_tier=chunk.customer_tier,
        ),
    )
    chunk.structured_tags = quality["structured_tags"]
    chunk.library_type = quality["library_type"]
    chunk.allowed_for_generation = bool(quality["allowed_for_generation"])
    chunk.usable_for_reply = bool(quality["usable_for_reply"])
    chunk.publishable = bool(quality["publishable"])
    chunk.topic_clarity_score = _decimal_score(quality["topic_clarity_score"])
    chunk.completeness_score = _decimal_score(quality["completeness_score"])
    chunk.reusability_score = _decimal_score(quality["reusability_score"])
    chunk.evidence_reliability_score = _decimal_score(quality["evidence_reliability_score"])
    chunk.useful_score = _decimal_score(quality["useful_score"])
    chunk.quality_notes = quality["quality_notes"]
    return quality

def _apply_candidate_governance(candidate: KnowledgeCandidate) -> dict:
    source_snapshot = dict(candidate.source_snapshot or {})
    tags = dict(source_snapshot.get("tags") or {})
    quality = _chunk_quality_payload(
        title=candidate.title,
        content=candidate.content,
        knowledge_type=candidate.knowledge_type,
        chunk_type=candidate.chunk_type,
        source_type=candidate.source_type,
        tags=tags,
        pricing_rule=candidate.pricing_rule,
        source_ref=candidate.source_ref,
        metadata=_governance_metadata(
            business_line=candidate.business_line,
            language_pair=candidate.language_pair,
            service_scope=candidate.service_scope,
            customer_tier=candidate.customer_tier,
        ),
    )
    source_snapshot["tags"] = quality["structured_tags"]
    source_snapshot["quality_notes"] = quality["quality_notes"]
    source_snapshot["mixed_knowledge"] = detect_mixed_knowledge(candidate.content, quality["structured_tags"])
    candidate.source_snapshot = source_snapshot
    candidate.library_type = quality["library_type"]
    candidate.allowed_for_generation = bool(quality["allowed_for_generation"])
    candidate.usable_for_reply = bool(quality["usable_for_reply"])
    candidate.publishable = bool(quality["publishable"])
    candidate.topic_clarity_score = _decimal_score(quality["topic_clarity_score"])
    candidate.completeness_score = _decimal_score(quality["completeness_score"])
    candidate.reusability_score = _decimal_score(quality["reusability_score"])
    candidate.evidence_reliability_score = _decimal_score(quality["evidence_reliability_score"])
    candidate.useful_score = _decimal_score(quality["useful_score"])
    candidate.quality_notes = quality["quality_notes"]
    return quality

def _thread_fact_to_dict(item: ThreadBusinessFact | None) -> dict | None:
    if not item:
        return None
    return {
        "fact_id": str(item.fact_id),
        "session_id": item.session_id,
        "thread_id": item.thread_id,
        "external_userid": item.external_userid,
        "sales_userid": item.sales_userid,
        "topic": item.topic,
        "core_demand": item.core_demand,
        "scenario_label": item.scenario_label,
        "intent_label": item.intent_label,
        "language_style": item.language_style,
        "business_state": item.business_state,
        "stage_signals": item.stage_signals,
        "merged_facts": item.merged_facts,
        "attachment_summary": item.attachment_summary,
        "fact_source": item.fact_source,
        "quality_score": float(item.quality_score) if item.quality_score is not None else None,
        "effect_score": float(item.effect_score) if item.effect_score is not None else None,
        "outcome_feedback": item.outcome_feedback,
        "usable_for_reply": item.usable_for_reply,
        "allowed_for_generation": item.allowed_for_generation,
        "reply_guard_reason": item.reply_guard_reason,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }

def _upsert_thread_business_fact(
    db: Session,
    *,
    session_id: str,
    summary_json: dict | None,
    crm_context: dict | None,
    messages: list[dict] | None,
    external_userid: str | None = None,
    sales_userid: str | None = None,
) -> ThreadBusinessFact:
    payload = build_thread_business_fact(
        session_id=session_id,
        summary=summary_json,
        crm_context=crm_context,
        messages=messages,
        external_userid=external_userid,
        sales_userid=sales_userid,
    )
    latest_log = None
    if session_id:
        latest_log = db.query(KnowledgeHitLog).filter(KnowledgeHitLog.session_id == session_id).order_by(KnowledgeHitLog.created_at.desc()).first()
    if latest_log:
        merged_facts = dict(payload.get("merged_facts") or {})
        merged_facts["latest_feedback_status"] = latest_log.feedback_status
        merged_facts["latest_final_response"] = sanitize_text((latest_log.final_response or "")[:280]) if latest_log.final_response else None
        merged_facts["latest_hit_chunk_ids"] = latest_log.hit_chunk_ids or []
        payload["merged_facts"] = merged_facts
        fact_source = dict(payload.get("fact_source") or {})
        fact_source["feedback_status"] = latest_log.feedback_status
        fact_source["has_final_response"] = bool(latest_log.final_response)
        fact_source["latest_log_at"] = latest_log.created_at.isoformat() if latest_log.created_at else None
        payload["fact_source"] = fact_source
    if session_id:
        email_assets = db.query(EmailThreadAsset).filter(
            or_(EmailThreadAsset.session_id == session_id, EmailThreadAsset.thread_id == session_id)
        ).order_by(EmailThreadAsset.created_at.desc()).limit(5).all()
        email_fragments = db.query(EmailFragmentAsset).filter(
            or_(EmailFragmentAsset.session_id == session_id, EmailFragmentAsset.thread_id == session_id)
        ).order_by(EmailFragmentAsset.created_at.desc()).limit(12).all()
        if email_assets or email_fragments:
            merged_facts = dict(payload.get("merged_facts") or {})
            if email_assets:
                merged_facts["email_subjects"] = [item.subject for item in email_assets if item.subject][:5]
                merged_facts["email_source_refs"] = [item.source_ref for item in email_assets if item.source_ref][:5]
            if email_fragments:
                merged_facts["email_fragment_count"] = len(email_fragments)
                merged_facts["email_fragment_types"] = sorted({item.function_fragment for item in email_fragments if item.function_fragment})
                merged_facts["email_fragment_samples"] = [
                    {
                        "fragment": item.function_fragment,
                        "content": sanitize_text((item.content or "")[:120]),
                    }
                    for item in email_fragments[:4]
                ]
            payload["merged_facts"] = merged_facts
            fact_source = dict(payload.get("fact_source") or {})
            fact_source["email_asset_count"] = len(email_assets)
            fact_source["email_fragment_count"] = len(email_fragments)
            payload["fact_source"] = fact_source
    item = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
    if not item:
        item = ThreadBusinessFact(session_id=session_id, thread_id=session_id)
        db.add(item)
        db.flush()
    for field in [
        "thread_id", "external_userid", "sales_userid", "topic", "core_demand",
        "scenario_label", "intent_label", "language_style", "business_state",
        "stage_signals", "merged_facts", "attachment_summary", "fact_source",
        "reply_guard_reason",
    ]:
        setattr(item, field, payload.get(field))
    item.quality_score = _decimal_score(payload.get("quality_score"))
    if latest_log:
        payload["outcome_feedback"] = {
            "feedback_status": latest_log.feedback_status,
            "manual_feedback": latest_log.manual_feedback,
            "final_response": sanitize_text((latest_log.final_response or "")[:280]) if latest_log.final_response else None,
        }
        if latest_log.feedback_status in {"useful", "adopted", "won", "advanced"}:
            payload["effect_score"] = 0.68
        elif latest_log.feedback_status in {"needs_fix", "rejected"}:
            payload["effect_score"] = 0.32
    item.effect_score = _decimal_score(payload.get("effect_score"))
    item.outcome_feedback = payload.get("outcome_feedback")
    item.usable_for_reply = bool(payload.get("usable_for_reply"))
    item.allowed_for_generation = bool(payload.get("allowed_for_generation"))
    return item


def _thread_fact_prompt_dict(item: ThreadBusinessFact | None) -> dict | None:
    if not item:
        return None
    return {
        "session_id": item.session_id,
        "scenario_label": item.scenario_label,
        "intent_label": item.intent_label,
        "language_style": item.language_style,
        "business_state": item.business_state,
        "stage_signals": item.stage_signals or {},
        "merged_facts": item.merged_facts or {},
        "reply_guard_reason": item.reply_guard_reason,
    }

def _parse_auto_split_flag(value) -> bool:
    if value is None or value == "":
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是", "需切分", "切分", "自动切分"}

def normalize_pricing_unit(value):
    if value is None:
        return None, None
    code = str(value).strip()
    if not code:
        return None, None
    if code in KB_PRICING_UNIT_LABELS:
        return code, None
    for item_code, label in KB_PRICING_UNIT_LABELS.items():
        if code == label:
            return item_code, None
    return None, code

def infer_faq_business_line(title: str, content: str) -> str:
    text = f"{title or ''}\n{content or ''}"
    if any(word in text for word in ["口译", "同传", "交传", "会议翻译", "会议服务"]):
        return "interpretation"
    if any(word in text for word in ["字幕", "配音", "译制", "多媒体", "音视频", "视频翻译"]):
        return "multimedia"
    if any(word in text for word in ["礼品", "赠品", "定制礼盒", "伴手礼"]):
        return "gifts"
    if any(word in text for word in ["翻译需求", "语种需求"]):
        return "translation"
    general_words = ["公司", "身份", "哪位", "没听说", "业务范围", "服务范围"]
    business_hits = sum(1 for word in ["翻译", "口译", "同传", "字幕", "配音", "语种", "印刷", "画册", "易拉宝", "样本", "手册", "展台", "展会", "搭建", "撤展", "礼品"] if word in text)
    if any(word in text for word in general_words) and business_hits >= 2:
        return "general"
    if any(word in text for word in ["印刷", "画册", "易拉宝", "样本", "手册"]):
        return "printing"
    if any(word in text for word in ["展台", "展会", "搭建", "撤展"]):
        return "exhibition"
    if any(word in text for word in ["翻译", "语种"]):
        return "translation"
    return "general"

def infer_faq_risk_level(title: str, content: str) -> str:
    text = f"{title or ''}\n{content or ''}"
    high_words = ["报价", "价格", "收费", "费用", "折扣", "合同", "赔付", "保证", "承诺", "税", "发票"]
    return "high" if any(word in text for word in high_words) else "medium"

DEFAULT_PRICING_EFFECTIVE_FROM = datetime(2026, 1, 1, 0, 0, 0)
DEFAULT_PRICING_EFFECTIVE_TO = datetime(2036, 12, 31, 23, 59, 59)
PRICING_TOPIC_WORDS = ["报价", "价格", "收费", "费用", "最低收费", "加急", "税点", "折扣", "发票"]
STRUCTURED_PRICING_PATTERNS = [
    r"(?:¥|￥|RMB|CNY|USD|EUR)?\s*\d+(?:\.\d+)?\s*(?:元|块|人民币|美元|欧元)",
    r"\d+(?:\.\d+)?\s*(?:元|块|人民币|美元|欧元)\s*(?:/|每)\s*(?:千字|千中文字符|千英文单词|页|项目|小时|人天|场|次)",
    r"(?:/|每)\s*(?:千字|千中文字符|千英文单词|页|项目|小时|人天|场|次)",
    r"最低收费\s*\d",
    r"加急倍率\s*\d",
    r"税点\s*\d",
    r"(?:含税|不含税|专票|普票)",
]

def mentions_pricing_topic(title: str | None, content: str | None) -> bool:
    text = f"{title or ''}\n{content or ''}"
    return any(word in text for word in PRICING_TOPIC_WORDS)

def has_structured_pricing_signal(title: str | None, content: str | None) -> bool:
    text = f"{title or ''}\n{content or ''}"
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in STRUCTURED_PRICING_PATTERNS)

def is_pricing_related_knowledge(
    knowledge_type: str | None = None,
    *,
    pricing_rule: dict | None = None,
) -> bool:
    return knowledge_type == "pricing" or bool(pricing_rule)

def default_pricing_effective_window(
    effective_from: datetime | None,
    effective_to: datetime | None,
    *,
    knowledge_type: str | None = None,
    pricing_rule: dict | None = None,
) -> tuple[datetime | None, datetime | None]:
    if is_pricing_related_knowledge(knowledge_type, pricing_rule=pricing_rule):
        effective_from = effective_from or DEFAULT_PRICING_EFFECTIVE_FROM
        effective_to = effective_to or DEFAULT_PRICING_EFFECTIVE_TO
    return effective_from, effective_to

def is_pricing_text(title: str | None, content: str | None, knowledge_type: str | None = None) -> bool:
    return knowledge_type == "pricing" or has_structured_pricing_signal(title, content)

def infer_pricing_rule_candidate(
    title: str,
    content: str,
    business_line: str,
    language_pair: str | None = None,
    service_scope: str | None = None,
    customer_tier: str | None = None,
    raw_pricing_rule: dict | None = None,
) -> dict | None:
    """从 LLM 结果或文本中生成结构化报价规则候选，候选仍随文档进入 draft/review。"""
    if raw_pricing_rule:
        candidate = {key: value for key, value in dict(raw_pricing_rule).items() if value is not None}
    elif is_pricing_text(title, content):
        candidate = {}
    else:
        return None

    text = f"{title or ''}\n{content or ''}"
    candidate.setdefault("business_line", business_line)
    candidate.setdefault("language_pair", language_pair)
    candidate.setdefault("service_scope", service_scope)
    candidate.setdefault("customer_tier", customer_tier)
    candidate.setdefault("currency", "CNY")

    if not candidate.get("unit"):
        if any(word in text for word in ["千字", "千字符", "每千字", "元/千字", "元每千字"]):
            candidate["unit"] = "per_1000_chars"
        elif "小时" in text:
            candidate["unit"] = "per_hour"
        elif "天" in text or "日" in text:
            candidate["unit"] = "per_day"
        else:
            candidate["unit"] = "per_project"

    price_match = re.search(r"(\d+(?:\.\d+)?)\s*元\s*/?\s*(?:每)?千字", text)
    if price_match and candidate.get("price_min") is None:
        candidate["price_min"] = price_match.group(1)

    min_charge_match = re.search(r"最低收费(?:为|是|:|：)?\s*(\d+(?:\.\d+)?)\s*元", text)
    if min_charge_match and candidate.get("min_charge") is None:
        candidate["min_charge"] = min_charge_match.group(1)

    urgent_multiplier_match = re.search(r"加急.*?(\d+(?:\.\d+)?)\s*倍", text)
    if urgent_multiplier_match and candidate.get("urgent_multiplier") is None:
        candidate["urgent_multiplier"] = urgent_multiplier_match.group(1)

    tax_match = re.search(r"(含税|不含税|税点\s*\d+(?:\.\d+)?%?|发票[^，。；;]*)", text)
    if tax_match and not candidate.get("tax_policy"):
        candidate["tax_policy"] = tax_match.group(1)

    candidate["source_ref"] = candidate.get("source_ref") or "auto_candidate_from_text"
    return candidate

def normalize_header(value) -> str:
    return str(value or "").strip().replace(" ", "").replace("\n", "")

def find_column_index(headers: list, candidates: list[str]) -> int | None:
    normalized = [normalize_header(h) for h in headers]
    for candidate in candidates:
        target = normalize_header(candidate)
        if target in normalized:
            return normalized.index(target)
    return None

KB_EXCEL_IMPORT_TYPES = {
    "unified": {"label": "统一知识与报价导入模板", "knowledge_class": None, "knowledge_type": "faq", "chunk_type": "faq", "risk_level": "medium"},
    "faq": {"label": "FAQ常见问答", "knowledge_class": "faq", "knowledge_type": "faq", "chunk_type": "faq", "risk_level": "medium"},
    "pricing": {"label": "结构化报价规则", "knowledge_class": None, "knowledge_type": "pricing", "chunk_type": "rule", "risk_level": "high"},
    "pricing_constraint": {"label": "报价限制条件", "knowledge_class": "pricing_constraint", "knowledge_type": "pricing", "chunk_type": "constraint", "risk_level": "high"},
    "process": {"label": "流程规则", "knowledge_class": "process", "knowledge_type": "process", "chunk_type": "rule", "risk_level": "medium"},
    "capability": {"label": "能力知识", "knowledge_class": "capability", "knowledge_type": "capability", "chunk_type": "rule", "risk_level": "medium"},
    "case": {"label": "案例", "knowledge_class": "example", "knowledge_type": "faq", "chunk_type": "example", "risk_level": "medium"},
    "script": {"label": "邮件模板", "knowledge_class": "email_template", "knowledge_type": "faq", "chunk_type": "template", "risk_level": "medium"},
    "definition": {"label": "名词定义", "knowledge_class": "definition", "knowledge_type": "faq", "chunk_type": "definition", "risk_level": "low"},
    "email_digest": {"label": "邮件整理", "knowledge_class": "faq", "knowledge_type": "faq", "chunk_type": "faq", "risk_level": "medium"},
}

KB_EXCEL_TEMPLATE_COLUMNS = [
    "标题",
    "内容",
    "知识分类",
    "切分",
    "服务",
    "语种",
    "服务范围",
    "客户层级",
    "优先级",
    "风险等级",
    "生效日期",
    "失效日期",
    "单位",
    "币种",
    "价格",
    "最高价格",
    "最低收费",
    "加急倍率",
    "税费",
    "标签",
]

KB_EXCEL_TEMPLATE_SAMPLE = {
    "unified": ["客户问公司主要做什么", "公司提供翻译、口译、多媒体、印刷、展会与礼品等企业服务。", "FAQ常见问答", 0, "通用", "", "普通资料", "", 50, "中", "", "", "", "", "", "", "", "", "", "公司介绍"],
    "faq": ["客户问公司主要做什么", "公司提供翻译、印刷、展台搭建等企业服务。", "FAQ常见问答", 0, "通用", "", "普通资料", "", 50, "中", "", "", "", "", "", "", "", "", "", "公司介绍"],
    "pricing": ["英译法普通商务资料报价", "英译法普通商务资料按220元/千字报价，最低收费300元。", "", 0, "翻译", "英译法", "普通资料", "", 95, "高", "2026-04-20", "2030-12-31", "每千中文字符", "CNY", 220, 220, 300, "", "不含税", "结构化报价"],
    "pricing_constraint": ["加急费用确认规则", "加急费用需根据交付时间和项目量另行确认，不能直接承诺固定加急价。", "报价限制条件", 0, "翻译", "", "普通资料", "", 90, "高", "", "", "", "", "", "", "", "", "", "报价限制"],
    "process": ["翻译下单流程", "收文件、确认语种用途、评估报价交期、付款立项、译审交付。", "流程规则", 1, "翻译", "", "普通资料", "", 80, "中", "", "", "", "", "", "", "", "", "", "流程"],
    "capability": ["英译法能力说明", "可承接英译法普通商务资料。", "能力知识", 0, "翻译", "英译法", "普通资料", "", 80, "中", "", "", "", "", "", "", "", "", "", "能力"],
    "case": ["价格异议案例", "客户质疑价格高，销售解释译审流程与交付保障后成交。", "案例", 0, "翻译", "", "普通资料", "", 60, "中", "", "", "", "", "", "", "", "", "", "案例"],
    "script": ["询价回复邮件模板", "您好，请您先发送需翻译文件，并说明语种、用途、交付时间和质量要求，我们收到后尽快评估报价。", "邮件模板", 0, "翻译", "", "普通资料", "", 60, "中", "", "", "", "", "", "", "", "", "", "邮件模板"],
    "definition": ["1000中文字符定义", "1000中文字符通常指中文原文字符数量，用于按千字计价的翻译报价口径。", "名词定义", 0, "翻译", "", "普通资料", "", 50, "低", "", "", "", "", "", "", "", "", "", "定义"],
    "email_digest": ["邮件询价整理", "客户邮件询问英译法报价和交期，需先确认文件字数和用途。", "FAQ常见问答", 0, "翻译", "英译法", "普通资料", "", 60, "中", "", "", "", "", "", "", "", "", "", "邮件"],
}

KB_UNIFIED_TEMPLATE_SAMPLE_ROWS = [
    ["客户问公司主要做什么", "公司提供翻译、口译、多媒体、印刷、展会与礼品等企业服务。", "FAQ常见问答", 0, "通用", "", "普通资料", "", 50, "中", "", "", "", "", "", "", "", "", "", "公司介绍"],
    ["翻译下单流程", "收文件、确认语种用途、评估报价交期、付款立项、译审交付。", "流程规则", 1, "翻译", "", "普通资料", "", 80, "中", "", "", "", "", "", "", "", "", "", "流程"],
    ["加急费用确认规则", "加急费用需根据交付时间和项目量另行确认，不能直接承诺固定加急价。", "报价限制条件", 0, "翻译", "", "普通资料", "", 90, "高", "", "", "", "", "", "", "", "", "", "报价限制"],
    ["英译法能力说明", "可承接英译法普通商务资料，支持标准译审流程和按需润色。", "能力知识", 0, "翻译", "英译法", "普通资料", "", 80, "中", "", "", "", "", "", "", "", "", "", "能力,英译法"],
    ["价格异议案例", "客户质疑价格高，销售先解释译审流程和交付保障，再补充历史交付经验，最终推进成交。", "案例", 0, "翻译", "", "普通资料", "重点客户", 65, "中", "", "", "", "", "", "", "", "", "", "案例,价格异议"],
    ["询价回复邮件模板", "您好，请您先发送需翻译文件，并说明语种、用途、交付时间和质量要求，我们收到后尽快评估报价。", "邮件模板", 0, "翻译", "", "普通资料", "", 70, "中", "", "", "", "", "", "", "", "", "", "邮件模板,询价"],
    ["1000中文字符定义", "1000中文字符通常指中文原文字符数量，用于按千字计价的翻译报价口径。", "名词定义", 0, "翻译", "", "普通资料", "", 55, "低", "", "", "", "", "", "", "", "", "", "定义,报价口径"],
    ["英译法普通商务资料报价", "英译法普通商务资料按220元/千字报价，最低收费300元。", "", 0, "翻译", "英译法", "普通资料", "", 95, "高", "2026-04-20", "2030-12-31", "每千中文字符", "CNY", 220, 220, 300, "", "不含税", "结构化报价"],
    ["邮件询价整理", "客户邮件询问英译法报价和交期，需先确认文件字数、用途、交付时间和是否需要盖章。", "FAQ常见问答", 1, "翻译", "英译法", "认证/盖章文件", "普通客户", 72, "中", "", "", "", "", "", "", "", "", "", "邮件,询价整理"],
]

KB_TEMPLATE_FIELD_GUIDE = [
    ["标题", "是", "全部知识", "建议写成可被审核人快速判断用途的主题，避免过短或过泛。", "翻译下单流程"],
    ["内容", "是", "全部知识", "填写可复用的业务说明、流程、问答、模板或规则正文。", "收文件、确认语种用途、评估报价交期、付款立项、译审交付。"],
    ["知识分类", "普通知识必填；结构化报价留空", "FAQ/流程/能力/案例/邮件模板/名词定义/报价限制条件", "统一模板用它判断落入哪类知识；结构化报价规则行请保持留空。", "流程规则"],
    ["切分", "否", "全部知识", "填 1 时会调用 LLM 按业务知识点拆成多条草稿切片；填 0 或留空则按单条导入。", "1"],
    ["服务", "建议填写", "全部知识", "用于限定业务线，常见值：翻译、印刷、同传、多媒体译制、展台搭建、礼品、通用。", "翻译"],
    ["语种", "按需填写", "语言相关知识或报价", "建议使用系统标准值，例如中译英、英译法、英译中。", "英译法"],
    ["服务范围", "按需填写", "全部知识", "限定资料范围，例如普通资料、法律资料、认证/盖章文件。", "普通资料"],
    ["客户层级", "按需填写", "全部知识", "用于限定知识适用对象，例如普通客户、重点客户、VIP客户。", "重点客户"],
    ["优先级", "否", "全部知识", "数值越大代表排序时越靠前，建议 1-100。", "80"],
    ["风险等级", "建议填写", "全部知识", "低/中/高；涉及报价承诺、限制条件、例外说明时建议填高。", "高"],
    ["生效日期", "结构化报价建议填写", "结构化报价规则", "价格开始生效日期，支持 yyyy-mm-dd。", "2026-04-20"],
    ["失效日期", "结构化报价建议填写", "结构化报价规则", "价格失效日期，支持 yyyy-mm-dd。", "2030-12-31"],
    ["单位", "结构化报价建议填写", "结构化报价规则", "建议使用模板示例口径，例如每千中文字符、每项目、每小时。", "每千中文字符"],
    ["币种", "结构化报价建议填写", "结构化报价规则", "如 CNY、USD。", "CNY"],
    ["价格", "结构化报价建议填写", "结构化报价规则", "最低价格或单价；若只有一个价格，最高价格可与之相同或留空。", "220"],
    ["最高价格", "否", "结构化报价规则", "用于价格区间上限；单点价格可留空或与价格相同。", "220"],
    ["最低收费", "按需填写", "结构化报价规则", "项目低于最低收费时按该金额执行。", "300"],
    ["加急倍率", "按需填写", "结构化报价规则", "只写明确可执行的倍率；不确定时应改写为报价限制条件。", "1.5"],
    ["税费", "按需填写", "结构化报价规则", "记录含税/不含税、发票要求等。", "不含税"],
    ["标签", "否", "全部知识", "多个标签用逗号分隔，方便后续筛选和审核。", "流程,下单"],
]

KB_TEMPLATE_RULE_GUIDE = [
    ["统一模板适用范围", "同一份模板同时支持 FAQ常见问答、流程规则、能力知识、案例、邮件模板、名词定义、报价限制条件，以及结构化报价规则。"],
    ["普通知识导入规则", "普通业务知识必须填写知识分类；系统会根据知识分类自动映射到底层 knowledge_type 和 chunk_type。"],
    ["结构化报价规则导入规则", "结构化报价规则行的知识分类留空，同时补齐价格、单位、币种、生效日期、失效日期等字段。"],
    ["切分规则", "切分=1 会调用 LLM 自动拆成多条草稿切片；结构化报价规则固定按单条导入。"],
    ["审核发布规则", "所有导入结果先进入 draft，仍需提交审核、审核同意并通过发布门禁后才能进入正式检索。"],
    ["样例使用方式", "模板首个工作表中的示例行可直接另存为内部样例，删除示例后再填正式数据。"],
]

KB_TEMPLATE_ENUM_GROUPS = [
    ("knowledge_class", "知识分类", "统一模板里普通知识必填；结构化报价规则留空。"),
    ("business_line", "服务", "用于限定业务线，推荐直接使用系统标准值。"),
    ("language_pair", "语种", "只在与语言方向强相关的知识或报价里填写。"),
    ("service_scope", "服务范围", "限定资料类型、项目范围或场景边界。"),
    ("customer_tier", "客户层级", "按客户重要度控制知识适用范围。"),
    ("risk_level", "风险等级", "影响审核与发布门禁，报价与限制条件一般为高风险。"),
]

KB_EXCEL_COLUMN_ALIASES = {
    "title": ["标题", "节点名", "问题", "场景", "主题", "案例标题", "话术标题", "邮件主题", "title"],
    "content": ["内容", "答案", "回复", "话术", "正文", "案例内容", "邮件内容", "处理方式", "流程", "content"],
    "knowledge_class": ["知识分类", "知识类别", "分类", "knowledge_class"],
    "auto_split": ["切分", "是否切分", "自动切分", "auto_split"],
    "knowledge_type": ["知识类型", "knowledge_type", "类型"],
    "chunk_type": ["切片类型", "chunk_type"],
    "business_line": ["服务", "业务线", "business_line", "业务"],
    "sub_service": ["子服务", "sub_service"],
    "language_pair": ["语种", "语言对", "language_pair"],
    "service_scope": ["服务范围", "资料类型", "service_scope"],
    "region": ["地区", "region"],
    "customer_tier": ["客户层级", "客户类型", "customer_tier"],
    "priority": ["优先级", "priority", "权重"],
    "risk_level": ["风险等级", "risk_level"],
    "effective_from": ["生效时间", "生效日期", "effective_from"],
    "effective_to": ["失效时间", "失效日期", "effective_to"],
    "unit": ["单位", "计价单位", "unit"],
    "currency": ["币种", "currency"],
    "price_min": ["最低价格", "价格", "单价", "price_min"],
    "price_max": ["最高价格", "price_max"],
    "min_charge": ["最低收费", "起步价", "min_charge"],
    "urgent_multiplier": ["加急倍率", "urgent_multiplier"],
    "tax_policy": ["税费", "税点", "发票", "tax_policy"],
    "tags": ["标签", "tags"],
}

def _set_worksheet_widths(sheet, rows: list[list[object]]):
    from openpyxl.utils import get_column_letter

    if not rows:
        return
    column_count = max(len(row) for row in rows)
    for column_index in range(1, column_count + 1):
        width = max(
            len(str(row[column_index - 1])) if column_index - 1 < len(row) and row[column_index - 1] is not None else 0
            for row in rows
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = max(12, min(width + 4, 40))

def _build_kb_import_template_workbook(template_type: str):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = KB_EXCEL_IMPORT_TYPES[template_type]["label"][:31]
    main_rows = [KB_EXCEL_TEMPLATE_COLUMNS]
    sample_rows = KB_UNIFIED_TEMPLATE_SAMPLE_ROWS if template_type == "unified" else [KB_EXCEL_TEMPLATE_SAMPLE[template_type]]
    main_rows.extend(sample_rows)
    for row in main_rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    _set_worksheet_widths(sheet, main_rows)

    if template_type == "unified":
        guide_sheet = workbook.create_sheet("填写说明")
        guide_rows = [["规则", "说明"], *KB_TEMPLATE_RULE_GUIDE]
        for row in guide_rows:
            guide_sheet.append(row)
        _set_worksheet_widths(guide_sheet, guide_rows)

        field_sheet = workbook.create_sheet("字段说明")
        field_rows = [["字段", "是否必填", "适用范围", "填写说明", "示例"], *KB_TEMPLATE_FIELD_GUIDE]
        for row in field_rows:
            field_sheet.append(row)
        _set_worksheet_widths(field_sheet, field_rows)

        enum_sheet = workbook.create_sheet("枚举参考")
        enum_rows = [["字段", "编码", "显示值", "说明"]]
        for dict_name, field_label, note in KB_TEMPLATE_ENUM_GROUPS:
            for code, label in KB_LABELS.get(dict_name, {}).items():
                enum_rows.append([field_label, code, label, note])
        for row in enum_rows:
            enum_sheet.append(row)
        _set_worksheet_widths(enum_sheet, enum_rows)

    workbook.active = 0
    return workbook

def _parse_datetime_or_none(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text_value = str(value).strip()
    if not text_value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(text_value, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text_value)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid datetime value: {value}")

def _cell(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value

def _safe_int(value, default: int = 50) -> int:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

def _has_structured_pricing_cells(row: tuple, columns: dict[str, int | None]) -> bool:
    pricing_fields = ["unit", "currency", "price_min", "price_max", "min_charge", "urgent_multiplier", "tax_policy"]
    return any(_cell(row, columns.get(field)) not in (None, "") for field in pricing_fields)

def _template_type_from_knowledge_class(knowledge_class: str | None) -> str:
    mapping = {
        "pricing_constraint": "pricing_constraint",
        "capability": "capability",
        "process": "process",
        "faq": "faq",
        "example": "case",
        "email_template": "script",
        "definition": "definition",
    }
    return mapping.get(knowledge_class or "", "faq")

def _resolve_excel_header(rows: list[tuple]) -> tuple[int, dict[str, int | None]]:
    best_index = None
    best_columns = None
    best_score = -1
    for idx, row in enumerate(rows[:10]):
        headers = list(row)
        columns = {
            field: find_column_index(headers, aliases)
            for field, aliases in KB_EXCEL_COLUMN_ALIASES.items()
        }
        score = sum(1 for value in columns.values() if value is not None)
        if columns["title"] is not None and columns["content"] is not None and score > best_score:
            best_index = idx
            best_columns = columns
            best_score = score
    if best_index is None or best_columns is None:
        raise HTTPException(status_code=400, detail="未找到必要列：标题/内容")
    return best_index, best_columns

def _load_excel_rows(raw: bytes, filename: str) -> tuple[str, list[tuple]]:
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl 未安装或不可用: {e}")
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件无法解析: {e}")
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 为空")
    return sheet.title, rows

def parse_kb_excel_import(raw: bytes, filename: str, import_type: str) -> dict:
    if import_type not in KB_EXCEL_IMPORT_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported import_type: {import_type}")
    sheet_title, rows = _load_excel_rows(raw, filename)
    header_row_index, columns = _resolve_excel_header(rows)
    import_config = KB_EXCEL_IMPORT_TYPES[import_type]
    valid_rows = []
    skipped = []
    failed = []

    for row_index, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        title = str(_cell(row, columns["title"]) or "").strip()
        content = str(_cell(row, columns["content"]) or "").strip()
        if not title and not content:
            skipped.append({"row": row_index, "reason": "empty"})
            continue
        row_errors = []
        if not title:
            row_errors.append("missing_title")
        if not content:
            row_errors.append("missing_content")

        requested_knowledge_class = _cell(row, columns["knowledge_class"])
        knowledge_class, raw_knowledge_class = normalize_knowledge_class(
            requested_knowledge_class or import_config.get("knowledge_class")
        )
        effective_import_type = import_type
        effective_import_config = import_config
        structured_pricing_row = _has_structured_pricing_cells(row, columns)
        if import_type == "unified":
            if structured_pricing_row or (not knowledge_class and has_structured_pricing_signal(title, content)):
                effective_import_type = "pricing"
                effective_import_config = KB_EXCEL_IMPORT_TYPES["pricing"]
            elif knowledge_class:
                effective_import_type = _template_type_from_knowledge_class(knowledge_class)
                effective_import_config = KB_EXCEL_IMPORT_TYPES[effective_import_type]
            else:
                row_errors.append("missing_knowledge_class")

        class_config = KB_KNOWLEDGE_CLASS_CONFIG.get(knowledge_class or "")
        knowledge_type, raw_knowledge_type = normalize_code(
            "knowledge_type",
            _cell(row, columns["knowledge_type"]) or (class_config or {}).get("knowledge_type") or effective_import_config["knowledge_type"],
        )
        business_line, raw_business_line = normalize_code(
            "business_line",
            _cell(row, columns["business_line"]) or infer_faq_business_line(title, content),
        )
        language_pair, raw_language_pair = normalize_code("language_pair", _cell(row, columns["language_pair"]))
        service_scope, raw_service_scope = normalize_code("service_scope", _cell(row, columns["service_scope"]))
        customer_tier, raw_customer_tier = normalize_code("customer_tier", _cell(row, columns["customer_tier"]))

        chunk_type, raw_chunk_type = normalize_code(
            "chunk_type",
            _cell(row, columns["chunk_type"]) or (class_config or {}).get("chunk_type") or effective_import_config["chunk_type"],
        )
        chunk_type = chunk_type or effective_import_config["chunk_type"]
        risk_level, raw_risk_level = normalize_code(
            "risk_level",
            _cell(row, columns["risk_level"]) or (class_config or {}).get("risk_level") or effective_import_config["risk_level"],
        )
        risk_level = risk_level or effective_import_config["risk_level"]
        if knowledge_type == "pricing" or mentions_pricing_topic(title, content):
            risk_level = "high"
        auto_split = _parse_auto_split_flag(_cell(row, columns["auto_split"]))
        if effective_import_type == "pricing":
            auto_split = False

        tags = {
            "import_type": effective_import_type,
            "request_import_type": import_type,
            "raw_source": "kb_excel_import",
            "knowledge_class": knowledge_class or knowledge_class_from_pair(knowledge_type, chunk_type),
            "auto_split": auto_split,
        }
        for key, value in {
            "raw_knowledge_class": raw_knowledge_class,
            "raw_knowledge_type": raw_knowledge_type,
            "raw_business_line": raw_business_line,
            "raw_language_pair": raw_language_pair,
            "raw_service_scope": raw_service_scope,
            "raw_customer_tier": raw_customer_tier,
            "raw_chunk_type": raw_chunk_type,
            "raw_risk_level": raw_risk_level,
        }.items():
            if value:
                tags[key] = value
        raw_tags = _cell(row, columns["tags"])
        if raw_tags:
            tags["source_tags"] = str(raw_tags)

        pricing_rule = None
        pricing_cells = [
            _cell(row, columns["unit"]),
            _cell(row, columns["currency"]),
            _cell(row, columns["price_min"]),
            _cell(row, columns["price_max"]),
            _cell(row, columns["min_charge"]),
            _cell(row, columns["urgent_multiplier"]),
            _cell(row, columns["tax_policy"]),
        ]
        should_build_pricing_rule = (
            effective_import_type == "pricing"
            or (
                knowledge_type == "pricing"
                and chunk_type != "constraint"
                and (any(value not in (None, "") for value in pricing_cells) or has_structured_pricing_signal(title, content))
            )
        )
        if should_build_pricing_rule:
            unit, raw_unit = normalize_pricing_unit(_cell(row, columns["unit"]) or "per_project")
            if raw_unit:
                tags["raw_unit"] = raw_unit
            pricing_rule = {
                "business_line": business_line,
                "sub_service": None,
                "language_pair": language_pair,
                "service_scope": service_scope,
                "unit": unit or "per_project",
                "currency": _cell(row, columns["currency"]) or "CNY",
                "price_min": _cell(row, columns["price_min"]),
                "price_max": _cell(row, columns["price_max"]),
                "min_charge": _cell(row, columns["min_charge"]),
                "urgent_multiplier": _cell(row, columns["urgent_multiplier"]),
                "tax_policy": _cell(row, columns["tax_policy"]),
            }
            if not any(pricing_rule.get(field) not in (None, "") for field in ["price_min", "price_max", "min_charge", "urgent_multiplier", "tax_policy"]):
                inferred_rule = infer_pricing_rule_candidate(title, content, business_line or "general", language_pair, service_scope, customer_tier)
                pricing_rule = inferred_rule or pricing_rule
            if not any(pricing_rule.get(field) not in (None, "") for field in ["price_min", "price_max", "min_charge", "urgent_multiplier", "tax_policy"]):
                row_errors.append("pricing_rule_missing_value")

        effective_from = _parse_datetime_or_none(_cell(row, columns["effective_from"]))
        effective_to = _parse_datetime_or_none(_cell(row, columns["effective_to"]))
        effective_from, effective_to = default_pricing_effective_window(
            effective_from,
            effective_to,
            knowledge_type=knowledge_type or effective_import_config["knowledge_type"],
            pricing_rule=pricing_rule,
        )

        item = {
            "row": row_index,
            "title": title,
            "content": content,
            "import_type": effective_import_type,
            "import_type_label": KB_EXCEL_IMPORT_TYPES[effective_import_type]["label"],
            "knowledge_type": knowledge_type or effective_import_config["knowledge_type"],
            "chunk_type": chunk_type,
            "business_line": business_line or "general",
            "sub_service": None,
            "language_pair": language_pair,
            "service_scope": service_scope,
            "region": _cell(row, columns["region"]),
            "customer_tier": customer_tier,
            "priority": _safe_int(_cell(row, columns["priority"]), 50),
            "risk_level": risk_level,
            "effective_from": effective_from,
            "effective_to": effective_to,
            "tags": tags,
            "pricing_rule": pricing_rule,
            "knowledge_class": tags["knowledge_class"],
            "auto_split": auto_split,
            "errors": row_errors,
        }
        if row_errors:
            failed.append({"row": row_index, "title": title, "errors": row_errors})
        else:
            valid_rows.append(item)

    return {
        "filename": filename,
        "sheet": sheet_title,
        "import_type": import_type,
        "import_type_label": KB_EXCEL_IMPORT_TYPES[import_type]["label"],
        "header_row": header_row_index + 1,
        "columns": columns,
        "valid_rows": valid_rows,
        "skipped": skipped,
        "failed": failed,
    }

def _excel_item_preview(item: dict) -> dict:
    return {
        "row": item["row"],
        "title": item["title"],
        "knowledge_class": item.get("knowledge_class"),
        "knowledge_class_label": label_for("knowledge_class", item.get("knowledge_class")),
        "auto_split": item.get("auto_split", False),
        "knowledge_type": item["knowledge_type"],
        "knowledge_type_label": label_for("knowledge_type", item["knowledge_type"]),
        "chunk_type": item["chunk_type"],
        "chunk_type_label": label_for("chunk_type", item["chunk_type"]),
        "business_line": item["business_line"],
        "business_line_label": label_for("business_line", item["business_line"]),
        "language_pair": item.get("language_pair"),
        "language_pair_label": label_for("language_pair", item.get("language_pair")),
        "service_scope": item.get("service_scope"),
        "service_scope_label": label_for("service_scope", item.get("service_scope")),
        "risk_level": item["risk_level"],
        "risk_level_label": label_for("risk_level", item["risk_level"]),
        "pricing_rule": item.get("pricing_rule"),
    }

def _doc_to_dict(doc: KnowledgeDocument) -> dict:
    def extract_knowledge_class(tags):
        if isinstance(tags, dict):
            return tags.get("knowledge_class")
        elif isinstance(tags, list):
            # 假设 list 里是 dict 或 key-value 对
            for item in tags:
                if isinstance(item, dict) and "knowledge_class" in item:
                    return item["knowledge_class"]
                if isinstance(item, (list, tuple)) and len(item) == 2 and item[0] == "knowledge_class":
                    return item[1]
        return None

    knowledge_class_val = extract_knowledge_class(doc.tags) or knowledge_class_from_pair(doc.knowledge_type, None)
    return {
        "document_id": str(doc.document_id),
        "title": doc.title,
        "knowledge_class": knowledge_class_val,
        "knowledge_class_label": label_for("knowledge_class", knowledge_class_val),
        "knowledge_type": doc.knowledge_type,
        "knowledge_type_label": label_for("knowledge_type", doc.knowledge_type),
        "business_line": doc.business_line,
        "business_line_label": label_for("business_line", doc.business_line),
        "sub_service": doc.sub_service,
        "sub_service_label": label_for("sub_service", doc.sub_service),
        "source_type": doc.source_type,
        "source_ref": doc.source_ref,
        "source_meta": doc.source_meta,
        "status": doc.status,
        "status_label": label_for("status", doc.status),
        "display_status": document_stage_code(doc),
        "display_status_label": document_stage_label(doc),
        "version_no": doc.version_no,
        "effective_from": doc.effective_from,
        "effective_to": doc.effective_to,
        "owner": doc.owner,
        "import_batch": doc.import_batch,
        "risk_level": doc.risk_level,
        "risk_level_label": label_for("risk_level", doc.risk_level),
        "review_required": doc.review_required,
        "review_status": doc.review_status,
        "review_status_label": label_for("review_status", doc.review_status),
        "library_type": doc.library_type,
        "allowed_actions": document_allowed_actions(doc),
        "tags": doc.tags,
        "created_at": doc.created_at,
        "updated_at": doc.updated_at,
    }

def _chunk_to_dict(chunk: KnowledgeChunk) -> dict:
    return {
        "chunk_id": str(chunk.chunk_id),
        "document_id": str(chunk.document_id),
        "chunk_no": chunk.chunk_no,
        "chunk_type": chunk.chunk_type,
        "title": chunk.title,
        "content": chunk.content,
        "knowledge_class": (chunk.structured_tags or {}).get("knowledge_class") or knowledge_class_from_pair(chunk.knowledge_type, chunk.chunk_type),
        "knowledge_class_label": knowledge_class_label(chunk.knowledge_type, chunk.chunk_type, chunk.structured_tags),
        "business_line": chunk.business_line,
        "business_line_label": label_for("business_line", chunk.business_line),
        "sub_service": chunk.sub_service,
        "sub_service_label": label_for("sub_service", chunk.sub_service),
        "knowledge_type": chunk.knowledge_type,
        "knowledge_type_label": label_for("knowledge_type", chunk.knowledge_type),
        "language_pair": chunk.language_pair,
        "language_pair_label": label_for("language_pair", chunk.language_pair),
        "service_scope": chunk.service_scope,
        "service_scope_label": label_for("service_scope", chunk.service_scope),
        "region": chunk.region,
        "customer_tier": chunk.customer_tier,
        "customer_tier_label": label_for("customer_tier", chunk.customer_tier),
        "priority": chunk.priority,
        "library_type": chunk.library_type,
        "allowed_for_generation": chunk.allowed_for_generation,
        "usable_for_reply": chunk.usable_for_reply,
        "publishable": chunk.publishable,
        "topic_clarity_score": float(chunk.topic_clarity_score) if chunk.topic_clarity_score is not None else None,
        "completeness_score": float(chunk.completeness_score) if chunk.completeness_score is not None else None,
        "reusability_score": float(chunk.reusability_score) if chunk.reusability_score is not None else None,
        "evidence_reliability_score": float(chunk.evidence_reliability_score) if chunk.evidence_reliability_score is not None else None,
        "useful_score": float(chunk.useful_score) if chunk.useful_score is not None else None,
        "effect_score": float(chunk.effect_score) if chunk.effect_score is not None else None,
        "feedback_count": chunk.feedback_count,
        "positive_feedback_count": chunk.positive_feedback_count,
        "last_feedback_at": chunk.last_feedback_at,
        "quality_notes": chunk.quality_notes,
        "structured_tags": chunk.structured_tags,
        "effective_from": chunk.effective_from,
        "effective_to": chunk.effective_to,
        "status": chunk.status,
        "status_label": label_for("status", chunk.status),
        "embedding_provider": chunk.embedding_provider,
        "embedding_model": chunk.embedding_model,
        "embedding_dim": chunk.embedding_dim,
        "created_at": chunk.created_at,
        "updated_at": chunk.updated_at,
    }

def _decimal_or_none(value):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise HTTPException(status_code=422, detail=f"Invalid decimal value: {value}")

def _pricing_rule_to_dict(rule: PricingRule) -> dict:
    return {
        "rule_id": str(rule.rule_id),
        "document_id": str(rule.document_id),
        "chunk_id": str(rule.chunk_id) if rule.chunk_id else None,
        "business_line": rule.business_line,
        "business_line_label": label_for("business_line", rule.business_line),
        "sub_service": rule.sub_service,
        "sub_service_label": label_for("sub_service", rule.sub_service),
        "language_pair": rule.language_pair,
        "language_pair_label": label_for("language_pair", rule.language_pair),
        "service_scope": rule.service_scope,
        "service_scope_label": label_for("service_scope", rule.service_scope),
        "unit": rule.unit,
        "currency": rule.currency,
        "price_min": float(rule.price_min) if rule.price_min is not None else None,
        "price_max": float(rule.price_max) if rule.price_max is not None else None,
        "urgent_multiplier": float(rule.urgent_multiplier) if rule.urgent_multiplier is not None else None,
        "tax_policy": rule.tax_policy,
        "min_charge": float(rule.min_charge) if rule.min_charge is not None else None,
        "customer_tier": rule.customer_tier,
        "customer_tier_label": label_for("customer_tier", rule.customer_tier),
        "region": rule.region,
        "status": rule.status,
        "status_label": label_for("status", rule.status),
        "effective_from": rule.effective_from,
        "effective_to": rule.effective_to,
        "version_no": rule.version_no,
        "source_ref": rule.source_ref,
        "created_at": rule.created_at,
        "updated_at": rule.updated_at,
    }

def _redact_value(value):
    if isinstance(value, str):
        return sanitize_text(value)
    if isinstance(value, list):
        return [_redact_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _redact_value(item) for key, item in value.items()}
    return value

def _hit_log_to_dict(log: KnowledgeHitLog, redact: bool = False) -> dict:
    query_text = log.query_text
    final_response = log.final_response
    manual_feedback = log.manual_feedback
    if redact:
        query_text = _redact_value(query_text)
        final_response = _redact_value(final_response)
        manual_feedback = _redact_value(manual_feedback)
    return {
        "log_id": str(log.log_id),
        "request_id": log.request_id,
        "session_id": log.session_id,
        "query_text": query_text,
        "query_features": log.query_features,
        "filters_used": log.filters_used,
        "hit_chunk_ids": log.hit_chunk_ids,
        "scores": log.scores,
        "no_hit_reason": log.no_hit_reason,
        "status": log.status,
        "retrieval_quality": log.retrieval_quality,
        "confidence_score": float(log.confidence_score) if log.confidence_score is not None else None,
        "insufficient_info": log.insufficient_info,
        "manual_review_required": log.manual_review_required,
        "final_response": final_response,
        "manual_feedback": manual_feedback,
        "feedback_status": log.feedback_status,
        "latency_ms": log.latency_ms,
        "created_at": log.created_at,
    }

def _candidate_to_dict(candidate: KnowledgeCandidate) -> dict:
    return {
        "candidate_id": str(candidate.candidate_id),
        "title": candidate.title,
        "content": candidate.content,
        "knowledge_class": (candidate.source_snapshot or {}).get("knowledge_class") or knowledge_class_from_pair(candidate.knowledge_type, candidate.chunk_type),
        "knowledge_class_label": label_for("knowledge_class", (candidate.source_snapshot or {}).get("knowledge_class") or knowledge_class_from_pair(candidate.knowledge_type, candidate.chunk_type)),
        "knowledge_type": candidate.knowledge_type,
        "knowledge_type_label": label_for("knowledge_type", candidate.knowledge_type),
        "chunk_type": candidate.chunk_type,
        "chunk_type_label": label_for("chunk_type", candidate.chunk_type),
        "business_line": candidate.business_line,
        "business_line_label": label_for("business_line", candidate.business_line),
        "sub_service": candidate.sub_service,
        "sub_service_label": label_for("sub_service", candidate.sub_service),
        "language_pair": candidate.language_pair,
        "language_pair_label": label_for("language_pair", candidate.language_pair),
        "service_scope": candidate.service_scope,
        "service_scope_label": label_for("service_scope", candidate.service_scope),
        "region": candidate.region,
        "customer_tier": candidate.customer_tier,
        "customer_tier_label": label_for("customer_tier", candidate.customer_tier),
        "priority": candidate.priority,
        "risk_level": candidate.risk_level,
        "risk_level_label": label_for("risk_level", candidate.risk_level),
        "effective_from": candidate.effective_from,
        "effective_to": candidate.effective_to,
        "pricing_rule": candidate.pricing_rule,
        "source_type": candidate.source_type,
        "source_type_label": label_for("candidate_source_type", candidate.source_type),
        "source_ref": candidate.source_ref,
        "source_snapshot": candidate.source_snapshot,
        "library_type": candidate.library_type,
        "allowed_for_generation": candidate.allowed_for_generation,
        "usable_for_reply": candidate.usable_for_reply,
        "publishable": candidate.publishable,
        "topic_clarity_score": float(candidate.topic_clarity_score) if candidate.topic_clarity_score is not None else None,
        "completeness_score": float(candidate.completeness_score) if candidate.completeness_score is not None else None,
        "reusability_score": float(candidate.reusability_score) if candidate.reusability_score is not None else None,
        "evidence_reliability_score": float(candidate.evidence_reliability_score) if candidate.evidence_reliability_score is not None else None,
        "useful_score": float(candidate.useful_score) if candidate.useful_score is not None else None,
        "effect_score": float(candidate.effect_score) if candidate.effect_score is not None else None,
        "feedback_count": candidate.feedback_count,
        "positive_feedback_count": candidate.positive_feedback_count,
        "last_feedback_at": candidate.last_feedback_at,
        "quality_notes": candidate.quality_notes,
        "status": candidate.status,
        "status_label": label_for("candidate_status", candidate.status),
        "owner": candidate.owner,
        "operator": candidate.operator,
        "review_notes": candidate.review_notes,
        "promoted_document_id": str(candidate.promoted_document_id) if candidate.promoted_document_id else None,
        "created_at": candidate.created_at,
        "updated_at": candidate.updated_at,
    }

def _email_thread_asset_to_dict(item: EmailThreadAsset) -> dict:
    return {
        "email_id": str(item.email_id),
        "source_type": item.source_type,
        "source_ref": item.source_ref,
        "import_batch": item.import_batch,
        "session_id": item.session_id,
        "thread_id": item.thread_id,
        "external_userid": item.external_userid,
        "sales_userid": item.sales_userid,
        "fact_id": str(item.fact_id) if item.fact_id else None,
        "subject": item.subject,
        "content": item.content,
        "sender": item.sender,
        "receiver": item.receiver,
        "sent_at": item.sent_at,
        "sent_at_raw": item.sent_at_raw,
        "scenario_label": item.scenario_label,
        "intent_label": item.intent_label,
        "language_style": item.language_style,
        "business_state": item.business_state,
        "library_type": item.library_type,
        "quality_score": float(item.quality_score) if item.quality_score is not None else None,
        "effect_score": float(item.effect_score) if item.effect_score is not None else None,
        "feedback_count": item.feedback_count,
        "positive_feedback_count": item.positive_feedback_count,
        "last_feedback_at": item.last_feedback_at,
        "usable_for_reply": item.usable_for_reply,
        "allowed_for_generation": item.allowed_for_generation,
        "publishable": item.publishable,
        "source_snapshot": item.source_snapshot,
        "status": item.status,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }

def _email_fragment_asset_to_dict(item: EmailFragmentAsset) -> dict:
    return {
        "fragment_id": str(item.fragment_id),
        "email_id": str(item.email_id) if item.email_id else None,
        "candidate_id": str(item.candidate_id) if item.candidate_id else None,
        "log_id": str(item.log_id) if item.log_id else None,
        "session_id": item.session_id,
        "thread_id": item.thread_id,
        "source_type": item.source_type,
        "source_ref": item.source_ref,
        "title": item.title,
        "content": item.content,
        "function_fragment": item.function_fragment,
        "scenario_label": item.scenario_label,
        "intent_label": item.intent_label,
        "language_style": item.language_style,
        "library_type": item.library_type,
        "allowed_for_generation": item.allowed_for_generation,
        "usable_for_reply": item.usable_for_reply,
        "publishable": item.publishable,
        "topic_clarity_score": float(item.topic_clarity_score) if item.topic_clarity_score is not None else None,
        "completeness_score": float(item.completeness_score) if item.completeness_score is not None else None,
        "reusability_score": float(item.reusability_score) if item.reusability_score is not None else None,
        "evidence_reliability_score": float(item.evidence_reliability_score) if item.evidence_reliability_score is not None else None,
        "useful_score": float(item.useful_score) if item.useful_score is not None else None,
        "effect_score": float(item.effect_score) if item.effect_score is not None else None,
        "feedback_count": item.feedback_count,
        "positive_feedback_count": item.positive_feedback_count,
        "last_feedback_at": item.last_feedback_at,
        "quality_notes": item.quality_notes,
        "source_snapshot": item.source_snapshot,
        "status": item.status,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }

def _email_effect_feedback_to_dict(item: EmailEffectFeedback) -> dict:
    return {
        "feedback_id": str(item.feedback_id),
        "email_id": str(item.email_id) if item.email_id else None,
        "fragment_id": str(item.fragment_id) if item.fragment_id else None,
        "candidate_id": str(item.candidate_id) if item.candidate_id else None,
        "log_id": str(item.log_id) if item.log_id else None,
        "session_id": item.session_id,
        "thread_id": item.thread_id,
        "feedback_status": item.feedback_status,
        "feedback_note": item.feedback_note,
        "feedback_payload": item.feedback_payload,
        "delta_score": float(item.delta_score) if item.delta_score is not None else None,
        "created_at": item.created_at,
    }

def _training_sample_to_dict(item: ModelTrainingSample) -> dict:
    return {
        "sample_id": str(item.sample_id),
        "sample_key": item.sample_key,
        "sample_type": item.sample_type,
        "source_table": item.source_table,
        "source_id": item.source_id,
        "source_type": item.source_type,
        "source_ref": item.source_ref,
        "instruction": item.instruction,
        "input_text": item.input_text,
        "target_text": item.target_text,
        "sample_metadata": item.sample_metadata,
        "quality_score": float(item.quality_score) if item.quality_score is not None else None,
        "effect_score": float(item.effect_score) if item.effect_score is not None else None,
        "exportable": item.exportable,
        "review_status": item.review_status,
        "export_status": item.export_status,
        "last_export_path": item.last_export_path,
        "exported_at": item.exported_at,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }

def _jsonable(value: Any) -> Any:
    return json.loads(json.dumps(value, ensure_ascii=False, default=str))

def _version_snapshot_to_dict(snapshot: KnowledgeVersionSnapshot) -> dict:
    return {
        "snapshot_id": str(snapshot.snapshot_id),
        "document_id": str(snapshot.document_id),
        "version_no": snapshot.version_no,
        "action": snapshot.action,
        "operator": snapshot.operator,
        "note": snapshot.note,
        "created_at": snapshot.created_at,
    }

def _job_to_dict(job: JobTask) -> dict:
    return {
        "job_id": str(job.job_id),
        "job_type": job.job_type,
        "status": job.status,
        "progress": job.progress,
        "summary": job.summary,
        "task_payload": job.task_payload,
        "result": job.result,
        "error_message": job.error_message,
        "retry_of": job.retry_of,
        "created_at": job.created_at,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "updated_at": job.updated_at,
    }

def _document_snapshot_payload(doc: KnowledgeDocument, chunks: list[KnowledgeChunk], pricing_rules: list[PricingRule]) -> dict:
    return _jsonable({
        "document": _doc_to_dict(doc),
        "chunks": [_chunk_to_dict(chunk) for chunk in chunks],
        "pricing_rules": [_pricing_rule_to_dict(rule) for rule in pricing_rules],
    })

def _next_publish_version_no(db: Session, document_id: str, current_version: int | None) -> int:
    latest_snapshot = db.query(func.max(KnowledgeVersionSnapshot.version_no)).filter(
        KnowledgeVersionSnapshot.document_id == document_id
    ).scalar()
    if latest_snapshot:
        return int(latest_snapshot) + 1
    return max(int(current_version or 1), 1)

def _create_version_snapshot(
    db: Session,
    doc: KnowledgeDocument,
    *,
    action: str = "publish",
    operator: str | None = None,
    note: str | None = None,
    version_no: int | None = None,
) -> KnowledgeVersionSnapshot:
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).order_by(KnowledgeChunk.chunk_no.asc()).all()
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id == doc.document_id).order_by(PricingRule.created_at.asc()).all()
    snapshot = KnowledgeVersionSnapshot(
        document_id=doc.document_id,
        version_no=version_no or int(doc.version_no or 1),
        action=action,
        operator=operator,
        note=note,
        snapshot_data=_document_snapshot_payload(doc, chunks, pricing_rules),
    )
    db.add(snapshot)
    db.flush()
    return snapshot

def _restore_document_from_snapshot(db: Session, doc: KnowledgeDocument, snapshot: KnowledgeVersionSnapshot) -> dict:
    payload = snapshot.snapshot_data or {}
    doc_payload = payload.get("document") or {}
    chunks_payload = payload.get("chunks") or []
    pricing_payload = payload.get("pricing_rules") or []

    for field in [
        "title", "knowledge_type", "business_line", "sub_service", "source_type",
        "source_ref", "risk_level", "owner", "import_batch", "review_required", "tags",
    ]:
        if field in doc_payload:
            setattr(doc, field, doc_payload.get(field))
    doc.source_meta = dict(doc_payload.get("source_meta") or {})
    doc.effective_from = _parse_datetime_or_none(doc_payload.get("effective_from"))
    doc.effective_to = _parse_datetime_or_none(doc_payload.get("effective_to"))
    doc.version_no = int(snapshot.version_no)
    doc.status = "review"
    doc.review_required = True
    doc.review_status = "in_review"

    db.query(PricingRule).filter(PricingRule.document_id == doc.document_id).delete(synchronize_session=False)
    db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).delete(synchronize_session=False)
    db.flush()

    chunk_id_map: dict[str, str] = {}
    recreated_chunks = []
    recreated_rules = []
    for chunk_payload in sorted(chunks_payload, key=lambda item: int(item.get("chunk_no") or 1)):
        retrieval_title = _strip_retrieval_title_prefix(chunk_payload.get("title") or doc.title)
        retrieval_text = _build_chunk_retrieval_text(chunk_payload.get("title") or doc.title, chunk_payload.get("content") or "")
        embedding = _optional_embedding_for_storage(
            retrieval_text,
            context=f"restore_chunk:{chunk_payload.get('title') or doc.title}",
        )
        chunk = KnowledgeChunk(
            document_id=doc.document_id,
            chunk_no=int(chunk_payload.get("chunk_no") or 1),
            chunk_type=chunk_payload.get("chunk_type") or "faq",
            title=retrieval_title,
            content=chunk_payload.get("content") or "",
            keyword_text=chunk_payload.get("keyword_text") or retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=int(chunk_payload.get("priority") or 50),
            retrieval_weight=_decimal_or_none(chunk_payload.get("retrieval_weight")) or Decimal("1.000"),
            business_line=chunk_payload.get("business_line") or doc.business_line,
            sub_service=chunk_payload.get("sub_service"),
            knowledge_type=chunk_payload.get("knowledge_type") or doc.knowledge_type,
            language_pair=chunk_payload.get("language_pair"),
            service_scope=chunk_payload.get("service_scope"),
            region=chunk_payload.get("region"),
            customer_tier=chunk_payload.get("customer_tier"),
            structured_tags=chunk_payload.get("structured_tags"),
            status="review",
            effective_from=doc.effective_from,
            effective_to=doc.effective_to,
        )
        db.add(chunk)
        db.flush()
        if chunk_payload.get("chunk_id"):
            chunk_id_map[str(chunk_payload["chunk_id"])] = str(chunk.chunk_id)
        recreated_chunks.append(chunk)

    for rule_payload in pricing_payload:
        rule = PricingRule(
            document_id=doc.document_id,
            chunk_id=chunk_id_map.get(str(rule_payload.get("chunk_id"))) if rule_payload.get("chunk_id") else None,
            business_line=rule_payload.get("business_line") or doc.business_line,
            sub_service=rule_payload.get("sub_service"),
            language_pair=rule_payload.get("language_pair"),
            service_scope=rule_payload.get("service_scope"),
            unit=rule_payload.get("unit") or "per_project",
            currency=rule_payload.get("currency") or "CNY",
            price_min=_decimal_or_none(rule_payload.get("price_min")),
            price_max=_decimal_or_none(rule_payload.get("price_max")),
            urgent_multiplier=_decimal_or_none(rule_payload.get("urgent_multiplier")),
            tax_policy=rule_payload.get("tax_policy"),
            min_charge=_decimal_or_none(rule_payload.get("min_charge")),
            customer_tier=rule_payload.get("customer_tier"),
            region=rule_payload.get("region"),
            status="review",
            effective_from=doc.effective_from,
            effective_to=doc.effective_to,
            version_no=doc.version_no,
            source_ref=rule_payload.get("source_ref") or doc.source_ref,
        )
        db.add(rule)
        recreated_rules.append(rule)

    for chunk in recreated_chunks:
        linked_rule_payload = next(
            (item for item in pricing_payload if str(item.get("chunk_id") or "") in {str(chunk.chunk_id), ""}),
            None,
        )
        _apply_chunk_governance(
            chunk,
            source_type=doc.source_type,
            pricing_rule=linked_rule_payload,
            source_ref=doc.source_ref,
        )
    doc.library_type = recreated_chunks[0].library_type if recreated_chunks else doc.library_type
    doc.tags = merge_tags(doc.tags, library_type=doc.library_type)
    restore_meta = dict(doc.source_meta or {})
    restore_meta["restored_from_version"] = snapshot.version_no
    restore_meta["restored_at"] = datetime.utcnow().isoformat()
    doc.source_meta = restore_meta
    return {
        "restored_version": snapshot.version_no,
        "chunk_count": len(recreated_chunks),
        "pricing_rule_count": len(recreated_rules),
    }

def _job_payload_dir() -> str:
    path = os.path.join(os.path.dirname(__file__), "..", "scratch", "job_payloads")
    os.makedirs(path, exist_ok=True)
    return os.path.abspath(path)

def _create_job_task(
    db: Session,
    *,
    job_type: str,
    task_payload: dict | None = None,
    summary: str | None = None,
    retry_of: str | None = None,
) -> JobTask:
    job = JobTask(
        job_type=job_type,
        status="queued",
        progress=0,
        summary=summary or "queued",
        task_payload=_jsonable(task_payload or {}),
        retry_of=retry_of,
    )
    db.add(job)
    db.flush()
    return job

def _create_single_document_and_chunk(
    db: Session,
    *,
    title: str,
    content: str,
    knowledge_class: str | None = None,
    knowledge_type: str,
    business_line: str,
    sub_service: str | None = None,
    chunk_type: str = "faq",
    language_pair: str | None = None,
    service_scope: str | None = None,
    region: str | None = None,
    customer_tier: str | None = None,
    source_type: str = "manual",
    source_ref: str | None = None,
    source_meta: dict | None = None,
    owner: str | None = None,
    priority: int = 50,
    risk_level: str = "medium",
    review_required: bool = True,
    tags: dict | None = None,
    effective_from: datetime | None = None,
    effective_to: datetime | None = None,
    pricing_rule: dict | None = None,
    import_batch: str | None = None,
):
    resolved_class, resolved_type, resolved_chunk_type, resolved_risk = resolve_knowledge_class_fields(
        knowledge_class=knowledge_class,
        knowledge_type=knowledge_type,
        chunk_type=chunk_type,
        risk_level=risk_level,
    )
    effective_from, effective_to = default_pricing_effective_window(
        effective_from,
        effective_to,
        knowledge_type=resolved_type,
        pricing_rule=pricing_rule,
    )
    merged_tags = merge_knowledge_class_tags(tags, resolved_class)
    retrieval_title = _strip_retrieval_title_prefix(title)
    retrieval_text = _build_chunk_retrieval_text(title, content)
    embedding = _optional_embedding_for_storage(retrieval_text, context=f"create_document:{title}")

    document = KnowledgeDocument(
        title=title,
        knowledge_type=resolved_type,
        business_line=business_line,
        sub_service=sub_service,
        source_type=source_type,
        source_ref=source_ref,
        source_meta=source_meta,
        status="draft",
        owner=owner,
        import_batch=import_batch,
        effective_from=effective_from,
        effective_to=effective_to,
        risk_level=resolved_risk,
        review_required=review_required,
        review_status="pending" if review_required else "auto_ready",
        library_type=infer_library_type(source_type=source_type, knowledge_type=resolved_type, chunk_type=resolved_chunk_type, tags=merged_tags),
        tags=merged_tags,
    )
    db.add(document)
    db.flush()

    chunk = KnowledgeChunk(
        document_id=document.document_id,
        chunk_no=1,
        chunk_type=resolved_chunk_type,
        title=retrieval_title,
        content=content,
        keyword_text=retrieval_text,
        embedding=embedding,
        **_embedding_metadata_fields(embedding),
        priority=priority,
        business_line=business_line,
        sub_service=sub_service,
        knowledge_type=resolved_type,
        language_pair=language_pair,
        service_scope=service_scope,
        region=region,
        customer_tier=customer_tier,
        structured_tags=merged_tags,
        status="draft",
        effective_from=effective_from,
        effective_to=effective_to,
    )
    db.add(chunk)
    db.flush()

    pricing_rule_payload = pricing_rule
    if not pricing_rule_payload and resolved_type == "pricing" and resolved_chunk_type != "constraint":
        pricing_rule_payload = infer_pricing_rule_candidate(
            title,
            content,
            business_line,
            language_pair,
            service_scope,
            customer_tier,
        )
    rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
    _apply_chunk_governance(chunk, source_type=source_type, pricing_rule=pricing_rule_payload, source_ref=source_ref)
    document.library_type = chunk.library_type
    document.tags = merge_tags(document.tags, library_type=document.library_type)
    return document, chunk, rule

def _extract_relevant_sentences(content: str, keywords: list[str], limit: int = 3) -> str:
    sentences = [
        sentence.strip()
        for sentence in re.split(r"[。！？!\n]+", content or "")
        if sentence and sentence.strip()
    ]
    matched = [sentence for sentence in sentences if any(keyword in sentence for keyword in keywords)]
    selected = matched[:limit] if matched else sentences[:limit]
    text = "。".join(selected).strip("。")
    return text or (content or "")[:500]


def _infer_historical_email_business_line(title: str, content: str) -> str:
    text = f"{title}\n{content}"
    if any(word in text for word in ["翻译", "译员", "译文", "稿件", "定稿"]):
        return "translation"
    if any(word in text for word in ["印刷", "打样", "彩页", "名片", "纸张"]):
        return "printing"
    if any(word in text for word in ["展台", "展会", "搭建"]):
        return "exhibition"
    if any(word in text for word in ["视频", "字幕", "配音"]):
        return "multimedia"
    if any(word in text for word in ["礼品", "台历", "笔记本"]):
        return "gifts"
    return "general"


def _historical_email_keywords(use_range_label: str) -> list[str]:
    mapping = {
        "商务沟通（方案+报价）": ["报价", "底价", "费用", "金额", "单价", "总金额", "折扣", "补充报价", "确认报价", "交稿", "提交"],
        "发送合同/NDA": ["合同", "NDA", "协议", "预付款", "金额", "确认", "签章"],
        "财务沟通": ["发票", "付款", "到账", "银行", "invoice", "payment", "结算", "清单"],
        "发送发票": ["发票", "付款", "invoice", "寄出", "盖章件", "结算"],
        "催款": ["催款", "付款", "到账", "invoice", "payment", "到期", "逾期"],
        "到账确认": ["到账", "收到款", "payment"],
        "成果交付": ["定稿", "下载", "评价", "修改", "交付", "完成", "反馈质量"],
        "修改/打样": ["修改", "打样", "源文件", "清晰度", "重新报价", "二维码", "不清晰"],
        "进度更新": ["进度", "完成", "安排", "修改", "重新报价", "源文件"],
        "售后支持": ["修改", "问题", "不满意", "免费修改", "售后", "补偿"],
        "催确认订单": ["确认订单", "签章合同", "当前报价", "价格上调", "本周", "风险"],
        "询问/确认需求": ["预算", "时间", "周期", "方案", "效率", "视觉效果", "需求", "倾向"],
    }
    return mapping.get(use_range_label, [])


def _historical_email_compact(text: str) -> str:
    return re.sub(r"\s+", "", str(text or "")).lower()


def _historical_email_has_any(text: str, tokens: list[str]) -> bool:
    return any(token in text for token in tokens)


def _historical_email_has_reusable_signal(record: dict, content: str) -> bool:
    use_range_label = str(record.get("use_range_label") or "").strip()
    compact = _historical_email_compact(content)

    reusable_tokens = [
        "银行手续费", "bankcharge", "含税", "净价", "单价", "工作日", "交稿", "排版参照原文",
        "免费修改", "评价", "订单中心", "下载界面", "产能", "交货期", "启动生产", "纸张",
        "成本增加", "重新报价", "是否需要翻译", "未包括", "模块化", "定制化", "有效期",
        "续签", "预付款", "到期", "结算", "每月", "付款进度", "到期", "逾期", "付款状态",
    ]
    if _historical_email_has_any(compact, [token.lower() for token in reusable_tokens]):
        return True

    if use_range_label == "商务沟通（方案+报价）":
        return _historical_email_has_any(compact, [
            "单价", "千中文字符", "千英文单词", "千字", "每页", "交稿时间", "报价未包括", "底价",
        ])
    if use_range_label == "催确认订单":
        return _historical_email_has_any(compact, [
            "本周", "否则", "风险", "po", "签章合同", "当前报价", "启动生产",
        ])
    if use_range_label == "发送合同/NDA":
        return _historical_email_has_any(compact, ["预付款", "续签", "有效期", "到期"])
    if use_range_label in {"财务沟通", "发送发票", "催款"}:
        return _historical_email_has_any(compact, [
            "bankcharge", "银行手续费", "付款进度", "付款状态", "逾期", "到期", "结算", "系统收货确认",
        ])
    if use_range_label == "成果交付":
        return _historical_email_has_any(compact, [
            "订单中心", "下载", "评价", "免费修改", "反馈质量", "上传和下载",
        ])
    return False


def _is_low_value_historical_email(record: dict, content: str) -> bool:
    use_range_label = str(record.get("use_range_label") or "").strip()
    compact = _historical_email_compact(content)
    if use_range_label in {"收到确认", "到账确认"}:
        return True
    if len(compact) <= 18 and any(token in compact for token in ["收到", "谢谢", "请查收", "好的"]):
        return True
    if use_range_label == "时间/物流/现场协调":
        logistics_tokens = ["快递", "单号", "收货地址", "收货人", "手机", "寄出", "派送", "顺丰", "中通"]
        if any(token in content for token in logistics_tokens):
            return True
    if use_range_label in {"发送发票", "财务沟通", "成果交付"} and len(compact) <= 24:
        return True
    attachment_status_patterns = [
        r"^附件.*请查收",
        r"^请查收附件",
        r"^发票请见附件",
        r"^报价单请查收",
        r"^合同.*请查收",
        r"^nda.*请查收",
        r"^已签字盖章寄出",
        r"^已安排快递",
        r"^thepaymenthasbeenreceived",
        r"^thisisthe.*translationfile.*plscheckit",
    ]
    if any(re.search(pattern, compact) for pattern in attachment_status_patterns):
        return not _historical_email_has_reusable_signal(record, content)
    if use_range_label in {"发送合同/NDA", "发送发票", "成果交付", "财务沟通"} and not _historical_email_has_reusable_signal(record, content):
        weak_action_tokens = ["请查收", "请安排", "已收到", "寄出", "附件", "发出", "invoice", "payment"]
        if len(compact) <= 48 or _historical_email_has_any(compact, weak_action_tokens):
            return True
    return False


def _build_deterministic_historical_email_items(title: str, content: str, record: dict) -> list[dict] | None:
    use_range_label = str(record.get("use_range_label") or "").strip()
    business_line = _infer_historical_email_business_line(title, content)
    compact = _historical_email_compact(content)
    pricing_rule = infer_pricing_rule_candidate(title, content, business_line) or {}
    if use_range_label in {"商务沟通（方案+报价）", "催确认订单"} and not pricing_rule.get("price_min"):
        generic_price_match = re.search(r"(?:RMB|人民币)?\s*(\d+(?:\.\d+)?)\s*元", content)
        if generic_price_match:
            pricing_rule = {
                **pricing_rule,
                "business_line": pricing_rule.get("business_line") or business_line,
                "language_pair": pricing_rule.get("language_pair"),
                "service_scope": pricing_rule.get("service_scope") or "general",
                "customer_tier": pricing_rule.get("customer_tier"),
                "currency": pricing_rule.get("currency") or "CNY",
                "unit": pricing_rule.get("unit") or "per_project",
                "price_min": pricing_rule.get("price_min") or generic_price_match.group(1),
                "source_ref": pricing_rule.get("source_ref") or "auto_candidate_from_text",
            }
    has_price = bool(pricing_rule.get("price_min") or pricing_rule.get("price_max") or pricing_rule.get("min_charge"))

    if use_range_label in {"财务沟通", "发送发票", "催款", "到账确认"}:
        if "bankcharge" in compact or "银行手续费" in compact:
            normalized_title = "国际付款场景下发票金额与银行手续费说明"
        elif any(token in compact for token in ["付款进度", "付款状态", "逾期", "到期"]):
            normalized_title = "逾期发票的付款状态跟进话术"
        elif any(token in compact for token in ["结算", "每月", "清单确认"]):
            normalized_title = "按月结算场景的清单确认与提交流程"
        else:
            return []
        return [{
            "title": normalized_title[:255],
            "content": content,
            "knowledge_type": "process",
            "chunk_type": "rule",
            "business_line": business_line,
            "sub_service": None,
            "language_pair": None,
            "service_scope": "general",
            "region": None,
            "customer_tier": None,
            "priority": 72,
            "risk_level": "high",
            "pricing_rule": None,
        }]

    if use_range_label == "发送合同/NDA":
        if any(token in compact for token in ["到期", "续签", "有效期"]):
            normalized_title = "长期协议到期后的续签提醒"
        elif "预付款" in compact:
            normalized_title = "合同中预付款条款的确认口径"
        else:
            return []
        return [{
            "title": normalized_title[:255],
            "content": content,
            "knowledge_type": "process",
            "chunk_type": "constraint",
            "business_line": business_line,
            "sub_service": None,
            "language_pair": None,
            "service_scope": "general",
            "region": None,
            "customer_tier": None,
            "priority": 78,
            "risk_level": "high",
            "pricing_rule": None,
        }]

    if use_range_label in {"商务沟通（方案+报价）", "催确认订单"}:
        if "产能" in compact or "启动生产" in compact or ("po" in compact and "交货期" in compact):
            normalized_title = "样稿确认后需PO启动生产并锁定产能"
            item_type = ("process", "constraint", None)
        elif any(token in compact for token in ["当前报价", "纸张", "成本增加", "否则", "本周"]):
            normalized_title = "印刷报价确认时限与材料成本风险提示"
            item_type = ("process", "constraint", None)
        elif has_price and any(token in compact for token in ["单价", "含税", "净价", "工作日", "交稿", "千中文字符", "千英文单词", "排版参照原文"]):
            normalized_title = "翻译报价模板：计价单位、含税净价、交付时效与文件格式"
            item_type = ("pricing", "rule", pricing_rule)
        else:
            return []
        return [{
            "title": normalized_title[:255],
            "content": content,
            "knowledge_type": item_type[0],
            "chunk_type": item_type[1],
            "business_line": business_line,
            "sub_service": None,
            "language_pair": None,
            "service_scope": "general",
            "region": None,
            "customer_tier": None,
            "priority": 78,
            "risk_level": "high",
            "pricing_rule": item_type[2],
        }]

    if use_range_label == "成果交付":
        if any(token in compact for token in ["订单中心", "下载", "评价", "免费修改", "反馈质量", "上传和下载"]):
            normalized_title = "交付后的下载、评价与免费修改引导"
        else:
            return []
        return [{
            "title": normalized_title[:255],
            "content": content,
            "knowledge_type": "process",
            "chunk_type": "rule",
            "business_line": business_line,
            "sub_service": None,
            "language_pair": None,
            "service_scope": "general",
            "region": None,
            "customer_tier": None,
            "priority": 70,
            "risk_level": "medium",
            "pricing_rule": None,
        }]

    return None


def _build_historical_email_candidate_entries(record: dict) -> list[dict]:
    use_range_label = str(record.get("use_range_label") or "").strip()
    intent_label = str(record.get("intent_type_label") or "").strip()
    title = str(record.get("title") or record.get("subject") or "历史邮件候选").strip()
    raw_content = sanitize_text(str(record.get("content") or "").strip())
    if not raw_content or _is_low_value_historical_email(record, raw_content):
        return []

    keywords = _historical_email_keywords(use_range_label)
    focused_content = _extract_relevant_sentences(raw_content, keywords, limit=5) if keywords else raw_content[:500]
    focused_content = focused_content[:800]
    if not focused_content.strip():
        return []

    deterministic_title = title
    if not str(record.get("subject") or "").strip():
        deterministic_title = f"{use_range_label or '历史邮件'}"
        if intent_label:
            deterministic_title += f"·{intent_label}"

    deterministic_items = _build_deterministic_historical_email_items(deterministic_title, focused_content, record)
    if deterministic_items is not None:
        return deterministic_items

    try:
        return _llm_extract_candidate_items_from_text(deterministic_title, focused_content, "email_excel")
    except Exception:
        return []


def _llm_extract_candidate_items_from_text(title: str, content: str, source_type: str) -> list[dict]:
    prompt = f"""
你是企业知识库候选抽取助手。只允许根据原文抽取已明确出现的业务知识，不允许猜测、补全、编造价格、能力、流程或客户信息。

请返回纯 JSON：
{{
  "items": [
    {{
      "title": "候选知识标题",
      "content": "可审核入库的原文依据或严谨改写",
      "knowledge_type": "faq|pricing|process|capability",
      "chunk_type": "faq|rule|example|constraint|definition",
      "business_line": "translation|printing|interpretation|multimedia|exhibition|gifts|general",
      "sub_service": null,
      "language_pair": null,
      "service_scope": null,
      "region": null,
      "customer_tier": null,
      "priority": 50,
      "risk_level": "low|medium|high",
      "pricing_rule": null
    }}
  ]
}}

要求：
- 没有明确知识点时返回 {{"items":[]}}，不要兜底生成常见问答。
- 报价、折扣、税费、合同、赔付、能力承诺必须 risk_level=high 或 medium，且 content 必须包含原文依据。
- pricing_rule 只能填原文中明确出现的数字、单位、币种、最低收费、税费或加急倍率；不确定则为 null。
- source_type={source_type}

标题：{title}
原文：
{content}
"""
    llm_result = IntentEngine.run_llm1_json_prompt(prompt, user_id="kb_candidate_extract")
    raw_items = llm_result.get("items") if isinstance(llm_result, dict) else None
    if raw_items is None:
        raise HTTPException(status_code=502, detail="LLM 候选抽取返回缺少 items 字段")
    if not isinstance(raw_items, list):
        raise HTTPException(status_code=502, detail="LLM 候选抽取 items 不是数组")

    items = []
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            raise HTTPException(status_code=502, detail="LLM 候选抽取返回了非对象 item")
        item_title = str(raw_item.get("title") or "").strip()
        item_content = str(raw_item.get("content") or "").strip()
        if not item_title or not item_content:
            continue
        knowledge_type, _raw_knowledge_type = normalize_code("knowledge_type", raw_item.get("knowledge_type") or "faq")
        business_line, _raw_business_line = normalize_code("business_line", raw_item.get("business_line") or "general")
        language_pair, _raw_language_pair = normalize_code("language_pair", raw_item.get("language_pair"))
        service_scope, _raw_service_scope = normalize_code("service_scope", raw_item.get("service_scope"))
        risk_level = str(raw_item.get("risk_level") or infer_faq_risk_level(item_title, item_content)).strip()
        if risk_level not in KB_LABELS["risk_level"]:
            raise HTTPException(status_code=502, detail=f"LLM 候选抽取返回非法 risk_level: {risk_level}")
        pricing_rule = raw_item.get("pricing_rule")
        if pricing_rule:
            pricing_rule = infer_pricing_rule_candidate(
                item_title,
                item_content,
                business_line or "general",
                language_pair,
                service_scope,
                raw_pricing_rule=pricing_rule,
            )
        items.append({
            "title": item_title[:255],
            "content": item_content,
            "knowledge_type": knowledge_type or "faq",
            "chunk_type": raw_item.get("chunk_type") or ("example" if source_type == "case_extract" else "faq"),
            "business_line": business_line or "general",
            "sub_service": raw_item.get("sub_service"),
            "language_pair": language_pair,
            "service_scope": service_scope,
            "region": raw_item.get("region"),
            "customer_tier": raw_item.get("customer_tier"),
            "priority": int(raw_item.get("priority") or 50),
            "risk_level": risk_level,
            "pricing_rule": pricing_rule,
        })
    return items

def _build_candidate_entries_from_record(record: dict, source_type: str) -> list[dict]:
    title = str(record.get("title") or record.get("subject") or "历史资料候选").strip()
    content = sanitize_text(str(record.get("content") or "").strip())
    if not content:
        return []
    if source_type == "email_excel":
        items = _build_historical_email_candidate_entries(record)
    else:
        items = _llm_extract_candidate_items_from_text(title, content, source_type)
    snapshot = {
        key: sanitize_text(str(value)) if isinstance(value, str) else value
        for key, value in record.items()
        if value not in (None, "")
    }
    for item in items:
        item["source_snapshot"] = snapshot
    return items

def _create_candidate_record(
    db: Session,
    *,
    title: str,
    content: str,
    knowledge_type: str,
    chunk_type: str,
    business_line: str,
    sub_service: str | None = None,
    language_pair: str | None = None,
    service_scope: str | None = None,
    region: str | None = None,
    customer_tier: str | None = None,
    priority: int = 50,
    risk_level: str = "medium",
    effective_from: datetime | None = None,
    effective_to: datetime | None = None,
    pricing_rule: dict | None = None,
    source_type: str = "feedback",
    source_ref: str | None = None,
    source_snapshot: dict | None = None,
    owner: str | None = None,
    operator: str | None = None,
    review_notes: str | None = None,
):
    source_snapshot = dict(source_snapshot or {})
    candidate = KnowledgeCandidate(
        title=title,
        content=content,
        knowledge_type=knowledge_type,
        chunk_type=chunk_type,
        business_line=business_line,
        sub_service=sub_service,
        language_pair=language_pair,
        service_scope=service_scope,
        region=region,
        customer_tier=customer_tier,
        priority=priority,
        risk_level=risk_level,
        effective_from=effective_from,
        effective_to=effective_to,
        pricing_rule=pricing_rule,
        source_type=source_type,
        source_ref=source_ref,
        source_snapshot=source_snapshot,
        library_type=infer_library_type(source_type=source_type, knowledge_type=knowledge_type, chunk_type=chunk_type, tags=source_snapshot.get("tags")),
        status="candidate",
        owner=owner,
        operator=operator,
        review_notes=review_notes,
    )
    db.add(candidate)
    db.flush()
    _apply_candidate_governance(candidate)
    return candidate

def _json_safe(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    return value


def _strip_retrieval_title_prefix(title: str | None) -> str:
    text = str(title or "").strip()
    if not text:
        return ""
    for prefix in ("开发邮件", "业务资料"):
        if not text.startswith(prefix):
            continue
        rest = text[len(prefix):].lstrip()
        index = 0
        while index < len(rest) and rest[index].isdigit():
            index += 1
        if index == 0:
            continue
        rest = rest[index:].lstrip()
        if rest and rest[0] in {"·", "•", "・", ".", "-", "_", ":", "：", "/"}:
            cleaned = rest[1:].strip()
            return cleaned or text
    return text


def _build_chunk_retrieval_text(title: str | None, content: str | None) -> str:
    retrieval_title = _strip_retrieval_title_prefix(title)
    body = str(content or "").strip()
    if retrieval_title and body:
        return f"{retrieval_title}\n{body}"
    return retrieval_title or body

def _email_asset_source_ref_from_candidate_source_ref(source_ref: str | None) -> str:
    value = str(source_ref or "").strip()
    if not value:
        return f"email_candidate_{uuid.uuid4().hex[:12]}"
    head, sep, tail = value.rpartition("_")
    if sep and tail.isdigit():
        return head
    return value

def _split_email_into_function_fragments(subject: str, content: str) -> list[dict]:
    text = sanitize_text(content or "")
    if not text:
        return []
    paragraphs = [part.strip() for part in re.split(r"\n{2,}", text) if part.strip()]
    if not paragraphs:
        paragraphs = [text]
    fragments: list[dict] = []
    for index, paragraph in enumerate(paragraphs[:8], start=1):
        fragment_kind = infer_function_fragment(
            title=subject,
            content=paragraph,
            source_type="email_excel",
            tags={"paragraph_index": index},
        ) or ("core_answer" if index == len(paragraphs) else "background")
        fragments.append(
            {
                "title": f"{subject[:80]} / {fragment_kind} / {index}",
                "content": paragraph,
                "function_fragment": fragment_kind,
            }
        )
    return fragments

def _parse_optional_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text_value = str(value).strip()
    if not text_value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text_value, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text_value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None

def _feedback_delta(feedback_status: str | None) -> float:
    if feedback_status in POSITIVE_FEEDBACK_STATUSES:
        return 0.12
    if feedback_status in NEGATIVE_FEEDBACK_STATUSES:
        return -0.10
    return 0.0

def _build_email_asset_source_ref(record: dict, *, import_batch: str, index: int) -> str:
    explicit = str(record.get("source_ref") or "").strip()
    if explicit:
        return explicit[:255]
    thread_id = str(record.get("thread_id") or record.get("session_id") or "").strip()
    if thread_id:
        return thread_id[:255]
    row_no = record.get("row") or index
    if str(record.get("use_range_label") or "").strip() == "开发推广触达":
        return f"dev_email_row_{row_no}"[:255]
    file_fingerprint = hashlib.md5(str(import_batch or "").strip().lower().encode("utf-8")).hexdigest()[:8]
    return f"email_excel_{file_fingerprint}_row_{row_no}"[:255]


DEV_EMAIL_ROW_SEQUENCE = [
    3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
]
DEV_EMAIL_SKIP_ROWS = {15, 18}
DEV_EMAIL_ROW_META = {
    3: {"name": "国际展会服务英文触达模板", "business_line": "exhibition", "service_scope": "marketing"},
    4: {"name": "国际化活动一站式服务模板", "business_line": "exhibition", "service_scope": "marketing"},
    5: {"name": "年终活动与礼品服务模板", "business_line": "exhibition", "service_scope": "marketing"},
    6: {"name": "年末感谢与多服务拓展模板", "business_line": "general", "service_scope": "marketing"},
    7: {"name": "生命科学语言服务介绍模板", "business_line": "translation", "service_scope": "medical"},
    8: {"name": "有限空间活动方案模板", "business_line": "exhibition", "service_scope": "marketing"},
    9: {"name": "新春问候与服务更新模板", "business_line": "general", "service_scope": "marketing"},
    10: {"name": "翻译+综合服务介绍模板", "business_line": "general", "service_scope": "marketing"},
    11: {"name": "新年致谢与新服务引导模板", "business_line": "general", "service_scope": "marketing"},
    12: {"name": "培训资料本地化能力模板", "business_line": "translation", "service_scope": "training"},
    13: {"name": "老客户回访与合作回顾模板", "business_line": "translation", "service_scope": "general"},
    14: {"name": "医药注册翻译能力模板", "business_line": "translation", "service_scope": "medical"},
    16: {"name": "设计印刷服务介绍模板", "business_line": "printing", "service_scope": "marketing"},
    17: {"name": "未接来电后跟进模板", "business_line": "translation", "service_scope": "medical"},
    19: {"name": "活动策划与多语沟通方案模板", "business_line": "exhibition", "service_scope": "marketing"},
    20: {"name": "老客户新年回访与赠礼模板", "business_line": "general", "service_scope": "marketing"},
    21: {"name": "英文综合服务开发模板", "business_line": "general", "service_scope": "marketing"},
    22: {"name": "展会筹备指南触达模板", "business_line": "exhibition", "service_scope": "marketing"},
    23: {"name": "医药翻译跟进与案例触达模板", "business_line": "translation", "service_scope": "medical"},
    24: {"name": "AI大会展会服务触达模板", "business_line": "exhibition", "service_scope": "marketing"},
    25: {"name": "AI大会展会服务触达模板（含账号引导）", "business_line": "exhibition", "service_scope": "marketing"},
    26: {"name": "AI大会展会服务触达模板", "business_line": "exhibition", "service_scope": "marketing"},
    27: {"name": "定制笔记本推广模板", "business_line": "gifts", "service_scope": "marketing"},
    28: {"name": "轨交标书翻译案例触达模板", "business_line": "translation", "service_scope": "railway"},
}
DEV_EMAIL_LINE_SPECS = {
    4: [("背景与行业痛点", 1, 2, "background", "template"), ("一站式活动方案", 3, 4, "core_answer", "template"), ("实战案例表现", 5, 10, "highlight", "example"), ("服务清单", 11, 16, "core_answer", "template"), ("SPEED三省优势", 17, 20, "highlight", "template"), ("合作场景与邀约", 21, 26, "cta", "template")],
    5: [("年终活动场景引入", 1, 1, "background", "template"), ("公司经验与整体价值", 2, 3, "highlight", "template"), ("活动策划与场地布置", 4, 6, "core_answer", "template"), ("印刷与视觉物料", 7, 8, "core_answer", "template"), ("口译与多媒体支持", 9, 12, "core_answer", "template"), ("礼品定制方案", 13, 14, "core_answer", "template"), ("档期提醒与合作邀约", 15, 18, "cta", "template")],
    6: [("年终问候与感谢", 1, 2, "background", "template"), ("年度服务成果", 3, 7, "highlight", "example"), ("客户信任与能力延展", 8, 9, "highlight", "template"), ("新年合作展望", 10, 12, "cta", "template")],
    7: [("HMOs项目服务背景", 1, 1, "background", "template"), ("HMOs领域理解", 2, 5, "highlight", "example"), ("翻译质量与合规要求", 6, 6, "core_answer", "template"), ("保密机制与沟通邀约", 7, 8, "cta", "template")],
    8: [("活动痛点引入", 1, 1, "background", "template"), ("一站式方案定位", 2, 3, "core_answer", "template"), ("案例效果证明", 4, 7, "highlight", "example"), ("服务清单", 8, 14, "core_answer", "template"), ("三省优势", 15, 18, "highlight", "template"), ("合作预期与伙伴定位", 19, 23, "cta", "template")],
    9: [("新春问候与合作回顾", 1, 2, "background", "template"), ("扩展服务一览", 3, 8, "core_answer", "template"), ("首批体验邀请", 9, 9, "cta", "template"), ("全流程合作伙伴定位", 10, 10, "highlight", "template"), ("推荐与新年祝福", 11, 12, "cta", "template")],
    10: [("服务组合价值引入", 1, 4, "background", "template"), ("同传服务能力", 5, 5, "core_answer", "template"), ("翻译设计印刷一体化方案", 6, 10, "core_answer", "template"), ("多媒体译制优势", 11, 11, "highlight", "example"), ("软件网站本地化能力", 12, 16, "core_answer", "template"), ("合作邀约与祝福", 17, 18, "cta", "template")],
    11: [("新年致谢与长期合作", 1, 2, "background", "template"), ("样册与案例附件说明", 3, 3, "highlight", "example"), ("AI视频面试平台引导", 4, 4, "core_answer", "template"), ("纪念笔记本赠送邀约", 5, 5, "cta", "template"), ("2026合作展望与致意", 6, 7, "cta", "template")],
    12: [("培训本地化挑战场景", 1, 1, "background", "template"), ("交付方案与客户结果", 2, 2, "highlight", "example"), ("AI培训系统价值延展", 3, 3, "core_answer", "template")],
    13: [("再次联系与背景铺垫", 1, 1, "background", "template"), ("公司简介与服务范围", 2, 4, "highlight", "template"), ("快速响应与技术质量", 5, 9, "core_answer", "template"), ("语言覆盖与行业案例", 10, 13, "highlight", "example"), ("项目管理与一站式方案", 14, 17, "core_answer", "template"), ("需求邀约与祝福", 18, 20, "cta", "template")],
    14: [("医药出海风险引入", 1, 1, "background", "template"), ("公司资历与多语种能力", 2, 3, "highlight", "template"), ("六大优势上半部分", 4, 7, "core_answer", "template"), ("六大优势下半部分", 8, 10, "core_answer", "template"), ("资料范围覆盖", 11, 17, "highlight", "example"), ("合作邀约", 18, 18, "cta", "template")],
    16: [("翻译印刷一体化定位", 1, 4, "background", "template"), ("省时优势", 5, 7, "highlight", "template"), ("省心优势", 8, 9, "highlight", "template"), ("省力优势与合作邀约", 10, 12, "cta", "template")],
    17: [("未接来电与合作延续", 1, 1, "background", "template"), ("历史合作基础", 2, 2, "highlight", "example"), ("医疗翻译能力", 3, 3, "core_answer", "template"), ("需求探询与联系", 4, 4, "cta", "template")],
    19: [("场景痛点与一站式定位", 1, 2, "background", "template"), ("三省优势", 3, 6, "highlight", "template"), ("客户案例证明", 7, 11, "highlight", "example"), ("适用诉求清单", 12, 15, "core_answer", "template"), ("合作定位与邀约", 16, 17, "cta", "template")],
    20: [("新年问候与对接变更", 1, 1, "background", "template"), ("化工行业案例能力", 2, 2, "highlight", "example"), ("服务范围与合作邀约", 3, 3, "core_answer", "template"), ("赠礼说明与祝福", 4, 5, "cta", "template")],
    21: [("Introduction and Connection", 1, 2, "background", "template"), ("Translation and Finance Credentials", 3, 3, "highlight", "example"), ("Gift and Merchandise Examples", 4, 4, "highlight", "example"), ("Service Breadth and CTA", 5, 7, "cta", "template")],
    22: [("指南价值说明", 1, 1, "highlight", "template"), ("查阅与咨询引导", 2, 2, "cta", "template")],
    23: [("沟通背景与现状理解", 1, 1, "background", "template"), ("合作基础与专业经验", 2, 2, "highlight", "template"), ("擅长方向示例", 3, 6, "core_answer", "template"), ("优势说明与后续跟进", 7, 8, "cta", "template"), ("医学翻译成功案例清单", 9, 22, "highlight", "example")],
    24: [("问候与大会引入", 1, 2, "background", "template"), ("已签约展会案例", 3, 8, "highlight", "example"), ("展会配套服务清单", 9, 14, "core_answer", "template"), ("公司积累与服务覆盖", 15, 16, "highlight", "template"), ("合作邀约", 17, 17, "cta", "template")],
    25: [("问候与大会引入", 1, 2, "background", "template"), ("已签约展会案例", 3, 8, "highlight", "example"), ("展会配套服务清单", 9, 14, "core_answer", "template"), ("公司积累与服务覆盖", 15, 16, "highlight", "template"), ("合作邀约", 17, 17, "cta", "template"), ("官方账号引导", 18, 21, "cta", "template")],
    26: [("问候与大会引入", 1, 2, "background", "template"), ("已签约展会案例", 3, 8, "highlight", "example"), ("展会配套服务清单", 9, 14, "core_answer", "template"), ("公司积累与服务覆盖", 15, 16, "highlight", "template"), ("合作邀约", 17, 17, "cta", "template")],
    27: [("定制笔记本整体价值", 1, 1, "core_answer", "template"), ("为什么选择我们", 2, 4, "highlight", "template"), ("材质选择卖点", 5, 8, "core_answer", "template"), ("工艺细节卖点", 9, 12, "core_answer", "template"), ("创意组合方案", 13, 17, "core_answer", "template"), ("行动引导与合作期待", 18, 20, "cta", "template")],
    28: [("轨交项目案例与交付结果", 1, 1, "highlight", "example"), ("专业领域与交付能力", 2, 2, "core_answer", "template"), ("专业语料积累", 3, 6, "highlight", "example"), ("轨交文档类型示例", 7, 14, "highlight", "example"), ("翻译支持邀约", 15, 16, "cta", "template")],
}


def _normalize_email_import_line(line: str) -> str:
    text = str(line or "").strip().replace("\u3000", " ")
    if not text:
        return ""
    if text.startswith("?") or text.startswith("？"):
        text = "○" + text[1:]
    text = re.sub(r"^[\-\*\u2022\u25cf\u25e6\u30fb]\s*", "○", text)
    text = re.sub(r"^[tC]\s*(?=[\u4e00-\u9fffA-Za-z0-9])", "○", text)
    return text


def _normalize_email_import_text(value: str) -> str:
    text = str(value or "").replace("_x000D_", "\n").replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(line for line in (_normalize_email_import_line(part) for part in text.split("\n")) if line).strip()


def _is_dev_email_template_record(record: dict) -> bool:
    return str(record.get("use_range_label") or "").strip() == "开发推广触达" and str(record.get("intent_type_label") or "").strip() == "售前触达"


def _dev_email_seq_no(row_no: int) -> int:
    return DEV_EMAIL_ROW_SEQUENCE.index(row_no) + 1


def _dev_email_doc_title(row_no: int) -> str:
    return f"开发邮件{_dev_email_seq_no(row_no):02d}·{DEV_EMAIL_ROW_META[row_no]['name']}"


def _email_chunk_priority(function_fragment: str, chunk_type: str) -> int:
    if chunk_type == "example":
        return 82
    return {"core_answer": 80, "highlight": 78, "cta": 74, "background": 72}.get(function_fragment, 70)


def _email_import_join_lines(lines: list[str], start: int, end: int) -> str:
    return "\n".join(lines[start - 1:end]).strip()


def _make_email_chunk(title: str, content: str, *, function_fragment: str, chunk_type: str, business_line: str, service_scope: str) -> dict:
    return {
        "title": title,
        "content": content.strip(),
        "function_fragment": function_fragment,
        "chunk_type": chunk_type,
        "business_line": business_line,
        "service_scope": service_scope,
    }


def _split_dev_email_row_3(content: str, *, business_line: str, service_scope: str) -> list[dict]:
    text = content.replace("''", "'")
    anchors = [
        "This is Joyce Sheng form SPEED.",
        "With the Chinaplas exhibition in Shanghai coming up this April,",
        "Beyond our past collaborations in printing and translation, our team also offers a range of other services that you may not have experienced before:",
        "Event Planning & Execution:",
        "If any of these services align with your current needs,",
    ]
    try:
        a1 = text.index(anchors[0]) + len(anchors[0])
        a2 = text.index(anchors[1])
        a3 = text.index(anchors[2])
        a4 = text.index(anchors[3])
        a5 = text.index(anchors[4])
    except ValueError:
        payloads = _split_email_into_function_fragments("development_email_01", text)
        return [
            _make_email_chunk(item["title"], item["content"], function_fragment=item["function_fragment"], chunk_type="template", business_line=business_line, service_scope=service_scope)
            for item in payloads
        ]
    return [
        _make_email_chunk("Greeting and Reconnection", text[:a1], function_fragment="background", chunk_type="template", business_line=business_line, service_scope=service_scope),
        _make_email_chunk("Past Collaboration and New Location", text[a1:a2], function_fragment="highlight", chunk_type="template", business_line=business_line, service_scope=service_scope),
        _make_email_chunk("Chinaplas Support Inquiry", text[a2:a3], function_fragment="cta", chunk_type="template", business_line=business_line, service_scope=service_scope),
        _make_email_chunk("Expanded Service Portfolio", text[a3:a4], function_fragment="core_answer", chunk_type="template", business_line=business_line, service_scope=service_scope),
        _make_email_chunk("Event and Gift Support", text[a4:a5], function_fragment="core_answer", chunk_type="template", business_line=business_line, service_scope=service_scope),
        _make_email_chunk("Call to Action", text[a5:], function_fragment="cta", chunk_type="template", business_line=business_line, service_scope=service_scope),
    ]


def _build_email_excel_document_title(record: dict, *, fallback_index: int) -> str:
    row_no = int(record.get("row") or fallback_index)
    if _is_dev_email_template_record(record) and row_no in DEV_EMAIL_ROW_META:
        return _dev_email_doc_title(row_no)
    subject = str(record.get("subject") or "").strip()
    if subject:
        return subject[:255]
    return f"邮件模板{fallback_index:02d}"


def _build_historical_email_extract_title(record: dict, *, fallback_index: int) -> str:
    subject = str(record.get("subject") or "").strip()
    if subject:
        return subject[:255]
    row_no = int(record.get("row") or fallback_index)
    use_range = str(record.get("use_range_label") or "").strip()
    intent = str(record.get("intent_type_label") or "").strip()
    parts = [part for part in [use_range, intent] if part]
    if parts:
        return f"历史邮件{row_no:03d}·{'·'.join(parts)}"[:255]
    return f"历史邮件{row_no:03d}"


def _build_email_excel_business_scope(record: dict) -> tuple[str, str]:
    row_no = int(record.get("row") or 0)
    if _is_dev_email_template_record(record) and row_no in DEV_EMAIL_ROW_META:
        meta = DEV_EMAIL_ROW_META[row_no]
        return meta["business_line"], meta["service_scope"]
    return "general", "marketing"


def _build_email_excel_chunks(record: dict, *, fallback_index: int) -> list[dict]:
    row_no = int(record.get("row") or fallback_index)
    content = _normalize_email_import_text(str(record.get("content") or ""))
    business_line, service_scope = _build_email_excel_business_scope(record)
    if _is_dev_email_template_record(record):
        if row_no == 3:
            return _split_dev_email_row_3(content, business_line=business_line, service_scope=service_scope)
        if row_no in DEV_EMAIL_LINE_SPECS:
            lines = [line.strip() for line in content.split("\n") if line.strip()]
            return [
                _make_email_chunk(
                    title,
                    _email_import_join_lines(lines, start, end),
                    function_fragment=function_fragment,
                    chunk_type=chunk_type,
                    business_line=business_line,
                    service_scope=service_scope,
                )
                for title, start, end, function_fragment, chunk_type in DEV_EMAIL_LINE_SPECS[row_no]
            ]
    subject = str(record.get("subject") or _build_email_excel_document_title(record, fallback_index=fallback_index)).strip() or f"email_{fallback_index}"
    payloads = _split_email_into_function_fragments(subject, content)
    chunks: list[dict] = []
    for idx, item in enumerate(payloads, start=1):
        chunks.append(
            _make_email_chunk(
                f"{subject[:40]} / {idx}",
                item["content"],
                function_fragment=item["function_fragment"],
                chunk_type="template" if item["function_fragment"] in {"background", "core_answer", "cta", "greeting", "closing"} else "example",
                business_line=business_line,
                service_scope=service_scope,
            )
        )
    return chunks


def _build_email_excel_doc_source_ref(record: dict, *, filename: str, fallback_index: int) -> str:
    row_no = int(record.get("row") or fallback_index)
    if _is_dev_email_template_record(record):
        return f"dev_email_row_{row_no}"[:255]
    file_fingerprint = hashlib.md5(str(filename or "").strip().lower().encode("utf-8")).hexdigest()[:8]
    return f"email_excel_{file_fingerprint}_row_{row_no}"[:255]


def _merge_stage_labels(values: list[str]) -> list[str]:
    merged: list[str] = []
    for raw in values:
        for part in [item.strip() for item in str(raw or "").split(",") if item and item.strip()]:
            if part not in merged:
                merged.append(part)
    return merged


def _prepare_email_excel_import_entries(records: list[dict], *, filename: str) -> tuple[list[dict], list[dict]]:
    prepared: list[dict] = []
    skipped: list[dict] = []
    seen_by_hash: dict[str, dict] = {}
    for index, original in enumerate(records, start=1):
        record = dict(original)
        record["content"] = _normalize_email_import_text(str(record.get("content") or ""))
        record["subject"] = sanitize_text(str(record.get("subject") or "").strip())
        row_no = int(record.get("row") or index)
        if _is_dev_email_template_record(record) and row_no in DEV_EMAIL_SKIP_ROWS:
            skipped.append({"row": row_no, "reason": "too_short_generic"})
            continue
        min_length = 60 if _is_dev_email_template_record(record) else 15
        if len(record["content"]) < min_length:
            skipped.append({"row": row_no, "reason": "too_short_generic"})
            continue
        fingerprint = hashlib.md5(record["content"].encode("utf-8")).hexdigest()
        if fingerprint in seen_by_hash:
            existing = seen_by_hash[fingerprint]
            existing["source_rows"].append(row_no)
            existing["stage_values"].append(str(record.get("email_stage_label") or "").strip())
            existing["duplicate_rows"].append(row_no)
            skipped.append({"row": row_no, "reason": f"duplicate_of_row_{existing['source_rows'][0]}"})
            continue
        entry = {
            "record": record,
            "source_rows": [row_no],
            "stage_values": [str(record.get("email_stage_label") or "").strip()],
            "duplicate_rows": [],
            "source_ref": _build_email_excel_doc_source_ref(record, filename=filename, fallback_index=index),
            "sequence_no": 0,
        }
        seen_by_hash[fingerprint] = entry
        prepared.append(entry)
    for seq, entry in enumerate(prepared, start=1):
        entry["sequence_no"] = seq
    return prepared, skipped


def _purge_email_excel_history(db: Session, *, filename: str, source_refs: list[str]) -> dict:
    prefixes = tuple(source_refs + [f"{filename}:"])
    doc_count = 0
    chunk_count = 0
    frag_count = 0
    cand_count = 0

    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.source_type == "email_excel").all()
    for doc in docs:
        source_meta = dict(doc.source_meta or {})
        if doc.source_ref in source_refs or str(source_meta.get("source_filename") or "").strip() == filename:
            chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
            for chunk in chunks:
                db.delete(chunk)
                chunk_count += 1
            db.delete(doc)
            doc_count += 1

    fragments = db.query(EmailFragmentAsset).filter(EmailFragmentAsset.source_type == "email_excel").all()
    for fragment in fragments:
        source_snapshot = dict(fragment.source_snapshot or {})
        if any(str(fragment.source_ref or "").startswith(prefix) for prefix in prefixes) or str(source_snapshot.get("source_filename") or "").strip() == filename:
            db.delete(fragment)
            frag_count += 1

    candidates = db.query(KnowledgeCandidate).filter(KnowledgeCandidate.source_type == "email_excel").all()
    for candidate in candidates:
        source_snapshot = dict(candidate.source_snapshot or {})
        if any(str(candidate.source_ref or "").startswith(prefix) for prefix in prefixes) or str(source_snapshot.get("source_filename") or "").strip() == filename:
            db.delete(candidate)
            cand_count += 1

    db.flush()
    return {
        "deleted_documents": doc_count,
        "deleted_chunks": chunk_count,
        "deleted_fragments": frag_count,
        "deleted_candidates": cand_count,
    }


def _create_email_excel_document(
    db: Session,
    *,
    email_asset: EmailThreadAsset,
    entry: dict,
    filename: str,
    import_batch: str,
    owner: str,
) -> tuple[KnowledgeDocument, list[KnowledgeChunk], list[EmailFragmentAsset]]:
    record = dict(entry["record"])
    row_no = int(record.get("row") or entry["sequence_no"])
    source_rows = list(entry["source_rows"])
    stage_labels = _merge_stage_labels(entry["stage_values"])
    document_title = _build_email_excel_document_title(record, fallback_index=entry["sequence_no"])
    if not str(record.get("subject") or "").strip():
        email_asset.subject = document_title[:255]
    chunks_payload = _build_email_excel_chunks(record, fallback_index=entry["sequence_no"])
    business_line, service_scope = _build_email_excel_business_scope(record)
    doc_tags = {
        "knowledge_class": "email_template",
        "scenario_label": email_asset.scenario_label or "development_outreach",
        "intent_label": email_asset.intent_label or "pre_sales_touch",
        "language_style": email_asset.language_style or "formal_email",
        "thread_id": email_asset.thread_id,
        "session_id": email_asset.session_id,
        "source_filename": filename,
        "source_row": row_no,
        "source_rows": source_rows,
        "sequence_no": entry["sequence_no"],
        "use_range_label": record.get("use_range_label"),
        "intent_type_label": record.get("intent_type_label"),
        "email_stage_labels": stage_labels,
        "import_batch": import_batch,
    }
    document = KnowledgeDocument(
        title=document_title,
        knowledge_type="faq",
        business_line=business_line,
        sub_service=None,
        source_type="email_excel",
        source_ref=entry["source_ref"],
        source_meta={
            "source_filename": filename,
            "source_row": row_no,
            "source_rows": source_rows,
            "sequence_no": entry["sequence_no"],
            "raw_labels": {
                "column_b": record.get("use_range_label"),
                "column_c": record.get("intent_type_label"),
                "column_d_list": stage_labels,
            },
            "duplicate_rows": entry["duplicate_rows"],
            "chunk_count": len(chunks_payload),
        },
        status="review",
        owner=owner,
        import_batch=import_batch,
        risk_level="medium",
        review_required=True,
        review_status="in_review",
        library_type=infer_library_type(source_type="email_excel", knowledge_type="faq", chunk_type="template", tags=doc_tags),
        tags=doc_tags,
    )
    db.add(document)
    db.flush()

    created_chunks: list[KnowledgeChunk] = []
    created_fragments: list[EmailFragmentAsset] = []
    for chunk_no, payload in enumerate(chunks_payload, start=1):
        retrieval_title = _strip_retrieval_title_prefix(payload["title"])
        retrieval_text = _build_chunk_retrieval_text(payload["title"], payload["content"])
        embedding = _optional_embedding_for_storage(
            retrieval_text,
            context=f"email_chunk:{chunk_no}:{payload['title']}",
        ) if payload["content"] else None
        chunk = KnowledgeChunk(
            document_id=document.document_id,
            chunk_no=chunk_no,
            chunk_type=payload["chunk_type"],
            title=retrieval_title,
            content=payload["content"],
            keyword_text=retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=_email_chunk_priority(payload["function_fragment"], payload["chunk_type"]),
            retrieval_weight=Decimal("1.000"),
            business_line=payload["business_line"],
            sub_service=None,
            knowledge_type="faq",
            language_pair=None,
            service_scope=payload["service_scope"],
            region=None,
            customer_tier=None,
            structured_tags=merge_tags(doc_tags, function_fragment=payload["function_fragment"], chunk_no=chunk_no, chunk_type=payload["chunk_type"], business_line=payload["business_line"], service_scope=payload["service_scope"]),
            status="review",
        )
        db.add(chunk)
        db.flush()
        quality = _apply_chunk_governance(chunk, source_type="email_excel", source_ref=document.source_ref)
        document.library_type = chunk.library_type
        document.tags = merge_tags(document.tags, library_type=document.library_type)
        created_chunks.append(chunk)

        fragment = EmailFragmentAsset(
            email_id=email_asset.email_id,
            session_id=email_asset.session_id,
            thread_id=email_asset.thread_id,
            source_type="email_excel",
            source_ref=f"dev_email_seq_{entry['sequence_no']:02d}__fragment__{chunk_no}" if _is_dev_email_template_record(record) else f"{entry['source_ref']}__fragment__{chunk_no}",
            title=payload["title"],
            content=payload["content"],
            function_fragment=payload["function_fragment"],
            scenario_label=email_asset.scenario_label,
            intent_label=email_asset.intent_label,
            language_style=email_asset.language_style,
            library_type=quality["library_type"],
            allowed_for_generation=bool(quality["allowed_for_generation"]),
            usable_for_reply=bool(quality["usable_for_reply"]),
            publishable=bool(quality["publishable"]),
            topic_clarity_score=_decimal_score(quality["topic_clarity_score"]),
            completeness_score=_decimal_score(quality["completeness_score"]),
            reusability_score=_decimal_score(quality["reusability_score"]),
            evidence_reliability_score=_decimal_score(quality["evidence_reliability_score"]),
            useful_score=_decimal_score(quality["useful_score"]),
            effect_score=_decimal_score(0.5),
            quality_notes=quality["quality_notes"],
            source_snapshot={
                "source_filename": filename,
                "source_row": row_no,
                "source_rows": source_rows,
                "sequence_no": entry["sequence_no"],
                "tags": chunk.structured_tags,
            },
            status="ready",
        )
        db.add(fragment)
        created_fragments.append(fragment)
    return document, created_chunks, created_fragments


def _extract_historical_email_candidates_for_entry(
    db: Session,
    *,
    entry: dict,
    email_asset: EmailThreadAsset,
    filename: str,
    owner: str,
    operator: str,
    max_candidates: int,
) -> tuple[list[KnowledgeCandidate], list[EmailFragmentAsset]]:
    if max_candidates <= 0:
        return [], []

    record = dict(entry["record"])
    row_no = int(record.get("row") or entry["sequence_no"])
    source_rows = list(entry["source_rows"])
    stage_labels = _merge_stage_labels(entry["stage_values"])
    extract_title = _build_historical_email_extract_title(record, fallback_index=entry["sequence_no"])
    extract_record = {
        **record,
        "title": extract_title,
        "subject": extract_title,
        "thread_id": email_asset.thread_id,
        "session_id": email_asset.session_id,
        "sent_at": email_asset.sent_at_raw or record.get("sent_at"),
        "source_ref": entry["source_ref"],
    }
    candidate_items = _build_candidate_entries_from_record(extract_record, "email_excel")
    created_candidates: list[KnowledgeCandidate] = []
    created_fragments: list[EmailFragmentAsset] = []

    for offset, item in enumerate(candidate_items[:max_candidates], start=1):
        source_snapshot = dict(item.get("source_snapshot") or {})
        source_snapshot.update({
            "source_filename": filename,
            "source_row": row_no,
            "source_rows": source_rows,
            "sequence_no": entry["sequence_no"],
            "email_stage_labels": stage_labels,
            "thread_id": email_asset.thread_id,
            "session_id": email_asset.session_id,
            "source_ref": entry["source_ref"],
        })
        source_snapshot["tags"] = merge_tags(
            source_snapshot.get("tags"),
            source_filename=filename,
            source_row=row_no,
            source_rows=source_rows,
            sequence_no=entry["sequence_no"],
            use_range_label=record.get("use_range_label"),
            intent_type_label=record.get("intent_type_label"),
            email_stage_labels=stage_labels,
            scenario_label=email_asset.scenario_label,
            intent_label=email_asset.intent_label,
            language_style=email_asset.language_style,
            thread_id=email_asset.thread_id,
            session_id=email_asset.session_id,
            email_source_kind="historical_transactional",
        )
        review_parts = [
            f"来源日常往来邮件 row {row_no}",
            f"用途={record.get('use_range_label')}" if record.get("use_range_label") else None,
            f"意图={record.get('intent_type_label')}" if record.get("intent_type_label") else None,
        ]
        candidate = _create_candidate_record(
            db,
            title=item["title"],
            content=item["content"],
            knowledge_type=item["knowledge_type"],
            chunk_type=item["chunk_type"],
            business_line=item["business_line"],
            sub_service=item.get("sub_service"),
            language_pair=item.get("language_pair"),
            service_scope=item.get("service_scope"),
            region=item.get("region"),
            customer_tier=item.get("customer_tier"),
            priority=int(item.get("priority") or 60),
            risk_level=item.get("risk_level") or "medium",
            effective_from=item.get("effective_from"),
            effective_to=item.get("effective_to"),
            pricing_rule=item.get("pricing_rule"),
            source_type="email_excel",
            source_ref=f"{entry['source_ref']}__candidate__{offset}",
            source_snapshot=source_snapshot,
            owner=owner,
            operator=operator,
            review_notes="；".join(part for part in review_parts if part),
        )
        created_candidates.append(candidate)
        fragment = _upsert_email_fragment_asset(
            db,
            source_type="email_excel",
            source_ref=f"{entry['source_ref']}__candidate_fragment__{offset}",
            title=candidate.title,
            content=candidate.content,
            session_id=email_asset.session_id,
            thread_id=email_asset.thread_id,
            email_asset=email_asset,
            candidate=candidate,
            source_snapshot=source_snapshot,
        )
        created_fragments.append(fragment)
    return created_candidates, created_fragments

def _rollup_email_thread_asset_stats(db: Session, email_ids: list[str]) -> None:
    if not email_ids:
        return
    items = db.query(EmailThreadAsset).filter(EmailThreadAsset.email_id.in_(email_ids)).all()
    for item in items:
        fragments = db.query(EmailFragmentAsset).filter(EmailFragmentAsset.email_id == item.email_id).all()
        if not fragments:
            continue
        scores = [float(fragment.effect_score) for fragment in fragments if fragment.effect_score is not None]
        item.effect_score = _decimal_score(sum(scores) / len(scores)) if scores else item.effect_score
        item.feedback_count = sum(int(fragment.feedback_count or 0) for fragment in fragments)
        item.positive_feedback_count = sum(int(fragment.positive_feedback_count or 0) for fragment in fragments)
        item.last_feedback_at = max((fragment.last_feedback_at for fragment in fragments if fragment.last_feedback_at), default=item.last_feedback_at)

def _upsert_email_thread_asset(
    db: Session,
    *,
    record: dict,
    source_type: str,
    source_ref: str,
    import_batch: str,
) -> EmailThreadAsset:
    subject = str(record.get("subject") or "未命名邮件").strip()[:255]
    content = sanitize_text(str(record.get("content") or "").strip())
    session_id = str(record.get("session_id") or record.get("thread_id") or source_ref).strip() or None
    thread_id = str(record.get("thread_id") or session_id or source_ref).strip() or source_ref
    scenario_label, intent_label, language_style = infer_scenario_intent(
        title=subject,
        content=content,
        tags={"thread_id": thread_id, "session_id": session_id},
    )
    governance = score_content_governance(
        title=subject,
        content=content,
        knowledge_type="faq",
        chunk_type="template",
        source_type=source_type,
        tags={"thread_id": thread_id, "session_id": session_id, "scenario_label": scenario_label, "intent_label": intent_label, "language_style": language_style},
        has_source_ref=True,
        metadata={"business_line": "general"},
    )
    business_state = {
        "quotation": "formal_quote",
        "payment": "payment",
        "shipment": "shipment",
        "after_sales": "after_sales",
    }.get(scenario_label, "inquiry")
    thread_fact = None
    if session_id:
        thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
    item = db.query(EmailThreadAsset).filter(EmailThreadAsset.source_ref == source_ref).first()
    if not item:
        item = EmailThreadAsset(source_ref=source_ref, source_type=source_type, thread_id=thread_id, subject=subject, content=content)
        db.add(item)
        db.flush()
    item.source_type = source_type
    item.import_batch = import_batch
    item.session_id = session_id
    item.thread_id = thread_id
    item.external_userid = str(record.get("external_userid") or "").strip() or None
    item.sales_userid = str(record.get("sales_userid") or "").strip() or None
    item.fact_id = thread_fact.fact_id if thread_fact else None
    item.subject = subject
    item.content = content
    item.sender = str(record.get("sender") or "").strip() or None
    item.receiver = str(record.get("receiver") or "").strip() or None
    item.sent_at_raw = str(record.get("sent_at") or "").strip() or None
    item.sent_at = _parse_optional_datetime(record.get("sent_at"))
    item.scenario_label = scenario_label
    item.intent_label = intent_label
    item.language_style = language_style
    item.business_state = business_state
    item.library_type = "fact"
    item.quality_score = _decimal_score(governance["useful_score"])
    item.effect_score = item.effect_score if item.effect_score is not None else _decimal_score(0.5)
    item.usable_for_reply = bool(float(governance["useful_score"]) >= 0.45)
    item.allowed_for_generation = False
    item.publishable = False
    item.source_snapshot = _json_safe(record)
    item.status = "ingested"
    return item

def _upsert_email_fragment_asset(
    db: Session,
    *,
    source_type: str,
    source_ref: str,
    title: str,
    content: str,
    session_id: str | None = None,
    thread_id: str | None = None,
    email_asset: EmailThreadAsset | None = None,
    candidate: KnowledgeCandidate | None = None,
    log_id: str | None = None,
    source_snapshot: dict | None = None,
) -> EmailFragmentAsset:
    raw_tags = {}
    if email_asset:
        raw_tags.update({
            "thread_id": email_asset.thread_id,
            "session_id": email_asset.session_id,
            "scenario_label": email_asset.scenario_label,
            "intent_label": email_asset.intent_label,
            "language_style": email_asset.language_style,
        })
    if candidate and isinstance(candidate.source_snapshot, dict):
        raw_tags.update((candidate.source_snapshot.get("tags") or {}))
    raw_tags.update((source_snapshot or {}).get("tags") or {})
    if session_id:
        raw_tags["session_id"] = session_id
    if thread_id:
        raw_tags["thread_id"] = thread_id
    fragment_kind = infer_function_fragment(title=title, content=content, source_type=source_type, tags=raw_tags)
    scenario_label, intent_label, language_style = infer_scenario_intent(title=title, content=content, tags=raw_tags)
    governance = score_content_governance(
        title=title,
        content=content,
        knowledge_type=candidate.knowledge_type if candidate else "faq",
        chunk_type=candidate.chunk_type if candidate else ("template" if fragment_kind in {"greeting", "cta", "closing", "core_answer"} else "example"),
        source_type=source_type,
        tags=merge_tags(raw_tags, scenario_label=scenario_label, intent_label=intent_label, language_style=language_style, function_fragment=fragment_kind),
        has_source_ref=True,
        metadata={
            "business_line": candidate.business_line if candidate else "general",
            "language_pair": candidate.language_pair if candidate else None,
            "service_scope": candidate.service_scope if candidate else None,
            "customer_tier": candidate.customer_tier if candidate else None,
        },
    )
    item = db.query(EmailFragmentAsset).filter(
        EmailFragmentAsset.source_type == source_type,
        EmailFragmentAsset.source_ref == source_ref,
    ).first()
    if not item:
        item = EmailFragmentAsset(source_type=source_type, source_ref=source_ref, thread_id=thread_id or (email_asset.thread_id if email_asset else source_ref), title=title, content=content)
        db.add(item)
        db.flush()
    item.email_id = email_asset.email_id if email_asset else item.email_id
    item.candidate_id = candidate.candidate_id if candidate else item.candidate_id
    item.log_id = log_id or item.log_id
    item.session_id = session_id or (email_asset.session_id if email_asset else None)
    item.thread_id = thread_id or (email_asset.thread_id if email_asset else item.thread_id)
    item.title = title[:255]
    item.content = content
    item.function_fragment = fragment_kind
    item.scenario_label = scenario_label
    item.intent_label = intent_label
    item.language_style = language_style
    item.library_type = governance["library_type"]
    item.allowed_for_generation = candidate.allowed_for_generation if candidate and candidate.allowed_for_generation is not None else governance["allowed_for_generation"]
    item.usable_for_reply = candidate.usable_for_reply if candidate and candidate.usable_for_reply is not None else governance["usable_for_reply"]
    item.publishable = candidate.publishable if candidate and candidate.publishable is not None else governance["publishable"]
    item.topic_clarity_score = candidate.topic_clarity_score if candidate and candidate.topic_clarity_score is not None else _decimal_score(governance["topic_clarity_score"])
    item.completeness_score = candidate.completeness_score if candidate and candidate.completeness_score is not None else _decimal_score(governance["completeness_score"])
    item.reusability_score = candidate.reusability_score if candidate and candidate.reusability_score is not None else _decimal_score(governance["reusability_score"])
    item.evidence_reliability_score = candidate.evidence_reliability_score if candidate and candidate.evidence_reliability_score is not None else _decimal_score(governance["evidence_reliability_score"])
    item.useful_score = candidate.useful_score if candidate and candidate.useful_score is not None else _decimal_score(governance["useful_score"])
    item.effect_score = candidate.effect_score if candidate and candidate.effect_score is not None else (item.effect_score if item.effect_score is not None else _decimal_score(0.5))
    item.feedback_count = candidate.feedback_count if candidate and candidate.feedback_count is not None else int(item.feedback_count or 0)
    item.positive_feedback_count = candidate.positive_feedback_count if candidate and candidate.positive_feedback_count is not None else int(item.positive_feedback_count or 0)
    item.last_feedback_at = candidate.last_feedback_at if candidate and candidate.last_feedback_at else item.last_feedback_at
    item.quality_notes = candidate.quality_notes if candidate and candidate.quality_notes else governance["quality_notes"]
    merged_snapshot = dict(source_snapshot or {})
    if candidate:
        merged_snapshot.setdefault("candidate_id", str(candidate.candidate_id))
    if email_asset:
        merged_snapshot.setdefault("email_id", str(email_asset.email_id))
    item.source_snapshot = _json_safe(merged_snapshot)
    item.status = "ready"
    return item

def _extract_email_assets_and_candidates(
    db: Session,
    *,
    records: list[dict],
    filename: str,
    owner: str,
    operator: str,
    max_candidates: int,
) -> dict:
    import_batch = f"email_excel_{uuid.uuid4().hex[:12]}"
    prepared_entries, skipped = _prepare_email_excel_import_entries(records, filename=filename)
    purge_summary = _purge_email_excel_history(db, filename=filename, source_refs=[item["source_ref"] for item in prepared_entries])

    created_assets: list[EmailThreadAsset] = []
    created_documents: list[KnowledgeDocument] = []
    created_chunks: list[KnowledgeChunk] = []
    created_fragments: list[EmailFragmentAsset] = []
    created_candidates: list[KnowledgeCandidate] = []
    seen_asset_ids: set[str] = set()
    remaining_candidate_slots = max_candidates

    for entry in prepared_entries:
        record = dict(entry["record"])
        record.setdefault("source_ref", entry["source_ref"])
        if not record.get("subject"):
            if _is_dev_email_template_record(record):
                record["subject"] = _build_email_excel_document_title(record, fallback_index=entry["sequence_no"])
            else:
                record["subject"] = _build_historical_email_extract_title(record, fallback_index=entry["sequence_no"])
        email_asset = _upsert_email_thread_asset(
            db,
            record=record,
            source_type="email_excel",
            source_ref=entry["source_ref"],
            import_batch=import_batch,
        )
        if str(email_asset.email_id) not in seen_asset_ids:
            created_assets.append(email_asset)
            seen_asset_ids.add(str(email_asset.email_id))

        if _is_dev_email_template_record(record):
            document, chunks, fragments = _create_email_excel_document(
                db,
                email_asset=email_asset,
                entry=entry,
                filename=filename,
                import_batch=import_batch,
                owner=owner,
            )
            created_documents.append(document)
            created_chunks.extend(chunks)
            created_fragments.extend(fragments)
            continue

        if remaining_candidate_slots > 0:
            candidates, fragments = _extract_historical_email_candidates_for_entry(
                db,
                entry=entry,
                email_asset=email_asset,
                filename=filename,
                owner=owner,
                operator=operator,
                max_candidates=remaining_candidate_slots,
            )
            created_candidates.extend(candidates)
            created_fragments.extend(fragments)
            remaining_candidate_slots = max(0, remaining_candidate_slots - len(candidates))
    return {
        "import_batch": import_batch,
        "email_assets": created_assets,
        "documents": created_documents,
        "chunks": created_chunks,
        "fragments": created_fragments,
        "candidates": created_candidates,
        "skipped": skipped,
        "purged": purge_summary,
    }

def _rebuild_candidate_governance(db: Session, *, source_type: str | None = None) -> dict:
    query = db.query(KnowledgeCandidate)
    if source_type:
        query = query.filter(KnowledgeCandidate.source_type == source_type)
    items = query.all()
    updated = 0
    for item in items:
        _apply_candidate_governance(item)
        updated += 1
    return {"updated": updated}


def _build_email_record_from_candidate(candidate: KnowledgeCandidate) -> dict:
    snapshot = dict(candidate.source_snapshot or {})
    content = sanitize_text(str(snapshot.get("content") or candidate.content or "").strip())
    subject = str(snapshot.get("subject") or candidate.title or "历史邮件资产").strip()[:255]
    thread_id = str(snapshot.get("thread_id") or snapshot.get("session_id") or _email_asset_source_ref_from_candidate_source_ref(candidate.source_ref)).strip()
    session_id = str(snapshot.get("session_id") or "").strip() or None
    return {
        "subject": subject,
        "content": content,
        "thread_id": thread_id,
        "session_id": session_id,
        "external_userid": snapshot.get("external_userid"),
        "sales_userid": snapshot.get("sales_userid"),
        "sender": snapshot.get("sender"),
        "receiver": snapshot.get("receiver"),
        "sent_at": snapshot.get("sent_at"),
        "attachment_names": snapshot.get("attachment_names"),
        "attachment_summary": snapshot.get("attachment_summary"),
        "attachment_content": snapshot.get("attachment_content"),
        "attachment_time": snapshot.get("attachment_time"),
    }


def _attachment_payloads_from_snapshot(snapshot: dict | None) -> list[dict[str, Any]]:
    snapshot = dict(snapshot or {})
    names = snapshot.get("attachment_names") or []
    if isinstance(names, str):
        names = [item.strip() for item in re.split(r"[;,；，]", names) if item.strip()]
    attachments: list[dict[str, Any]] = []
    text_payload = snapshot.get("attachment_content") or snapshot.get("attachment_summary")
    if names:
        for name in names:
            attachments.append(
                {
                    "filename": str(name).strip()[:120],
                    "summary": text_payload,
                    "file_time": snapshot.get("attachment_time"),
                }
            )
    elif text_payload:
        attachments.append(
            {
                "summary": text_payload,
                "file_time": snapshot.get("attachment_time"),
            }
        )
    return attachments


def _build_summary_from_email_assets(
    email_assets: list[EmailThreadAsset],
    email_fragments: list[EmailFragmentAsset],
) -> dict | None:
    if not email_assets and not email_fragments:
        return None
    latest_asset = email_assets[0] if email_assets else None
    fragment_types = sorted({item.function_fragment for item in email_fragments if item.function_fragment})
    key_facts: dict[str, Any] = {}
    if latest_asset:
        key_facts["latest_email_subject"] = latest_asset.subject
    if fragment_types:
        key_facts["email_fragment_types"] = fragment_types
    attachment_names: list[str] = []
    for asset in email_assets:
        for attachment in _attachment_payloads_from_snapshot(asset.source_snapshot):
            filename = str(attachment.get("filename") or "").strip()
            if filename and filename not in attachment_names:
                attachment_names.append(filename)
    if attachment_names:
        key_facts["attachment_names"] = attachment_names[:10]
    demand_parts = [asset.subject for asset in email_assets if asset.subject][:3]
    fragment_samples = [sanitize_text((fragment.content or "")[:120]) for fragment in email_fragments[:3] if fragment.content]
    if fragment_samples:
        key_facts["fragment_samples"] = fragment_samples
    return {
        "topic": latest_asset.subject if latest_asset else "email_thread",
        "core_demand": " / ".join(demand_parts) if demand_parts else None,
        "key_facts": key_facts,
        "todo_items": None,
        "risks": None,
        "to_be_confirmed": None,
        "status": "email_backfill",
    }


def _collect_thread_fact_session_ids(db: Session, session_ids: list[str] | None = None) -> list[str]:
    if session_ids:
        return sorted({str(item).strip() for item in session_ids if str(item).strip()})

    collected: set[str] = set()
    for value, in db.query(KnowledgeHitLog.session_id).filter(KnowledgeHitLog.session_id.isnot(None)).distinct().all():
        if value:
            collected.add(str(value).strip())
    for value, in db.query(IntentSummary.user_id).filter(IntentSummary.user_id.isnot(None)).distinct().all():
        if value:
            collected.add(str(value).strip())
    for value, in db.query(MessageLog.user_id).filter(MessageLog.user_id.isnot(None)).distinct().all():
        if value:
            collected.add(str(value).strip())
    for value, in db.query(EmailThreadAsset.session_id).filter(EmailThreadAsset.session_id.isnot(None)).distinct().all():
        if value:
            collected.add(str(value).strip())
    for value, in db.query(EmailThreadAsset.thread_id).filter(EmailThreadAsset.thread_id.isnot(None)).distinct().all():
        if value:
            collected.add(str(value).strip())
    return sorted(collected)


def _backfill_email_assets_from_candidates(db: Session, *, source_type: str = "email_excel") -> dict:
    groups: dict[str, list[KnowledgeCandidate]] = {}
    candidates = db.query(KnowledgeCandidate).filter(KnowledgeCandidate.source_type == source_type).order_by(KnowledgeCandidate.created_at.asc()).all()
    for candidate in candidates:
        asset_ref = _email_asset_source_ref_from_candidate_source_ref(candidate.source_ref)
        groups.setdefault(asset_ref, []).append(candidate)

    created_assets = 0
    created_fragments = 0
    for asset_ref, grouped_candidates in groups.items():
        first = grouped_candidates[0]
        email_asset = _upsert_email_thread_asset(
            db,
            record=_build_email_record_from_candidate(first),
            source_type=source_type,
            source_ref=asset_ref,
            import_batch="historical_email_candidate_backfill",
        )
        created_assets += 1

        raw_fragments = _split_email_into_function_fragments(email_asset.subject, email_asset.content)
        for index, fragment_payload in enumerate(raw_fragments, start=1):
            _upsert_email_fragment_asset(
                db,
                source_type=source_type,
                source_ref=f"{asset_ref}__fragment__{index}",
                title=fragment_payload["title"],
                content=fragment_payload["content"],
                session_id=email_asset.session_id,
                thread_id=email_asset.thread_id,
                email_asset=email_asset,
                source_snapshot={
                    "email_asset_ref": asset_ref,
                    "tags": {
                        "thread_id": email_asset.thread_id,
                        "session_id": email_asset.session_id,
                        "scenario_label": email_asset.scenario_label,
                        "intent_label": email_asset.intent_label,
                        "language_style": email_asset.language_style,
                        "function_fragment": fragment_payload["function_fragment"],
                    },
                },
            )
            created_fragments += 1

    return {
        "candidate_groups": len(groups),
        "assets_touched": created_assets,
        "raw_fragments_touched": created_fragments,
    }


def _backfill_thread_business_facts(db: Session, *, session_ids: list[str] | None = None) -> dict:
    candidate_session_ids = _collect_thread_fact_session_ids(db, session_ids)
    touched = 0
    for session_id in candidate_session_ids:
        if not session_id:
            continue
        summary_model = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
        summary_json = None
        if summary_model:
            summary_json = {
                "topic": summary_model.topic,
                "core_demand": summary_model.core_demand,
                "key_facts": summary_model.key_facts,
                "todo_items": summary_model.todo_items,
                "risks": summary_model.risks,
                "to_be_confirmed": summary_model.to_be_confirmed,
                "status": summary_model.status,
            }
        email_assets = db.query(EmailThreadAsset).filter(
            or_(EmailThreadAsset.session_id == session_id, EmailThreadAsset.thread_id == session_id)
        ).order_by(EmailThreadAsset.created_at.desc()).all()
        email_fragments = db.query(EmailFragmentAsset).filter(
            or_(EmailFragmentAsset.session_id == session_id, EmailFragmentAsset.thread_id == session_id)
        ).order_by(EmailFragmentAsset.created_at.desc()).all()
        if not summary_json:
            summary_json = _build_summary_from_email_assets(email_assets, email_fragments)
        messages = db.query(MessageLog).filter(MessageLog.user_id == session_id).order_by(MessageLog.timestamp.asc()).all()
        payload_messages = [
            {
                "content": item.content,
                "sender_type": item.sender_type,
                "timestamp": item.timestamp.isoformat() if item.timestamp else None,
            }
            for item in messages
        ]
        for asset in email_assets[:5]:
            payload_messages.append(
                {
                    "content": asset.content,
                    "sender_type": "email",
                    "timestamp": asset.sent_at.isoformat() if asset.sent_at else None,
                    "attachments": _attachment_payloads_from_snapshot(asset.source_snapshot),
                }
            )
        _upsert_thread_business_fact(
            db,
            session_id=session_id,
            summary_json=summary_json,
            crm_context={},
            messages=payload_messages,
            external_userid=extract_external_userid(session_id),
        )
        touched += 1
    return {"touched": touched}

def _sync_email_effect_rollup_for_log(
    db: Session,
    *,
    log: KnowledgeHitLog,
    feedback_status: str,
    manual_feedback: dict | None = None,
) -> None:
    delta = _feedback_delta(feedback_status)
    conditions = []
    if log.session_id:
        conditions.append(EmailFragmentAsset.session_id == log.session_id)
    if log.log_id:
        conditions.append(EmailFragmentAsset.log_id == log.log_id)
    fragments: list[EmailFragmentAsset] = []
    if conditions:
        fragments = db.query(EmailFragmentAsset).filter(or_(*conditions)).all()
    now = datetime.utcnow()
    touched_email_ids: list[str] = []
    for fragment in fragments:
        current = float(fragment.effect_score) if fragment.effect_score is not None else 0.5
        fragment.feedback_count = int(fragment.feedback_count or 0) + 1
        if feedback_status in POSITIVE_FEEDBACK_STATUSES:
            fragment.positive_feedback_count = int(fragment.positive_feedback_count or 0) + 1
        fragment.effect_score = _decimal_score(max(0.0, min(1.0, current + delta)))
        fragment.last_feedback_at = now
        db.add(EmailEffectFeedback(
            email_id=fragment.email_id,
            fragment_id=fragment.fragment_id,
            candidate_id=fragment.candidate_id,
            log_id=log.log_id,
            session_id=log.session_id,
            thread_id=fragment.thread_id,
            feedback_status=feedback_status,
            feedback_note=(manual_feedback or {}).get("note") if isinstance(manual_feedback, dict) else None,
            feedback_payload=_json_safe(manual_feedback or {}),
            delta_score=_decimal_score(delta) if delta else _decimal_score(0.0),
        ))
        if fragment.email_id:
            touched_email_ids.append(str(fragment.email_id))
    if not fragments and log.session_id:
        assets = db.query(EmailThreadAsset).filter(EmailThreadAsset.session_id == log.session_id).all()
        for asset in assets:
            current = float(asset.effect_score) if asset.effect_score is not None else 0.5
            asset.feedback_count = int(asset.feedback_count or 0) + 1
            if feedback_status in POSITIVE_FEEDBACK_STATUSES:
                asset.positive_feedback_count = int(asset.positive_feedback_count or 0) + 1
            asset.effect_score = _decimal_score(max(0.0, min(1.0, current + delta)))
            asset.last_feedback_at = now
            db.add(EmailEffectFeedback(
                email_id=asset.email_id,
                log_id=log.log_id,
                session_id=log.session_id,
                thread_id=asset.thread_id,
                feedback_status=feedback_status,
                feedback_note=(manual_feedback or {}).get("note") if isinstance(manual_feedback, dict) else None,
                feedback_payload=_json_safe(manual_feedback or {}),
                delta_score=_decimal_score(delta) if delta else _decimal_score(0.0),
            ))
            touched_email_ids.append(str(asset.email_id))
    _rollup_email_thread_asset_stats(db, touched_email_ids)

def _normalize_training_sample_types(sample_types: list[str] | None) -> list[str]:
    normalized = []
    for item in sample_types or []:
        value = str(item or "").strip()
        if value in TRAINING_SAMPLE_TYPES and value not in normalized:
            normalized.append(value)
    return normalized or sorted(TRAINING_SAMPLE_TYPES)

def _build_training_sample_key(*parts: Any) -> str:
    payload = json.dumps([_json_safe(part) for part in parts], ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()

def _upsert_training_sample(
    db: Session,
    *,
    sample_type: str,
    source_table: str,
    source_id: str | None,
    source_type: str | None,
    source_ref: str | None,
    instruction: str | None,
    input_text: str,
    target_text: str | None,
    sample_metadata: dict | None,
    quality_score: float | None,
    effect_score: float | None,
    exportable: bool,
) -> tuple[ModelTrainingSample, bool]:
    sample_key = _build_training_sample_key(sample_type, source_table, source_id, source_ref, instruction, input_text, target_text)
    item = db.query(ModelTrainingSample).filter(ModelTrainingSample.sample_key == sample_key).first()
    created = False
    if not item:
        item = ModelTrainingSample(sample_key=sample_key)
        db.add(item)
        created = True
    item.sample_type = sample_type
    item.source_table = source_table
    item.source_id = source_id
    item.source_type = source_type
    item.source_ref = source_ref
    item.instruction = instruction
    item.input_text = input_text
    item.target_text = target_text
    item.sample_metadata = _json_safe(sample_metadata or {})
    item.quality_score = _decimal_score(quality_score) if quality_score is not None else None
    item.effect_score = _decimal_score(effect_score) if effect_score is not None else None
    item.exportable = exportable
    item.review_status = "ready" if exportable else "hold"
    if item.export_status == "exported":
        item.export_status = "stale"
    elif item.export_status not in {"pending", "stale"}:
        item.export_status = "pending"
    db.flush()
    return item, created

def _prepare_training_samples(db: Session, payload: TrainingSamplePrepareRequest) -> dict:
    sample_types = _normalize_training_sample_types(payload.sample_types)
    limit = max(1, min(int(payload.limit_per_source or 200), 1000))
    min_quality = max(0.0, min(float(payload.min_quality_score), 1.0))
    min_effect = max(0.0, min(float(payload.min_effect_score), 1.0))
    stats = {sample_type: {"created": 0, "updated": 0} for sample_type in sample_types}

    if "embedding_corpus" in sample_types:
        chunks = db.query(KnowledgeChunk).filter(
            KnowledgeChunk.status == "active",
            KnowledgeChunk.publishable.is_(True),
            KnowledgeChunk.allowed_for_generation.is_(True),
        ).order_by(KnowledgeChunk.useful_score.desc(), KnowledgeChunk.effect_score.desc()).limit(limit).all()
        for chunk in chunks:
            quality = float(chunk.useful_score) if chunk.useful_score is not None else 0.0
            effect = float(chunk.effect_score) if chunk.effect_score is not None else 0.5
            if quality < min_quality:
                continue
            sample, created = _upsert_training_sample(
                db,
                sample_type="embedding_corpus",
                source_table="knowledge_chunk",
                source_id=str(chunk.chunk_id),
                source_type=chunk.library_type,
                source_ref=str(chunk.document_id),
                instruction=None,
                input_text=f"{chunk.title}\n{chunk.content}".strip(),
                target_text=None,
                sample_metadata={
                    "knowledge_type": chunk.knowledge_type,
                    "chunk_type": chunk.chunk_type,
                    "business_line": chunk.business_line,
                    "structured_tags": chunk.structured_tags,
                },
                quality_score=quality,
                effect_score=effect,
                exportable=True,
            )
            stats["embedding_corpus"]["created" if created else "updated"] += 1

    if "reply_fragment_sft" in sample_types:
        fragments = db.query(EmailFragmentAsset).filter(
            EmailFragmentAsset.publishable.is_(True),
            EmailFragmentAsset.allowed_for_generation.is_(True),
            EmailFragmentAsset.usable_for_reply.is_(True),
        ).order_by(EmailFragmentAsset.useful_score.desc(), EmailFragmentAsset.effect_score.desc()).limit(limit).all()
        for fragment in fragments:
            quality = float(fragment.useful_score) if fragment.useful_score is not None else 0.0
            effect = float(fragment.effect_score) if fragment.effect_score is not None else 0.5
            if quality < min_quality or effect < min_effect:
                continue
            sample, created = _upsert_training_sample(
                db,
                sample_type="reply_fragment_sft",
                source_table="email_fragment_asset",
                source_id=str(fragment.fragment_id),
                source_type=fragment.source_type,
                source_ref=fragment.source_ref,
                instruction="根据业务场景、意图、风格和片段类型输出一个可直接复用的回复片段。",
                input_text=(
                    f"场景:{fragment.scenario_label or 'general'}\n"
                    f"意图:{fragment.intent_label or 'general'}\n"
                    f"风格:{fragment.language_style or 'general'}\n"
                    f"片段:{fragment.function_fragment or 'core_answer'}"
                ),
                target_text=fragment.content,
                sample_metadata={
                    "title": fragment.title,
                    "thread_id": fragment.thread_id,
                    "session_id": fragment.session_id,
                    "quality_notes": fragment.quality_notes,
                },
                quality_score=quality,
                effect_score=effect,
                exportable=True,
            )
            stats["reply_fragment_sft"]["created" if created else "updated"] += 1

    positive_logs = []
    if {"thread_reply_sft", "retrieval_pair"} & set(sample_types):
        positive_logs = db.query(KnowledgeHitLog).filter(
            KnowledgeHitLog.feedback_status.in_(tuple(POSITIVE_FEEDBACK_STATUSES)),
            KnowledgeHitLog.final_response.isnot(None),
        ).order_by(KnowledgeHitLog.created_at.desc()).limit(limit).all()

    if "thread_reply_sft" in sample_types:
        for log in positive_logs:
            thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == log.session_id).first() if log.session_id else None
            quality = float(thread_fact.quality_score) if thread_fact and thread_fact.quality_score is not None else 0.6
            effect = 0.8 if log.feedback_status in {"won", "advanced"} else 0.7
            if quality < min_quality or effect < min_effect:
                continue
            sample, created = _upsert_training_sample(
                db,
                sample_type="thread_reply_sft",
                source_table="knowledge_hit_logs",
                source_id=str(log.log_id),
                source_type="assist_feedback",
                source_ref=log.session_id,
                instruction="根据客户问题、线程业务事实和当前阶段生成安全、可执行的销售回复。",
                input_text=(
                    f"客户问题:{log.query_text}\n"
                    f"线程事实:{json.dumps(_json_safe(thread_fact.merged_facts if thread_fact else {}), ensure_ascii=False)}\n"
                    f"当前阶段:{thread_fact.business_state if thread_fact else 'unknown'}"
                ),
                target_text=log.final_response,
                sample_metadata={
                    "feedback_status": log.feedback_status,
                    "thread_fact_id": str(thread_fact.fact_id) if thread_fact else None,
                    "hit_chunk_ids": log.hit_chunk_ids,
                },
                quality_score=quality,
                effect_score=effect,
                exportable=True,
            )
            stats["thread_reply_sft"]["created" if created else "updated"] += 1

    if "retrieval_pair" in sample_types:
        for log in positive_logs:
            chunk_ids = [item for item in (log.hit_chunk_ids or []) if item]
            if not chunk_ids:
                continue
            chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id == chunk_ids[0]).first()
            if not chunk:
                continue
            quality = float(chunk.useful_score) if chunk.useful_score is not None else 0.0
            effect = float(chunk.effect_score) if chunk.effect_score is not None else 0.7
            if quality < min_quality or effect < min_effect:
                continue
            sample, created = _upsert_training_sample(
                db,
                sample_type="retrieval_pair",
                source_table="knowledge_hit_logs",
                source_id=str(log.log_id),
                source_type="retrieval_feedback",
                source_ref=log.session_id,
                instruction="为检索训练提供高质量 query-positive pair。",
                input_text=log.query_text,
                target_text=f"{chunk.title}\n{chunk.content}".strip(),
                sample_metadata={
                    "chunk_id": str(chunk.chunk_id),
                    "feedback_status": log.feedback_status,
                    "business_line": chunk.business_line,
                    "knowledge_type": chunk.knowledge_type,
                },
                quality_score=quality,
                effect_score=effect,
                exportable=True,
            )
            stats["retrieval_pair"]["created" if created else "updated"] += 1

    return {
        "sample_types": sample_types,
        "min_quality_score": min_quality,
        "min_effect_score": min_effect,
        "limit_per_source": limit,
        "stats": stats,
    }

def _training_sample_export_line(item: ModelTrainingSample) -> dict:
    if item.sample_type == "embedding_corpus":
        return {
            "id": item.sample_key,
            "text": item.input_text,
            "metadata": item.sample_metadata,
        }
    if item.sample_type in {"reply_fragment_sft", "thread_reply_sft"}:
        user_text = item.input_text
        if item.instruction:
            user_text = f"{item.instruction}\n\n{item.input_text}"
        return {
            "id": item.sample_key,
            "messages": [
                {"role": "system", "content": "你是企业销售回复训练数据整理助手。"},
                {"role": "user", "content": user_text},
                {"role": "assistant", "content": item.target_text or ""},
            ],
            "metadata": item.sample_metadata,
        }
    return {
        "id": item.sample_key,
        "query": item.input_text,
        "positive_passage": item.target_text,
        "metadata": item.sample_metadata,
    }

def _export_training_samples(db: Session, payload: TrainingSampleExportRequest) -> dict:
    sample_types = _normalize_training_sample_types(payload.sample_types)
    max_samples = max(1, min(int(payload.max_samples or 1000), 5000))
    run_id = f"train_export_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    export_dir = os.path.join(os.getcwd(), "scratch", "training_exports", run_id)
    os.makedirs(export_dir, exist_ok=True)
    files = []
    exported_counts = {}
    now = datetime.utcnow()
    for sample_type in sample_types:
        items = db.query(ModelTrainingSample).filter(
            ModelTrainingSample.sample_type == sample_type,
            ModelTrainingSample.exportable.is_(True),
            ModelTrainingSample.review_status == "ready",
        ).order_by(ModelTrainingSample.quality_score.desc(), ModelTrainingSample.effect_score.desc(), ModelTrainingSample.created_at.desc()).limit(max_samples).all()
        if not items:
            continue
        filepath = os.path.join(export_dir, f"{sample_type}.jsonl")
        with open(filepath, "w", encoding="utf-8") as fh:
            for item in items:
                fh.write(json.dumps(_training_sample_export_line(item), ensure_ascii=False) + "\n")
                item.export_status = "exported"
                item.last_export_path = filepath
                item.exported_at = now
        files.append({"sample_type": sample_type, "path": filepath, "count": len(items)})
        exported_counts[sample_type] = len(items)
    manifest_path = os.path.join(export_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as fh:
        json.dump(
            {
                "run_id": run_id,
                "created_at": now.isoformat(),
                "files": files,
                "sample_types": sample_types,
            },
            fh,
            ensure_ascii=False,
            indent=2,
        )
    return {
        "run_id": run_id,
        "export_dir": export_dir,
        "manifest_path": manifest_path,
        "files": files,
        "exported_counts": exported_counts,
    }


def _analysis_completion_counts(db: Session) -> dict[str, Any]:
    return {
        "thread_business_fact": db.query(func.count(ThreadBusinessFact.fact_id)).scalar(),
        "email_thread_asset": db.query(func.count(EmailThreadAsset.email_id)).scalar(),
        "email_fragment_asset": db.query(func.count(EmailFragmentAsset.fragment_id)).scalar(),
        "email_effect_feedback": db.query(func.count(EmailEffectFeedback.feedback_id)).scalar(),
        "model_training_sample": db.query(func.count(ModelTrainingSample.sample_id)).scalar(),
        "knowledge_chunk_library_types": {
            str(key): count
            for key, count in db.query(KnowledgeChunk.library_type, func.count(KnowledgeChunk.chunk_id))
            .group_by(KnowledgeChunk.library_type)
            .all()
        },
        "knowledge_candidate_library_types": {
            str(key): count
            for key, count in db.query(KnowledgeCandidate.library_type, func.count(KnowledgeCandidate.candidate_id))
            .group_by(KnowledgeCandidate.library_type)
            .all()
        },
    }


def _run_training_pipeline(db: Session, payload: TrainingExecutionRequest, report: Any = None) -> dict:
    prepare_payload = TrainingSamplePrepareRequest(
        sample_types=payload.sample_types,
        min_quality_score=payload.min_quality_score,
        min_effect_score=payload.min_effect_score,
        limit_per_source=payload.limit_per_source,
        operator=payload.operator,
    )
    if report:
        report(progress=10, summary="prepare training samples")
    prepare_result = _prepare_training_samples(db, prepare_payload)

    export_payload = TrainingSampleExportRequest(
        sample_types=prepare_result["sample_types"],
        max_samples=payload.max_samples,
        operator=payload.operator,
    )
    if report:
        report(progress=45, summary="export training samples")
    export_result = _export_training_samples(db, export_payload)

    runner_command = (payload.runner_command or settings.TRAINING_RUNNER_COMMAND or "").strip()
    runner_workdir = payload.runner_workdir or settings.TRAINING_RUNNER_WORKDIR or os.getcwd()
    timeout_seconds = payload.runner_timeout_seconds or settings.TRAINING_RUNNER_TIMEOUT_SECONDS
    runner_result = {
        "executed": False,
        "command": runner_command,
        "workdir": runner_workdir,
        "timeout_seconds": timeout_seconds,
    }
    if payload.execute_runner and runner_command:
        stdout_path = os.path.join(export_result["export_dir"], "runner.stdout.log")
        stderr_path = os.path.join(export_result["export_dir"], "runner.stderr.log")
        env = os.environ.copy()
        env.update(
            {
                "TRAINING_EXPORT_DIR": export_result["export_dir"],
                "TRAINING_MANIFEST_PATH": export_result["manifest_path"],
                "TRAINING_RUN_ID": export_result["run_id"],
                "TRAINING_SAMPLE_TYPES": ",".join(export_result["exported_counts"].keys()),
            }
        )
        if report:
            report(progress=70, summary="execute training runner")
        with open(stdout_path, "w", encoding="utf-8") as stdout_fh, open(stderr_path, "w", encoding="utf-8") as stderr_fh:
            completed = subprocess.run(
                runner_command,
                cwd=runner_workdir,
                shell=True,
                env=env,
                stdout=stdout_fh,
                stderr=stderr_fh,
                timeout=max(60, int(timeout_seconds)),
                check=False,
            )
        runner_result = {
            "executed": True,
            "command": runner_command,
            "workdir": runner_workdir,
            "timeout_seconds": timeout_seconds,
            "returncode": completed.returncode,
            "stdout_path": stdout_path,
            "stderr_path": stderr_path,
        }
    elif report:
        report(progress=70, summary="skip external training runner")

    if report:
        report(progress=90, summary="training pipeline ready")
    return {
        "prepare": prepare_result,
        "export": export_result,
        "runner": runner_result,
    }


def _run_training_pipeline_job(report: Any, payload_data: dict) -> dict:
    payload = TrainingExecutionRequest(**payload_data)
    db = SessionLocal()
    try:
        result = _run_training_pipeline(db, payload, report=report)
        db.commit()
        return result
    finally:
        db.close()


def _complete_analysis_items(db: Session, payload: AnalysisCompletionRequest, report: Any = None) -> dict:
    summary: dict[str, Any] = {
        "before_counts": _analysis_completion_counts(db),
        "source_type": payload.source_type,
    }

    if payload.rebuild_candidate_governance:
        if report:
            report(progress=10, summary="rebuild candidate governance")
        summary["candidate_governance"] = _rebuild_candidate_governance(db, source_type=payload.source_type)
    if payload.backfill_email_assets:
        if report:
            report(progress=25, summary="backfill email assets")
        summary["email_asset_backfill"] = _backfill_email_assets_from_candidates(db, source_type=payload.source_type)
    if payload.backfill_thread_facts:
        if report:
            report(progress=40, summary="backfill thread facts")
        summary["thread_fact_backfill_before_feedback"] = _backfill_thread_business_facts(db)

    positive_limit = max(0, min(int(payload.bootstrap_positive_feedback_count or 0), 20))
    positive_logs = (
        db.query(KnowledgeHitLog)
        .filter(KnowledgeHitLog.final_response.isnot(None), KnowledgeHitLog.status == "ok")
        .order_by(KnowledgeHitLog.created_at.desc())
        .limit(positive_limit)
        .all()
    )
    touched_sessions = sorted({log.session_id for log in positive_logs if log.session_id})
    feedback_updated = 0
    extracted_reply_fragments = 0
    if positive_logs:
        if report:
            report(progress=55, summary="bootstrap positive feedback")
        for log in positive_logs:
            bootstrap_applied = False
            if log.feedback_status not in POSITIVE_FEEDBACK_STATUSES:
                log.feedback_status = payload.bootstrap_feedback_status
                log.manual_feedback = {
                    "note": "system bootstrap: historical positive feedback backfill",
                    "operator": payload.operator or "analysis_completion",
                }
                _update_effect_rollup_for_log(
                    db,
                    log=log,
                    feedback_status=log.feedback_status,
                    manual_feedback=log.manual_feedback,
                )
                feedback_updated += 1
                bootstrap_applied = True
            if payload.extract_excellent_replies and log.session_id:
                created = _extract_excellent_reply_candidates(
                    db,
                    session_id=log.session_id,
                    owner="system_backfill",
                    operator=payload.operator or "analysis_completion",
                    force=True,
                )
                extracted_reply_fragments += len(created)
                existing_email_feedback = db.query(EmailEffectFeedback.feedback_id).filter(
                    EmailEffectFeedback.log_id == log.log_id
                ).first()
                if bootstrap_applied or not existing_email_feedback:
                    _sync_email_effect_rollup_for_log(
                        db,
                        log=log,
                        feedback_status=log.feedback_status or payload.bootstrap_feedback_status,
                        manual_feedback=log.manual_feedback,
                    )
    summary["positive_feedback_bootstrap"] = {
        "logs_touched": feedback_updated,
        "sessions": touched_sessions,
        "excellent_reply_candidates_created": extracted_reply_fragments,
    }

    if payload.backfill_thread_facts:
        if report:
            report(progress=70, summary="refresh thread facts")
        summary["thread_fact_backfill_after_feedback"] = _backfill_thread_business_facts(db, session_ids=touched_sessions or None)

    if payload.prepare_training_samples:
        if report:
            report(progress=82, summary="prepare training samples")
        prepare_payload = TrainingSamplePrepareRequest(
            sample_types=payload.sample_types,
            min_quality_score=payload.min_quality_score,
            min_effect_score=payload.min_effect_score,
            limit_per_source=payload.limit_per_source,
            operator=payload.operator,
        )
        summary["training_prepare"] = _prepare_training_samples(db, prepare_payload)
    if payload.export_training_samples:
        if report:
            report(progress=92, summary="export training samples")
        export_payload = TrainingSampleExportRequest(
            sample_types=payload.sample_types,
            max_samples=payload.max_samples,
            operator=payload.operator,
        )
        summary["training_export"] = _export_training_samples(db, export_payload)

    summary["after_counts"] = _analysis_completion_counts(db)
    return summary


def _run_complete_analysis_items_job(report: Any, payload_data: dict) -> dict:
    payload = AnalysisCompletionRequest(**payload_data)
    db = SessionLocal()
    try:
        result = _complete_analysis_items(db, payload, report=report)
        db.commit()
        return result
    finally:
        db.close()

def _create_candidate_from_feedback_log(
    db: Session,
    log: KnowledgeHitLog,
    *,
    owner: str | None = None,
    operator: str | None = None,
    note: str | None = None,
    title: str | None = None,
):
    existing = db.query(KnowledgeCandidate).filter(
        KnowledgeCandidate.source_type == "feedback",
        KnowledgeCandidate.source_ref == str(log.log_id),
        KnowledgeCandidate.status == "candidate",
    ).first()
    if existing:
        if note and note not in (existing.review_notes or ""):
            existing.review_notes = (existing.review_notes or "").strip()
            existing.review_notes = f"{existing.review_notes}\n{note}".strip()
        return existing

    query_features = log.query_features or {}
    requested_types = query_features.get("knowledge_type") or []
    knowledge_type = requested_types[0] if requested_types else "faq"
    business_line = query_features.get("business_line") or "general"
    feedback_note = note or ((log.manual_feedback or {}).get("note") if isinstance(log.manual_feedback, dict) else None)
    candidate_title = title or f"销售反馈修正：{(log.query_text or '未命名问题')[:40]}"
    content_parts = [
        f"客户问题：{log.query_text or ''}",
        f"当前建议：{log.final_response or '暂无'}",
    ]
    if feedback_note:
        content_parts.append(f"销售修正意见：{feedback_note}")
    candidate_content = "\n".join(part for part in content_parts if part).strip()
    pricing_rule = infer_pricing_rule_candidate(candidate_title, candidate_content, business_line)
    source_snapshot = {
        "log_id": str(log.log_id),
        "query_text": sanitize_text(log.query_text or ""),
        "query_features": log.query_features,
        "manual_feedback": _redact_value(log.manual_feedback),
        "final_response": sanitize_text(log.final_response or "") if log.final_response else None,
        "hit_chunk_ids": log.hit_chunk_ids,
        "status": log.status,
        "feedback_status": log.feedback_status,
    }
    return _create_candidate_record(
        db,
        title=candidate_title,
        content=candidate_content,
        knowledge_type=knowledge_type if knowledge_type in KB_LABELS["knowledge_type"] else "faq",
        chunk_type="rule" if knowledge_type in {"pricing", "process", "capability"} else "faq",
        business_line=business_line if business_line in KB_LABELS["business_line"] else "general",
        language_pair=query_features.get("language_pair"),
        service_scope=query_features.get("service_scope"),
        customer_tier=query_features.get("customer_tier"),
        priority=70,
        risk_level="high" if knowledge_type == "pricing" or mentions_pricing_topic(candidate_title, candidate_content) else "medium",
        pricing_rule=pricing_rule,
        source_type="feedback",
        source_ref=str(log.log_id),
        source_snapshot=source_snapshot,
        owner=owner or "sales_feedback",
        operator=operator,
        review_notes=feedback_note,
    )

def _extract_candidates_from_records(
    db: Session,
    records: list[dict],
    *,
    source_type: str,
    owner: str | None = None,
    operator: str | None = None,
    source_ref_prefix: str | None = None,
    max_candidates: int = 20,
) -> list[KnowledgeCandidate]:
    created = []
    for index, record in enumerate(records, start=1):
        if len(created) >= max_candidates:
            break
        entries = _build_candidate_entries_from_record(record, source_type)
        for offset, item in enumerate(entries, start=1):
            if len(created) >= max_candidates:
                break
            candidate = _create_candidate_record(
                db,
                title=item["title"],
                content=item["content"],
                knowledge_type=item["knowledge_type"],
                chunk_type=item["chunk_type"],
                business_line=item["business_line"],
                sub_service=item.get("sub_service"),
                language_pair=item.get("language_pair"),
                service_scope=item.get("service_scope"),
                region=item.get("region"),
                customer_tier=item.get("customer_tier"),
                priority=int(item.get("priority") or 60),
                risk_level=item.get("risk_level") or "medium",
                effective_from=item.get("effective_from"),
                effective_to=item.get("effective_to"),
                pricing_rule=item.get("pricing_rule"),
                source_type=source_type,
                source_ref=f"{source_ref_prefix or source_type}_{index}_{offset}",
                source_snapshot=item.get("source_snapshot"),
                owner=owner,
                operator=operator,
                review_notes=item.get("review_notes"),
            )
            created.append(candidate)
    return created

def _update_effect_rollup_for_log(
    db: Session,
    *,
    log: KnowledgeHitLog,
    feedback_status: str,
    manual_feedback: dict | None = None,
) -> None:
    hit_chunk_ids = [item for item in (log.hit_chunk_ids or []) if item]
    if not hit_chunk_ids:
        _sync_email_effect_rollup_for_log(db, log=log, feedback_status=feedback_status, manual_feedback=manual_feedback)
        return
    delta = _feedback_delta(feedback_status)
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id.in_(hit_chunk_ids)).all()
    now = datetime.utcnow()
    for chunk in chunks:
        current = float(chunk.effect_score) if chunk.effect_score is not None else 0.5
        chunk.feedback_count = int(chunk.feedback_count or 0) + 1
        if feedback_status in POSITIVE_FEEDBACK_STATUSES:
            chunk.positive_feedback_count = int(chunk.positive_feedback_count or 0) + 1
        chunk.effect_score = _decimal_score(max(0.0, min(1.0, current + delta)))
        chunk.last_feedback_at = now
    if log.session_id:
        thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == log.session_id).first()
        if thread_fact:
            outcome = dict(thread_fact.outcome_feedback or {})
            outcome.update({
                "last_feedback_status": feedback_status,
                "last_feedback_note": (manual_feedback or {}).get("note") if isinstance(manual_feedback, dict) else None,
                "last_feedback_at": now.isoformat(),
                "hit_chunk_ids": hit_chunk_ids,
            })
            thread_fact.outcome_feedback = outcome
            current_effect = float(thread_fact.effect_score) if thread_fact.effect_score is not None else 0.5
            thread_fact.effect_score = _decimal_score(max(0.0, min(1.0, current_effect + delta)))
    _sync_email_effect_rollup_for_log(db, log=log, feedback_status=feedback_status, manual_feedback=manual_feedback)

def _split_reply_into_candidate_fragments(reply_text: str) -> list[dict]:
    raw_parts = [part.strip() for part in re.split(r"\n{2,}|[。！？]\s*", reply_text or "") if part.strip()]
    if not raw_parts and reply_text:
        raw_parts = [reply_text.strip()]
    fragments = []
    for index, part in enumerate(raw_parts[:6], start=1):
        fragment = infer_function_fragment(title=f"reply_fragment_{index}", content=part, source_type="excellent_reply", tags={})
        fragments.append({
            "title": f"优秀回复片段{index}",
            "content": part,
            "function_fragment": fragment or "core_answer",
        })
    return fragments

def _extract_excellent_reply_candidates(
    db: Session,
    *,
    session_id: str,
    owner: str | None = None,
    operator: str | None = None,
    force: bool = False,
) -> list[KnowledgeCandidate]:
    log = db.query(KnowledgeHitLog).filter(
        KnowledgeHitLog.session_id == session_id,
        KnowledgeHitLog.final_response.isnot(None),
    ).order_by(KnowledgeHitLog.created_at.desc()).first()
    if not log or not (log.final_response or "").strip():
        raise HTTPException(status_code=404, detail="当前会话没有可抽取的最终回复")
    if not force and log.feedback_status not in {"useful", "adopted", "won", "advanced"}:
        raise HTTPException(status_code=422, detail="仅在反馈为 useful/adopted/won/advanced 时允许抽取优秀回复；如需跳过请使用 force=true")
    summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
    thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
    query_features = log.query_features or {}
    knowledge_type = (query_features.get("knowledge_type") or ["faq"])[0] if isinstance(query_features.get("knowledge_type"), list) else (query_features.get("knowledge_type") or "faq")
    business_line = query_features.get("business_line") or "general"
    created: list[KnowledgeCandidate] = []
    for idx, item in enumerate(_split_reply_into_candidate_fragments(log.final_response or ""), start=1):
        source_ref = f"{session_id}_{idx}"
        source_snapshot = {
            "session_id": session_id,
            "log_id": str(log.log_id),
            "summary_topic": summary.topic if summary else None,
            "summary_core_demand": summary.core_demand if summary else None,
            "feedback_status": log.feedback_status,
            "hit_chunk_ids": log.hit_chunk_ids,
            "tags": {
                "thread_id": session_id,
                "function_fragment": item["function_fragment"],
                "scenario_label": thread_fact.scenario_label if thread_fact else None,
                "intent_label": thread_fact.intent_label if thread_fact else None,
                "language_style": "concise_wechat",
            },
        }
        candidate = db.query(KnowledgeCandidate).filter(
            KnowledgeCandidate.source_type == "excellent_reply",
            KnowledgeCandidate.source_ref == source_ref,
        ).first()
        if candidate:
            candidate.title = f"{item['title']} · {item['function_fragment']}"
            candidate.content = item["content"]
            candidate.knowledge_type = knowledge_type if knowledge_type in KB_LABELS["knowledge_type"] else "faq"
            candidate.chunk_type = "template" if item["function_fragment"] in {"greeting", "cta", "closing", "core_answer"} else "example"
            candidate.business_line = business_line if business_line in KB_LABELS["business_line"] else "general"
            candidate.language_pair = query_features.get("language_pair")
            candidate.service_scope = query_features.get("service_scope")
            candidate.customer_tier = query_features.get("customer_tier")
            candidate.priority = 85
            candidate.risk_level = "medium"
            candidate.source_snapshot = source_snapshot
            candidate.owner = owner or candidate.owner or "excellent_reply_extract"
            candidate.operator = operator or "excellent_reply_extract"
            candidate.review_notes = f"来源会话 {session_id} 的优秀回复片段抽取"
            _apply_candidate_governance(candidate)
        else:
            candidate = _create_candidate_record(
                db,
                title=f"{item['title']} · {item['function_fragment']}",
                content=item["content"],
                knowledge_type=knowledge_type if knowledge_type in KB_LABELS["knowledge_type"] else "faq",
                chunk_type="template" if item["function_fragment"] in {"greeting", "cta", "closing", "core_answer"} else "example",
                business_line=business_line if business_line in KB_LABELS["business_line"] else "general",
                language_pair=query_features.get("language_pair"),
                service_scope=query_features.get("service_scope"),
                customer_tier=query_features.get("customer_tier"),
                priority=85,
                risk_level="medium",
                source_type="excellent_reply",
                source_ref=source_ref,
                source_snapshot=source_snapshot,
                owner=owner or "excellent_reply_extract",
                operator=operator or "excellent_reply_extract",
                review_notes=f"来源会话 {session_id} 的优秀回复片段抽取",
            )
        _upsert_email_fragment_asset(
            db,
            source_type="excellent_reply",
            source_ref=source_ref,
            title=candidate.title,
            content=candidate.content,
            session_id=session_id,
            thread_id=session_id,
            candidate=candidate,
            log_id=str(log.log_id),
            source_snapshot=source_snapshot,
        )
        created.append(candidate)
    return created

def _load_regression_cases() -> list[dict]:
    filepath = os.path.join(os.path.dirname(__file__), "kb_regression_cases.json")
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)

def _filter_regression_cases(
    cases: list[dict],
    *,
    case_ids: list[str] | None = None,
    groups: list[str] | None = None,
    category: str | None = None,
    risk_level: str | None = None,
    business_line: str | None = None,
) -> list[dict]:
    selected_ids = set(case_ids or [])
    selected_groups = set(groups or [])
    filtered = []
    for case in cases:
        if selected_ids and case.get("case_id") not in selected_ids:
            continue
        if selected_groups and case.get("group") not in selected_groups:
            continue
        if category and case.get("category") != category:
            continue
        if risk_level and case.get("risk_level") != risk_level:
            continue
        case_business_line = (case.get("query_features") or {}).get("business_line") or case.get("business_line")
        if business_line and case_business_line != business_line:
            continue
        filtered.append(case)
    return filtered

def _evaluate_regression_case(case: dict, result: dict) -> dict:
    expected = case.get("expected") or {}
    hits = result.get("hits") or []
    failures = []
    warnings = []

    def hit_text(hit):
        return f"{hit.get('title') or ''}\n{hit.get('content') or ''}".lower()

    def label(dict_name: str, code: str | None) -> str:
        if not code:
            return "-"
        return label_for(dict_name, code)

    def hit_has_tax_or_invoice_evidence(hit: dict) -> bool:
        text = hit_text(hit)
        if any(word in text for word in ["税", "发票", "invoice", "tax"]):
            return True
        for rule in hit.get("pricing_rules") or []:
            raw_tax_policy = rule.get("tax_policy") or ""
            if str(raw_tax_policy).strip():
                return True
            tax_policy = str(raw_tax_policy).lower()
            source_ref = (rule.get("source_ref") or "").lower()
            if any(word in tax_policy for word in ["税", "发票", "invoice", "tax"]):
                return True
            if any(word in source_ref for word in ["税", "发票", "invoice", "tax"]):
                return True
        return False

    must_type = expected.get("must_hit_knowledge_type")
    if must_type and not any(hit.get("knowledge_type") == must_type for hit in hits):
        failures.append(f"未命中要求的知识类型：{label('knowledge_type', must_type)}")

    must_class = expected.get("must_hit_knowledge_class")
    if must_class and not any(hit.get("knowledge_class") == must_class for hit in hits):
        failures.append(f"未命中要求的知识分类：{label('knowledge_class', must_class)}")

    if expected.get("must_hit_pricing_rule") and not any(hit.get("pricing_rules") for hit in hits):
        failures.append("未命中结构化报价规则")

    if expected.get("manual_review_required") is True and not result.get("manual_review_required"):
        failures.append("本用例要求触发人工复核，但检索结果没有触发")

    if expected.get("must_clarify_query") and result.get("status") not in {"ambiguous_query", "low_confidence", "manual_review_required"}:
        failures.append("宽泛问题未触发澄清、低置信或人工复核状态")

    if expected.get("manual_review_allowed") is False and result.get("manual_review_required"):
        failures.append("本用例不允许触发人工复核，但检索结果触发了人工复核")

    service_scope = expected.get("service_scope_filter")
    if service_scope and hits and not any(hit.get("service_scope") == service_scope for hit in hits):
        failures.append(f"未命中要求的服务范围：{label('service_scope', service_scope)}")

    business_line = expected.get("business_line_filter")
    if business_line and hits and not all(hit.get("business_line") == business_line for hit in hits):
        failures.append(f"命中了指定服务之外的知识，要求服务：{label('business_line', business_line)}")

    customer_tier = expected.get("customer_tier_filter")
    if customer_tier and not result.get("filters_used", {}).get("customer_tier") == customer_tier:
        failures.append(f"客户层级过滤未生效：{label('customer_tier', customer_tier)}")

    if expected.get("must_not_use_translation_price") and any(
        hit.get("business_line") == "translation" and hit.get("knowledge_type") == "pricing" for hit in hits
    ):
        failures.append("非翻译业务问题误用了翻译报价知识")

    if expected.get("must_hit_tax_or_invoice") and not any(hit_has_tax_or_invoice_evidence(hit) for hit in hits):
        failures.append("未找到税费或发票相关证据")

    if expected.get("should_return_process_or_faq") and not any(
        hit.get("knowledge_type") in {"process", "faq"} for hit in hits
    ):
        failures.append("未命中流程知识或常见问答证据")

    if expected.get("must_check_effective_time") and result.get("filters_used", {}).get("effective_time") != "now":
        failures.append("未按当前有效期过滤知识")

    if expected.get("must_not_use_general_if_legal_exists"):
        legal_hits = [hit for hit in hits if hit.get("service_scope") == "legal"]
        general_hits = [hit for hit in hits if hit.get("service_scope") == "general"]
        if general_hits and not legal_hits:
            failures.append("请求法律范围时仅命中了通用范围证据")

    if expected.get("should_include_rules_before_expression") and not result.get("evidence_context", {}).get("rules"):
        warnings.append("建议优先命中规则类证据，但当前未命中规则证据")

    if expected.get("must_not_fabricate_price") and result.get("status") == "ok" and not hits:
        failures.append("价格类问题没有价格证据时不得返回正常结果")

    if expected.get("must_not_fabricate_capability") and result.get("status") == "ok" and not hits:
        failures.append("能力类问题没有能力证据时不得返回正常结果")

    return {
        "passed": not failures,
        "failures": failures,
        "warnings": warnings,
    }

def _select_publish_gate_case_ids(doc: KnowledgeDocument, chunks: list[KnowledgeChunk], db: Session) -> list[str]:
    cases = _load_regression_cases()
    rules = db.query(PricingRule).filter(
        PricingRule.document_id == doc.document_id,
        PricingRule.status != "archived",
    ).all()
    knowledge_types = {doc.knowledge_type}
    knowledge_types.update(chunk.knowledge_type for chunk in chunks if chunk.knowledge_type)
    chunk_types = {chunk.chunk_type for chunk in chunks if chunk.chunk_type}
    language_pairs = {chunk.language_pair for chunk in chunks if chunk.language_pair}
    language_pairs.update(rule.language_pair for rule in rules if rule.language_pair)
    service_scopes = {chunk.service_scope for chunk in chunks if chunk.service_scope}
    service_scopes.update(rule.service_scope for rule in rules if rule.service_scope)
    customer_tiers = {chunk.customer_tier for chunk in chunks if chunk.customer_tier}
    customer_tiers.update(rule.customer_tier for rule in rules if rule.customer_tier)
    business_line = doc.business_line
    is_high_risk = doc.risk_level == "high" or "pricing" in knowledge_types

    selected = []
    for case in cases:
        query_features = case.get("query_features") or {}
        raw_case_types = query_features.get("knowledge_type") or []
        case_types = set(raw_case_types if isinstance(raw_case_types, list) else [raw_case_types])
        case_business_line = query_features.get("business_line")
        category = case.get("category")

        business_match = (
            not case_business_line
            or case_business_line == business_line
            or category == business_line
        )
        if not business_match:
            continue

        knowledge_match = bool(case_types & knowledge_types)
        if not knowledge_match:
            if category in {"low_confidence", "versioning"} and is_high_risk:
                knowledge_match = True
            elif category == "customer_tier" and is_high_risk and customer_tiers:
                knowledge_match = True
            elif category == "template" and ("template" in chunk_types or {"pricing", "faq"} & knowledge_types):
                knowledge_match = True
        if not knowledge_match:
            continue

        language_pair = query_features.get("language_pair")
        if language_pair and language_pairs and language_pair not in language_pairs:
            continue

        service_scope = query_features.get("service_scope")
        if service_scope and service_scopes and service_scope not in service_scopes and category != "capability":
            continue

        customer_tier = query_features.get("customer_tier")
        if customer_tier and customer_tiers and customer_tier not in customer_tiers and category != "customer_tier":
            continue

        selected.append(case["case_id"])

    if not selected and is_high_risk:
        selected = [
            case["case_id"]
            for case in cases
            if ((case.get("query_features") or {}).get("business_line") in {None, business_line})
            and (case.get("category") in {"pricing", "low_confidence", "versioning"})
        ]
    return list(dict.fromkeys(selected))

def _governance_gate_failures_for_doc(doc: KnowledgeDocument, chunks: list[KnowledgeChunk]) -> list[str]:
    def auxiliary_reference_chunk(chunk: KnowledgeChunk) -> bool:
        tags = chunk.structured_tags or {}
        knowledge_class = str(tags.get("knowledge_class") or "").strip()
        if knowledge_class in {"pricing_constraint", "definition", "process_constraint"}:
            return True
        if chunk.chunk_type == "definition":
            return True
        if chunk.chunk_type == "constraint" and chunk.knowledge_type in {"pricing", "process"}:
            return True
        return False

    failures: list[str] = []
    if not chunks:
        failures.append("文档没有可发布切片")
        return failures
    replyability_required_chunks = [
        chunk for chunk in chunks
        if chunk.library_type == "reference" and not auxiliary_reference_chunk(chunk)
    ]
    replyable_reference_chunks = [
        chunk for chunk in replyability_required_chunks
        if bool(chunk.usable_for_reply)
    ]
    if doc.library_type == "reference" and replyability_required_chunks and not replyable_reference_chunks:
        failures.append("参考型文档没有任何可自动回复切片")
    for chunk in chunks:
        mixed = detect_mixed_knowledge(chunk.content, chunk.structured_tags)
        if mixed.get("mixed"):
            failures.append(f"切片《{chunk.title}》存在混合知识点，分类过多：{', '.join(mixed.get('categories') or [])}")
        if not chunk.publishable:
            failures.append(f"切片《{chunk.title}》未通过 publishable 门禁")
        if chunk.library_type == "reference" and not auxiliary_reference_chunk(chunk) and not chunk.allowed_for_generation:
            failures.append(f"切片《{chunk.title}》未通过自动回复许可门禁")
    return list(dict.fromkeys(failures))

async def _run_publish_gate_for_documents(docs: list[KnowledgeDocument], db: Session) -> dict:
    case_ids = []
    docs_report = []
    case_meta = {case["case_id"]: case for case in _load_regression_cases()}
    governance_failures = []
    for doc in docs:
        chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
        selected_ids = _select_publish_gate_case_ids(doc, chunks, db)
        doc_failures = _governance_gate_failures_for_doc(doc, chunks)
        case_ids.extend(selected_ids)
        category_counts: dict[str, int] = {}
        for case_id in selected_ids:
            category = case_meta.get(case_id, {}).get("category") or "uncategorized"
            category_counts[category] = category_counts.get(category, 0) + 1
        docs_report.append({
            "document_id": str(doc.document_id),
            "title": doc.title,
            "knowledge_type": doc.knowledge_type,
            "business_line": doc.business_line,
            "selected_case_ids": selected_ids,
            "selected_case_count": len(selected_ids),
            "selected_case_category_counts": category_counts,
            "governance_failures": doc_failures,
        })
        if doc_failures:
            governance_failures.append({
                "case_id": f"governance_{doc.document_id}",
                "document_id": str(doc.document_id),
                "title": doc.title,
                "failures": doc_failures,
                "passed": False,
            })

    case_ids = list(dict.fromkeys(case_ids))
    selected_case_category_counts: dict[str, int] = {}
    for case_id in case_ids:
        category = case_meta.get(case_id, {}).get("category") or "uncategorized"
        selected_case_category_counts[category] = selected_case_category_counts.get(category, 0) + 1
    governance_failed_documents = len(governance_failures)
    if not case_ids:
        return {
            "selected_case_ids": [],
            "target_document_ids": [str(doc.document_id) for doc in docs],
            "target_document_count": len(docs),
            "documents": docs_report,
            "passed": not governance_failures,
            "summary": {
                "selected": 0,
                "passed": 0,
                "failed": governance_failed_documents,
                "regression_selected": 0,
                "regression_passed": 0,
                "regression_failed": 0,
                "governance_failed_documents": governance_failed_documents,
            },
            "selected_case_category_counts": {},
            "results": [],
            "failed_cases": governance_failures,
        }

    regression = await run_kb_regression_cases(
        KnowledgeRegressionRunRequest(
            case_ids=case_ids,
            top_k=5,
            cleanup_logs=True,
            include_hits=False,
            run_id=f"publish_gate_{uuid.uuid4().hex[:10]}",
        )
    )
    regression_failed_cases = [item for item in regression.get("results", []) if not item.get("passed")]
    regression_failed_count = len(regression_failed_cases)
    regression_passed_count = regression.get("passed", 0)
    failed_cases = list(regression_failed_cases)
    failed_cases.extend(governance_failures)
    return {
        "selected_case_ids": case_ids,
        "target_document_ids": [str(doc.document_id) for doc in docs],
        "target_document_count": len(docs),
        "documents": docs_report,
        "passed": not failed_cases,
        "summary": {
            "selected": len(case_ids),
            "passed": regression_passed_count,
            "failed": regression_failed_count + governance_failed_documents,
            "regression_selected": len(case_ids),
            "regression_passed": regression_passed_count,
            "regression_failed": regression_failed_count,
            "governance_failed_documents": governance_failed_documents,
        },
        "selected_case_category_counts": selected_case_category_counts,
        "results": regression.get("results", []),
        "failed_cases": failed_cases,
        "run_id": regression.get("run_id"),
    }

def _ensure_documents_approved_for_publish(docs: list[KnowledgeDocument]) -> None:
    invalid = []
    for doc in docs:
        if document_stage_code(doc) != "approved":
            invalid.append({
                "document_id": str(doc.document_id),
                "title": doc.title,
                "stage_label": document_stage_label(doc),
            })
    if invalid:
        raise HTTPException(status_code=422, detail={
            "message": "只有审核同意的知识才能发布",
            "documents": {
                item["document_id"]: {
                    "title": item["title"],
                    "errors": [f"当前状态为{item['stage_label']}，请先完成审核同意后再发布"]
                }
                for item in invalid
            },
        })

def _append_publish_gate_audit(
    doc: KnowledgeDocument,
    *,
    gate_report: dict,
    operator: str | None,
    force: bool,
    force_reason: str | None,
):
    meta = dict(doc.source_meta or {})
    audits = list(meta.get("publish_gate_audit") or [])
    audits.append({
        "timestamp": datetime.utcnow().isoformat(),
        "operator": operator,
        "force": force,
        "force_reason": force_reason,
        "selected_case_ids": gate_report.get("selected_case_ids") or [],
        "passed": gate_report.get("passed", False),
        "failed_case_ids": [item.get("case_id") for item in gate_report.get("failed_cases", [])],
    })
    meta["publish_gate_audit"] = audits[-20:]
    meta["last_publish_gate"] = {
        "timestamp": audits[-1]["timestamp"],
        "passed": gate_report.get("passed", False),
        "selected_case_ids": gate_report.get("selected_case_ids") or [],
    }
    doc.source_meta = meta

async def _ensure_publish_gate(
    docs: list[KnowledgeDocument],
    db: Session,
    *,
    force: bool = False,
    force_reason: str | None = None,
    operator: str | None = None,
) -> dict:
    doc_ids = [doc.document_id for doc in docs]
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id.in_(doc_ids)).all() if doc_ids else []
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id.in_(doc_ids)).all() if doc_ids else []
    original_doc_states = {
        doc.document_id: (doc.status, doc.review_status)
        for doc in docs
    }
    original_chunk_states = {
        chunk.chunk_id: chunk.status
        for chunk in chunks
    }
    original_rule_states = {
        rule.rule_id: rule.status
        for rule in pricing_rules
    }
    try:
        # 发布门禁必须检索到“本次待发布”的知识；失败后会恢复原状态，不污染线上 active 集合。
        for doc in docs:
            doc.status = "active"
            doc.review_status = "approved"
        for chunk in chunks:
            doc = next((item for item in docs if item.document_id == chunk.document_id), None)
            chunk.status = "active"
            if doc:
                chunk.effective_from = doc.effective_from
                chunk.effective_to = doc.effective_to
        for rule in pricing_rules:
            doc = next((item for item in docs if item.document_id == rule.document_id), None)
            rule.status = "active"
            if doc:
                rule.effective_from = doc.effective_from
                rule.effective_to = doc.effective_to
                rule.version_no = doc.version_no
        db.commit()
        gate_report = await _run_publish_gate_for_documents(docs, db)
    finally:
        for doc in docs:
            old_status, old_review_status = original_doc_states.get(doc.document_id, (doc.status, doc.review_status))
            doc.status = old_status
            doc.review_status = old_review_status
        for chunk in chunks:
            chunk.status = original_chunk_states.get(chunk.chunk_id, chunk.status)
        for rule in pricing_rules:
            rule.status = original_rule_states.get(rule.rule_id, rule.status)
        db.commit()

    if not gate_report.get("passed") and not force:
        raise HTTPException(status_code=422, detail={
            "message": "知识发布门禁未通过",
            "suggestion": "请查看 failed_cases 中的失败用例，修正知识内容、结构化报价规则或检索条件后再发布；如确需上线，请填写强制发布原因。",
            "gate_report": gate_report,
        })
    if force and not force_reason:
        raise HTTPException(status_code=422, detail="强制发布时必须填写强制发布原因")
    for doc in docs:
        _append_publish_gate_audit(
            doc,
            gate_report=gate_report,
            operator=operator,
            force=force,
            force_reason=force_reason,
        )
    return gate_report

def _pricing_signature(rule: PricingRule) -> tuple:
    return (
        str(rule.price_min) if rule.price_min is not None else None,
        str(rule.price_max) if rule.price_max is not None else None,
        str(rule.min_charge) if rule.min_charge is not None else None,
        str(rule.urgent_multiplier) if rule.urgent_multiplier is not None else None,
        rule.tax_policy,
    )

def _pricing_scope(rule: PricingRule) -> tuple:
    return (
        rule.business_line,
        rule.language_pair,
        rule.service_scope,
        rule.customer_tier,
        rule.region,
        rule.unit,
        rule.currency,
    )

def _scope_to_dict(scope: tuple) -> dict:
    keys = ["business_line", "language_pair", "service_scope", "customer_tier", "region", "unit", "currency"]
    return {key: value for key, value in zip(keys, scope)}

def build_kb_quality_report(db: Session, days: int = 30, limit: int = 300) -> dict:
    now = datetime.utcnow()
    deadline = now + timedelta(days=max(1, min(days, 365)))
    issues = []

    pricing_chunks = db.query(KnowledgeChunk).filter(
        KnowledgeChunk.status == "active",
        KnowledgeChunk.knowledge_type == "pricing",
    ).limit(max(1, min(limit, 1000))).all()
    for chunk in pricing_chunks:
        rule_count = db.query(PricingRule).filter(
            PricingRule.status == "active",
            PricingRule.document_id == chunk.document_id,
            or_(PricingRule.chunk_id == chunk.chunk_id, PricingRule.chunk_id.is_(None)),
        ).count()
        if rule_count == 0:
            issues.append({
                "type": "missing_pricing_rule",
                "severity": "high",
                "message": "已发布报价切片缺少可用的结构化报价规则",
                "document_id": str(chunk.document_id),
                "chunk_id": str(chunk.chunk_id),
                "title": chunk.title,
                "scope": {
                    "business_line": chunk.business_line,
                    "language_pair": chunk.language_pair,
                    "service_scope": chunk.service_scope,
                    "customer_tier": chunk.customer_tier,
                },
                "actions": ["view_document", "view_chunk", "add_pricing_rule"],
            })

    active_rules = db.query(PricingRule).filter(PricingRule.status == "active").limit(max(1, min(limit, 1000))).all()
    grouped_rules: dict[tuple, list[PricingRule]] = {}
    for rule in active_rules:
        grouped_rules.setdefault(_pricing_scope(rule), []).append(rule)

    for scope, rules in grouped_rules.items():
        if len(rules) < 2:
            continue
        signatures = {_pricing_signature(rule) for rule in rules}
        issue_type = "pricing_conflict" if len(signatures) > 1 else "duplicate_pricing_rule"
        severity = "high" if issue_type == "pricing_conflict" else "medium"
        issues.append({
            "type": issue_type,
            "severity": severity,
            "message": "同一报价范围存在冲突价格" if issue_type == "pricing_conflict" else "同一报价范围存在重复价格",
            "scope": _scope_to_dict(scope),
            "rule_ids": [str(rule.rule_id) for rule in rules],
            "document_ids": sorted({str(rule.document_id) for rule in rules}),
            "signatures": [list(signature) for signature in sorted(signatures, key=lambda item: str(item))],
            "actions": ["view_rules", "archive_duplicate_rules" if issue_type == "duplicate_pricing_rule" else "review_conflict_rules"],
        })

    expiring_docs = db.query(KnowledgeDocument).filter(
        KnowledgeDocument.status == "active",
        KnowledgeDocument.effective_to.isnot(None),
        KnowledgeDocument.effective_to <= deadline,
    ).limit(max(1, min(limit, 1000))).all()
    for doc in expiring_docs:
        expired = doc.effective_to < now
        issues.append({
            "type": "expired_document" if expired else "expiring_document",
            "severity": "high" if expired else "medium",
            "message": "已发布知识已过期" if expired else "已发布知识即将过期",
            "document_id": str(doc.document_id),
            "title": doc.title,
            "effective_to": doc.effective_to,
            "days_left": (doc.effective_to - now).days,
            "actions": ["view_document", "extend_effective_to", "archive_document"],
        })

    expiring_rules = db.query(PricingRule).filter(
        PricingRule.status == "active",
        PricingRule.effective_to.isnot(None),
        PricingRule.effective_to <= deadline,
    ).limit(max(1, min(limit, 1000))).all()
    for rule in expiring_rules:
        expired = rule.effective_to < now
        issues.append({
            "type": "expired_pricing_rule" if expired else "expiring_pricing_rule",
            "severity": "high" if expired else "medium",
            "message": "已发布报价规则已过期" if expired else "已发布报价规则即将过期",
            "rule_id": str(rule.rule_id),
            "document_id": str(rule.document_id),
            "scope": _scope_to_dict(_pricing_scope(rule)),
            "effective_to": rule.effective_to,
            "days_left": (rule.effective_to - now).days,
            "actions": ["view_rules", "archive_pricing_rule"],
        })

    by_type: dict[str, int] = {}
    by_severity: dict[str, int] = {}
    for issue in issues:
        by_type[issue["type"]] = by_type.get(issue["type"], 0) + 1
        by_severity[issue["severity"]] = by_severity.get(issue["severity"], 0) + 1

    severity_order = {"high": 0, "medium": 1, "low": 2}
    issues.sort(key=lambda issue: (severity_order.get(issue["severity"], 9), issue["type"]))
    return {
        "status": "success",
        "generated_at": now.isoformat(),
        "window_days": days,
        "total_issues": len(issues),
        "by_type": by_type,
        "by_severity": by_severity,
        "issues": issues[:max(1, min(limit, 1000))],
    }

def _create_pricing_rule(
    db: Session,
    document: KnowledgeDocument,
    chunk: KnowledgeChunk | None,
    payload: dict | PricingRulePayload | None,
):
    if not payload:
        return None
    rule_payload = payload if isinstance(payload, PricingRulePayload) else PricingRulePayload(**payload)
    rule = PricingRule(
        document_id=document.document_id,
        chunk_id=chunk.chunk_id if chunk else None,
        business_line=rule_payload.business_line or (chunk.business_line if chunk else document.business_line),
        sub_service=rule_payload.sub_service or (chunk.sub_service if chunk else document.sub_service),
        language_pair=rule_payload.language_pair or (chunk.language_pair if chunk else None),
        service_scope=rule_payload.service_scope or (chunk.service_scope if chunk else None),
        unit=rule_payload.unit,
        currency=rule_payload.currency,
        price_min=_decimal_or_none(rule_payload.price_min),
        price_max=_decimal_or_none(rule_payload.price_max),
        urgent_multiplier=_decimal_or_none(rule_payload.urgent_multiplier),
        tax_policy=rule_payload.tax_policy,
        min_charge=_decimal_or_none(rule_payload.min_charge),
        customer_tier=rule_payload.customer_tier or (chunk.customer_tier if chunk else None),
        region=rule_payload.region or (chunk.region if chunk else None),
        status=document.status,
        effective_from=document.effective_from,
        effective_to=document.effective_to,
        version_no=document.version_no,
        source_ref=rule_payload.source_ref or document.source_ref,
    )
    db.add(rule)
    return rule

def _document_publish_errors(db: Session, doc: KnowledgeDocument) -> list[str]:
    errors = []
    if not doc.version_no:
        errors.append("版本号不能为空")
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
    if not chunks:
        errors.append("至少需要一个知识切片")

    pricing_chunks = [
        chunk for chunk in chunks
        if chunk.chunk_type != "constraint" and is_pricing_text(chunk.title, chunk.content, chunk.knowledge_type)
    ]
    if pricing_chunks:
        if not doc.effective_from or not doc.effective_to:
            errors.append("报价类知识必须填写生效时间和失效时间")
        for chunk in pricing_chunks:
            rule_count = db.query(PricingRule).filter(
                PricingRule.document_id == doc.document_id,
                or_(PricingRule.chunk_id == chunk.chunk_id, PricingRule.chunk_id.is_(None)),
            ).count()
            if rule_count == 0:
                errors.append(f"报价切片缺少结构化报价规则：{chunk.title}")
    return errors

def _validate_documents_publishable(db: Session, docs) -> None:
    _ensure_documents_publishable(db, docs, force=False)
    return None

def _ensure_documents_publishable(db: Session, docs, *, force: bool = False) -> dict:
    all_errors = {}
    for doc in docs:
        errors = _document_publish_errors(db, doc)
        if errors:
            all_errors[str(doc.document_id)] = {"title": doc.title, "errors": errors}
    if all_errors:
        report = {"passed": False, "documents": all_errors}
        if not force:
            raise HTTPException(status_code=422, detail={
                "message": "知识发布前结构校验未通过",
                "documents": all_errors,
                "force_allowed": True,
                "publish_validation": report,
                "suggestion": "请补齐控制项后重新发布；如确认风险可接受，可填写强制发布原因后人工通过。",
            })
        return report
    return {"passed": True, "documents": {}}

def _sync_related_status(db: Session, doc: KnowledgeDocument, status: str, review_status: str | None = None):
    doc.status = status
    if review_status:
        doc.review_status = review_status
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
    for chunk in chunks:
        chunk.status = status
        chunk.effective_from = doc.effective_from
        chunk.effective_to = doc.effective_to
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id == doc.document_id).all()
    for rule in pricing_rules:
        rule.status = status
        rule.effective_from = doc.effective_from
        rule.effective_to = doc.effective_to
        rule.version_no = doc.version_no
    return len(chunks), len(pricing_rules)

def _sync_single_chunk_document_from_chunk(db: Session, chunk: KnowledgeChunk) -> None:
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == chunk.document_id).first()
    if not doc:
        return
    chunk_count = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == chunk.document_id).count()
    if chunk_count != 1:
        return
    class_code = (chunk.structured_tags or {}).get("knowledge_class") or knowledge_class_from_pair(chunk.knowledge_type, chunk.chunk_type)
    doc.title = chunk.title
    doc.knowledge_type = chunk.knowledge_type
    doc.business_line = chunk.business_line
    doc.tags = merge_tags(merge_knowledge_class_tags(doc.tags, class_code), library_type=chunk.library_type)
    doc.library_type = chunk.library_type
    if class_code:
        config = KB_KNOWLEDGE_CLASS_CONFIG.get(class_code)
        if config:
            doc.risk_level = config["risk_level"]

@app.post("/api/kb/documents/manual")
async def create_manual_knowledge(payload: KnowledgeManualCreate, db: Session = Depends(get_db)):
    """手工新增一条知识，默认进入 draft，不直接发布。"""
    knowledge_class = normalize_knowledge_class_or_raise(payload.knowledge_class)
    resolved_class, resolved_type, resolved_chunk_type, resolved_risk = resolve_knowledge_class_fields(
        knowledge_class=knowledge_class,
        knowledge_type=payload.knowledge_type,
        chunk_type=payload.chunk_type,
        risk_level=payload.risk_level,
    )
    merged_tags = merge_knowledge_class_tags(payload.tags, resolved_class)
    embedding_text = _build_chunk_retrieval_text(payload.title, payload.content)
    embedding = _optional_embedding_for_storage(embedding_text, context=f"manual_create:{payload.title}")

    effective_from, effective_to = default_pricing_effective_window(
        payload.effective_from,
        payload.effective_to,
        knowledge_type=resolved_type,
        pricing_rule=payload.pricing_rule,
    )
    now_status = "draft"
    document = KnowledgeDocument(
        title=payload.title,
        knowledge_type=resolved_type,
        business_line=payload.business_line,
        sub_service=payload.sub_service,
        source_type=payload.source_type,
        source_ref=payload.source_ref,
        status=now_status,
        owner=payload.owner,
        effective_from=effective_from,
        effective_to=effective_to,
        risk_level=resolved_risk,
        review_required=payload.review_required,
        review_status="pending" if payload.review_required else "auto_ready",
        library_type=infer_library_type(source_type=payload.source_type, knowledge_type=resolved_type, chunk_type=resolved_chunk_type, tags=merged_tags),
        tags=merged_tags,
    )
    db.add(document)
    db.flush()

    chunk = KnowledgeChunk(
        document_id=document.document_id,
        chunk_no=1,
        chunk_type=resolved_chunk_type,
        title=_strip_retrieval_title_prefix(payload.title),
        content=payload.content,
        keyword_text=_build_chunk_retrieval_text(payload.title, payload.content),
        embedding=embedding,
        **_embedding_metadata_fields(embedding),
        priority=payload.priority,
        business_line=payload.business_line,
        sub_service=payload.sub_service,
        knowledge_type=resolved_type,
        language_pair=payload.language_pair,
        service_scope=payload.service_scope,
        region=payload.region,
        customer_tier=payload.customer_tier,
        structured_tags=merged_tags,
        status=now_status,
        effective_from=effective_from,
        effective_to=effective_to,
    )
    db.add(chunk)
    db.flush()
    pricing_rule_payload = payload.pricing_rule
    if not pricing_rule_payload and resolved_type == "pricing" and resolved_chunk_type != "constraint":
        pricing_rule_payload = infer_pricing_rule_candidate(
            payload.title,
            payload.content,
            payload.business_line,
            payload.language_pair,
            payload.service_scope,
            payload.customer_tier,
        )
    pricing_rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
    _apply_chunk_governance(chunk, source_type=payload.source_type, pricing_rule=pricing_rule_payload, source_ref=payload.source_ref)
    document.library_type = chunk.library_type
    document.tags = merge_tags(document.tags, library_type=document.library_type)
    db.commit()
    db.refresh(document)
    db.refresh(chunk)
    if pricing_rule:
        db.refresh(pricing_rule)
    return {
        "status": "success",
        "document": _doc_to_dict(document),
        "chunk": _chunk_to_dict(chunk),
        "pricing_rule": _pricing_rule_to_dict(pricing_rule) if pricing_rule else None,
    }

@app.post("/api/kb/documents/manual_multi")
async def create_manual_multi_knowledge(payload: KnowledgeMultiChunkCreate, db: Session = Depends(get_db)):
    """手工新增一个文档、多条切片，默认 draft。"""
    if not payload.chunks:
        raise HTTPException(status_code=400, detail="chunks 不能为空")

    document_knowledge_class = normalize_knowledge_class_or_raise(payload.knowledge_class)
    resolved_doc_class, resolved_doc_type, _resolved_doc_chunk_type, resolved_doc_risk = resolve_knowledge_class_fields(
        knowledge_class=document_knowledge_class,
        knowledge_type=payload.knowledge_type,
        chunk_type=None,
        risk_level=payload.risk_level,
    )
    has_pricing_chunk = resolved_doc_type == "pricing" or any(
        normalize_knowledge_class_or_raise(item.knowledge_class) == "pricing_constraint"
        or item.knowledge_type == "pricing"
        or bool(item.pricing_rule)
        for item in payload.chunks
    )
    effective_from, effective_to = default_pricing_effective_window(
        payload.effective_from,
        payload.effective_to,
        knowledge_type="pricing" if has_pricing_chunk else resolved_doc_type,
    )
    merged_doc_tags = merge_knowledge_class_tags(payload.tags, resolved_doc_class)
    document = KnowledgeDocument(
        title=payload.title,
        knowledge_type=resolved_doc_type,
        business_line=payload.business_line,
        sub_service=payload.sub_service,
        source_type=payload.source_type,
        source_ref=payload.source_ref,
        status="draft",
        owner=payload.owner,
        effective_from=effective_from,
        effective_to=effective_to,
        risk_level=resolved_doc_risk,
        review_required=payload.review_required,
        review_status="pending" if payload.review_required else "auto_ready",
        library_type=infer_library_type(source_type=payload.source_type, knowledge_type=resolved_doc_type, chunk_type=None, tags=merged_doc_tags),
        tags=merged_doc_tags,
    )
    db.add(document)
    db.flush()

    created_chunks = []
    created_pricing_rules = []
    for idx, item in enumerate(payload.chunks, start=1):
        chunk_knowledge_class = normalize_knowledge_class_or_raise(item.knowledge_class)
        resolved_chunk_class, resolved_chunk_type, resolved_chunk_chunk_type, _resolved_chunk_risk = resolve_knowledge_class_fields(
            knowledge_class=chunk_knowledge_class,
            knowledge_type=item.knowledge_type or resolved_doc_type,
            chunk_type=item.chunk_type,
            risk_level=None,
        )
        merged_chunk_tags = merge_knowledge_class_tags(item.tags, resolved_chunk_class)
        retrieval_title = _strip_retrieval_title_prefix(item.title)
        retrieval_text = _build_chunk_retrieval_text(item.title, item.content)
        embedding = _optional_embedding_for_storage(retrieval_text, context=f"multi_chunk_create:{idx}:{item.title}")
        chunk = KnowledgeChunk(
            document_id=document.document_id,
            chunk_no=idx,
            chunk_type=resolved_chunk_chunk_type,
            title=retrieval_title,
            content=item.content,
            keyword_text=retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=item.priority,
            business_line=item.business_line or payload.business_line,
            sub_service=item.sub_service or payload.sub_service,
            knowledge_type=resolved_chunk_type,
            language_pair=item.language_pair,
            service_scope=item.service_scope,
            region=item.region,
            customer_tier=item.customer_tier,
            structured_tags=merged_chunk_tags,
            status="draft",
            effective_from=effective_from,
            effective_to=effective_to,
        )
        db.add(chunk)
        db.flush()
        pricing_rule_payload = item.pricing_rule
        if not pricing_rule_payload and resolved_chunk_type == "pricing" and resolved_chunk_chunk_type != "constraint":
            pricing_rule_payload = infer_pricing_rule_candidate(
                item.title,
                item.content,
                item.business_line or payload.business_line,
                item.language_pair,
                item.service_scope,
                item.customer_tier,
            )
        pricing_rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
        _apply_chunk_governance(
            chunk,
            source_type=payload.source_type,
            pricing_rule=pricing_rule_payload,
            source_ref=payload.source_ref,
        )
        if pricing_rule:
            created_pricing_rules.append(pricing_rule)
        created_chunks.append(chunk)

    if created_chunks:
        document.library_type = created_chunks[0].library_type if all(item.library_type == created_chunks[0].library_type for item in created_chunks) else "reference"
        document.tags = merge_tags(document.tags, library_type=document.library_type)
    db.commit()
    db.refresh(document)
    for chunk in created_chunks:
        db.refresh(chunk)
    for rule in created_pricing_rules:
        db.refresh(rule)
    return {
        "status": "success",
        "document": _doc_to_dict(document),
        "chunks": [_chunk_to_dict(chunk) for chunk in created_chunks],
        "pricing_rules": [_pricing_rule_to_dict(rule) for rule in created_pricing_rules],
    }

@app.get("/api/kb/dictionaries")
async def get_kb_dictionaries():
    return KB_LABELS

@app.get("/api/kb/import/templates")
async def list_kb_import_templates():
    default_template = {
        "type": "unified",
        "label": KB_EXCEL_IMPORT_TYPES["unified"]["label"],
        "knowledge_class": KB_EXCEL_IMPORT_TYPES["unified"]["knowledge_class"],
        "knowledge_type": KB_EXCEL_IMPORT_TYPES["unified"]["knowledge_type"],
        "chunk_type": KB_EXCEL_IMPORT_TYPES["unified"]["chunk_type"],
        "risk_level": KB_EXCEL_IMPORT_TYPES["unified"]["risk_level"],
        "required_columns": ["标题", "内容"],
        "guide_sheets": ["填写说明", "字段说明", "枚举参考"],
        "sample_row_count": len(KB_UNIFIED_TEMPLATE_SAMPLE_ROWS),
        "supports": [item["label"] for key, item in KB_EXCEL_IMPORT_TYPES.items() if key != "unified"],
    }
    return {
        "status": "success",
        "columns": KB_EXCEL_TEMPLATE_COLUMNS,
        "default_template": default_template,
        "templates": [default_template],
    }

@app.get("/api/kb/import/template/download")
async def download_default_kb_import_template():
    return await download_kb_import_template("unified")

@app.get("/api/kb/import/templates/{template_type}/download")
async def download_kb_import_template(template_type: str):
    if template_type not in KB_EXCEL_IMPORT_TYPES:
        raise HTTPException(status_code=404, detail="Template type not found")
    try:
        workbook = _build_kb_import_template_workbook(template_type)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl 未安装或不可用: {e}")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"kb_{template_type}_template.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.get("/api/kb/candidates")
async def list_kb_candidates(
    status: str | None = None,
    source_type: str | None = None,
    business_line: str | None = None,
    knowledge_type: str | None = None,
    title_keyword: str | None = None,
    content_keyword: str | None = None,
    limit: int = 200,
    offset: int = 0,
    include_meta: bool = False,
    db: Session = Depends(get_db),
):
    query = db.query(KnowledgeCandidate)
    if status:
        query = query.filter(KnowledgeCandidate.status == status)
    if source_type:
        query = query.filter(KnowledgeCandidate.source_type == source_type)
    if business_line:
        query = query.filter(KnowledgeCandidate.business_line == business_line)
    if knowledge_type:
        query = query.filter(KnowledgeCandidate.knowledge_type == knowledge_type)
    if title_keyword:
        query = query.filter(KnowledgeCandidate.title.ilike(f"%{title_keyword.strip()}%"))
    if content_keyword:
        query = query.filter(KnowledgeCandidate.content.ilike(f"%{content_keyword.strip()}%"))
    query_limit = max(1, min(limit, 500))
    query_offset = max(0, int(offset or 0))
    total = query.order_by(None).count()
    items = (
        query.order_by(KnowledgeCandidate.created_at.desc())
        .offset(query_offset)
        .limit(query_limit)
        .all()
    )
    result = [_candidate_to_dict(item) for item in items]
    if include_meta:
        return {
            "items": result,
            "total": total,
            "limit": query_limit,
            "offset": query_offset,
            "has_more": query_offset + len(result) < total,
            "next_offset": query_offset + len(result),
        }
    return result

@app.get("/api/kb/candidates/{candidate_id}")
async def get_kb_candidate(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(KnowledgeCandidate).filter(KnowledgeCandidate.candidate_id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Knowledge candidate not found")
    return {"status": "success", "candidate": _candidate_to_dict(candidate)}

@app.patch("/api/kb/candidates/{candidate_id}")
async def update_kb_candidate(candidate_id: str, payload: KnowledgeCandidateUpdate, db: Session = Depends(get_db)):
    candidate = db.query(KnowledgeCandidate).filter(KnowledgeCandidate.candidate_id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Knowledge candidate not found")
    for field in [
        "title", "content", "knowledge_type", "chunk_type", "business_line", "sub_service",
        "language_pair", "service_scope", "region", "customer_tier", "priority", "risk_level",
        "effective_from", "effective_to", "pricing_rule", "owner", "operator", "review_notes", "status",
    ]:
        value = getattr(payload, field)
        if value is not None:
            setattr(candidate, field, value)
    _apply_candidate_governance(candidate)
    db.commit()
    db.refresh(candidate)
    return {"status": "success", "candidate": _candidate_to_dict(candidate)}

@app.post("/api/kb/candidates/from_feedback")
async def create_kb_candidate_from_feedback(payload: KnowledgeCandidateFromFeedbackRequest, db: Session = Depends(get_db)):
    log = db.query(KnowledgeHitLog).filter(KnowledgeHitLog.log_id == payload.log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Knowledge hit log not found")
    candidate = _create_candidate_from_feedback_log(
        db,
        log,
        owner=payload.owner,
        operator=payload.operator,
        note=payload.note,
        title=payload.title,
    )
    db.commit()
    db.refresh(candidate)
    return {"status": "success", "candidate": _candidate_to_dict(candidate)}

def _promote_candidate_record(
    db: Session,
    candidate: KnowledgeCandidate,
    *,
    owner: str | None = None,
    operator: str | None = None,
    title: str | None = None,
    content: str | None = None,
    knowledge_type: str | None = None,
    chunk_type: str | None = None,
    business_line: str | None = None,
    sub_service: str | None = None,
    language_pair: str | None = None,
    service_scope: str | None = None,
    region: str | None = None,
    customer_tier: str | None = None,
    priority: int | None = None,
    risk_level: str | None = None,
    effective_from: datetime | None = None,
    effective_to: datetime | None = None,
    pricing_rule: dict | None = None,
    review_notes: str | None = None,
):
    if candidate.promoted_document_id:
        doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == candidate.promoted_document_id).first()
        return {
            "status": "success",
            "candidate": candidate,
            "document": doc,
            "chunk": None,
            "pricing_rule": None,
            "message": "Candidate already promoted",
        }

    promoted_title = title or candidate.title
    promoted_content = content or candidate.content
    promoted_knowledge_type = knowledge_type or candidate.knowledge_type
    promoted_business_line = business_line or candidate.business_line
    promoted_chunk_type = chunk_type or candidate.chunk_type
    doc, chunk, rule = _create_single_document_and_chunk(
        db,
        title=promoted_title,
        content=promoted_content,
        knowledge_type=promoted_knowledge_type,
        business_line=promoted_business_line,
        sub_service=sub_service if sub_service is not None else candidate.sub_service,
        chunk_type=promoted_chunk_type,
        language_pair=language_pair if language_pair is not None else candidate.language_pair,
        service_scope=service_scope if service_scope is not None else candidate.service_scope,
        region=region if region is not None else candidate.region,
        customer_tier=customer_tier if customer_tier is not None else candidate.customer_tier,
        source_type=f"candidate_{candidate.source_type}",
        source_ref=str(candidate.candidate_id),
        source_meta={"candidate_source_type": candidate.source_type, "candidate_source_ref": candidate.source_ref},
        owner=owner or candidate.owner or "kb_candidate_promote",
        priority=priority or candidate.priority,
        risk_level=risk_level or candidate.risk_level,
        review_required=True,
        tags={"candidate_id": str(candidate.candidate_id), "candidate_source_type": candidate.source_type},
        effective_from=effective_from if effective_from is not None else candidate.effective_from,
        effective_to=effective_to if effective_to is not None else candidate.effective_to,
        pricing_rule=pricing_rule if pricing_rule is not None else candidate.pricing_rule,
    )
    candidate.status = "promoted"
    candidate.promoted_document_id = doc.document_id
    if operator is not None:
        candidate.operator = operator
    if review_notes is not None:
        candidate.review_notes = review_notes
    return {
        "status": "success",
        "candidate": candidate,
        "document": doc,
        "chunk": chunk,
        "pricing_rule": rule,
        "message": "Candidate promoted",
    }

@app.post("/api/kb/candidates/{candidate_id}/promote")
async def promote_kb_candidate(candidate_id: str, payload: KnowledgeCandidatePromoteRequest, db: Session = Depends(get_db)):
    candidate = db.query(KnowledgeCandidate).filter(KnowledgeCandidate.candidate_id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Knowledge candidate not found")
    result = _promote_candidate_record(
        db,
        candidate,
        owner=payload.owner,
        operator=payload.operator,
        title=payload.title,
        content=payload.content,
        knowledge_type=payload.knowledge_type,
        chunk_type=payload.chunk_type,
        business_line=payload.business_line,
        sub_service=payload.sub_service,
        language_pair=payload.language_pair,
        service_scope=payload.service_scope,
        region=payload.region,
        customer_tier=payload.customer_tier,
        priority=payload.priority,
        risk_level=payload.risk_level,
        effective_from=payload.effective_from,
        effective_to=payload.effective_to,
        pricing_rule=payload.pricing_rule,
        review_notes=payload.review_notes,
    )
    db.commit()
    db.refresh(candidate)
    if result["document"]:
        db.refresh(result["document"])
    if result["chunk"]:
        db.refresh(result["chunk"])
    if result["pricing_rule"]:
        db.refresh(result["pricing_rule"])
    return {
        "status": "success",
        "candidate": _candidate_to_dict(candidate),
        "document": _doc_to_dict(result["document"]) if result["document"] else None,
        "chunk": _chunk_to_dict(result["chunk"]) if result["chunk"] else None,
        "pricing_rule": _pricing_rule_to_dict(result["pricing_rule"]) if result["pricing_rule"] else None,
        "message": result["message"],
    }

@app.post("/api/kb/candidates/batch_promote")
async def batch_promote_kb_candidates(payload: KnowledgeCandidateBatchPromoteRequest, db: Session = Depends(get_db)):
    candidate_ids: list[str] = []
    seen_ids: set[str] = set()
    for raw in payload.candidate_ids:
        candidate_id = str(raw or "").strip()
        if candidate_id and candidate_id not in seen_ids:
            candidate_ids.append(candidate_id)
            seen_ids.add(candidate_id)
    if not candidate_ids:
        raise HTTPException(status_code=400, detail="candidate_ids is required")

    candidates = (
        db.query(KnowledgeCandidate)
        .filter(KnowledgeCandidate.candidate_id.in_(candidate_ids))
        .all()
    )
    candidate_map = {str(item.candidate_id): item for item in candidates}
    results: list[dict[str, Any]] = []
    promoted_documents: list[KnowledgeDocument] = []

    for candidate_id in candidate_ids:
        candidate = candidate_map.get(candidate_id)
        if not candidate:
            results.append({"candidate_id": candidate_id, "status": "not_found"})
            continue
        if candidate.status != "candidate" and not candidate.promoted_document_id:
            results.append({
                "candidate_id": candidate_id,
                "status": "skipped",
                "reason": f"status={candidate.status}",
                "candidate": _candidate_to_dict(candidate),
            })
            continue
        result = _promote_candidate_record(
            db,
            candidate,
            owner=payload.owner,
            operator=payload.operator,
        )
        if result["document"]:
            promoted_documents.append(result["document"])
        results.append({
            "candidate_id": candidate_id,
            "status": "promoted" if result["chunk"] else "already_promoted",
            "candidate": _candidate_to_dict(candidate),
            "document": _doc_to_dict(result["document"]) if result["document"] else None,
            "message": result["message"],
        })

    db.commit()
    for item in candidates:
        db.refresh(item)
    for doc in promoted_documents:
        db.refresh(doc)

    promoted_count = sum(1 for item in results if item["status"] == "promoted")
    already_promoted_count = sum(1 for item in results if item["status"] == "already_promoted")
    skipped_count = sum(1 for item in results if item["status"] == "skipped")
    not_found_count = sum(1 for item in results if item["status"] == "not_found")
    return {
        "status": "success",
        "requested_count": len(candidate_ids),
        "promoted_count": promoted_count,
        "already_promoted_count": already_promoted_count,
        "skipped_count": skipped_count,
        "not_found_count": not_found_count,
        "results": results,
    }

@app.get("/api/kb/documents")
async def list_kb_documents(
    status: str | None = None,
    stage: str | None = None,
    business_line: str | None = None,
    knowledge_class: str | None = None,
    knowledge_type: str | None = None,
    import_batch: str | None = None,
    review_status: str | None = None,
    source_type: str | None = None,
    risk_level: str | None = None,
    title_keyword: str | None = None,
    content_keyword: str | None = None,
    limit: int = 200,
    offset: int = 0,
    include_meta: bool = False,
    db: Session = Depends(get_db),
):
    query = db.query(KnowledgeDocument)
    knowledge_class_code = normalize_knowledge_class_or_raise(knowledge_class) if knowledge_class else None
    if status:
        if status == "approved":
            query = query.filter(KnowledgeDocument.status == "review", KnowledgeDocument.review_status == "approved")
        else:
            query = query.filter(KnowledgeDocument.status == status)
    if stage:
        if stage == "draft":
            query = query.filter(KnowledgeDocument.status == "draft")
        elif stage == "review":
            query = query.filter(KnowledgeDocument.status == "review", KnowledgeDocument.review_status != "approved")
        elif stage == "approved":
            query = query.filter(KnowledgeDocument.status == "review", KnowledgeDocument.review_status == "approved")
        elif stage in {"review_queue", "pending_review"}:
            query = query.filter(
                or_(
                    KnowledgeDocument.status == "draft",
                    and_(KnowledgeDocument.status == "review", KnowledgeDocument.review_status != "approved"),
                    and_(KnowledgeDocument.status == "review", KnowledgeDocument.review_status == "approved"),
                )
            )
        elif stage == "active":
            query = query.filter(KnowledgeDocument.status == "active")
        elif stage == "archived":
            query = query.filter(KnowledgeDocument.status == "archived")
        elif stage == "rejected":
            query = query.filter(or_(KnowledgeDocument.status == "rejected", KnowledgeDocument.review_status == "rejected"))
        else:
            raise HTTPException(status_code=400, detail="stage 参数不合法")
    if business_line:
        query = query.filter(KnowledgeDocument.business_line == business_line)
    if knowledge_class_code:
        config = KB_KNOWLEDGE_CLASS_CONFIG.get(knowledge_class_code)
        if config:
            query = query.filter(KnowledgeDocument.knowledge_type == config["knowledge_type"])
    if knowledge_type:
        query = query.filter(KnowledgeDocument.knowledge_type == knowledge_type)
    if import_batch:
        query = query.filter(KnowledgeDocument.import_batch == import_batch)
    if review_status:
        query = query.filter(KnowledgeDocument.review_status == review_status)
    if source_type:
        query = query.filter(KnowledgeDocument.source_type == source_type)
    if risk_level:
        query = query.filter(KnowledgeDocument.risk_level == risk_level)
    if title_keyword:
        query = query.filter(KnowledgeDocument.title.ilike(f"%{title_keyword.strip()}%"))
    if content_keyword:
        content_pattern = f"%{content_keyword.strip()}%"
        matched_document_ids = (
            db.query(KnowledgeChunk.document_id)
            .filter(KnowledgeChunk.content.ilike(content_pattern))
            .distinct()
        )
        query = query.filter(KnowledgeDocument.document_id.in_(matched_document_ids))

    query_limit = max(1, min(limit, 500))
    query_offset = max(0, int(offset or 0))
    ordered_query = query.order_by(KnowledgeDocument.created_at.desc())

    if knowledge_class_code:
        result = [
            item for item in (_doc_to_dict(doc) for doc in ordered_query.all())
            if item.get("knowledge_class") == knowledge_class_code
        ]
        total = len(result)
        items = result[query_offset:query_offset + query_limit]
    else:
        total = query.order_by(None).count()
        docs = ordered_query.offset(query_offset).limit(query_limit).all()
        items = [_doc_to_dict(doc) for doc in docs]

    if include_meta:
        return {
            "items": items,
            "total": total,
            "limit": query_limit,
            "offset": query_offset,
            "has_more": query_offset + len(items) < total,
            "next_offset": query_offset + len(items),
        }
    return items

@app.get("/api/kb/documents/{document_id}")
async def get_kb_document(document_id: str, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == document_id).order_by(KnowledgeChunk.chunk_no.asc()).all()
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id == document_id).order_by(PricingRule.created_at.asc()).all()
    snapshots = db.query(KnowledgeVersionSnapshot).filter(
        KnowledgeVersionSnapshot.document_id == document_id
    ).order_by(KnowledgeVersionSnapshot.version_no.desc(), KnowledgeVersionSnapshot.created_at.desc()).all()
    return {
        "document": _doc_to_dict(doc),
        "chunks": [_chunk_to_dict(chunk) for chunk in chunks],
        "pricing_rules": [_pricing_rule_to_dict(rule) for rule in pricing_rules],
        "version_snapshots": [_version_snapshot_to_dict(item) for item in snapshots],
    }

@app.get("/api/kb/chunks")
async def list_kb_chunks(
    status: str | None = None,
    business_line: str | None = None,
    knowledge_class: str | None = None,
    knowledge_type: str | None = None,
    chunk_type: str | None = None,
    language_pair: str | None = None,
    service_scope: str | None = None,
    customer_tier: str | None = None,
    document_id: str | None = None,
    title_keyword: str | None = None,
    content_keyword: str | None = None,
    limit: int = 200,
    offset: int = 0,
    include_meta: bool = False,
    db: Session = Depends(get_db),
):
    query = db.query(KnowledgeChunk)
    knowledge_class_code = normalize_knowledge_class_or_raise(knowledge_class) if knowledge_class else None
    if status:
        query = query.filter(KnowledgeChunk.status == status)
    if business_line:
        query = query.filter(KnowledgeChunk.business_line == business_line)
    if knowledge_class_code:
        config = KB_KNOWLEDGE_CLASS_CONFIG.get(knowledge_class_code)
        if config:
            query = query.filter(
                KnowledgeChunk.knowledge_type == config["knowledge_type"],
                KnowledgeChunk.chunk_type == config["chunk_type"],
            )
    if knowledge_type:
        query = query.filter(KnowledgeChunk.knowledge_type == knowledge_type)
    if chunk_type:
        query = query.filter(KnowledgeChunk.chunk_type == chunk_type)
    if language_pair:
        query = query.filter(KnowledgeChunk.language_pair == language_pair)
    if service_scope:
        query = query.filter(KnowledgeChunk.service_scope == service_scope)
    if customer_tier:
        query = query.filter(KnowledgeChunk.customer_tier == customer_tier)
    if document_id:
        query = query.filter(KnowledgeChunk.document_id == document_id)
    if title_keyword:
        query = query.filter(KnowledgeChunk.title.ilike(f"%{title_keyword.strip()}%"))
    if content_keyword:
        query = query.filter(KnowledgeChunk.content.ilike(f"%{content_keyword.strip()}%"))

    query_limit = max(1, min(limit, 500))
    query_offset = max(0, int(offset or 0))
    ordered_query = query.order_by(KnowledgeChunk.created_at.desc(), KnowledgeChunk.chunk_no.asc())
    total = query.order_by(None).count()
    chunks = ordered_query.offset(query_offset).limit(query_limit).all()
    items = [_chunk_to_dict(chunk) for chunk in chunks]

    if include_meta:
        return {
            "items": items,
            "total": total,
            "limit": query_limit,
            "offset": query_offset,
            "has_more": query_offset + len(items) < total,
            "next_offset": query_offset + len(items),
        }
    return items

@app.get("/api/kb/chunks/{chunk_id}")
async def get_kb_chunk(chunk_id: str, db: Session = Depends(get_db)):
    chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id == chunk_id).first()
    if not chunk:
        raise HTTPException(status_code=404, detail="Knowledge chunk not found")
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == chunk.document_id).first()
    pricing_rules = db.query(PricingRule).filter(
        PricingRule.document_id == chunk.document_id,
        or_(PricingRule.chunk_id == chunk.chunk_id, PricingRule.chunk_id.is_(None)),
    ).order_by(PricingRule.created_at.asc()).all()
    return {
        "chunk": _chunk_to_dict(chunk),
        "document": _doc_to_dict(doc) if doc else None,
        "pricing_rules": [_pricing_rule_to_dict(rule) for rule in pricing_rules],
    }

@app.patch("/api/kb/documents/{document_id}")
async def update_kb_document(document_id: str, payload: KnowledgeDocumentUpdate, db: Session = Depends(get_db)):
    """人工审核时修改文档级信息，保持 draft/pending。"""
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    if payload.title is not None:
        doc.title = payload.title
    if payload.business_line is not None:
        doc.business_line = payload.business_line
    if payload.sub_service is not None:
        doc.sub_service = payload.sub_service
    if payload.review_required is not None:
        doc.review_required = payload.review_required
    if payload.effective_from is not None:
        doc.effective_from = payload.effective_from
    if payload.effective_to is not None:
        doc.effective_to = payload.effective_to

    knowledge_class = normalize_knowledge_class_or_raise(payload.knowledge_class)
    merged_tags = payload.tags if payload.tags is not None else doc.tags
    if knowledge_class:
        _class_code, resolved_type, resolved_chunk_type, resolved_risk = resolve_knowledge_class_fields(
            knowledge_class=knowledge_class,
            knowledge_type=payload.knowledge_type or doc.knowledge_type,
            chunk_type=None,
            risk_level=payload.risk_level or doc.risk_level,
        )
        doc.knowledge_type = resolved_type
        doc.risk_level = resolved_risk if payload.risk_level is None else payload.risk_level
        doc.tags = merge_knowledge_class_tags(merged_tags, knowledge_class)
        chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
        if len(chunks) == 1:
            chunks[0].knowledge_type = resolved_type
            chunks[0].chunk_type = resolved_chunk_type
            chunks[0].structured_tags = merge_knowledge_class_tags(chunks[0].structured_tags, knowledge_class)
    else:
        if payload.knowledge_type is not None:
            doc.knowledge_type = payload.knowledge_type
        if payload.risk_level is not None:
            doc.risk_level = payload.risk_level
        if payload.tags is not None:
            doc.tags = payload.tags
    doc.effective_from, doc.effective_to = default_pricing_effective_window(
        doc.effective_from,
        doc.effective_to,
        knowledge_type=doc.knowledge_type,
    )
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == doc.document_id).all()
    for chunk in chunks:
        chunk.effective_from = doc.effective_from
        chunk.effective_to = doc.effective_to
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id == doc.document_id).all()
    for rule in pricing_rules:
        rule.effective_from = doc.effective_from
        rule.effective_to = doc.effective_to
    if doc.status == "draft":
        doc.review_status = "pending" if doc.review_required else "auto_ready"
    db.commit()
    db.refresh(doc)
    return {"status": "success", "document": _doc_to_dict(doc)}

@app.patch("/api/kb/chunks/{chunk_id}")
async def update_kb_chunk(chunk_id: str, payload: KnowledgeChunkUpdate, db: Session = Depends(get_db)):
    """人工审核时修改切片，标题或正文变化会重算 embedding。"""
    chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id == chunk_id).first()
    if not chunk:
        raise HTTPException(status_code=404, detail="Knowledge chunk not found")

    old_title = chunk.title
    old_content = chunk.content
    if payload.title is not None:
        chunk.title = _strip_retrieval_title_prefix(payload.title)
    if payload.content is not None:
        chunk.content = payload.content
    if payload.business_line is not None:
        chunk.business_line = payload.business_line
    if payload.sub_service is not None:
        chunk.sub_service = payload.sub_service
    if payload.language_pair is not None:
        chunk.language_pair = payload.language_pair
    if payload.service_scope is not None:
        chunk.service_scope = payload.service_scope
    if payload.region is not None:
        chunk.region = payload.region
    if payload.customer_tier is not None:
        chunk.customer_tier = payload.customer_tier
    if payload.priority is not None:
        chunk.priority = payload.priority

    knowledge_class = normalize_knowledge_class_or_raise(payload.knowledge_class)
    if knowledge_class:
        _class_code, resolved_type, resolved_chunk_type, _resolved_risk = resolve_knowledge_class_fields(
            knowledge_class=knowledge_class,
            knowledge_type=payload.knowledge_type or chunk.knowledge_type,
            chunk_type=payload.chunk_type or chunk.chunk_type,
            risk_level=None,
        )
        chunk.knowledge_type = resolved_type
        chunk.chunk_type = resolved_chunk_type
        chunk.structured_tags = merge_knowledge_class_tags(
            payload.structured_tags if payload.structured_tags is not None else chunk.structured_tags,
            knowledge_class,
        )
    else:
        if payload.chunk_type is not None:
            chunk.chunk_type = payload.chunk_type
        if payload.knowledge_type is not None:
            chunk.knowledge_type = payload.knowledge_type
        if payload.structured_tags is not None:
            chunk.structured_tags = payload.structured_tags

    if chunk.title != old_title or chunk.content != old_content:
        retrieval_text = _build_chunk_retrieval_text(chunk.title, chunk.content)
        embedding = _optional_embedding_for_storage(retrieval_text, context=f"update_chunk:{chunk.chunk_id}")
        chunk.keyword_text = retrieval_text
        chunk.embedding = embedding
        chunk.embedding_provider = settings.EMBEDDING_PROVIDER if embedding else None
        chunk.embedding_model = settings.EMBEDDING_MODEL if embedding else None
        chunk.embedding_dim = len(embedding) if embedding else None

    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == chunk.document_id).first()
    _apply_chunk_governance(
        chunk,
        source_type=doc.source_type if doc else "manual",
        source_ref=doc.source_ref if doc else None,
    )
    _sync_single_chunk_document_from_chunk(db, chunk)
    db.commit()
    db.refresh(chunk)
    return {"status": "success", "chunk": _chunk_to_dict(chunk)}

@app.post("/api/kb/chunks/{chunk_id}/pricing_rule")
async def create_kb_chunk_pricing_rule(chunk_id: str, payload: PricingRulePayload, db: Session = Depends(get_db)):
    """为某条报价切片补充结构化报价规则。"""
    chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id == chunk_id).first()
    if not chunk:
        raise HTTPException(status_code=404, detail="Knowledge chunk not found")
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == chunk.document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    rule = _create_pricing_rule(db, doc, chunk, payload)
    _apply_chunk_governance(chunk, source_type=doc.source_type, pricing_rule=payload.model_dump(), source_ref=doc.source_ref)
    _sync_single_chunk_document_from_chunk(db, chunk)
    db.commit()
    db.refresh(rule)
    return {"status": "success", "pricing_rule": _pricing_rule_to_dict(rule)}

@app.get("/api/kb/pricing_rules")
async def list_kb_pricing_rules(
    status: str | None = None,
    business_line: str | None = None,
    language_pair: str | None = None,
    service_scope: str | None = None,
    customer_tier: str | None = None,
    document_id: str | None = None,
    rule_ids: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    query = db.query(PricingRule)
    if status:
        query = query.filter(PricingRule.status == status)
    if business_line:
        query = query.filter(PricingRule.business_line == business_line)
    if language_pair:
        query = query.filter(PricingRule.language_pair == language_pair)
    if service_scope:
        query = query.filter(PricingRule.service_scope == service_scope)
    if customer_tier:
        query = query.filter(PricingRule.customer_tier == customer_tier)
    if document_id:
        query = query.filter(PricingRule.document_id == document_id)
    if rule_ids:
        parsed_rule_ids = [item.strip() for item in rule_ids.split(",") if item.strip()]
        if parsed_rule_ids:
            query = query.filter(PricingRule.rule_id.in_(parsed_rule_ids))
    rules = query.order_by(PricingRule.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return [_pricing_rule_to_dict(rule) for rule in rules]

@app.patch("/api/kb/pricing_rules/{rule_id}")
async def update_kb_pricing_rule(rule_id: str, payload: PricingRulePayload, db: Session = Depends(get_db)):
    rule = db.query(PricingRule).filter(PricingRule.rule_id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Pricing rule not found")
    for field in [
        "business_line", "sub_service", "language_pair", "service_scope", "unit", "currency",
        "tax_policy", "customer_tier", "region", "source_ref"
    ]:
        value = getattr(payload, field)
        if value is not None:
            setattr(rule, field, value)
    for field in ["price_min", "price_max", "urgent_multiplier", "min_charge"]:
        value = getattr(payload, field)
        if value is not None:
            setattr(rule, field, _decimal_or_none(value))
    chunk = None
    if rule.chunk_id:
        chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.chunk_id == rule.chunk_id).first()
    elif rule.document_id:
        chunk = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id == rule.document_id).order_by(KnowledgeChunk.chunk_no.asc()).first()
    if chunk:
        doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == chunk.document_id).first()
        _apply_chunk_governance(
            chunk,
            source_type=doc.source_type if doc else "manual",
            pricing_rule=payload.model_dump(),
            source_ref=doc.source_ref if doc else None,
        )
        _sync_single_chunk_document_from_chunk(db, chunk)
    db.commit()
    db.refresh(rule)
    return {"status": "success", "pricing_rule": _pricing_rule_to_dict(rule)}

@app.post("/api/kb/pricing_rules/{rule_id}/archive")
async def archive_kb_pricing_rule(rule_id: str, db: Session = Depends(get_db)):
    rule = db.query(PricingRule).filter(PricingRule.rule_id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="Pricing rule not found")
    rule.status = "archived"
    db.commit()
    db.refresh(rule)
    return {"status": "success", "pricing_rule": _pricing_rule_to_dict(rule)}

@app.get("/api/kb/hit_logs")
async def list_kb_hit_logs(
    status: str | None = None,
    session_id: str | None = None,
    request_id: str | None = None,
    redact: bool = False,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    query = db.query(KnowledgeHitLog)
    if status:
        query = query.filter(KnowledgeHitLog.status == status)
    if session_id:
        query = query.filter(KnowledgeHitLog.session_id == session_id)
    if request_id:
        query = query.filter(KnowledgeHitLog.request_id == request_id)
    logs = query.order_by(KnowledgeHitLog.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return [_hit_log_to_dict(log, redact=redact) for log in logs]

@app.get("/api/kb/hit_logs/stats")
async def get_kb_hit_log_stats(db: Session = Depends(get_db)):
    from sqlalchemy import func
    total = db.query(func.count(KnowledgeHitLog.log_id)).scalar() or 0
    by_status = {
        row.status: row.count
        for row in db.query(KnowledgeHitLog.status, func.count(KnowledgeHitLog.log_id).label("count"))
        .group_by(KnowledgeHitLog.status)
        .all()
    }
    by_quality = {
        row.retrieval_quality or "unknown": row.count
        for row in db.query(KnowledgeHitLog.retrieval_quality, func.count(KnowledgeHitLog.log_id).label("count"))
        .group_by(KnowledgeHitLog.retrieval_quality)
        .all()
    }
    manual_review_count = db.query(func.count(KnowledgeHitLog.log_id)).filter(KnowledgeHitLog.manual_review_required == True).scalar() or 0
    insufficient_count = db.query(func.count(KnowledgeHitLog.log_id)).filter(KnowledgeHitLog.insufficient_info == True).scalar() or 0
    avg_latency = db.query(func.avg(KnowledgeHitLog.latency_ms)).scalar()
    return {
        "total": total,
        "by_status": by_status,
        "by_retrieval_quality": by_quality,
        "manual_review_count": manual_review_count,
        "insufficient_count": insufficient_count,
        "avg_latency_ms": round(float(avg_latency), 2) if avg_latency is not None else None,
    }

@app.post("/api/kb/hit_logs/{log_id}/feedback")
async def update_kb_hit_log_feedback(log_id: str, payload: KnowledgeHitFeedback, db: Session = Depends(get_db)):
    log = db.query(KnowledgeHitLog).filter(KnowledgeHitLog.log_id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Knowledge hit log not found")
    log.feedback_status = payload.feedback_status
    log.manual_feedback = payload.manual_feedback
    candidate = None
    if payload.feedback_status == "needs_fix":
        candidate = _create_candidate_from_feedback_log(
            db,
            log,
            owner="sales_feedback",
            operator="kb_hit_log_feedback",
            note=(payload.manual_feedback or {}).get("note") if isinstance(payload.manual_feedback, dict) else None,
        )
    _update_effect_rollup_for_log(db, log=log, feedback_status=payload.feedback_status, manual_feedback=payload.manual_feedback)
    db.commit()
    db.refresh(log)
    if candidate:
        db.refresh(candidate)
    return {"status": "success", "log": _hit_log_to_dict(log), "candidate": _candidate_to_dict(candidate) if candidate else None}

@app.post("/api/assist/feedback")
async def submit_assist_feedback(payload: AssistFeedback, db: Session = Depends(get_db)):
    log = None
    if payload.snapshot_id and payload.snapshot_id != "current_tail":
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == payload.snapshot_id,
            ReplyChainSnapshot.session_id == payload.session_id,
        ).first()
        if snapshot and snapshot.knowledge_log_id:
            log = db.query(KnowledgeHitLog).filter(
                KnowledgeHitLog.log_id == snapshot.knowledge_log_id
            ).first()
    if not log:
        log = db.query(KnowledgeHitLog).filter(
            KnowledgeHitLog.session_id == payload.session_id
        ).order_by(KnowledgeHitLog.created_at.desc()).first()
    if not log:
        raise HTTPException(status_code=404, detail="No knowledge hit log found for session")
    log.feedback_status = payload.feedback_status
    log.manual_feedback = payload.manual_feedback
    if payload.final_response:
        log.final_response = payload.final_response
    candidate = None
    if payload.feedback_status == "needs_fix":
        candidate = _create_candidate_from_feedback_log(
            db,
            log,
            owner="sales_feedback",
            operator="frontend_sidebar",
            note=(payload.manual_feedback or {}).get("note") if isinstance(payload.manual_feedback, dict) else None,
        )
    _update_effect_rollup_for_log(db, log=log, feedback_status=payload.feedback_status, manual_feedback=payload.manual_feedback)
    excellent_reply_candidates: list[KnowledgeCandidate] = []
    auto_training = None
    if payload.feedback_status in POSITIVE_FEEDBACK_STATUSES:
        _backfill_thread_business_facts(db, session_ids=[payload.session_id])
        excellent_reply_candidates = _extract_excellent_reply_candidates(
            db,
            session_id=payload.session_id,
            owner="assist_feedback_auto",
            operator="assist_feedback_auto",
            force=True,
        )
        auto_training = _prepare_training_samples(
            db,
            TrainingSamplePrepareRequest(
                sample_types=["reply_fragment_sft", "thread_reply_sft", "retrieval_pair"],
                min_quality_score=0.45,
                min_effect_score=0.45,
                limit_per_source=50,
                operator="assist_feedback_auto",
            ),
        )
    db.commit()
    db.refresh(log)
    if candidate:
        db.refresh(candidate)
    for item in excellent_reply_candidates:
        db.refresh(item)
    return {
        "status": "success",
        "log": _hit_log_to_dict(log, redact=True),
        "candidate": _candidate_to_dict(candidate) if candidate else None,
        "excellent_reply_candidates": [_candidate_to_dict(item) for item in excellent_reply_candidates],
        "auto_training": auto_training,
    }

@app.post("/api/assist/extract_excellent_reply")
async def extract_excellent_reply(payload: ExcellentReplyExtractRequest, db: Session = Depends(get_db)):
    created = _extract_excellent_reply_candidates(
        db,
        session_id=payload.session_id,
        owner=payload.owner,
        operator=payload.operator,
        force=payload.force,
    )
    db.commit()
    for item in created:
        db.refresh(item)
    return {
        "status": "success",
        "created_count": len(created),
        "candidates": [_candidate_to_dict(item) for item in created],
    }

def _execute_regression_cases(payload: KnowledgeRegressionRunRequest, report: Any = None) -> dict:
    cases = _filter_regression_cases(
        _load_regression_cases(),
        case_ids=payload.case_ids,
        groups=payload.groups,
        category=payload.category,
        risk_level=payload.risk_level,
        business_line=payload.business_line,
    )
    if not cases:
        raise HTTPException(status_code=404, detail="No regression cases matched")

    run_id = payload.run_id or f"kb_reg_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    results = []
    request_ids = []
    total_cases = len(cases)
    for index, case in enumerate(cases, start=1):
        if report:
            report(
                progress=min(95, round(index * 100 / max(total_cases, 1))),
                summary=f"running {case.get('case_id')}",
                result_patch={"current_case_id": case.get("case_id"), "total_cases": total_cases},
            )
        case_id = case["case_id"]
        request_id = f"{run_id}_{case_id}"
        request_ids.append(request_id)
        query_features = dict(case.get("query_features") or {})
        if payload.min_score is not None:
            query_features["min_score"] = payload.min_score
        result = IntentEngine.retrieve_knowledge_v2(
            query_text=case.get("query_text") or "",
            query_features=query_features,
            top_k=max(1, min(payload.top_k, 20)),
            request_id=request_id,
            session_id=run_id,
        )
        evaluation = _evaluate_regression_case(case, result)
        hit_summary = [
            {
                "chunk_id": hit.get("chunk_id"),
                "document_id": hit.get("document_id"),
                "knowledge_type": hit.get("knowledge_type"),
                "business_line": hit.get("business_line"),
                "language_pair": hit.get("language_pair"),
                "service_scope": hit.get("service_scope"),
                "score": hit.get("score"),
                "pricing_rule_count": len(hit.get("pricing_rules") or []),
            }
            for hit in result.get("hits", [])
        ]
        item = {
            "case_id": case_id,
            "query_text": case.get("query_text"),
            "group": case.get("group"),
            "risk_level": case.get("risk_level"),
            "category": case.get("category"),
            "passed": evaluation["passed"],
            "status": result.get("status"),
            "retrieval_quality": result.get("retrieval_quality"),
            "confidence_score": result.get("confidence_score"),
            "manual_review_required": result.get("manual_review_required"),
            "failures": evaluation["failures"],
            "warnings": evaluation["warnings"],
            "log_id": result.get("log_id"),
            "hit_count": len(result.get("hits", [])),
            "hits": hit_summary if payload.include_hits else [],
        }
        results.append(item)

    passed_count = sum(1 for item in results if item["passed"])
    failed_count = len(results) - passed_count
    cleanup = {"enabled": payload.cleanup_logs, "deleted_hit_logs": 0}
    if payload.cleanup_logs:
        db = SessionLocal()
        try:
            cleanup["deleted_hit_logs"] = db.query(KnowledgeHitLog).filter(KnowledgeHitLog.request_id.in_(request_ids)).delete(
                synchronize_session=False
            )
            db.commit()
        finally:
            db.close()

    summary_by_group = {}
    for item in results:
        key = item.get("group") or "default"
        bucket = summary_by_group.setdefault(key, {"total": 0, "passed": 0, "failed": 0})
        bucket["total"] += 1
        if item["passed"]:
            bucket["passed"] += 1
        else:
            bucket["failed"] += 1

    return {
        "status": "success" if failed_count == 0 else "failed",
        "run_id": run_id,
        "total": len(results),
        "passed": passed_count,
        "failed": failed_count,
        "pass_rate": round((passed_count / len(results)) * 100, 2) if results else 0,
        "cleanup": cleanup,
        "summary_by_group": summary_by_group,
        "results": results,
    }

def _run_regression_job(report: Any, payload_data: dict) -> dict:
    payload = KnowledgeRegressionRunRequest(**payload_data)
    return _execute_regression_cases(payload, report=report)

@app.get("/api/kb/regression_cases")
async def list_kb_regression_cases(
    group: str | None = None,
    category: str | None = None,
    risk_level: str | None = None,
    business_line: str | None = None,
):
    cases = _filter_regression_cases(
        _load_regression_cases(),
        groups=[group] if group else None,
        category=category,
        risk_level=risk_level,
        business_line=business_line,
    )
    meta = {
        "groups": sorted({case.get("group") for case in _load_regression_cases() if case.get("group")}),
        "categories": sorted({case.get("category") for case in _load_regression_cases() if case.get("category")}),
        "risk_levels": sorted({case.get("risk_level") for case in _load_regression_cases() if case.get("risk_level")}),
        "business_lines": sorted({
            ((case.get("query_features") or {}).get("business_line") or case.get("business_line"))
            for case in _load_regression_cases()
            if ((case.get("query_features") or {}).get("business_line") or case.get("business_line"))
        }),
    }
    return {
        "status": "success",
        "count": len(cases),
        "meta": meta,
        "cases": cases,
    }

@app.post("/api/kb/regression_cases/run")
async def run_kb_regression_cases(payload: KnowledgeRegressionRunRequest):
    return _execute_regression_cases(payload)

@app.post("/api/kb/regression_cases/run_async")
async def run_kb_regression_cases_async(payload: KnowledgeRegressionRunRequest, db: Session = Depends(get_db)):
    job = _create_job_task(
        db,
        job_type="kb_regression",
        task_payload=payload.model_dump(),
        summary="queued regression run",
    )
    db.commit()
    db.refresh(job)
    start_job(str(job.job_id), _run_regression_job, payload.model_dump())
    return {"status": "queued", "job": _job_to_dict(job)}

@app.get("/api/jobs")
async def list_jobs(job_type: str | None = None, status: str | None = None, limit: int = 100, db: Session = Depends(get_db)):
    query = db.query(JobTask)
    if job_type:
        query = query.filter(JobTask.job_type == job_type)
    if status:
        query = query.filter(JobTask.status == status)
    jobs = query.order_by(JobTask.created_at.desc()).limit(max(1, min(limit, 300))).all()
    return {
        "status": "success",
        "count": len(jobs),
        "jobs": [_job_to_dict(job) for job in jobs],
    }

@app.get("/api/jobs/{job_id}")
async def get_job(job_id: str, db: Session = Depends(get_db)):
    job = db.query(JobTask).filter(JobTask.job_id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"status": "success", "job": _job_to_dict(job)}

@app.post("/api/jobs/{job_id}/retry")
async def retry_job(job_id: str, payload: JobRetryRequest | None = None, db: Session = Depends(get_db)):
    job = db.query(JobTask).filter(JobTask.job_id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    retry_payload = dict(job.task_payload or {})
    if payload and payload.operator:
        retry_payload["operator"] = payload.operator
    retry_job_record = _create_job_task(
        db,
        job_type=job.job_type,
        task_payload=retry_payload,
        summary=f"retry {job.job_type}",
        retry_of=str(job.job_id),
    )
    db.commit()
    db.refresh(retry_job_record)
    if job.job_type == "kb_regression":
        start_job(str(retry_job_record.job_id), _run_regression_job, retry_payload)
    elif job.job_type == "kb_excel_import":
        start_job(str(retry_job_record.job_id), _run_excel_import_job, retry_payload)
    elif job.job_type == "training_pipeline":
        start_job(str(retry_job_record.job_id), _run_training_pipeline_job, retry_payload)
    elif job.job_type == "analysis_completion":
        start_job(str(retry_job_record.job_id), _run_complete_analysis_items_job, retry_payload)
    else:
        raise HTTPException(status_code=422, detail=f"Retry not supported for job_type={job.job_type}")
    return {"status": "queued", "job": _job_to_dict(retry_job_record)}

@app.get("/api/kb/quality/report")
async def get_kb_quality_report(days: int = 30, limit: int = 300, db: Session = Depends(get_db)):
    return build_kb_quality_report(db, days=days, limit=limit)

@app.get("/api/kb/quality/conflicts")
async def get_kb_quality_conflicts(limit: int = 300, db: Session = Depends(get_db)):
    report = build_kb_quality_report(db, days=30, limit=limit)
    report["issues"] = [
        issue for issue in report["issues"]
        if issue["type"] in {"pricing_conflict", "duplicate_pricing_rule", "missing_pricing_rule"}
    ]
    report["total_issues"] = len(report["issues"])
    return report

@app.get("/api/kb/quality/expiring")
async def get_kb_quality_expiring(days: int = 30, limit: int = 300, db: Session = Depends(get_db)):
    report = build_kb_quality_report(db, days=days, limit=limit)
    report["issues"] = [
        issue for issue in report["issues"]
        if issue["type"] in {"expired_document", "expiring_document", "expired_pricing_rule", "expiring_pricing_rule"}
    ]
    report["total_issues"] = len(report["issues"])
    return report

@app.get("/api/kb/performance/status")
async def get_kb_performance_status(db: Session = Depends(get_db)):
    index_rows = db.execute(text(
        "SELECT tablename, indexname FROM pg_indexes "
        "WHERE tablename IN ('knowledge_chunk','knowledge_document','pricing_rule','knowledge_hit_logs') "
        "ORDER BY tablename, indexname"
    )).mappings().all()
    extensions = {
        row["extname"]: True
        for row in db.execute(text("SELECT extname FROM pg_extension WHERE extname IN ('vector','pg_trgm')")).mappings().all()
    }
    chunk_count = db.query(func.count(KnowledgeChunk.chunk_id)).scalar() or 0
    active_chunk_count = db.query(func.count(KnowledgeChunk.chunk_id)).filter(KnowledgeChunk.status == "active").scalar() or 0
    active_pricing_rule_count = db.query(func.count(PricingRule.rule_id)).filter(PricingRule.status == "active").scalar() or 0
    return {
        "status": "success",
        "retrieval_strategy": {
            "pgvector_enabled": settings.PGVECTOR_ENABLED,
            "pgvector_required": settings.PGVECTOR_REQUIRED,
            "pgvector_extension_available": bool(extensions.get("vector")),
            "embedding_storage": "json",
            "fulltext_index_enabled": settings.KB_FULLTEXT_INDEX_ENABLED,
            "keyword_prefilter_enabled": settings.KB_KEYWORD_PREFILTER_ENABLED,
            "candidate_limit": settings.KB_CANDIDATE_LIMIT,
        },
        "counts": {
            "knowledge_chunks": chunk_count,
            "active_knowledge_chunks": active_chunk_count,
            "active_pricing_rules": active_pricing_rule_count,
        },
        "indexes": [dict(row) for row in index_rows],
    }

def _is_placeholder_value(value: str | None) -> bool:
    text_value = str(value or "").strip()
    if not text_value:
        return True
    placeholders = {"your-corp-id", "your-corp-secret", "your-agent-id", "your-token", "your-token",
                    "your-encoding-aes-key", "your-chatdata-secret", "your-db-name", "your-db-user",
                    "your-db-password", "your-llm1-api-key", "your-llm2-api-key"}
    return text_value in placeholders or text_value.startswith("your-")

def _mask_config_value(value: str | None, keep: int = 4) -> str | None:
    if value is None:
        return None
    text_value = str(value)
    if len(text_value) <= keep:
        return "*" * len(text_value)
    return f"{text_value[:keep]}***"

def _load_runtime_env_file() -> dict[str, str]:
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    try:
        mtime = os.path.getmtime(env_path)
    except OSError:
        return {}
    if _RUNTIME_ENV_CACHE.get("mtime") == mtime:
        return dict(_RUNTIME_ENV_CACHE.get("values") or {})
    values: dict[str, str] = {}
    try:
        with open(env_path, "r", encoding="utf-8-sig") as env_file:
            for raw_line in env_file:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()
                if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
                    value = value[1:-1]
                values[key] = value
    except OSError:
        return {}
    _RUNTIME_ENV_CACHE["mtime"] = mtime
    _RUNTIME_ENV_CACHE["values"] = values
    return dict(values)

def _runtime_setting(name: str, current: Any = None) -> str:
    env_value = os.environ.get(name)
    if str(env_value or "").strip():
        return str(env_value).strip()
    file_value = str(_load_runtime_env_file().get(name) or "").strip()
    if file_value and not _is_placeholder_value(file_value):
        return file_value
    current_value = str(current or "").strip()
    if current_value and not _is_placeholder_value(current_value):
        return current_value
    return current_value

def _runtime_int_setting(name: str, current: Any = None, default: int = 100) -> int:
    value = _runtime_setting(name, current)
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(current)
        except (TypeError, ValueError):
            return default

def _runtime_bool_setting(name: str, current: Any = None, default: bool = False) -> bool:
    value = str(_runtime_setting(name, current)).strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    try:
        return bool(current)
    except Exception:
        return default

def _external_api_base_url() -> str:
    return str(_runtime_setting("EXTERNAL_API_BASE_URL", settings.EXTERNAL_API_BASE_URL) or "").strip().rstrip("/")


def _sales_kb_api_base_url() -> str:
    return str(_runtime_setting("SALES_KB_API_BASE_URL", settings.SALES_KB_API_BASE_URL) or "").strip().rstrip("/")


def _sales_kb_api_timeout_seconds() -> int:
    return max(3, _runtime_int_setting("SALES_KB_API_TIMEOUT_SECONDS", settings.SALES_KB_API_TIMEOUT_SECONDS, default=8))


def _sales_kb_probe_and_suggestions() -> tuple[dict, list[str]]:
    base_url = _sales_kb_api_base_url()
    endpoint = f"{base_url}/kb-units/qa-retrieve" if base_url else ""
    probe = {
        "status": "error",
        "configured": bool(base_url),
        "reachable": False,
        "base_url": base_url,
        "probe_url": endpoint,
        "timeout_seconds": settings.KB_HEALTHCHECK_TIMEOUT_SECONDS,
    }
    suggestions: list[str] = []
    if not base_url:
        suggestions.append("补齐 SALES_KB_API_BASE_URL，并指向反向代理后的 HTTPS 域名。")
        return probe, suggestions

    parsed = urlparse(base_url)
    host = str(parsed.hostname or "").strip().lower()
    scheme = str(parsed.scheme or "").strip().lower()
    probe["scheme"] = scheme or "-"
    probe["host"] = host or "-"
    if scheme != "https":
        suggestions.append("知识库2当前应统一走 HTTPS 反向代理域名，不要继续使用 http。")
    if host in LEGACY_SALES_KB_HOSTS:
        suggestions.append("知识库2不应再指向旧内网 IP，请改为 https://knowledgebase.speedasia.net。")
    if parsed.port:
        suggestions.append("知识库2当前反向代理入口不需要显式端口，请去掉端口号。")
        probe["port"] = parsed.port
    if suggestions:
        probe["error"] = "知识库2基址仍为旧格式"
        return probe, suggestions

    try:
        session = requests.Session()
        session.trust_env = settings.HTTP_TRUST_ENV
        response = session.post(
            endpoint,
            json={"question": "付款"},
            timeout=settings.KB_HEALTHCHECK_TIMEOUT_SECONDS,
        )
        probe["http_status"] = response.status_code
        if response.status_code != 200:
            probe["error"] = f"HTTP {response.status_code}"
            suggestions.append("检查 knowledgebase.speedasia.net 的反向代理、证书和 /kb-units/qa-retrieve 是否正常。")
            return probe, suggestions
        payload = response.json()
        answers = payload.get("answers") if isinstance(payload, dict) else None
        if not isinstance(answers, list):
            probe["error"] = "响应格式不是 {answers:[...]}"
            suggestions.append("检查知识库2 Q&A 接口返回结构，当前页面和后台都依赖 {answers:[...]}。")
            return probe, suggestions
        probe["status"] = "ok"
        probe["reachable"] = True
        probe["answer_count"] = len(answers)
        if answers:
            top_answer = answers[0] if isinstance(answers[0], dict) else {}
            probe["top_kb_unit_id"] = top_answer.get("kb_unit_id")
            probe["top_score"] = top_answer.get("score")
        return probe, suggestions
    except Exception as exc:
        probe["error"] = sanitize_text(str(exc))
        suggestions.append("检查 knowledgebase.speedasia.net 是否可从当前服务器直连，以及反向代理目标服务是否在线。")
        return probe, suggestions


def _sales_kb_unit_type_label(unit_type: str | None) -> str:
    mapping = {
        "customer_profile": "客户画像",
        "product_intent": "产品意向",
        "quotation": "报价信息",
        "order": "订单/PO",
        "shipment": "发货/物流",
        "after_sales": "售后问题",
    }
    return mapping.get(str(unit_type or "").strip(), str(unit_type or "外部知识"))


def _normalize_sales_kb_qa_answer(item: dict, idx: int) -> dict:
    title = sanitize_text(str(item.get("title") or f"参考回复 {idx + 1}").strip())
    seller_response_text = sanitize_text(str(item.get("seller_response_text") or "").strip())
    kb_unit_id = item.get("kb_unit_id")
    score = item.get("score")
    try:
        useful_score = round(float(score) * 100, 2) if score is not None else None
    except (TypeError, ValueError):
        useful_score = None
    content_parts = []
    if seller_response_text:
        content_parts.append(f"销售回复参考：{seller_response_text}")
    if kb_unit_id not in (None, ""):
        content_parts.append(f"来源知识单元：{kb_unit_id}")
    content = "\n".join(content_parts).strip() or title
    return {
        "title": title,
        "content": content,
        "score": score,
        "knowledge_type": "qa_answer",
        "knowledge_type_label": "Q&A回复",
        "chunk_type": "qa_answer",
        "chunk_type_label": "Q&A直答",
        "business_line": "external_api",
        "business_line_label": "外部销售知识库",
        "service_scope": None,
        "service_scope_label": "Q&A直答",
        "usable_for_reply": bool(seller_response_text),
        "allowed_for_generation": bool(seller_response_text),
        "manual_review_required": False,
        "library_type": "external_api",
        "useful_score": useful_score,
        "pricing_rules": [],
        "summary": None,
        "seller_response_text": seller_response_text or None,
        "hit_reason": None,
        "confidence": score,
        "unit_id": kb_unit_id,
    }


def _search_sales_kb_api(query_text: str | None, top_k: int = 5) -> dict:
    normalized_query = sanitize_text(str(query_text or "").strip())
    if not normalized_query:
        return {
            "status": "skipped_no_query",
            "query_text": "",
            "hits": [],
            "replyable_hits": [],
            "human_only_hits": [],
            "manual_review_required": False,
            "confidence_score": None,
            "filters_used": {},
            "evidence_refs": [],
            "source": "sales_kb_api",
            "how": "直接调用销售知识库 API 的 POST /kb-units/qa-retrieve 接口，返回候选销售回复。",
        }

    base_url = _sales_kb_api_base_url()
    if not base_url:
        return {
            "status": "not_configured",
            "query_text": normalized_query,
            "hits": [],
            "replyable_hits": [],
            "human_only_hits": [],
            "manual_review_required": False,
            "confidence_score": None,
            "filters_used": {},
            "evidence_refs": [],
            "source": "sales_kb_api",
            "error": "未配置 SALES_KB_API_BASE_URL",
            "how": "直接调用销售知识库 API 的 POST /kb-units/qa-retrieve 接口，返回候选销售回复。",
        }

    started = perf_counter()
    endpoint = f"{base_url}/kb-units/qa-retrieve"
    limit = max(1, min(int(top_k or 5), 20))
    try:
        session = requests.Session()
        session.trust_env = settings.HTTP_TRUST_ENV
        response = session.post(
            endpoint,
            json={
                "question": normalized_query,
            },
            timeout=_sales_kb_api_timeout_seconds(),
        )
        latency_ms = round((perf_counter() - started) * 1000, 2)
        if response.status_code != 200:
            logger.warning("销售知识库 API 不可用，已跳过。 endpoint=%s status=%s", endpoint, response.status_code)
            return {
                "status": "skipped_unavailable",
                "query_text": normalized_query,
                "hits": [],
                "replyable_hits": [],
                "human_only_hits": [],
                "manual_review_required": False,
                "confidence_score": None,
                "filters_used": {
                    "method": "POST",
                    "response_trim_limit": limit,
                    "endpoint": endpoint,
                },
                "evidence_refs": [],
                "source": "sales_kb_api",
                "latency_ms": latency_ms,
                "skipped": True,
                "skip_reason": f"HTTP {response.status_code}",
                "error": f"HTTP {response.status_code}",
                "how": "直接调用销售知识库 API 的 POST /kb-units/qa-retrieve 接口，返回候选销售回复。",
            }
        payload = response.json()
        answers = payload.get("answers") if isinstance(payload, dict) else None
        if not isinstance(answers, list):
            raise RuntimeError("销售知识库 Q&A API 返回格式不是 {answers:[...]} 对象")
        hits = [_normalize_sales_kb_qa_answer(item, idx) for idx, item in enumerate(answers[:limit]) if isinstance(item, dict)]
        replyable_hits = [item for item in hits if item.get("usable_for_reply")]
        human_only_hits = [item for item in hits if not item.get("usable_for_reply")]
        top_score = None
        if hits:
            try:
                top_score = float(hits[0].get("score")) if hits[0].get("score") is not None else None
            except (TypeError, ValueError):
                top_score = None
        evidence_refs = [
            {
                "ref_id": f"kb_unit:{hit.get('unit_id')}",
                "ref_type": "kb_unit",
                "title": hit.get("title"),
                "snippet": hit.get("seller_response_text") or hit.get("content"),
            }
            for hit in hits
            if hit.get("unit_id")
        ]
        return {
            "status": "ok" if hits else "no_hits",
            "query_text": normalized_query,
            "hits": hits,
            "replyable_hits": replyable_hits,
            "human_only_hits": human_only_hits,
            "manual_review_required": False,
            "confidence_score": round(top_score, 4) if top_score is not None else None,
            "filters_used": {
                "method": "POST",
                "response_trim_limit": limit,
                "endpoint": endpoint,
            },
            "evidence_refs": evidence_refs,
            "source": "sales_kb_api",
            "latency_ms": latency_ms,
            "no_hit_reason": "外部销售知识库 Q&A 接口未返回可用答案" if not hits else "",
            "how": "直接调用销售知识库 API 的 POST /kb-units/qa-retrieve 接口，返回候选销售回复。",
        }
    except Exception as exc:
        logger.warning("销售知识库 API 连接失败，已跳过。 endpoint=%s error=%s", endpoint, exc)
        return {
            "status": "skipped_unavailable",
            "query_text": normalized_query,
            "hits": [],
            "replyable_hits": [],
            "human_only_hits": [],
            "manual_review_required": False,
            "confidence_score": None,
            "filters_used": {
                "method": "POST",
                "response_trim_limit": limit,
                "endpoint": endpoint,
            },
            "evidence_refs": [],
            "source": "sales_kb_api",
            "latency_ms": round((perf_counter() - started) * 1000, 2),
            "skipped": True,
            "skip_reason": "connection_failed",
            "error": sanitize_text(str(exc)),
            "how": "直接调用销售知识库 API 的 POST /kb-units/qa-retrieve 接口，返回候选销售回复。",
        }

def _public_endpoint_map() -> dict[str, str]:
    base_url = _external_api_base_url()
    sales_kb_base_url = _sales_kb_api_base_url()
    if not base_url:
        return {
            "base_url": "",
            "docs_url": "",
            "openapi_url": "",
            "static_url": "",
            "health_url": "",
            "readiness_url": "",
            "config_check_url": "",
            "public_endpoints_url": "",
            "ai_scripts_url": "",
            "sessions_url": "",
            "qywx_callback_url": "",
            "sales_kb_base_url": sales_kb_base_url,
            "sales_kb_qa_url": f"{sales_kb_base_url}/kb-units/qa-retrieve" if sales_kb_base_url else "",
            "sales_kb_docs_url": f"{sales_kb_base_url}/docs" if sales_kb_base_url else "",
        }
    return {
        "base_url": base_url,
        "docs_url": f"{base_url}/docs",
        "openapi_url": f"{base_url}/openapi.json",
        "static_url": f"{base_url}/static",
        "health_url": f"{base_url}/health",
        "readiness_url": f"{base_url}/api/health/ready",
        "config_check_url": f"{base_url}/api/system/config_check",
        "public_endpoints_url": f"{base_url}/api/system/public_endpoints",
        "ai_scripts_url": f"{base_url}/api/system/ai_scripts",
        "sessions_url": f"{base_url}/api/sessions",
        "qywx_callback_url": f"{base_url}/cb/qywx",
        "sales_kb_base_url": sales_kb_base_url,
        "sales_kb_qa_url": f"{sales_kb_base_url}/kb-units/qa-retrieve" if sales_kb_base_url else "",
        "sales_kb_docs_url": f"{sales_kb_base_url}/docs" if sales_kb_base_url else "",
    }

def _public_access_runtime() -> dict[str, Any]:
    return {
        "external_api_base_url": _external_api_base_url(),
        "cloudflared_public_hostname": _runtime_setting("CLOUDFLARED_PUBLIC_HOSTNAME", settings.CLOUDFLARED_PUBLIC_HOSTNAME),
        "cloudflared_target_url": _runtime_setting("CLOUDFLARED_TARGET_URL", settings.CLOUDFLARED_TARGET_URL),
        "cloudflared_bin": _runtime_setting("CLOUDFLARED_BIN", settings.CLOUDFLARED_BIN),
        "cloudflared_log_file": _runtime_setting("CLOUDFLARED_LOG_FILE", settings.CLOUDFLARED_LOG_FILE),
        "cloudflared_token_masked": _mask_config_value(
            _runtime_setting("CLOUDFLARED_TUNNEL_TOKEN", settings.CLOUDFLARED_TUNNEL_TOKEN)
        ),
        "local_https_enabled": _runtime_bool_setting("ENABLE_LOCAL_HTTPS", settings.ENABLE_LOCAL_HTTPS),
        "local_https_cert_file": _runtime_setting("LOCAL_HTTPS_CERT_FILE", settings.LOCAL_HTTPS_CERT_FILE),
        "local_https_key_file": _runtime_setting("LOCAL_HTTPS_KEY_FILE", settings.LOCAL_HTTPS_KEY_FILE),
    }

def _runtime_llm_values(prefix: str) -> dict:
    return {
        "api_url": _runtime_setting(f"{prefix}_API_URL", getattr(settings, f"{prefix}_API_URL", "")),
        "api_key": _runtime_setting(f"{prefix}_API_KEY", getattr(settings, f"{prefix}_API_KEY", "")),
        "model": _runtime_setting(f"{prefix}_MODEL", getattr(settings, f"{prefix}_MODEL", "")),
        "timeout_seconds": _runtime_int_setting(
            f"{prefix}_TIMEOUT_SECONDS",
            getattr(settings, f"{prefix}_TIMEOUT_SECONDS", 100),
            100,
        ),
    }

def _runtime_stage1_llm_values() -> dict:
    if _runtime_bool_setting("STAGE1_USE_LLM2", settings.STAGE1_USE_LLM2):
        return _runtime_llm_values("LLM2")
    return _runtime_llm_values("LLM1")

def _runtime_llm_configured(config: dict) -> bool:
    return not any(
        _is_placeholder_value(config.get(field))
        for field in ("api_url", "api_key", "model")
    )

def _set_llm_compare_status(session_id: str, status: str, reason: str = "", config: dict | None = None):
    payload = {
        "status": status,
        "reason": sanitize_text(reason or ""),
        "updated_at": datetime.utcnow().isoformat(),
    }
    if config:
        payload.update({
            "model": config.get("model") or "",
            "provider": _llm_provider_label(config.get("api_key")),
            "url": config.get("api_url") or "",
        })
    LLM_COMPARE_RUNTIME_STATUS[session_id] = payload

def _llm_provider_label(api_key: str | None) -> str:
    if not str(api_key or "").strip():
        return "未配置"
    return "Dify" if str(api_key or "").startswith("app-") else "OpenAI-compatible"

def _llm_runtime_config() -> dict:
    llm1 = _runtime_stage1_llm_values()
    llm1_compare = _runtime_llm_values("LLM1_COMPARE")
    llm2 = _runtime_llm_values("LLM2")
    llm2_compare = _runtime_llm_values("LLM2_COMPARE")
    llm1_display_prefix = "LLM-2" if _runtime_bool_setting("STAGE1_USE_LLM2", settings.STAGE1_USE_LLM2) else "LLM-1"
    llm1_role = "第一阶段特征提取（复用 LLM-2）" if _runtime_bool_setting("STAGE1_USE_LLM2", settings.STAGE1_USE_LLM2) else "第一阶段特征提取"
    return {
        "llm1": {
            "id": "LLM-1",
            "role": llm1_role,
            "model": llm1["model"],
            "provider": _llm_provider_label(llm1["api_key"]),
            "url": llm1["api_url"],
            "display_name": f"{llm1_display_prefix} / {llm1['model'] or '未配置'}",
        },
        "llm1_compare": {
            "id": "LLM-1 对比",
            "role": "第一阶段对比特征提取",
            "model": llm1_compare["model"],
            "provider": _llm_provider_label(llm1_compare["api_key"]),
            "url": llm1_compare["api_url"],
            "display_name": f"LLM-1 对比 / {llm1_compare['model'] or '未配置'}",
        },
        "llm2": {
            "id": "LLM-2",
            "role": "第二阶段主回复生成",
            "model": llm2["model"],
            "provider": _llm_provider_label(llm2["api_key"]),
            "url": llm2["api_url"],
            "display_name": f"LLM-2 / {llm2['model'] or '未配置'}",
        },
        "llm2_compare": {
            "id": "LLM-2 对比",
            "role": "第二阶段对比回复生成",
            "model": llm2_compare["model"],
            "provider": _llm_provider_label(llm2_compare["api_key"]),
            "url": llm2_compare["api_url"],
            "display_name": f"LLM-2 对比 / {llm2_compare['model'] or '未配置'}",
        },
    }

def _probe_http_status(url: str, headers: dict | None = None) -> dict:
    try:
        session = requests.Session()
        session.trust_env = settings.HTTP_TRUST_ENV
        response = session.get(url, headers=headers or {}, timeout=settings.KB_HEALTHCHECK_TIMEOUT_SECONDS)
        reachable = response.status_code < 500
        return {
            "status": "ok" if reachable else "error",
            "reachable": reachable,
            "http_status": response.status_code,
        }
    except Exception as exc:
        return {
            "status": "error",
            "reachable": False,
            "error": sanitize_text(str(exc)),
        }


def _embedding_probe_and_suggestions() -> tuple[dict, list[str]]:
    probe_target = EmbeddingService.healthcheck_probe()
    probe = _probe_http_status(probe_target["url"], probe_target.get("headers") or {})
    provider = probe_target.get("provider") or EmbeddingService.provider_name()
    if provider == "ollama":
        suggestions = [] if probe.get("reachable") else ["检查 Ollama 服务是否启动、URL 是否正确、模型是否已拉取。"]
    elif provider == "dify":
        suggestions = [] if probe.get("reachable") else ["检查 Dify Service API 地址、应用发布状态和 API Key 是否正确。"]
        if probe.get("reachable"):
            try:
                app_info = EmbeddingService._dify_app_info()
                app_mode = str(app_info.get("mode") or "").strip().lower()
                probe["app_mode"] = app_mode
                probe["app_name"] = app_info.get("name")
                if app_mode == "chat":
                    probe["status"] = "warning"
                    probe["vector_capable"] = False
                    probe["fallback_mode"] = "keyword_only"
                    suggestions = [
                        "当前 embedding 与其他节点统一复用同一个 Dify 对话型应用；知识库会自动降级为关键词检索，不再强依赖原始向量输出。"
                    ]
                else:
                    probe["vector_capable"] = True
            except Exception as exc:
                probe["app_info_error"] = sanitize_text(str(exc))
    else:
        suggestions = [] if probe.get("reachable") else ["检查 Embedding 网关 URL、API Key 和 /models 或 /embeddings 能力是否可用。"]
    probe["mode"] = probe_target.get("mode")
    probe["probe_url"] = probe_target.get("url")
    return probe, suggestions

def _system_config_check() -> dict:
    checks = {}

    db_configured = bool(settings.DATABASE_URL and not _is_placeholder_value(settings.DATABASE_URL)) or not any(
        _is_placeholder_value(value)
        for value in [settings.DB_HOST, settings.DB_NAME, settings.DB_USER, settings.DB_PASSWORD]
    )
    db_host = settings.DB_HOST
    db_name = settings.DB_NAME
    db_user = settings.DB_USER
    if settings.DATABASE_URL and not _is_placeholder_value(settings.DATABASE_URL):
        try:
            from sqlalchemy.engine import make_url
            parsed_db_url = make_url(settings.DATABASE_URL)
            db_host = parsed_db_url.host or db_host
            db_name = parsed_db_url.database or db_name
            db_user = parsed_db_url.username or db_user
        except Exception:
            pass
    db_check = {
        "configured": db_configured,
        "url_configured": bool(settings.DATABASE_URL and not _is_placeholder_value(settings.DATABASE_URL)),
        "host": db_host,
        "database": db_name,
        "user": db_user,
        "suggestions": [],
    }
    if db_configured:
        try:
            db = SessionLocal()
            db.execute(text("SELECT 1"))
            db_check["status"] = "ok"
        except Exception as exc:
            db_check["status"] = "error"
            db_check["error"] = sanitize_text(str(exc))
            db_check["suggestions"].append("检查 PostgreSQL 地址、库名、账号密码，以及数据库是否允许当前主机连接。")
        finally:
            try:
                db.close()
            except Exception:
                pass
    else:
        db_check["status"] = "error"
        db_check["suggestions"].append("补齐 DATABASE_URL，或 DB_HOST / DB_NAME / DB_USER / DB_PASSWORD。")
    checks["postgres"] = db_check

    embedding_probe, embedding_suggestions = _embedding_probe_and_suggestions()
    checks["embedding"] = {
        "status": embedding_probe["status"] if not _is_placeholder_value(settings.EMBEDDING_API_URL) else "error",
        "configured": not _is_placeholder_value(settings.EMBEDDING_API_URL),
        "provider": settings.EMBEDDING_PROVIDER,
        "url": settings.EMBEDDING_API_URL,
        "model": settings.EMBEDDING_MODEL,
        "probe": embedding_probe,
        "suggestions": embedding_suggestions,
    }

    sales_kb_probe, sales_kb_suggestions = _sales_kb_probe_and_suggestions()
    checks["sales_kb"] = {
        "status": sales_kb_probe.get("status") or "error",
        "configured": sales_kb_probe.get("configured"),
        "url": _sales_kb_api_base_url(),
        "endpoint": sales_kb_probe.get("probe_url"),
        "timeout_seconds": _sales_kb_api_timeout_seconds(),
        "probe": sales_kb_probe,
        "suggestions": sales_kb_suggestions,
    }

    def build_llm_check(name: str, api_url: str, api_key: str, model: str):
        configured = not any(_is_placeholder_value(value) for value in [api_url, api_key, model])
        headers = {"Authorization": f"Bearer {api_key}"} if api_key and not api_key.startswith("app-") else {}
        probe_url = api_url.rstrip("/")
        if api_key.startswith("app-"):
            probe_url = probe_url + "/info"
            headers = {"Authorization": f"Bearer {api_key}"}
        else:
            probe_url = probe_url + "/models"
        probe = _probe_http_status(probe_url, headers)
        suggestions = []
        if not configured:
            suggestions.append(f"补齐 {name.upper()} 的 API URL / API KEY / MODEL。")
        elif not probe.get("reachable"):
            suggestions.append(f"检查 {name.upper()} 的网关地址、代理和 API KEY。")
        return {
            "status": "ok" if configured and probe.get("reachable") else "error",
            "configured": configured,
            "url": api_url,
            "api_key_masked": _mask_config_value(api_key),
            "model": model,
            "provider": _llm_provider_label(api_key),
            "role": _llm_runtime_config().get(name, {}).get("role"),
            "probe": probe,
            "suggestions": suggestions,
        }

    llm1 = _runtime_stage1_llm_values()
    llm1_compare = _runtime_llm_values("LLM1_COMPARE")
    llm2 = _runtime_llm_values("LLM2")
    llm2_compare = _runtime_llm_values("LLM2_COMPARE")
    checks["llm1"] = build_llm_check("llm1", llm1["api_url"], llm1["api_key"], llm1["model"])
    checks["llm1_compare"] = build_llm_check("llm1_compare", llm1_compare["api_url"], llm1_compare["api_key"], llm1_compare["model"])
    checks["llm2"] = build_llm_check("llm2", llm2["api_url"], llm2["api_key"], llm2["model"])
    checks["llm2_compare"] = build_llm_check(
        "llm2_compare",
        llm2_compare["api_url"],
        llm2_compare["api_key"],
        llm2_compare["model"],
    )

    crm_configured = not any(
        _is_placeholder_value(value)
        for value in [settings.CRM_DBHost, settings.CRM_DBName, settings.CRM_DBUserId, settings.CRM_DBPassword]
    )
    crm_check = {
        "configured": crm_configured,
        "host": settings.CRM_DBHost,
        "database": settings.CRM_DBName,
        "user": settings.CRM_DBUserId,
        "suggestions": [],
        "fallback_enabled": True,
    }
    if crm_configured:
        crm_db = None
        try:
            from crm_database import CRMSessionLocal, get_crm_connection_debug_info
            crm_db = CRMSessionLocal()
            crm_db.execute(text("SELECT 1"))
            crm_check["status"] = "ok"
            crm_check.update(get_crm_connection_debug_info())
        except Exception as exc:
            crm_check["status"] = "error"
            crm_check["error"] = sanitize_text(str(exc))
            crm_check["suggestions"].append("检查 SQL Server 地址、ODBC Driver、账号密码和网络连通性。")
            try:
                from crm_database import get_crm_connection_debug_info
                crm_check.update(get_crm_connection_debug_info())
            except Exception:
                pass
        finally:
            try:
                crm_db.close()
            except Exception:
                pass
    else:
        crm_check["status"] = "error"
        crm_check["suggestions"].append("补齐 CRM_DBHost / CRM_DBName / CRM_DBUserId / CRM_DBPassword。")
    checks["crm"] = crm_check

    qywx_credentials_ready = not any(
        _is_placeholder_value(value)
        for value in [settings.CORP_ID, settings.CORP_SECRET, settings.AGENT_ID, settings.TOKEN, settings.ENCODING_AES_KEY]
    )
    checks["qywx"] = {
        "status": "ok" if qywx_credentials_ready else "error",
        "configured": qywx_credentials_ready,
        "corp_id_masked": _mask_config_value(settings.CORP_ID),
        "agent_id_masked": _mask_config_value(settings.AGENT_ID),
        "suggestions": [] if qywx_credentials_ready else ["补齐企微应用 CorpID / Secret / AgentID / Token / EncodingAESKey。"],
    }

    archive_status = ArchiveService.config_status()
    checks["archive_sdk"] = {
        "status": "ok" if archive_status["ready"] else "error",
        "sdk_present": archive_status["sdk_present"],
        "private_key_present": archive_status["private_key_present"],
        "chatdata_secret_configured": archive_status["chatdata_secret_configured"],
        "private_key_path": archive_status["private_key_path"],
        "archive_polling_enabled": archive_status["archive_polling_enabled"],
        "suggestions": [] if archive_status["ready"] else ["企微会话存档未完整配置；未启用时不影响知识库后台启动。"],
    }
    checks["runtime"] = {
        "status": "ok",
        "external_api_base_url": _external_api_base_url(),
        "sales_kb_api_base_url": _sales_kb_api_base_url(),
        "http_trust_env": settings.HTTP_TRUST_ENV,
        "api_reply_single_model_single_style": settings.API_REPLY_SINGLE_MODEL_SINGLE_STYLE,
        "api_reply_enable_scoring": settings.API_REPLY_ENABLE_SCORING,
        "public_access": _public_access_runtime(),
    }

    overall = "ok" if all(item.get("status") == "ok" for item in checks.values()) else "degraded"
    return {
        "status": overall,
        "generated_at": datetime.utcnow().isoformat(),
        "checks": checks,
        "llm_runtime": _llm_runtime_config(),
        "deployment_doc": "docs/生产部署与配置清单.md",
        "external_endpoints": _public_endpoint_map(),
    }

def _runtime_health() -> dict:
    checks = {}
    overall = "ok"

    db = SessionLocal()
    try:
        db.execute(text("SELECT 1"))
        checks["postgres"] = {"status": "ok"}
    except Exception as e:
        checks["postgres"] = {"status": "error", "error": sanitize_text(str(e))}
        overall = "degraded"
    finally:
        db.close()

    if EmbeddingService.provider_name() == "ollama":
        try:
            session = requests.Session()
            session.trust_env = settings.HTTP_TRUST_ENV
            response = session.get(
                settings.EMBEDDING_API_URL.rstrip("/") + "/api/tags",
                timeout=settings.KB_HEALTHCHECK_TIMEOUT_SECONDS,
            )
            response.raise_for_status()
            models = [item.get("name") for item in response.json().get("models", [])]
            checks["embedding"] = {
                "status": "ok" if settings.EMBEDDING_MODEL in models else "warning",
                "provider": settings.EMBEDDING_PROVIDER,
                "url": settings.EMBEDDING_API_URL,
                "model": settings.EMBEDDING_MODEL,
                "model_available": settings.EMBEDDING_MODEL in models,
            }
            if settings.EMBEDDING_MODEL not in models and overall == "ok":
                overall = "degraded"
        except Exception as e:
            checks["embedding"] = {
                "status": "error",
                "provider": settings.EMBEDDING_PROVIDER,
                "url": settings.EMBEDDING_API_URL,
                "model": settings.EMBEDDING_MODEL,
                "error": sanitize_text(str(e)),
            }
            overall = "degraded"
    else:
        embedding_probe, _ = _embedding_probe_and_suggestions()
        checks["embedding"] = {
            "status": embedding_probe.get("status") or ("ok" if embedding_probe.get("reachable") else "error"),
            "provider": settings.EMBEDDING_PROVIDER,
            "url": settings.EMBEDDING_API_URL,
            "model": settings.EMBEDDING_MODEL,
            "probe": embedding_probe,
        }
        if (embedding_probe.get("status") or "ok") != "ok":
            overall = "degraded"

    sales_kb_probe, sales_kb_suggestions = _sales_kb_probe_and_suggestions()
    checks["sales_kb"] = {
        "status": sales_kb_probe.get("status") or "error",
        "url": _sales_kb_api_base_url(),
        "endpoint": sales_kb_probe.get("probe_url"),
        "timeout_seconds": _sales_kb_api_timeout_seconds(),
        "probe": sales_kb_probe,
        "suggestions": sales_kb_suggestions,
    }
    if (sales_kb_probe.get("status") or "error") != "ok":
        overall = "degraded"

    checks["runtime"] = {
        "status": "ok",
        "external_api_base_url": _external_api_base_url(),
        "sales_kb_api_base_url": _sales_kb_api_base_url(),
        "http_trust_env": settings.HTTP_TRUST_ENV,
        "llm1_timeout_seconds": settings.LLM1_TIMEOUT_SECONDS,
        "llm1_compare_timeout_seconds": settings.LLM1_COMPARE_TIMEOUT_SECONDS,
        "llm2_timeout_seconds": settings.LLM2_TIMEOUT_SECONDS,
        "llm2_compare_timeout_seconds": settings.LLM2_COMPARE_TIMEOUT_SECONDS,
        "embedding_timeout_seconds": settings.EMBEDDING_TIMEOUT_SECONDS,
        "slow_request_ms": settings.SLOW_REQUEST_MS,
        "log_desensitize_enabled": settings.LOG_DESENSITIZE_ENABLED,
        "api_reply_single_model_single_style": settings.API_REPLY_SINGLE_MODEL_SINGLE_STYLE,
        "api_reply_enable_scoring": settings.API_REPLY_ENABLE_SCORING,
        "public_access": _public_access_runtime(),
    }
    return {
        "status": overall,
        "service": "qw-ai-sales-assist",
        "generated_at": datetime.utcnow().isoformat(),
        "checks": checks,
    }

@app.get("/health")
async def health_check():
    return _runtime_health()

@app.get("/api/health")
async def health_check_api_alias():
    return _runtime_health()

@app.get("/api/health/ready")
async def readiness_check():
    return _runtime_health()

@app.get("/api/version")
async def get_api_version():
    return {
        "status": "ok",
        "service": "qw-ai-sales-assist",
        "version": "local-dev",
        "generated_at": datetime.utcnow().isoformat(),
    }

@app.get("/api/system/config_check")
async def get_system_config_check():
    return _system_config_check()

@app.get("/api/system/public_endpoints")
async def get_public_endpoints():
    return {
        "status": "success",
        "generated_at": datetime.utcnow().isoformat(),
        "public_access": _public_access_runtime(),
        "external_endpoints": _public_endpoint_map(),
    }

def _archive_sync_db_summary(db: Session, chat_date: str | None = None) -> dict:
    latest_row = db.query(
        MessageLog.timestamp,
        MessageLog.user_id,
        MessageLog.archive_seq,
    ).filter(
        MessageLog.is_mock.is_(False)
    ).order_by(
        MessageLog.timestamp.desc(),
        MessageLog.id.desc(),
    ).first()

    recent_rows = db.query(
        func.date(MessageLog.timestamp).label("chat_day"),
        func.count(MessageLog.id).label("message_count"),
        func.count(func.distinct(MessageLog.user_id)).label("session_count"),
    ).filter(
        MessageLog.is_mock.is_(False)
    ).group_by(
        text("chat_day")
    ).order_by(
        text("chat_day DESC")
    ).limit(5).all()

    selected_date_summary = None
    date_range = _parse_filter_date(chat_date) if chat_date else None
    if date_range:
        day_start, day_end = date_range
        row = db.query(
            func.count(MessageLog.id).label("message_count"),
            func.count(func.distinct(MessageLog.user_id)).label("session_count"),
            func.min(MessageLog.timestamp).label("first_message_at"),
            func.max(MessageLog.timestamp).label("last_message_at"),
        ).filter(
            MessageLog.is_mock.is_(False),
            MessageLog.timestamp >= day_start,
            MessageLog.timestamp < day_end,
        ).first()
        selected_date_summary = {
            "date": chat_date,
            "message_count": int(row.message_count or 0) if row else 0,
            "session_count": int(row.session_count or 0) if row else 0,
            "first_message_at": row.first_message_at if row else None,
            "last_message_at": row.last_message_at if row else None,
        }

    return {
        "latest_message": {
            "timestamp": latest_row.timestamp if latest_row else None,
            "session_id": latest_row.user_id if latest_row else None,
            "archive_seq": latest_row.archive_seq if latest_row else None,
        },
        "recent_days": [
            {
                "date": str(item.chat_day),
                "message_count": int(item.message_count or 0),
                "session_count": int(item.session_count or 0),
            }
            for item in recent_rows
        ],
        "selected_date": selected_date_summary,
    }

@app.get("/api/system/archive_sync_status")
async def get_archive_sync_status(chat_date: str | None = Query(default=None), db: Session = Depends(get_db)):
    from archive_service import ArchiveService

    diagnostics = ArchiveService.get_runtime_diagnostics()
    diagnostics["db_summary"] = _archive_sync_db_summary(db, chat_date=chat_date)
    return {
        "status": "success",
        "generated_at": datetime.utcnow().isoformat(),
        "diagnostics": diagnostics,
    }

@app.post("/api/kb/documents/{document_id}/publish")
async def publish_kb_document(
    document_id: str,
    payload: KnowledgePublishRequest | None = None,
    db: Session = Depends(get_db),
):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    _ensure_document_action_allowed(doc, "publish")
    _ensure_documents_approved_for_publish([doc])
    publish_payload = payload or KnowledgePublishRequest()
    publish_validation = _ensure_documents_publishable(db, [doc], force=publish_payload.force)
    gate_report = await _ensure_publish_gate(
        [doc],
        db,
        force=publish_payload.force,
        force_reason=publish_payload.force_reason,
        operator=publish_payload.operator,
    )
    doc.version_no = _next_publish_version_no(db, str(doc.document_id), doc.version_no)
    _create_version_snapshot(
        db,
        doc,
        action="publish",
        operator=publish_payload.operator,
        note=publish_payload.force_reason if publish_payload.force else None,
        version_no=doc.version_no,
    )
    updated_chunks, updated_pricing_rules = _sync_related_status(db, doc, "active", "approved")
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
        "gate_report": gate_report,
        "publish_validation": publish_validation,
        "force_published": publish_payload.force,
    }

@app.post("/api/kb/documents/{document_id}/submit_review")
async def submit_review_kb_document(document_id: str, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    _ensure_document_action_allowed(doc, "submit_review")
    updated_chunks, updated_pricing_rules = _sync_related_status(db, doc, "review", "in_review")
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/documents/{document_id}/approve")
async def approve_kb_document(document_id: str, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    _ensure_document_action_allowed(doc, "approve")
    updated_chunks, updated_pricing_rules = _sync_related_status(db, doc, "review", "approved")
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/documents/{document_id}/archive")
async def archive_kb_document(document_id: str, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    _ensure_document_action_allowed(doc, "archive")
    updated_chunks, updated_pricing_rules = _sync_related_status(db, doc, "archived")
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/documents/{document_id}/reject")
async def reject_kb_document(document_id: str, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    _ensure_document_action_allowed(doc, "reject")
    updated_chunks, updated_pricing_rules = _sync_related_status(db, doc, "rejected", "rejected")
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/documents/{document_id}/restore/{version_no}")
async def restore_kb_document_version(document_id: str, version_no: int, db: Session = Depends(get_db)):
    doc = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Knowledge document not found")
    snapshot = db.query(KnowledgeVersionSnapshot).filter(
        KnowledgeVersionSnapshot.document_id == document_id,
        KnowledgeVersionSnapshot.version_no == version_no,
    ).order_by(KnowledgeVersionSnapshot.created_at.desc()).first()
    if not snapshot:
        raise HTTPException(status_code=404, detail="Knowledge version snapshot not found")
    restore_report = _restore_document_from_snapshot(db, doc, snapshot)
    db.commit()
    db.refresh(doc)
    return {
        "status": "success",
        "document": _doc_to_dict(doc),
        "restore_report": restore_report,
        "version_snapshot": _version_snapshot_to_dict(snapshot),
    }

def _load_bulk_docs(db: Session, document_ids: list[str]):
    unique_ids = [doc_id for doc_id in dict.fromkeys(document_ids or []) if doc_id]
    if not unique_ids:
        raise HTTPException(status_code=400, detail="document_ids 不能为空")
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.document_id.in_(unique_ids)).all()
    found = {str(doc.document_id) for doc in docs}
    missing = [doc_id for doc_id in unique_ids if doc_id not in found]
    if missing:
        raise HTTPException(status_code=404, detail={"message": "Some documents not found", "missing": missing})
    return docs

def _bulk_set_documents_status(db: Session, docs, status: str, review_status: str | None = None):
    updated_chunks = 0
    updated_pricing_rules = 0
    for doc in docs:
        chunk_count, rule_count = _sync_related_status(db, doc, status, review_status)
        updated_chunks += chunk_count
        updated_pricing_rules += rule_count
    db.commit()
    return {
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/bulk/documents/submit_review")
async def submit_review_kb_documents(payload: KnowledgeBulkAction, db: Session = Depends(get_db)):
    docs = _load_bulk_docs(db, payload.document_ids)
    _ensure_documents_action_allowed(docs, "submit_review")
    result = _bulk_set_documents_status(db, docs, "review", "in_review")
    return {"status": "success", **result}

@app.post("/api/kb/bulk/documents/publish")
async def publish_kb_documents(payload: KnowledgeBulkAction, db: Session = Depends(get_db)):
    docs = _load_bulk_docs(db, payload.document_ids)
    _ensure_documents_action_allowed(docs, "publish")
    _ensure_documents_approved_for_publish(docs)
    publish_validation = _ensure_documents_publishable(db, docs, force=payload.force)
    gate_report = await _ensure_publish_gate(
        docs,
        db,
        force=payload.force,
        force_reason=payload.force_reason,
        operator=payload.operator,
    )
    for doc in docs:
        doc.version_no = _next_publish_version_no(db, str(doc.document_id), doc.version_no)
        _create_version_snapshot(
            db,
            doc,
            action="publish",
            operator=payload.operator,
            note=payload.force_reason if payload.force else None,
            version_no=doc.version_no,
        )
    result = _bulk_set_documents_status(db, docs, "active", "approved")
    return {
        "status": "success",
        **result,
        "gate_report": gate_report,
        "publish_validation": publish_validation,
        "force_published": payload.force,
    }

@app.post("/api/kb/bulk/documents/archive")
async def archive_kb_documents(payload: KnowledgeBulkAction, db: Session = Depends(get_db)):
    docs = _load_bulk_docs(db, payload.document_ids)
    _ensure_documents_action_allowed(docs, "archive")
    result = _bulk_set_documents_status(db, docs, "archived")
    return {"status": "success", **result}

@app.post("/api/kb/bulk/documents/approve")
async def approve_kb_documents(payload: KnowledgeBulkAction, db: Session = Depends(get_db)):
    docs = _load_bulk_docs(db, payload.document_ids)
    _ensure_documents_action_allowed(docs, "approve")
    result = _bulk_set_documents_status(db, docs, "review", "approved")
    return {"status": "success", **result}

@app.get("/api/kb/import_batches")
async def list_kb_import_batches(db: Session = Depends(get_db)):
    from sqlalchemy import func
    rows = db.query(
        KnowledgeDocument.import_batch,
        KnowledgeDocument.status,
        func.count(KnowledgeDocument.document_id).label("count"),
        func.min(KnowledgeDocument.created_at).label("created_from"),
        func.max(KnowledgeDocument.created_at).label("created_to"),
    ).filter(KnowledgeDocument.import_batch.isnot(None)).group_by(
        KnowledgeDocument.import_batch,
        KnowledgeDocument.status,
    ).order_by(func.max(KnowledgeDocument.created_at).desc()).all()
    batches = {}
    for row in rows:
        item = batches.setdefault(row.import_batch, {
            "import_batch": row.import_batch,
            "total_count": 0,
            "status_counts": {},
            "created_from": row.created_from,
            "created_to": row.created_to,
        })
        item["total_count"] += row.count
        item["status_counts"][row.status] = row.count
        item["created_from"] = min(item["created_from"], row.created_from)
        item["created_to"] = max(item["created_to"], row.created_to)
    return list(batches.values())

def _build_import_batch_summary(docs: list[KnowledgeDocument]) -> dict:
    status_counts = {}
    risk_counts = {}
    business_line_counts = {}
    knowledge_type_counts = {}
    for doc in docs:
        status_counts[doc.status] = status_counts.get(doc.status, 0) + 1
        risk_counts[doc.risk_level] = risk_counts.get(doc.risk_level, 0) + 1
        business_line_counts[doc.business_line] = business_line_counts.get(doc.business_line, 0) + 1
        knowledge_type_counts[doc.knowledge_type] = knowledge_type_counts.get(doc.knowledge_type, 0) + 1
    return {
        "total_count": len(docs),
        "status_counts": status_counts,
        "risk_counts": risk_counts,
        "business_line_counts": business_line_counts,
        "knowledge_type_counts": knowledge_type_counts,
    }

@app.get("/api/kb/import_batches/{import_batch}")
async def get_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).order_by(KnowledgeDocument.created_at.asc()).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    doc_ids = [doc.document_id for doc in docs]
    chunks = db.query(KnowledgeChunk).filter(KnowledgeChunk.document_id.in_(doc_ids)).all()
    pricing_rules = db.query(PricingRule).filter(PricingRule.document_id.in_(doc_ids)).all()
    chunk_counts = {}
    for chunk in chunks:
        key = str(chunk.document_id)
        chunk_counts[key] = chunk_counts.get(key, 0) + 1
    pricing_counts = {}
    for rule in pricing_rules:
        key = str(rule.document_id)
        pricing_counts[key] = pricing_counts.get(key, 0) + 1
    documents = []
    for doc in docs:
        item = _doc_to_dict(doc)
        doc_id = str(doc.document_id)
        item["chunk_count"] = chunk_counts.get(doc_id, 0)
        item["pricing_rule_count"] = pricing_counts.get(doc_id, 0)
        documents.append(item)
    return {
        "import_batch": import_batch,
        "summary": _build_import_batch_summary(docs),
        "documents": documents,
    }

@app.get("/api/kb/import_batches/{import_batch}/documents")
async def list_kb_import_batch_documents(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).order_by(KnowledgeDocument.created_at.asc()).all()
    return [_doc_to_dict(doc) for doc in docs]

def set_documents_status(db: Session, docs, status: str, review_status: str | None = None):
    updated_chunks = 0
    updated_pricing_rules = 0
    for doc in docs:
        chunk_count, rule_count = _sync_related_status(db, doc, status, review_status)
        updated_chunks += chunk_count
        updated_pricing_rules += rule_count
    db.commit()
    return updated_chunks, updated_pricing_rules

@app.post("/api/kb/import_batches/{import_batch}/submit_review")
async def submit_review_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    _ensure_documents_action_allowed(docs, "submit_review")
    updated_chunks, updated_pricing_rules = set_documents_status(db, docs, "review", "in_review")
    return {
        "status": "success",
        "import_batch": import_batch,
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/import_batches/{import_batch}/publish")
async def publish_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    _ensure_documents_action_allowed(docs, "publish")
    _ensure_documents_approved_for_publish(docs)
    _validate_documents_publishable(db, docs)
    gate_report = await _ensure_publish_gate(docs, db, force=False, operator="import_batch_publish")
    for doc in docs:
        doc.version_no = _next_publish_version_no(db, str(doc.document_id), doc.version_no)
        _create_version_snapshot(
            db,
            doc,
            action="publish",
            operator="import_batch_publish",
            version_no=doc.version_no,
        )
    updated_chunks, updated_pricing_rules = set_documents_status(db, docs, "active", "approved")
    return {
        "status": "success",
        "import_batch": import_batch,
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
        "gate_report": gate_report,
    }

@app.post("/api/kb/import_batches/{import_batch}/archive")
async def archive_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    _ensure_documents_action_allowed(docs, "archive")
    updated_chunks, updated_pricing_rules = set_documents_status(db, docs, "archived")
    return {
        "status": "success",
        "import_batch": import_batch,
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/import_batches/{import_batch}/approve")
async def approve_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    _ensure_documents_action_allowed(docs, "approve")
    updated_chunks, updated_pricing_rules = set_documents_status(db, docs, "review", "approved")
    return {
        "status": "success",
        "import_batch": import_batch,
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/import_batches/{import_batch}/rollback")
async def rollback_kb_import_batch(import_batch: str, db: Session = Depends(get_db)):
    """非破坏性回滚：将该批次全部归档，并标记 review_status=rolled_back。"""
    docs = db.query(KnowledgeDocument).filter(KnowledgeDocument.import_batch == import_batch).all()
    if not docs:
        raise HTTPException(status_code=404, detail="Import batch not found")
    updated_chunks, updated_pricing_rules = set_documents_status(db, docs, "archived", "rolled_back")
    return {
        "status": "success",
        "import_batch": import_batch,
        "updated_documents": len(docs),
        "updated_chunks": updated_chunks,
        "updated_pricing_rules": updated_pricing_rules,
    }

@app.post("/api/kb/retrieve")
async def retrieve_kb(payload: KnowledgeRetrieveRequest):
    """知识库 V2 检索接口：只返回知识证据包，不生成回复。"""
    top_k = max(1, min(payload.top_k, 20))
    result = IntentEngine.retrieve_knowledge_v2(
        query_text=payload.query_text,
        query_features=payload.query_features or {},
        top_k=top_k,
        request_id=payload.request_id,
        session_id=payload.session_id,
    )
    for hit in (result.get("hits", []) + result.get("supporting_chunks", []) + result.get("related_chunks", [])):
        hit["knowledge_class_label"] = label_for("knowledge_class", hit.get("knowledge_class"))
        hit["business_line_label"] = label_for("business_line", hit.get("business_line"))
        hit["knowledge_type_label"] = label_for("knowledge_type", hit.get("knowledge_type"))
        hit["chunk_type_label"] = label_for("chunk_type", hit.get("chunk_type"))
        hit["language_pair_label"] = label_for("language_pair", hit.get("language_pair"))
        hit["service_scope_label"] = label_for("service_scope", hit.get("service_scope"))
    return result

@app.post("/api/kb/import/excel/preview")
async def preview_kb_excel_import(
    file: UploadFile = File(...),
    import_type: str = Form("unified"),
):
    """统一知识 Excel 预览：只解析和校验，不写入数据库。"""
    filename = file.filename or "kb_import.xlsx"
    raw = await file.read()
    parsed = parse_kb_excel_import(raw, filename, import_type)
    return {
        "status": "success",
        "filename": parsed["filename"],
        "sheet": parsed["sheet"],
        "import_type": parsed["import_type"],
        "import_type_label": parsed["import_type_label"],
        "header_row": parsed["header_row"],
        "valid_count": len(parsed["valid_rows"]),
        "skipped_count": len(parsed["skipped"]),
        "failed_count": len(parsed["failed"]),
        "rows": [_excel_item_preview(item) for item in parsed["valid_rows"][:200]],
        "skipped": parsed["skipped"],
        "failed": parsed["failed"],
    }

def _llm_split_excel_item_to_chunks(item: dict) -> list[KnowledgeChunkCreate]:
    class_options = "、".join(KB_LABELS["knowledge_class"].values())
    prompt = f"""
你是企业知识库切分助手。请只根据原文拆分知识，不要编造价格、能力、交期、承诺或客户信息。

业务侧知识分类只能从以下 7 类中选择：{class_options}。
注意：报价规则不是知识分类；明确价格、单位、最低收费、税费、加急倍率应由结构化 pricing_rule 管理。这里仅可输出报价限制条件，例如“需另行确认”“不能直接承诺固定价”。

输出纯 JSON：
{{
  "chunks": [
    {{
      "title": "简短标题",
      "content": "只表达一个业务可调用知识点",
      "knowledge_class": "报价限制条件|能力知识|流程规则|FAQ常见问答|案例|邮件模板|名词定义",
      "business_line": "translation|printing|interpretation|multimedia|exhibition|gifts|general",
      "language_pair": null,
      "service_scope": null,
      "customer_tier": null,
      "priority": 50,
      "risk_level": "low|medium|high",
      "tags": {{}}
    }}
  ]
}}

规则：
- 一条 chunk 只表达一个知识点；流程步骤、能力说明、案例、模板、定义混在一起时必须拆开。
- 报价、费用、合同、承诺、交付边界类内容 risk_level 必须为 high 或 medium。
- 没有可拆分知识时返回 {{"chunks":[]}}，不要兜底编造 FAQ。
- 不确定的结构字段填 null。

默认知识分类：{label_for("knowledge_class", item.get("knowledge_class")) or "未指定"}
默认服务：{item.get("business_line") or "general"}
原始标题：{item.get("title")}
原文：
{item.get("content")}
"""
    llm_result = IntentEngine.run_llm1_json_prompt(prompt, user_id="kb_excel_auto_split")
    raw_chunks = llm_result.get("chunks") if isinstance(llm_result, dict) else None
    if not isinstance(raw_chunks, list):
        raise HTTPException(status_code=502, detail="LLM 自动切分返回格式错误")

    chunks = []
    default_class = item.get("knowledge_class") or knowledge_class_from_pair(item.get("knowledge_type"), item.get("chunk_type")) or "faq"
    for raw_chunk in raw_chunks:
        if not isinstance(raw_chunk, dict):
            continue
        title = str(raw_chunk.get("title") or "").strip()
        content = str(raw_chunk.get("content") or "").strip()
        if not title or not content:
            continue
        knowledge_class, raw_knowledge_class = normalize_knowledge_class(raw_chunk.get("knowledge_class") or default_class)
        knowledge_class = knowledge_class or default_class
        class_config = KB_KNOWLEDGE_CLASS_CONFIG.get(knowledge_class, KB_KNOWLEDGE_CLASS_CONFIG["faq"])
        business_line, raw_business_line = normalize_code("business_line", raw_chunk.get("business_line") or item.get("business_line") or "general")
        language_pair, raw_language_pair = normalize_code("language_pair", raw_chunk.get("language_pair") or item.get("language_pair"))
        service_scope, raw_service_scope = normalize_code("service_scope", raw_chunk.get("service_scope") or item.get("service_scope"))
        customer_tier, raw_customer_tier = normalize_code("customer_tier", raw_chunk.get("customer_tier") or item.get("customer_tier"))
        risk_level, raw_risk_level = normalize_code("risk_level", raw_chunk.get("risk_level") or class_config["risk_level"])
        tags = dict(raw_chunk.get("tags") or {})
        tags.update({
            "knowledge_class": knowledge_class,
            "auto_split": True,
            "source_row": item.get("row"),
        })
        for key, value in {
            "raw_knowledge_class": raw_knowledge_class,
            "raw_business_line": raw_business_line,
            "raw_language_pair": raw_language_pair,
            "raw_service_scope": raw_service_scope,
            "raw_customer_tier": raw_customer_tier,
            "raw_risk_level": raw_risk_level,
        }.items():
            if value:
                tags[key] = value
        chunks.append(KnowledgeChunkCreate(
            title=title[:255],
            content=content,
            knowledge_type=class_config["knowledge_type"],
            chunk_type=class_config["chunk_type"],
            business_line=business_line or item.get("business_line") or "general",
            language_pair=language_pair,
            service_scope=service_scope,
            customer_tier=customer_tier,
            priority=int(raw_chunk.get("priority") or item.get("priority") or 50),
            tags=tags,
            pricing_rule=None,
        ))
    return chunks

def _commit_kb_excel_import_raw(
    db: Session,
    *,
    raw: bytes,
    filename: str,
    import_type: str,
    owner: str,
    report: Any = None,
) -> dict:
    parsed = parse_kb_excel_import(raw, filename, import_type)
    import_batch = f"{import_type}_excel_{uuid.uuid4().hex[:12]}"
    created = []
    failed = list(parsed["failed"])
    valid_rows = parsed["valid_rows"]
    total_rows = max(len(valid_rows), 1)

    for index, item in enumerate(valid_rows, start=1):
        if report:
            report(
                progress=min(95, round(index * 100 / total_rows)),
                summary=f"importing row {item['row']}",
                result_patch={"current_row": item["row"], "total_rows": len(valid_rows)},
            )
        title = item["title"]
        content = item["content"]
        item_import_type = item.get("import_type") or import_type
        if item.get("auto_split"):
            try:
                split_chunks = _llm_split_excel_item_to_chunks(item)
            except Exception as exc:
                failed.append({"row": item["row"], "title": title, "errors": [f"auto_split_failed: {exc}"]})
                continue
            if not split_chunks:
                failed.append({"row": item["row"], "title": title, "errors": ["auto_split_empty"]})
                continue

            chunk_knowledge_types = {chunk_payload.knowledge_type for chunk_payload in split_chunks}
            document_knowledge_type = split_chunks[0].knowledge_type if len(chunk_knowledge_types) == 1 else item["knowledge_type"]
            document_risk_level = "high" if any(
                chunk_payload.knowledge_type == "pricing" or mentions_pricing_topic(chunk_payload.title, chunk_payload.content)
                for chunk_payload in split_chunks
            ) else item["risk_level"]
            document_tags = dict(item["tags"] or {})
            document_tags.update({
                "knowledge_class": item.get("knowledge_class"),
                "auto_split": True,
                "split_chunk_count": len(split_chunks),
            })
            document = KnowledgeDocument(
                title=title,
                knowledge_type=document_knowledge_type,
                business_line=item["business_line"],
                sub_service=item.get("sub_service"),
                source_type=f"excel_{item_import_type}",
                source_ref=f"{filename} / {parsed['sheet']} / row {item['row']}",
                source_meta={
                    "filename": filename,
                    "sheet": parsed["sheet"],
                    "row": item["row"],
                    "import_type": item_import_type,
                    "request_import_type": import_type,
                    "auto_split": True,
                },
                status="draft",
                owner=owner,
                import_batch=import_batch,
                effective_from=item.get("effective_from"),
                effective_to=item.get("effective_to"),
                risk_level=document_risk_level,
                review_required=True,
                review_status="pending",
                tags=document_tags,
            )
            db.add(document)
            db.flush()

            created_chunk_ids = []
            split_pricing_count = 0
            for chunk_index, chunk_payload in enumerate(split_chunks, start=1):
                retrieval_title = _strip_retrieval_title_prefix(chunk_payload.title)
                retrieval_text = _build_chunk_retrieval_text(chunk_payload.title, chunk_payload.content)
                embedding = _optional_embedding_for_storage(
                    retrieval_text,
                    context=f"publish_split_chunk:row{item['row']}:chunk{chunk_index}:{chunk_payload.title}",
                )
                chunk = KnowledgeChunk(
                    document_id=document.document_id,
                    chunk_no=chunk_index,
                    chunk_type=chunk_payload.chunk_type,
                    title=retrieval_title,
                    content=chunk_payload.content,
                    keyword_text=retrieval_text,
                    embedding=embedding,
                    **_embedding_metadata_fields(embedding),
                    priority=chunk_payload.priority,
                    business_line=chunk_payload.business_line or item["business_line"],
                    sub_service=chunk_payload.sub_service or item.get("sub_service"),
                    knowledge_type=chunk_payload.knowledge_type,
                    language_pair=chunk_payload.language_pair,
                    service_scope=chunk_payload.service_scope,
                    region=chunk_payload.region,
                    customer_tier=chunk_payload.customer_tier,
                    structured_tags=chunk_payload.tags,
                    status="draft",
                    effective_from=item.get("effective_from"),
                    effective_to=item.get("effective_to"),
                )
                db.add(chunk)
                db.flush()
                pricing_rule_payload = None
                if item_import_type == "pricing" and chunk_payload.chunk_type != "constraint":
                    pricing_rule_payload = chunk_payload.pricing_rule or infer_pricing_rule_candidate(
                        chunk_payload.title,
                        chunk_payload.content,
                        chunk_payload.business_line or item["business_line"],
                        chunk_payload.language_pair,
                        chunk_payload.service_scope,
                        chunk_payload.customer_tier,
                    )
                pricing_rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
                _apply_chunk_governance(
                    chunk,
                    source_type=document.source_type,
                    pricing_rule=pricing_rule_payload,
                    source_ref=document.source_ref,
                )
                document.library_type = chunk.library_type
                document.tags = merge_tags(document.tags, library_type=document.library_type)
                if pricing_rule:
                    split_pricing_count += 1
                created_chunk_ids.append(str(chunk.chunk_id))

            if not created_chunk_ids:
                db.delete(document)
                failed.append({"row": item["row"], "title": title, "errors": ["auto_split_no_valid_chunks"]})
                continue

            created.append({
                "row": item["row"],
                "document_id": str(document.document_id),
                "chunk_id": created_chunk_ids[0] if created_chunk_ids else None,
                "chunk_ids": created_chunk_ids,
                "split_chunk_count": len(created_chunk_ids),
                "pricing_rule_created": bool(split_pricing_count),
                "title": title,
                "import_type": item_import_type,
                "knowledge_type": document_knowledge_type,
                "business_line": item["business_line"],
                "risk_level": document_risk_level,
                "auto_split": True,
            })
            continue
        retrieval_title = _strip_retrieval_title_prefix(title)
        retrieval_text = _build_chunk_retrieval_text(title, content)
        embedding = _optional_embedding_for_storage(retrieval_text, context=f"excel_import:row{item['row']}:{title}")

        document = KnowledgeDocument(
            title=title,
            knowledge_type=item["knowledge_type"],
            business_line=item["business_line"],
            sub_service=item.get("sub_service"),
            source_type=f"excel_{item_import_type}",
            source_ref=f"{filename} / {parsed['sheet']} / row {item['row']}",
            source_meta={
                "filename": filename,
                "sheet": parsed["sheet"],
                "row": item["row"],
                "import_type": item_import_type,
                "request_import_type": import_type,
            },
            status="draft",
            owner=owner,
            import_batch=import_batch,
            effective_from=item.get("effective_from"),
            effective_to=item.get("effective_to"),
            risk_level=item["risk_level"],
            review_required=True,
            review_status="pending",
            tags=item["tags"],
        )
        db.add(document)
        db.flush()

        chunk = KnowledgeChunk(
            document_id=document.document_id,
            chunk_no=1,
            chunk_type=item["chunk_type"],
            title=retrieval_title,
            content=content,
            keyword_text=retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=item["priority"],
            business_line=item["business_line"],
            sub_service=item.get("sub_service"),
            knowledge_type=item["knowledge_type"],
            language_pair=item.get("language_pair"),
            service_scope=item.get("service_scope"),
            region=item.get("region"),
            customer_tier=item.get("customer_tier"),
            structured_tags=item["tags"],
            status="draft",
            effective_from=item.get("effective_from"),
            effective_to=item.get("effective_to"),
        )
        db.add(chunk)
        db.flush()

        pricing_rule = None
        if item.get("pricing_rule"):
            pricing_rule_payload = dict(item["pricing_rule"])
            pricing_rule_payload["source_ref"] = pricing_rule_payload.get("source_ref") or document.source_ref
            pricing_rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
        else:
            pricing_rule_payload = None
        _apply_chunk_governance(
            chunk,
            source_type=document.source_type,
            pricing_rule=pricing_rule_payload,
            source_ref=document.source_ref,
        )
        document.library_type = chunk.library_type
        document.tags = merge_tags(document.tags, library_type=document.library_type)

        created.append({
            "row": item["row"],
            "document_id": str(document.document_id),
            "chunk_id": str(chunk.chunk_id),
            "pricing_rule_created": bool(pricing_rule),
            "title": title,
            "import_type": item_import_type,
            "knowledge_type": item["knowledge_type"],
            "business_line": item["business_line"],
            "risk_level": item["risk_level"],
        })

    db.commit()
    logger.info("KB_EXCEL_IMPORT %s", json.dumps({
        "filename": filename,
        "import_type": import_type,
        "import_batch": import_batch,
        "created": len(created),
        "skipped": len(parsed["skipped"]),
        "failed": len(failed),
    }, ensure_ascii=False))

    return {
        "status": "success",
        "filename": filename,
        "sheet": parsed["sheet"],
        "import_type": import_type,
        "import_batch": import_batch,
        "created_count": len(created),
        "skipped_count": len(parsed["skipped"]),
        "failed_count": len(failed),
        "created": created,
        "skipped": parsed["skipped"],
        "failed": failed,
    }

def _run_excel_import_job(report: Any, payload_data: dict) -> dict:
    temp_path = payload_data["temp_path"]
    filename = payload_data["filename"]
    import_type = payload_data["import_type"]
    owner = payload_data.get("owner") or "kb_excel_import"
    with open(temp_path, "rb") as f:
        raw = f.read()
    db = SessionLocal()
    try:
        result = _commit_kb_excel_import_raw(
            db,
            raw=raw,
            filename=filename,
            import_type=import_type,
            owner=owner,
            report=report,
        )
    finally:
        db.close()
        try:
            os.remove(temp_path)
        except OSError:
            pass
    return result

@app.post("/api/kb/import/excel/commit")
async def commit_kb_excel_import(
    file: UploadFile = File(...),
    import_type: str = Form("unified"),
    owner: str = Form("kb_excel_import"),
    db: Session = Depends(get_db),
):
    """统一知识 Excel 确认导入：合法行写入 draft，等待审核发布。"""
    filename = file.filename or "kb_import.xlsx"
    raw = await file.read()
    return _commit_kb_excel_import_raw(db, raw=raw, filename=filename, import_type=import_type, owner=owner)

@app.post("/api/kb/import/excel/commit_async")
async def commit_kb_excel_import_async(
    file: UploadFile = File(...),
    import_type: str = Form("unified"),
    owner: str = Form("kb_excel_import"),
    db: Session = Depends(get_db),
):
    filename = file.filename or "kb_import.xlsx"
    raw = await file.read()
    temp_filename = f"{uuid.uuid4().hex}_{filename}"
    temp_path = os.path.join(_job_payload_dir(), temp_filename)
    with open(temp_path, "wb") as f:
        f.write(raw)
    payload = {
        "temp_path": temp_path,
        "filename": filename,
        "import_type": import_type,
        "owner": owner,
    }
    job = _create_job_task(
        db,
        job_type="kb_excel_import",
        task_payload=payload,
        summary=f"queued excel import {filename}",
    )
    db.commit()
    db.refresh(job)
    start_job(str(job.job_id), _run_excel_import_job, payload)
    return {"status": "queued", "job": _job_to_dict(job)}

@app.post("/api/kb/extract/from_messages")
async def extract_kb_candidates_from_messages(payload: KnowledgeExtractFromMessagesRequest, db: Session = Depends(get_db)):
    records = []
    if payload.session_id:
        logs = (
            db.query(MessageLog)
            .filter(MessageLog.user_id == payload.session_id, MessageLog.is_mock.is_(False))
            .order_by(MessageLog.timestamp.asc(), MessageLog.id.asc())
            .limit(max(1, min(payload.max_messages, 200)))
            .all()
        )
        if not logs:
            raise HTTPException(status_code=404, detail="No messages found for session")
        transcript = "\n".join(
            f"{'客户' if log.sender_type == 'customer' else '我方'}：{sanitize_text(log.content or '')}"
            for log in logs
        )
        records.append({
            "title": payload.title or f"会话抽取 {payload.session_id}",
            "content": transcript,
            "session_id": payload.session_id,
            "message_count": len(logs),
        })
    elif payload.messages:
        for index, message in enumerate(payload.messages, start=1):
            text_value = sanitize_text((message or "").strip())
            if not text_value:
                continue
            records.append({
                "title": f"{payload.title or '历史资料抽取'} #{index}",
                "content": text_value,
            })
    else:
        raise HTTPException(status_code=400, detail="session_id 或 messages 至少提供一个")

    created = _extract_candidates_from_records(
        db,
        records,
        source_type=payload.source_type if payload.source_type in KB_LABELS["candidate_source_type"] else "message_extract",
        owner=payload.owner or "kb_extract",
        operator=payload.operator or "frontend_extract",
        source_ref_prefix=payload.session_id or payload.source_type,
        max_candidates=max(1, min(payload.max_candidates, 100)),
    )
    db.commit()
    for item in created:
        db.refresh(item)
    return {
        "status": "success",
        "source_type": payload.source_type,
        "created_count": len(created),
        "candidates": [_candidate_to_dict(item) for item in created],
    }

@app.post("/api/kb/extract/from_email_excel")
async def extract_kb_candidates_from_email_excel(
    file: UploadFile = File(...),
    owner: str = Form("kb_extract"),
    operator: str = Form("frontend_extract"),
    max_candidates: int = Form(20),
    db: Session = Depends(get_db),
):
    filename = file.filename or "email_extract.xlsx"
    raw = await file.read()
    try:
        records = EmailImportService.parse_file(raw, filename)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not records:
        raise HTTPException(status_code=400, detail="未解析到可抽取邮件记录")
    result = _extract_email_assets_and_candidates(
        db,
        records=records,
        filename=filename,
        owner=owner,
        operator=operator,
        max_candidates=max(1, min(int(max_candidates or 20), 100)),
    )
    db.commit()
    for item in result["email_assets"]:
        db.refresh(item)
    for item in result["documents"]:
        db.refresh(item)
    for item in result["chunks"]:
        db.refresh(item)
    for item in result["fragments"]:
        db.refresh(item)
    return {
        "status": "success",
        "filename": filename,
        "import_batch": result["import_batch"],
        "source_rows": len(records),
        "email_asset_count": len(result["email_assets"]),
        "created_count": len(result["candidates"]),
        "document_count": len(result["documents"]),
        "chunk_count": len(result["chunks"]),
        "fragment_count": len(result["fragments"]),
        "purged": result["purged"],
        "skipped": result["skipped"],
        "email_assets": [_email_thread_asset_to_dict(item) for item in result["email_assets"]],
        "documents": [_knowledge_document_to_dict(item, db) for item in result["documents"]],
        "candidates": [_candidate_to_dict(item) for item in result["candidates"]],
        "fragments": [_email_fragment_asset_to_dict(item) for item in result["fragments"]],
    }

@app.get("/api/email_assets")
async def list_email_assets(
    source_type: str | None = None,
    status: str | None = None,
    session_id: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    query = db.query(EmailThreadAsset)
    if source_type:
        query = query.filter(EmailThreadAsset.source_type == source_type)
    if status:
        query = query.filter(EmailThreadAsset.status == status)
    if session_id:
        query = query.filter(EmailThreadAsset.session_id == session_id)
    items = query.order_by(EmailThreadAsset.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return {"status": "success", "items": [_email_thread_asset_to_dict(item) for item in items]}

@app.get("/api/email_fragments")
async def list_email_fragments(
    source_type: str | None = None,
    function_fragment: str | None = None,
    session_id: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    query = db.query(EmailFragmentAsset)
    if source_type:
        query = query.filter(EmailFragmentAsset.source_type == source_type)
    if function_fragment:
        query = query.filter(EmailFragmentAsset.function_fragment == function_fragment)
    if session_id:
        query = query.filter(EmailFragmentAsset.session_id == session_id)
    items = query.order_by(EmailFragmentAsset.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return {"status": "success", "items": [_email_fragment_asset_to_dict(item) for item in items]}

@app.get("/api/email_effect_feedback")
async def list_email_effect_feedback(
    session_id: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    query = db.query(EmailEffectFeedback)
    if session_id:
        query = query.filter(EmailEffectFeedback.session_id == session_id)
    items = query.order_by(EmailEffectFeedback.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return {"status": "success", "items": [_email_effect_feedback_to_dict(item) for item in items]}

@app.post("/api/ml/prepare_training_samples")
async def prepare_training_samples(payload: TrainingSamplePrepareRequest, db: Session = Depends(get_db)):
    result = _prepare_training_samples(db, payload)
    db.commit()
    return {"status": "success", **result}

@app.get("/api/ml/training_samples")
async def list_training_samples(
    sample_type: str | None = None,
    review_status: str | None = None,
    export_status: str | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    query = db.query(ModelTrainingSample)
    if sample_type:
        query = query.filter(ModelTrainingSample.sample_type == sample_type)
    if review_status:
        query = query.filter(ModelTrainingSample.review_status == review_status)
    if export_status:
        query = query.filter(ModelTrainingSample.export_status == export_status)
    items = query.order_by(ModelTrainingSample.created_at.desc()).limit(max(1, min(limit, 500))).all()
    return {"status": "success", "items": [_training_sample_to_dict(item) for item in items]}

@app.post("/api/ml/export_training_samples")
async def export_training_samples(payload: TrainingSampleExportRequest, db: Session = Depends(get_db)):
    result = _export_training_samples(db, payload)
    db.commit()
    return {"status": "success", **result}


@app.post("/api/ml/run_training_pipeline")
async def run_training_pipeline(payload: TrainingExecutionRequest, db: Session = Depends(get_db)):
    result = _run_training_pipeline(db, payload)
    db.commit()
    return {"status": "success", **result}


@app.post("/api/ml/run_training_pipeline_async")
async def run_training_pipeline_async(payload: TrainingExecutionRequest, db: Session = Depends(get_db)):
    job = _create_job_task(
        db,
        job_type="training_pipeline",
        task_payload=payload.model_dump(),
        summary="queued training pipeline",
    )
    db.commit()
    db.refresh(job)
    start_job(str(job.job_id), _run_training_pipeline_job, payload.model_dump())
    return {"status": "queued", "job": _job_to_dict(job)}


@app.post("/api/kb/complete_partial_items")
async def complete_partial_items(payload: AnalysisCompletionRequest, db: Session = Depends(get_db)):
    result = _complete_analysis_items(db, payload)
    db.commit()
    return {"status": "success", **result}


@app.post("/api/kb/complete_partial_items_async")
async def complete_partial_items_async(payload: AnalysisCompletionRequest, db: Session = Depends(get_db)):
    job = _create_job_task(
        db,
        job_type="analysis_completion",
        task_payload=payload.model_dump(),
        summary="queued analysis completion",
    )
    db.commit()
    db.refresh(job)
    start_job(str(job.job_id), _run_complete_analysis_items_job, payload.model_dump())
    return {"status": "queued", "job": _job_to_dict(job)}

@app.post("/api/kb/import/business_csv")
async def import_business_csv(
    file: UploadFile = File(...),
    start_row: int = Form(DEFAULT_BUSINESS_CSV_ROW_START),
    end_row: int = Form(DEFAULT_BUSINESS_CSV_ROW_END),
    db: Session = Depends(get_db),
):
    filename = file.filename or DEFAULT_BUSINESS_CSV_FILENAME
    if os.path.basename(filename) != DEFAULT_BUSINESS_CSV_FILENAME:
        raise HTTPException(status_code=400, detail=f"当前正式导入口径仅适用于 {DEFAULT_BUSINESS_CSV_FILENAME}")
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="仅支持 .csv 文件")
    if start_row < 1 or end_row < start_row:
        raise HTTPException(status_code=400, detail=f"导入范围非法: {start_row}-{end_row}")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="CSV 文件为空")

    try:
        result = run_business_csv_import(
            db,
            raw=raw,
            filename=os.path.basename(filename),
            start_row=start_row,
            end_row=end_row,
            source_type=BUSINESS_CSV_SOURCE_TYPE,
            owner="business_csv_import_api",
        )
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"CSV 编码解析失败，请使用 UTF-8/UTF-8-SIG: {exc}") from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"business_csv 导入失败: {exc}") from exc

    logger.info("KB_BUSINESS_CSV_IMPORT %s", json.dumps({
        "filename": os.path.basename(filename),
        "start_row": start_row,
        "end_row": end_row,
        "row_limit": end_row - start_row + 1,
        "import_batch": result.get("import_batch"),
        "created_documents": result.get("created_documents"),
        "created_chunks": result.get("created_chunks"),
        "skipped_count": result.get("skipped_count"),
        "purged": result.get("purged"),
    }, ensure_ascii=False))
    return result


@app.post("/api/kb/import/faq_excel")
async def import_faq_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入常见问答 Excel：按“节点名/内容”列生成 draft 知识。"""
    filename = file.filename or "faq_import.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl 未安装或不可用: {e}")

    raw = await file.read()
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件无法解析: {e}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 为空")

    header_row_index = None
    title_col = None
    content_col = None
    for idx, row in enumerate(rows[:10]):
        headers = list(row)
        title_col = find_column_index(headers, ["节点名", "标题", "问题", "场景"])
        content_col = find_column_index(headers, ["内容", "答案", "回复", "话术"])
        if title_col is not None and content_col is not None:
            header_row_index = idx
            break

    if header_row_index is None:
        raise HTTPException(status_code=400, detail="未找到必要列：节点名/内容")

    import_batch = f"faq_excel_{uuid.uuid4().hex[:12]}"
    created = []
    skipped = []
    failed = []

    for row_index, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        title = str(row[title_col] or "").strip() if title_col < len(row) else ""
        content = str(row[content_col] or "").strip() if content_col < len(row) else ""
        if not title and not content:
            skipped.append({"row": row_index, "reason": "empty"})
            continue
        if not title or not content:
            skipped.append({"row": row_index, "reason": "missing_title_or_content", "title": title})
            continue

        business_line = infer_faq_business_line(title, content)
        risk_level = infer_faq_risk_level(title, content)
        retrieval_title = _strip_retrieval_title_prefix(title)
        retrieval_text = _build_chunk_retrieval_text(title, content)
        embedding = _optional_embedding_for_storage(retrieval_text, context=f"faq_excel_import:row{row_index}:{title}")

        source_ref = f"{filename} / {sheet.title} / row {row_index}"
        document = KnowledgeDocument(
            title=title,
            knowledge_type="faq",
            business_line=business_line,
            source_type="excel",
            source_ref=source_ref,
            source_meta={"filename": filename, "sheet": sheet.title, "row": row_index},
            status="draft",
            owner="faq_excel_import",
            import_batch=import_batch,
            risk_level=risk_level,
            review_required=True,
            review_status="pending",
            tags={"scenario": title, "raw_source": "faq_excel"},
        )
        db.add(document)
        db.flush()

        chunk = KnowledgeChunk(
            document_id=document.document_id,
            chunk_no=1,
            chunk_type="faq",
            title=retrieval_title,
            content=content,
            keyword_text=retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=50,
            business_line=business_line,
            knowledge_type="faq",
            structured_tags={"scenario": title, "raw_source": "faq_excel"},
            status="draft",
        )
        db.add(chunk)
        db.flush()
        _apply_chunk_governance(
            chunk,
            source_type=document.source_type,
            source_ref=document.source_ref,
        )
        document.library_type = chunk.library_type
        document.tags = merge_tags(document.tags, library_type=document.library_type)
        created.append({
            "row": row_index,
            "document_id": str(document.document_id),
            "title": title,
            "business_line": business_line,
            "risk_level": risk_level,
        })

    db.commit()
    logger.info("KB_COMMON_QA_EXCEL_IMPORT %s", json.dumps({
        "filename": filename,
        "sheet": sheet.title,
        "import_batch": import_batch,
        "created": len(created),
        "skipped": len(skipped),
        "failed": len(failed),
    }, ensure_ascii=False))

    return {
        "status": "success",
        "filename": filename,
        "sheet": sheet.title,
        "import_batch": import_batch,
        "created_count": len(created),
        "skipped_count": len(skipped),
        "failed_count": len(failed),
        "created": created,
        "skipped": skipped,
        "failed": failed,
    }

@app.post("/api/kb/import/assisted_text")
async def import_assisted_text(payload: KnowledgeAssistTextImport, db: Session = Depends(get_db)):
    """使用 LLM-1 辅助把长文本拆成候选知识切片，全部保存为 draft。"""
    prompt = f"""
你是企业知识库整理助手。请只根据原文整理候选知识切片，不要编造事实、价格、客户名称或承诺。

输出必须是 JSON，格式如下：
{{
  "chunks": [
    {{
      "title": "简短标题",
      "content": "只包含一个业务知识点的原文或规范化表述",
      "knowledge_class": "报价限制条件|能力知识|流程规则|FAQ常见问答|案例|邮件模板|名词定义",
      "business_line": "translation|printing|interpretation|multimedia|exhibition|gifts|general",
      "sub_service": null,
      "language_pair": null,
      "service_scope": null,
      "priority": 50,
      "risk_level": "low|medium|high",
      "review_required": true,
      "pricing_rule": {
        "unit": "per_1000_chars|per_project|per_hour|per_day",
        "currency": "CNY",
        "price_min": null,
        "price_max": null,
        "min_charge": null,
        "urgent_multiplier": null,
        "tax_policy": null
      },
      "tags": {{}}
    }}
  ]
}}

字段约束：
- knowledge_class 只能使用 7 类业务分类：报价限制条件、能力知识、流程规则、FAQ常见问答、案例、邮件模板、名词定义。
- 报价规则不是知识分类；明确价格和单位必须进入结构化 pricing_rule，普通 chunk 只承载报价限制条件或说明证据。
- business_line 只能使用 translation、printing、interpretation、multimedia、exhibition、gifts、general。
- 价格、费用、折扣、税费、合同、承诺、赔付类内容 risk_level 必须为 high 且 review_required=true。
- 一条 chunk 只能表达一个知识点；如果一句话里有报价、最低收费、加急规则，必须拆成多条。
- 确有原文价格数字时必须尽量输出 pricing_rule；无法确定价格数字时保留 null，但不能编造数字。
- 不确定的字段填 null，不要自造 code。

资料标题：{payload.title}
默认服务：{payload.business_line}
原文：
{payload.content}
"""
    llm_result = IntentEngine.run_llm1_json_prompt(prompt, user_id="kb_assisted_text_import")
    if not llm_result or not isinstance(llm_result.get("chunks"), list):
        raise HTTPException(status_code=502, detail="LLM 辅助拆分失败或返回格式错误")

    chunks_payload = []
    for idx, raw_chunk in enumerate(llm_result["chunks"], start=1):
        title = str(raw_chunk.get("title") or "").strip()
        content = str(raw_chunk.get("content") or "").strip()
        if not title or not content:
            continue
        knowledge_class, raw_knowledge_class = normalize_knowledge_class(raw_chunk.get("knowledge_class"))
        class_config = KB_KNOWLEDGE_CLASS_CONFIG.get(knowledge_class or "")
        knowledge_type, raw_knowledge_type = normalize_code("knowledge_type", raw_chunk.get("knowledge_type") or (class_config or {}).get("knowledge_type") or "faq")
        legacy_chunk_type, raw_legacy_chunk_type = normalize_code("chunk_type", raw_chunk.get("chunk_type"))
        business_line, raw_business_line = normalize_code("business_line", raw_chunk.get("business_line") or payload.business_line)
        language_pair, raw_language_pair = normalize_code("language_pair", raw_chunk.get("language_pair"))
        service_scope, raw_service_scope = normalize_code("service_scope", raw_chunk.get("service_scope"))
        tags = raw_chunk.get("tags") or {}
        raw_codes = {
            "raw_knowledge_type": raw_knowledge_type,
            "raw_knowledge_class": raw_knowledge_class,
            "raw_chunk_type": raw_legacy_chunk_type,
            "raw_business_line": raw_business_line,
            "raw_language_pair": raw_language_pair,
            "raw_service_scope": raw_service_scope,
        }
        tags.update({key: value for key, value in raw_codes.items() if value})
        tags["knowledge_class"] = knowledge_class or knowledge_class_from_pair(knowledge_type, legacy_chunk_type) or "faq"
        class_config = KB_KNOWLEDGE_CLASS_CONFIG.get(tags["knowledge_class"], class_config or KB_KNOWLEDGE_CLASS_CONFIG["faq"])
        pricing_rule = infer_pricing_rule_candidate(
            title=title,
            content=content,
            business_line=business_line or payload.business_line,
            language_pair=language_pair,
            service_scope=service_scope,
            raw_pricing_rule=raw_chunk.get("pricing_rule"),
        )
        chunks_payload.append(KnowledgeChunkCreate(
            title=title,
            content=content,
            knowledge_type=class_config["knowledge_type"],
            chunk_type=class_config["chunk_type"],
            business_line=business_line or payload.business_line,
            sub_service=raw_chunk.get("sub_service"),
            language_pair=language_pair,
            service_scope=service_scope,
            priority=int(raw_chunk.get("priority") or 50),
            tags=tags,
            pricing_rule=pricing_rule,
        ))

    if not chunks_payload:
        raise HTTPException(status_code=422, detail="LLM 未返回可入库的有效切片")

    import_batch = f"assist_text_{uuid.uuid4().hex[:12]}"
    document_knowledge_type = "pricing" if all(item.knowledge_type == "pricing" for item in chunks_payload) else "faq"
    document_risk_level = "high" if any(item.knowledge_type == "pricing" or mentions_pricing_topic(item.title, item.content) for item in chunks_payload) else "medium"
    document = KnowledgeDocument(
        title=payload.title,
        knowledge_type=document_knowledge_type,
        business_line=payload.business_line,
        source_type=payload.source_type,
        source_ref=payload.source_ref,
        source_meta={"assist_provider": settings.KB_LLM_ASSIST_PROVIDER, "raw_content_chars": len(payload.content)},
        status="draft",
        owner=payload.owner or "kb_assisted_text",
        import_batch=import_batch,
        risk_level=document_risk_level,
        review_required=True,
        review_status="pending",
        tags={"assist_import": True},
    )
    db.add(document)
    db.flush()

    created_chunks = []
    created_pricing_rules = []
    for idx, item in enumerate(chunks_payload, start=1):
        retrieval_title = _strip_retrieval_title_prefix(item.title)
        retrieval_text = _build_chunk_retrieval_text(item.title, item.content)
        embedding = _optional_embedding_for_storage(retrieval_text, context=f"assist_chunk:{idx}:{item.title}")
        chunk = KnowledgeChunk(
            document_id=document.document_id,
            chunk_no=idx,
            chunk_type=item.chunk_type,
            title=retrieval_title,
            content=item.content,
            keyword_text=retrieval_text,
            embedding=embedding,
            **_embedding_metadata_fields(embedding),
            priority=item.priority,
            business_line=item.business_line,
            sub_service=item.sub_service,
            knowledge_type=item.knowledge_type,
            language_pair=item.language_pair,
            service_scope=item.service_scope,
            structured_tags=item.tags,
            status="draft",
        )
        db.add(chunk)
        db.flush()
        pricing_rule_payload = item.pricing_rule
        if not pricing_rule_payload and item.knowledge_type == "pricing" and item.chunk_type != "constraint":
            pricing_rule_payload = infer_pricing_rule_candidate(
                item.title,
                item.content,
                item.business_line,
                item.language_pair,
                item.service_scope,
                item.customer_tier,
            )
        pricing_rule = _create_pricing_rule(db, document, chunk, pricing_rule_payload)
        _apply_chunk_governance(
            chunk,
            source_type=document.source_type,
            pricing_rule=pricing_rule_payload,
            source_ref=document.source_ref,
        )
        if pricing_rule:
            created_pricing_rules.append(pricing_rule)
        created_chunks.append(chunk)

    if created_chunks:
        document.library_type = created_chunks[0].library_type if all(item.library_type == created_chunks[0].library_type for item in created_chunks) else "reference"
        document.tags = merge_tags(document.tags, library_type=document.library_type)
    db.commit()
    db.refresh(document)
    for chunk in created_chunks:
        db.refresh(chunk)
    for rule in created_pricing_rules:
        db.refresh(rule)
    logger.info("KB_ASSISTED_TEXT_IMPORT %s", json.dumps({
        "title": payload.title,
        "import_batch": import_batch,
        "chunks": len(created_chunks),
    }, ensure_ascii=False))
    return {
        "status": "success",
        "import_batch": import_batch,
        "document": _doc_to_dict(document),
        "chunks": [_chunk_to_dict(chunk) for chunk in created_chunks],
        "pricing_rules": [_pricing_rule_to_dict(rule) for rule in created_pricing_rules],
        "raw_llm_chunks_count": len(llm_result["chunks"]),
    }

# --- 系统运行时脚本管理 API ---

@app.get("/api/system/ai_scripts")
async def get_ai_scripts():
    import os, json
    filepath = os.path.join(os.path.dirname(__file__), "ai_settings.json")
    defaults = {
        "WECOM_RECENT_MESSAGE_LIMIT": 6,
    }
    try:
        if os.path.exists(filepath):
            with open(filepath, "r", encoding="utf-8") as f:
                stored = json.load(f)
                if not isinstance(stored, dict):
                    raise HTTPException(status_code=500, detail="ai_settings.json 格式错误")
                return {**defaults, **stored}
        raise HTTPException(status_code=500, detail="ai_settings.json 不存在")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"读取 ai_settings.json 失败: {exc}") from exc

@app.post("/api/system/ai_scripts")
async def save_ai_scripts(payload: dict):
    import os, json
    filepath = os.path.join(os.path.dirname(__file__), "ai_settings.json")
    defaults = {
        "WECOM_RECENT_MESSAGE_LIMIT": 6,
    }
    try:
        existing = {}
        if os.path.exists(filepath):
            with open(filepath, "r", encoding="utf-8") as f:
                existing = json.load(f) or {}
                if not isinstance(existing, dict):
                    existing = {}
        merged = {**defaults, **existing, **(payload or {})}
        try:
            merged["WECOM_RECENT_MESSAGE_LIMIT"] = max(1, min(30, int(merged.get("WECOM_RECENT_MESSAGE_LIMIT", 6))))
        except (TypeError, ValueError):
            merged["WECOM_RECENT_MESSAGE_LIMIT"] = 6
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=4)
        return {"status": "success", "saved": merged}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- 会话查询 API (供前端使用) ---

def _parse_filter_date(chat_date: str | None) -> tuple[datetime, datetime] | None:
    raw = str(chat_date or "").strip()
    if not raw:
        return None
    try:
        day = datetime.strptime(raw, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="chat_date 必须为 YYYY-MM-DD") from exc
    return day, day + timedelta(days=1)

def _build_api_day_stats(db: Session, *, day_start: datetime, day_end: datetime) -> dict:
    invocations = db.query(ApiAssistInvocation).filter(
        ApiAssistInvocation.triggered_at >= day_start,
        ApiAssistInvocation.triggered_at < day_end,
    ).order_by(ApiAssistInvocation.triggered_at.asc()).all()
    if not invocations:
        return {
            "invocation_count": 0,
            "avg_quality_score": None,
            "quality_score_rule": "调用次数按同会话 + 同客户节点 + 同一天去重；贴合代理分均分仅统计最后一次有效分",
        }

    deduped: dict[tuple[str, int | None], ApiAssistInvocation] = {}
    for item in invocations:
        key = (item.session_id, item.anchor_message_id)
        existing = deduped.get(key)
        if existing is None or (item.triggered_at or datetime.min) >= (existing.triggered_at or datetime.min):
            deduped[key] = item

    valid_items = [
        item for item in deduped.values()
        if item.quality_status == "scored" and item.quality_score is not None
    ]
    avg_score = None
    if valid_items:
        avg_score = round(sum(float(item.quality_score) for item in valid_items) / len(valid_items), 2)
    return {
        "invocation_count": len(deduped),
        "avg_quality_score": avg_score,
        "quality_score_rule": "调用次数按同会话 + 同客户节点 + 同一天去重；贴合代理分均分仅统计最后一次有效分，且单次贴合代理分取知识库1/知识库2两路较高分",
    }


def _build_api_session_marks(db: Session, *, day_start: datetime, day_end: datetime) -> dict[str, dict]:
    invocations = db.query(ApiAssistInvocation).filter(
        ApiAssistInvocation.triggered_at >= day_start,
        ApiAssistInvocation.triggered_at < day_end,
    ).order_by(ApiAssistInvocation.triggered_at.asc()).all()

    session_stats: dict[str, dict[str, Any]] = {}
    deduped: dict[tuple[str, int | None], ApiAssistInvocation] = {}
    for item in invocations:
        session_key = str(item.session_id or "")
        if not session_key:
            continue
        stats = session_stats.setdefault(session_key, {
            "invocation_count": 0,
            "last_triggered_at": None,
        })
        stats["invocation_count"] += 1
        last_triggered_at = stats.get("last_triggered_at")
        if last_triggered_at is None or (item.triggered_at or datetime.min) >= (last_triggered_at or datetime.min):
            stats["last_triggered_at"] = item.triggered_at
        dedupe_key = (session_key, item.anchor_message_id)
        existing = deduped.get(dedupe_key)
        if existing is None or (item.triggered_at or datetime.min) >= (existing.triggered_at or datetime.min):
            deduped[dedupe_key] = item

    quality_by_session: dict[str, list[float]] = {}
    for item in deduped.values():
        if item.quality_status != "scored" or item.quality_score is None:
            continue
        quality_by_session.setdefault(str(item.session_id), []).append(float(item.quality_score))

    result: dict[str, dict] = {}
    for session_id, stats in session_stats.items():
        quality_scores = quality_by_session.get(session_id) or []
        avg_quality = round(sum(quality_scores) / len(quality_scores), 2) if quality_scores else None
        result[session_id] = {
            "has_api_invocation": True,
            "api_invocation_count": int(stats.get("invocation_count") or 0),
            "api_last_triggered_at": stats.get("last_triggered_at"),
            "api_avg_quality_score": avg_quality,
        }
    return result

@app.get("/api/sessions")
async def get_sessions(chat_date: str | None = Query(default=None), db: Session = Depends(get_db)):
    """获取真实会话列表；模拟数据不会进入日常运行视图。"""
    from sqlalchemy import func, cast, Integer, text
    total_started = perf_counter()
    timings_ms: dict[str, float] = {}

    stage_started = perf_counter()
    date_range = _parse_filter_date(chat_date)
    timings_ms["parse_filter_date_ms"] = round((perf_counter() - stage_started) * 1000, 2)

    stage_started = perf_counter()
    results = db.query(
        MessageLog.user_id,
        func.max(MessageLog.timestamp).label("last_msg"),
        func.max(cast(MessageLog.is_mock, Integer)).label("is_mock_int")
    ).filter(MessageLog.is_mock.is_(False))
    if date_range:
        day_start, day_end = date_range
        results = results.filter(
            MessageLog.timestamp >= day_start,
            MessageLog.timestamp < day_end,
        )
    results = results.group_by(MessageLog.user_id).order_by(text("last_msg DESC")).all()
    timings_ms["query_sessions_ms"] = round((perf_counter() - stage_started) * 1000, 2)

    session_marks: dict[str, dict] = {}
    stats_payload = None
    if date_range:
        try:
            stage_started = perf_counter()
            stats_payload = _build_api_day_stats(db, day_start=day_start, day_end=day_end)
            timings_ms["build_day_stats_ms"] = round((perf_counter() - stage_started) * 1000, 2)
        except Exception as exc:
            logger.warning("构建当日 API 统计失败，已降级跳过。 date=%s error=%s", chat_date, exc)
            stats_payload = {
                "invocation_count": 0,
                "avg_quality_score": None,
                "quality_score_rule": "统计降级：当日 API 质量统计构建失败，已跳过",
                "status": "degraded",
                "error": sanitize_text(str(exc)),
            }
        try:
            stage_started = perf_counter()
            session_marks = _build_api_session_marks(db, day_start=day_start, day_end=day_end)
            timings_ms["build_session_marks_ms"] = round((perf_counter() - stage_started) * 1000, 2)
        except Exception as exc:
            logger.warning("构建会话级 API 标记失败，已降级跳过。 date=%s error=%s", chat_date, exc)
            session_marks = {}

    payload = {
        "sessions": [
            {
                "user_id": r.user_id,
                "last_msg": r.last_msg,
                "is_mock": bool(r.is_mock_int),
                **(session_marks.get(str(r.user_id)) or {
                    "has_api_invocation": False,
                    "api_invocation_count": 0,
                    "api_last_triggered_at": None,
                    "api_avg_quality_score": None,
                }),
            }
            for r in results
        ],
        "filter_date": chat_date or None,
        "stats": stats_payload if date_range else None,
    }
    timings_ms["total_ms"] = round((perf_counter() - total_started) * 1000, 2)
    log_session_view_event("SESSIONS", {
        "filter_date": chat_date or None,
        "session_count": len(payload["sessions"]),
        "has_date_stats": bool(date_range),
        "api_invocation_count": (stats_payload or {}).get("invocation_count") if isinstance(stats_payload, dict) else None,
        "stats_status": (stats_payload or {}).get("status") if isinstance(stats_payload, dict) else None,
        "timings_ms": timings_ms,
    })
    return payload

@app.get("/api/sessions/{user_id}/messages")
async def get_messages(user_id: str):
    from time import perf_counter

    total_started = perf_counter()
    db = SessionLocal()
    try:
        messages = db.query(MessageLog).filter(
            MessageLog.user_id == user_id,
            MessageLog.is_mock.is_(False),
        ).order_by(MessageLog.id.asc()).all()
        logger.info(f"查询到用户 {user_id} 的消息，数量: {len(messages)}")
        log_session_view_event("MESSAGES", {
            "user_id": user_id,
            "message_count": len(messages),
            "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
        })
        return [{"id": m.id, "sender": m.sender_type, "content": m.content, "time": m.timestamp} for m in messages]
    finally:
        db.close()

def build_single_session_id(userid: str, external_userid: str) -> str:
    """保持与会话存档 1v1 session_id 生成规则一致"""
    participants = sorted([userid.strip(), external_userid.strip()])
    return f"single_{'_'.join(participants)}"

def _archive_session_candidates(session_id: str) -> list[str]:
    candidate_ids = [session_id]
    legacy_truncated_session_id = session_id[:50]
    if legacy_truncated_session_id not in candidate_ids:
        candidate_ids.append(legacy_truncated_session_id)
    return candidate_ids

def find_existing_single_session_id(db: Session, userid: str, external_userid: str, limit: int = 15) -> tuple[str, list[MessageLog], str]:
    """Prefer the full session_id, but fall back to historical 50-char-truncated ids."""
    session_id = build_single_session_id(userid, external_userid)
    candidate_ids = _archive_session_candidates(session_id)

    for index, candidate_id in enumerate(candidate_ids):
        logs = db.query(MessageLog).filter(
            MessageLog.user_id == candidate_id,
            MessageLog.is_mock.is_(False),
        ).order_by(MessageLog.id.desc()).limit(limit).all()
        if logs:
            return candidate_id, logs, ("exact" if index == 0 else "legacy_truncated_fallback")
    return session_id, [], "exact"

async def _sync_archive_for_session(requested_session_id: str, timeout_seconds: int | None = None) -> dict:
    from time import perf_counter

    resolved_timeout_seconds = max(30, int(timeout_seconds or settings.ARCHIVE_SYNC_TIMEOUT_SECONDS))
    total_started = perf_counter()
    archive_status = ArchiveService.config_status()
    if not archive_status["ready"]:
        return {
            "status": "skipped",
            "reason": "archive_not_ready",
            "msg": "企微会话存档未完整配置",
            "config": archive_status,
            "session_id": requested_session_id,
            "timeout_seconds": resolved_timeout_seconds,
            "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
        }
    try:
        result = await asyncio.wait_for(
            asyncio.to_thread(ArchiveService.sync_today_data, requested_session_id),
            timeout=resolved_timeout_seconds,
        )
    except asyncio.TimeoutError:
        return {
            "status": "timeout",
            "reason": "archive_sync_timeout",
            "msg": f"企微会话存档同步超过 {resolved_timeout_seconds} 秒未返回",
            "session_id": requested_session_id,
            "timeout_seconds": resolved_timeout_seconds,
            "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
        }
    result["requested_session_id"] = requested_session_id
    result["timeout_seconds"] = resolved_timeout_seconds
    result["timings_ms"] = _jsonable(result.get("timings_ms") or {})
    result["timings_ms"]["wait_ms"] = round((perf_counter() - total_started) * 1000, 2)
    return result


async def _sync_archive_for_session_background(requested_session_id: str) -> None:
    try:
        result = await asyncio.to_thread(ArchiveService.sync_today_data, requested_session_id)
        if result.get("status") != "success":
            logger.warning(
                "侧边栏辅助后台企微同步未成功 session_id=%s msg=%s",
                requested_session_id,
                result.get("msg"),
            )
    except Exception as exc:
        logger.error("侧边栏辅助后台企微同步异常 session_id=%s err=%s", requested_session_id, exc)


def _schedule_archive_sync_for_session(requested_session_id: str) -> bool:
    archive_status = ArchiveService.config_status()
    if not archive_status["ready"]:
        return False
    asyncio.create_task(_sync_archive_for_session_background(requested_session_id))
    return True

def extract_external_userid(value: str) -> str:
    """从单聊 session_id 中取真实外部联系人 ID；群聊没有 CRM external_userid。"""
    import re

    text_value = (value or "").strip()
    if text_value.startswith(("wm", "wo", "wb")):
        return text_value
    if text_value.startswith("group_"):
        return ""
    if text_value.startswith("single_"):
        body = text_value.replace("single_", "", 1)
        match = re.search(r"(?:^|_)((?:wm|wo|wb)[A-Za-z0-9_-]+)$", body)
        if match:
            return match.group(1)
    return ""

def collect_fast_track_signals(logs) -> list:
    """扫描最近消息命中的强信号规则，规则异常不阻断主流程"""
    import re
    fast_track_signals = []
    for log in logs:
        content = log.content or ""
        for rule in IntentEngine.get_rules():
            name = rule.get("name", "未命名规则")
            pattern = rule.get("pattern", "")
            exclude_pattern = rule.get("exclude_pattern", "")
            try:
                if not pattern:
                    continue
                if not re.search(pattern, content, re.IGNORECASE):
                    continue
                if exclude_pattern and re.search(exclude_pattern, content, re.IGNORECASE):
                    continue
                if name not in fast_track_signals:
                    fast_track_signals.append(name)
            except re.error as e:
                logger.error(f"Fast-Track 正则规则异常 [{name}]: {e}")
    return fast_track_signals

def log_sidebar_result(event: str, payload: dict):
    """Log one compact JSON line for each sidebar request result."""
    logger.info(
        "SIDEBAR_ASSIST_%s %s",
        event,
        json.dumps(payload, ensure_ascii=False, default=str),
    )

def log_reply_chain_event(event: str, payload: dict):
    """Log one compact JSON line for reply-generation chain UI actions."""
    logger.info(
        "REPLY_CHAIN_%s %s",
        event,
        json.dumps(payload, ensure_ascii=False, default=str),
    )

def log_session_view_event(event: str, payload: dict):
    """Log one compact JSON line for session list/detail reads."""
    logger.info(
        "SESSION_VIEW_%s %s",
        event,
        json.dumps(payload, ensure_ascii=False, default=str),
    )

def _reply_chain_executor_task_key(session_id: str, snapshot_id: str | None = None) -> str:
    return f"{session_id}::{snapshot_id or 'current_tail'}"

def _reply_chain_task_runner(task_key: str, fn, *args):
    try:
        return fn(*args)
    finally:
        with REPLY_CHAIN_ACTIVE_LOCK:
            REPLY_CHAIN_ACTIVE_FUTURES.pop(task_key, None)

def _submit_named_reply_chain_task(task_key: str, fn, *args) -> tuple[Future, bool]:
    with REPLY_CHAIN_ACTIVE_LOCK:
        active = REPLY_CHAIN_ACTIVE_FUTURES.get(task_key)
        if active and not active.done():
            return active, False
        future = REPLY_CHAIN_EXECUTOR.submit(_reply_chain_task_runner, task_key, fn, *args)
        REPLY_CHAIN_ACTIVE_FUTURES[task_key] = future
        return future, True

def _submit_reply_chain_task(
    session_id: str,
    snapshot_id: str | None,
    step: int,
    analytics_record_id: str | None = None,
) -> tuple[Future, bool]:
    task_key = _reply_chain_executor_task_key(session_id, snapshot_id)
    target_fn = reanalyze_snapshot_task if snapshot_id and snapshot_id != "current_tail" else reanalyze_session_task
    target_args = (
        (session_id, snapshot_id, step, analytics_record_id)
        if target_fn is reanalyze_snapshot_task
        else (session_id, step, analytics_record_id)
    )
    return _submit_named_reply_chain_task(task_key, target_fn, *target_args)

def _knowledge_refresh_executor_task_key(session_id: str, snapshot_id: str | None, channel: str) -> str:
    return f"{_reply_chain_executor_task_key(session_id, snapshot_id)}::knowledge::{channel}"

def _submit_knowledge_refresh_task(
    session_id: str,
    snapshot_id: str | None,
    channel: str,
    analytics_record_id: str | None = None,
) -> tuple[Future, bool]:
    task_key = _knowledge_refresh_executor_task_key(session_id, snapshot_id, channel)
    target_fn = refresh_snapshot_knowledge_task if snapshot_id and snapshot_id != "current_tail" else refresh_session_knowledge_task
    target_args = (
        (session_id, snapshot_id, channel, analytics_record_id)
        if target_fn is refresh_snapshot_knowledge_task
        else (session_id, channel, analytics_record_id)
    )
    return _submit_named_reply_chain_task(task_key, target_fn, *target_args)

def _reply_chain_runtime_key(session_id: str, snapshot_id: str | None = None) -> str:
    return snapshot_id or session_id

def _normalize_trigger_source(value: str | None, default: str = "web_manual") -> str:
    raw = str(value or "").strip().lower()
    alias_map = {
        "api": "api",
        "api_trigger": "api",
        "sidebar_api": "api",
        "web": "web_manual",
        "web_manual": "web_manual",
        "frontend": "web_manual",
        "manual": "web_manual",
        "test": "test",
        "qa": "test",
    }
    return alias_map.get(raw, default)

def _is_test_user_agent(user_agent: str | None) -> bool:
    text_value = str(user_agent or "").strip().lower()
    if not text_value:
        return False
    return any(keyword in text_value for keyword in [
        "postman",
        "apifox",
        "insomnia",
        "curl",
        "python-requests",
        "python/",
        "pytest",
        "selenium",
        "playwright",
    ])

def _resolve_trigger_source(
    *,
    request: Request | None,
    explicit: str | None,
    default: str,
) -> str:
    normalized = _normalize_trigger_source(explicit, default=default)
    if normalized == "test":
        return "test"
    user_agent = request.headers.get("user-agent", "") if request else ""
    header_value = request.headers.get("x-trigger-source", "") if request else ""
    if _normalize_trigger_source(header_value, default="") == "test":
        return "test"
    if _is_test_user_agent(user_agent):
        return "test"
    return normalized

def _message_analytics_dict(item: MessageLog) -> dict:
    return {
        "id": item.id,
        "sender_type": item.sender_type,
        "content": sanitize_text(item.content or ""),
        "time": item.timestamp.isoformat() if item.timestamp else None,
    }

def _build_trigger_record_snapshot(
    db: Session,
    *,
    session_id: str,
    snapshot_id: str | None,
) -> dict:
    all_messages = _load_session_messages(db, session_id)
    anchor_message = None
    visible_messages = all_messages
    if snapshot_id and snapshot_id != "current_tail":
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if snapshot and snapshot.anchor_message_id:
            anchor_message = next((item for item in all_messages if item.id == snapshot.anchor_message_id), None)
            visible_messages = _visible_messages_until_anchor(all_messages, snapshot.anchor_message_id)
    if anchor_message is None:
        anchor_message = _find_latest_customer_anchor(all_messages)
        if anchor_message:
            visible_messages = _visible_messages_until_anchor(all_messages, anchor_message.id)
    input_messages = visible_messages[-15:]
    recent_customer_messages = [item for item in input_messages if item.sender_type == "customer"]
    actual_sales_replies = _collect_actual_sales_replies(all_messages, int(anchor_message.id)) if anchor_message else []
    return {
        "anchor_message_id": int(anchor_message.id) if anchor_message else None,
        "anchor_message_time": anchor_message.timestamp if anchor_message else None,
        "anchor_message_text": sanitize_text(anchor_message.content or "") if anchor_message else None,
        "visible_message_ids": [item.id for item in visible_messages],
        "input_messages": [_message_analytics_dict(item) for item in input_messages],
        "recent_customer_messages": [_message_analytics_dict(item) for item in recent_customer_messages],
        "latest_dialog_count": len(visible_messages),
        "actual_sales_replies": _jsonable(actual_sales_replies),
        "actual_sales_reply_text": _reply_block_text(actual_sales_replies, strip_noise=True),
    }

def _create_wecom_trigger_record(
    *,
    session_id: str,
    snapshot_id: str | None,
    run_id: str | None,
    trigger_source: str,
    trigger_kind: str,
    requested_step: int | None = None,
    requested_channel: str | None = None,
    request_status: str = "queued",
    request_payload: dict | None = None,
) -> str | None:
    db = SessionLocal()
    try:
        snapshot = _build_trigger_record_snapshot(db, session_id=session_id, snapshot_id=snapshot_id)
        item = WecomTriggerRecord(
            session_id=session_id,
            snapshot_id=snapshot_id,
            run_id=run_id,
            trigger_source=trigger_source,
            trigger_kind=trigger_kind,
            requested_step=requested_step,
            requested_channel=requested_channel,
            request_status=request_status,
            anchor_message_id=snapshot.get("anchor_message_id"),
            anchor_message_time=snapshot.get("anchor_message_time"),
            anchor_message_text=snapshot.get("anchor_message_text"),
            visible_message_ids=snapshot.get("visible_message_ids"),
            input_messages=snapshot.get("input_messages"),
            recent_customer_messages=snapshot.get("recent_customer_messages"),
            latest_dialog_count=int(snapshot.get("latest_dialog_count") or 0),
            actual_sales_replies=snapshot.get("actual_sales_replies"),
            actual_sales_reply_text=snapshot.get("actual_sales_reply_text"),
            request_payload=_jsonable(request_payload or {}),
            stage_status={},
            triggered_at=datetime.utcnow(),
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return str(item.record_id)
    except Exception as exc:
        logger.error("创建企微触发统计记录失败: %s", exc)
        db.rollback()
        return None
    finally:
        db.close()

def _update_wecom_trigger_record(
    record_id: str | None,
    *,
    request_status: str,
    result_payload: dict | None = None,
    stage_status: dict | None = None,
    error_message: str | None = None,
) -> None:
    if not record_id:
        return
    db = SessionLocal()
    try:
        item = db.query(WecomTriggerRecord).filter(WecomTriggerRecord.record_id == record_id).first()
        if not item:
            return
        item.request_status = str(request_status or item.request_status or "unknown")
        if stage_status is not None:
            item.stage_status = _jsonable(stage_status)
        if result_payload is not None:
            item.result_payload = _jsonable(result_payload)
            item.stage_status = _jsonable((result_payload or {}).get("stage_status") or stage_status or {})
            item.actual_sales_replies = _jsonable((result_payload or {}).get("actual_sales_replies") or item.actual_sales_replies or [])
            item.actual_sales_reply_text = _reply_block_text(item.actual_sales_replies or [])
            item.latest_dialog_count = int((result_payload or {}).get("latest_dialog_count") or item.latest_dialog_count or 0)
        if error_message is not None:
            item.error_message = sanitize_text(error_message)
        item.finished_at = datetime.utcnow()
        db.commit()
    except Exception as exc:
        logger.error("回写企微触发统计记录失败: %s", exc)
        db.rollback()
    finally:
        db.close()

def _build_analysis_result_for_target(
    db: Session,
    *,
    session_id: str,
    snapshot_id: str | None,
) -> dict:
    if snapshot_id and snapshot_id != "current_tail":
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if not snapshot:
            return {}
        all_messages = _load_session_messages(db, session_id)
        visible_messages = _visible_messages_until_anchor(all_messages, snapshot.anchor_message_id)
        result = _snapshot_result_payload(snapshot=snapshot, visible_messages=visible_messages)
        if not result.get("fast_track"):
            result["fast_track"] = collect_fast_track_signals(visible_messages[-15:])
        if not result.get("actual_sales_replies"):
            result["actual_sales_replies"] = _collect_actual_sales_replies(all_messages, snapshot.anchor_message_id)
        result["node_timings_ms"] = _extract_node_timings(result.get("stage_status"))
        return _jsonable(result)

    recent_logs = db.query(MessageLog).filter(
        MessageLog.user_id == session_id,
        MessageLog.is_mock.is_(False),
    ).order_by(MessageLog.id.desc()).limit(15).all()
    fast_track_signals = collect_fast_track_signals(recent_logs)
    summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
    result = _read_only_current_tail_result(
        db=db,
        user_id=session_id,
        summary=summary,
        recent_logs=recent_logs,
        fast_track_signals=fast_track_signals,
    )
    result["node_timings_ms"] = _extract_node_timings(result.get("stage_status"))
    return _jsonable(result)

def _summary_payload_from_result(result: dict | None) -> dict:
    payload = result or {}
    return {
        "topic": payload.get("topic"),
        "core_demand": payload.get("core_demand"),
        "key_facts": payload.get("key_facts"),
        "todo_items": payload.get("todo_items"),
        "risks": payload.get("risks"),
        "to_be_confirmed": payload.get("to_be_confirmed"),
        "status": payload.get("status"),
    }

def _sync_wecom_trigger_record_result(
    *,
    record_id: str | None,
    session_id: str,
    snapshot_id: str | None,
    request_status: str,
    error_message: str | None = None,
) -> None:
    if not record_id:
        return
    db = SessionLocal()
    try:
        item = db.query(WecomTriggerRecord).filter(WecomTriggerRecord.record_id == record_id).first()
        if not item:
            return
        result = _build_analysis_result_for_target(
            db,
            session_id=session_id,
            snapshot_id=snapshot_id,
        )
        if result:
            actual_sales_replies = result.get("actual_sales_replies") or item.actual_sales_replies or []
            similarity_payload, best_score, quality_status = _build_api_similarity_payload(
                result_payload=result,
                actual_sales_replies=actual_sales_replies,
                triggered_at=item.triggered_at,
            )
            result["summary"] = _summary_payload_from_result(result)
            result["api_quality_similarity"] = similarity_payload
            result["api_quality_score"] = float(best_score) if best_score is not None else None
            result["api_quality_status"] = quality_status
            item.result_payload = _jsonable(result)
            item.stage_status = _jsonable(result.get("stage_status") or {})
            item.actual_sales_replies = _jsonable(actual_sales_replies)
            item.actual_sales_reply_text = _reply_block_text(actual_sales_replies, strip_noise=True)
        item.request_status = str(request_status or item.request_status or "unknown")
        if error_message is not None:
            item.error_message = sanitize_text(error_message)
        item.finished_at = datetime.utcnow()
        db.commit()
    except Exception as exc:
        logger.error("同步企微触发统计结果失败 record_id=%s err=%s", record_id, exc)
        db.rollback()
    finally:
        db.close()

def _mark_reply_chain_requested(session_id: str, snapshot_id: str | None, step: int, run_id: str) -> None:
    db = SessionLocal()
    try:
        requested_at = _utc_now_iso()
        if snapshot_id and snapshot_id != "current_tail":
            snapshot = db.query(ReplyChainSnapshot).filter(
                ReplyChainSnapshot.snapshot_id == snapshot_id,
                ReplyChainSnapshot.session_id == session_id,
            ).first()
            if not snapshot:
                return
            stage_status = dict(snapshot.stage_status or {})
            stage_status["last_requested_run_id"] = run_id
            stage_status["last_requested_step"] = step
            if step == 1:
                stage_status["llm1"] = "queued"
                stage_status["llm1_requested_at"] = requested_at
                stage_status["llm1_display_reset"] = True
            elif step == 2:
                stage_status["llm2"] = "queued"
                stage_status["llm2_compare"] = "queued"
            snapshot.stage_status = stage_status
            db.commit()
            return

        summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
        if not summary:
            return
        stage_status = dict(summary.stage_status or {})
        stage_status["last_requested_run_id"] = run_id
        stage_status["last_requested_step"] = step
        if step == 1:
            stage_status["llm1"] = "queued"
            stage_status["llm1_requested_at"] = requested_at
            stage_status["llm1_display_reset"] = True
        elif step == 2:
            stage_status["llm2"] = "queued"
            stage_status["llm2_compare"] = "queued"
        summary.stage_status = stage_status
        db.commit()
    finally:
        db.close()

def _empty_knowledge_evidence_context() -> dict:
    return {"rules": [], "faqs": [], "cases": []}

def _knowledge_channel_pending_payload(channel: str, *, query_text: str = "") -> dict:
    source = "sales_kb_api" if channel == "knowledge_external_api" else "knowledge_v2"
    return {
        "status": "queued",
        "query_text": sanitize_text(query_text or ""),
        "hits": [],
        "replyable_hits": [],
        "human_only_hits": [],
        "manual_review_required": False,
        "confidence_score": None,
        "filters_used": {},
        "evidence_refs": [],
        "evidence_context": _empty_knowledge_evidence_context(),
        "query_features": {},
        "source": source,
    }

def _knowledge_v2_error_payload(
    *,
    query_text: str,
    query_features: dict | None,
    error: str,
) -> dict:
    return {
        "status": "error",
        "query_text": sanitize_text(query_text or ""),
        "hits": [],
        "manual_review_required": False,
        "confidence_score": None,
        "filters_used": {},
        "evidence_refs": [],
        "evidence_context": _empty_knowledge_evidence_context(),
        "query_features": query_features or {},
        "source": "knowledge_v2",
        "error": sanitize_text(error or "知识库1检索失败"),
    }

def _knowledge_generation_placeholder(channel: str, *, reason: str) -> dict:
    source = "sales_kb_api" if channel == "knowledge_external_api" else "knowledge_v2"
    payload = {
        "status": "skipped_missing_knowledge",
        "query_text": "",
        "hits": [],
        "replyable_hits": [],
        "human_only_hits": [],
        "manual_review_required": False,
        "confidence_score": None,
        "filters_used": {},
        "evidence_refs": [],
        "source": source,
        "error": sanitize_text(reason),
    }
    if channel == "knowledge_v2":
        payload["query_features"] = {}
        payload["evidence_context"] = _empty_knowledge_evidence_context()
    return payload

def _knowledge_payload_ready(payload: dict | None) -> bool:
    if not isinstance(payload, dict) or not payload:
        return False
    status = str(payload.get("status") or "").strip() or "unknown"
    return status not in {"not_started", "queued", "running"}

def _mark_knowledge_refresh_requested(session_id: str, snapshot_id: str | None, channel: str, run_id: str) -> None:
    db = SessionLocal()
    try:
        if snapshot_id and snapshot_id != "current_tail":
            snapshot = db.query(ReplyChainSnapshot).filter(
                ReplyChainSnapshot.snapshot_id == snapshot_id,
                ReplyChainSnapshot.session_id == session_id,
            ).first()
            if not snapshot:
                return
            stage_status = dict(snapshot.stage_status or {})
            stage_status["last_requested_run_id"] = run_id
            stage_status["last_requested_action"] = channel
            stage_status[channel] = "queued"
            if channel == "knowledge_v2":
                snapshot.knowledge_v2 = _knowledge_channel_pending_payload(channel)
                snapshot.knowledge_status = "queued"
                snapshot.knowledge_log_id = None
                snapshot.knowledge_confidence_score = None
                snapshot.knowledge_manual_review_required = False
            else:
                snapshot.knowledge_external_api = _knowledge_channel_pending_payload(channel)
            snapshot.stage_status = stage_status
            db.commit()
            return
        summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
        if not summary:
            return
        stage_status = dict(summary.stage_status or {})
        stage_status["last_requested_run_id"] = run_id
        stage_status["last_requested_action"] = channel
        stage_status[channel] = "queued"
        if channel == "knowledge_v2":
            summary.knowledge_v2 = _knowledge_channel_pending_payload(channel)
            summary.knowledge_status = "queued"
            summary.knowledge_log_id = None
            summary.knowledge_confidence_score = None
            summary.knowledge_manual_review_required = False
        else:
            summary.knowledge_external_api = _knowledge_channel_pending_payload(channel)
        summary.stage_status = stage_status
        db.commit()
    finally:
        db.close()

def _message_ui_dict(item: MessageLog) -> dict:
    return {
        "id": item.id,
        "sender": item.sender_type,
        "content": item.content,
        "time": item.timestamp,
    }

def _message_chain_dict(item: MessageLog) -> dict:
    return {
        "id": item.id,
        "sender_type": item.sender_type,
        "content": item.content,
        "timestamp": item.timestamp,
    }

def _load_session_messages(db: Session, session_id: str) -> list[MessageLog]:
    return db.query(MessageLog).filter(
        MessageLog.user_id == session_id,
        MessageLog.is_mock.is_(False),
    ).order_by(MessageLog.id.asc()).all()

def _visible_messages_until_anchor(messages: list[MessageLog], anchor_message_id: int | None = None) -> list[MessageLog]:
    if not anchor_message_id:
        return list(messages)
    return [item for item in messages if item.id <= anchor_message_id]

def _collect_actual_sales_replies(messages: list[MessageLog], anchor_message_id: int) -> list[dict]:
    replies: list[dict] = []
    collecting = False
    for item in messages:
        if item.id <= anchor_message_id:
            continue
        if item.sender_type == "sales":
            collecting = True
            replies.append({
                "id": item.id,
                "sender": item.sender_type,
                "content": item.content,
                "time": item.timestamp.isoformat() if item.timestamp else None,
            })
            continue
        if collecting and item.sender_type == "customer":
            break
    return replies

_REPLY_BLOCK_MESSAGE_PLACEHOLDER_RE = re.compile(r"^\[[^\]]+消息\].*$")
_REPLY_BLOCK_ACK_LINES = {
    "好",
    "好的",
    "好的好的",
    "嗯",
    "嗯嗯",
    "哦",
    "哦哦",
    "收到",
    "收到啦",
    "好滴",
    "好滴好滴",
    "ok",
    "okay",
    "okk",
    "yes",
}

def _reply_block_lines_from_text(text: str | None) -> list[str]:
    return [
        sanitize_text(line.strip())
        for line in str(text or "").splitlines()
        if sanitize_text(line.strip())
    ]

def _reply_block_line_is_placeholder(text: str) -> bool:
    normalized = re.sub(r"\s+", "", str(text or ""))
    if not normalized:
        return True
    if _REPLY_BLOCK_MESSAGE_PLACEHOLDER_RE.fullmatch(normalized):
        return True
    return bool(re.fullmatch(r"\d{1,4}", normalized))

def _reply_block_line_is_ack(text: str) -> bool:
    normalized = re.sub(r"\s+", "", str(text or "")).lower()
    if not normalized:
        return False
    return normalized in _REPLY_BLOCK_ACK_LINES

def _clean_reply_block_text(text: str | None) -> str:
    raw_lines = _reply_block_lines_from_text(text)
    if not raw_lines:
        return ""
    substantive_lines: list[str] = []
    fallback_lines: list[str] = []
    for line in raw_lines:
        if _reply_block_line_is_placeholder(line):
            continue
        fallback_lines.append(line)
        if _reply_block_line_is_ack(line):
            continue
        substantive_lines.append(line)
    if substantive_lines:
        return "\n".join(substantive_lines).strip()
    if fallback_lines:
        return "\n".join(fallback_lines).strip()
    return "\n".join(raw_lines).strip()

def _reply_block_text(replies: list[dict] | None, *, strip_noise: bool = False) -> str:
    text = "\n".join(
        str(item.get("content") or "").strip()
        for item in (replies or [])
        if str(item.get("content") or "").strip()
    ).strip()
    return _clean_reply_block_text(text) if strip_noise else text

def _reply_block_hash(replies: list[dict] | None) -> str | None:
    text = _reply_block_text(replies, strip_noise=False)
    if not text:
        return None
    return hashlib.sha1(text.encode("utf-8")).hexdigest()

def _message_id_signature(message_ids: list[Any] | None) -> str:
    normalized: list[str] = []
    for raw_value in message_ids or []:
        if raw_value is None:
            continue
        try:
            normalized.append(str(int(raw_value)))
        except (TypeError, ValueError):
            normalized.append(str(raw_value))
    payload = ",".join(normalized)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()

def _find_latest_customer_anchor(messages: list[MessageLog]) -> MessageLog | None:
    for item in reversed(messages or []):
        if item.sender_type == "customer":
            return item
    return None

def _short_anchor_text(text_value: str | None, max_len: int = 18) -> str:
    text_value = sanitize_text(str(text_value or "").strip())
    if len(text_value) <= max_len:
        return text_value
    return text_value[:max_len] + "..."

def _api_invocation_label(item: ApiAssistInvocation) -> str:
    ts = item.anchor_message_time.strftime("%m-%d %H:%M") if item.anchor_message_time else (
        item.triggered_at.strftime("%m-%d %H:%M") if item.triggered_at else "API节点"
    )
    return f"API {ts} {_short_anchor_text(item.anchor_message_text) or '未命名节点'}"

def _api_invocation_meta(item: ApiAssistInvocation) -> dict:
    return {
        "snapshot_id": str(item.invocation_id),
        "kind": "api_sidebar_assist",
        "label": _api_invocation_label(item),
        "anchor_message_id": item.anchor_message_id,
        "anchor_message_time": item.anchor_message_time,
        "anchor_message_text": item.anchor_message_text,
        "triggered_at": item.triggered_at,
        "has_v1": bool((item.result_payload or {}).get("has_v1")),
        "has_v2": bool((item.result_payload or {}).get("has_v2")),
        "source_label": "API调用",
    }

def _api_invocation_identity_key(item: ApiAssistInvocation) -> str:
    return (
        f"{item.session_id or ''}"
        f"::{int(item.anchor_message_id or 0)}"
        f"::{_message_id_signature(item.visible_message_ids)}"
    )

def _sidebar_assist_request_key(
    *,
    session_id: str,
    messages: list[MessageLog],
    force_refresh: bool,
) -> str:
    anchor_message = _find_latest_customer_anchor(messages)
    return (
        f"{session_id or ''}"
        f"::{1 if force_refresh else 0}"
        f"::{int(anchor_message.id) if anchor_message else 0}"
        f"::{_message_id_signature([item.id for item in messages])}"
    )

def _decorate_sidebar_assist_result(
    payload: dict,
    *,
    status: str,
    reason: str,
    request_key: str | None = None,
) -> dict:
    result = dict(_jsonable(payload) or {})
    result["request_dedupe"] = {
        "status": status,
        "reason": reason,
        "ttl_seconds": SIDEBAR_ASSIST_RECENT_DEDUPE_TTL_SECONDS,
        "request_key": request_key,
    }
    return result

def _get_cached_sidebar_assist_result(request_key: str) -> dict | None:
    now_epoch = datetime.utcnow().timestamp()
    with SIDEBAR_ASSIST_CACHE_LOCK:
        cached = SIDEBAR_ASSIST_RESULT_CACHE.get(request_key)
        if not cached:
            return None
        cached_at_epoch = float(cached.get("cached_at_epoch") or 0)
        if now_epoch - cached_at_epoch > SIDEBAR_ASSIST_RECENT_DEDUPE_TTL_SECONDS:
            SIDEBAR_ASSIST_RESULT_CACHE.pop(request_key, None)
            return None
        return dict(cached.get("payload") or {})

def _remember_sidebar_assist_result(request_key: str, payload: dict) -> None:
    now_epoch = datetime.utcnow().timestamp()
    with SIDEBAR_ASSIST_CACHE_LOCK:
        expired_keys = [
            key
            for key, value in SIDEBAR_ASSIST_RESULT_CACHE.items()
            if now_epoch - float(value.get("cached_at_epoch") or 0) > SIDEBAR_ASSIST_RECENT_DEDUPE_TTL_SECONDS
        ]
        for key in expired_keys:
            SIDEBAR_ASSIST_RESULT_CACHE.pop(key, None)
        SIDEBAR_ASSIST_RESULT_CACHE[request_key] = {
            "cached_at_epoch": now_epoch,
            "payload": dict(_jsonable(payload) or {}),
        }

def _acquire_sidebar_assist_future(request_key: str) -> tuple[Future, bool]:
    with SIDEBAR_ASSIST_ACTIVE_LOCK:
        current_future = SIDEBAR_ASSIST_ACTIVE_FUTURES.get(request_key)
        if current_future and not current_future.done():
            return current_future, False
        new_future: Future = Future()
        SIDEBAR_ASSIST_ACTIVE_FUTURES[request_key] = new_future
        return new_future, True

def _resolve_sidebar_assist_future(
    request_key: str,
    future: Future,
    *,
    result: dict | None = None,
    error: Exception | None = None,
) -> None:
    with SIDEBAR_ASSIST_ACTIVE_LOCK:
        if SIDEBAR_ASSIST_ACTIVE_FUTURES.get(request_key) is future:
            SIDEBAR_ASSIST_ACTIVE_FUTURES.pop(request_key, None)
    if future.done():
        return
    if error is not None:
        future.set_exception(error)
        return
    future.set_result(dict(_jsonable(result) or {}))

def _is_same_calendar_day(left: datetime | None, right: datetime | None) -> bool:
    return bool(left and right and left.date() == right.date())

def _build_api_similarity_payload(
    *,
    result_payload: dict,
    actual_sales_replies: list[dict],
    triggered_at: datetime | None,
) -> tuple[dict, Decimal | None, str]:
    if not actual_sales_replies:
        return ({
            "status": "pending_no_sales_reply",
            "reason": "触发后尚无我方连续回复块，暂不能计算贴合代理分。",
            "scores": [],
            "actual_sales_reply_text": "",
        }, None, "pending_no_sales_reply")

    first_reply_time = None
    first_reply_raw = actual_sales_replies[0].get("time")
    if first_reply_raw:
        try:
            first_reply_time = datetime.fromisoformat(str(first_reply_raw).replace("Z", "+00:00"))
        except ValueError:
            first_reply_time = None
    if triggered_at and first_reply_time and not _is_same_calendar_day(triggered_at, first_reply_time):
        return ({
            "status": "ignored_cross_day",
            "reason": "触发后的第一条我方回复发生在次日，该次 API 调用按误触发处理，不纳入统计。",
            "scores": [],
            "actual_sales_reply_text": _reply_block_text(actual_sales_replies, strip_noise=True),
        }, None, "ignored_cross_day")

    actual_reply_text = _reply_block_text(actual_sales_replies, strip_noise=True)
    similarity = IntentEngine.score_reply_similarity_to_sales_block(
        kb1_reply_text=result_payload.get("reply_reference"),
        kb2_reply_text=result_payload.get("reply_reference_compare"),
        actual_sales_reply_text=actual_reply_text,
        summary_json={
            "topic": result_payload.get("topic"),
            "core_demand": result_payload.get("core_demand"),
            "status": result_payload.get("status"),
        },
    )
    similarity["actual_sales_reply_text"] = actual_reply_text
    similarity["quality_score_rule"] = "取知识库1/知识库2两路贴合代理分较高者作为单次贴合代理分"
    best_score = similarity.get("best_score")
    return similarity, _decimal_score(best_score) if best_score is not None else None, similarity.get("status") or "scored"

def _refresh_api_invocation_quality(
    db: Session,
    item: ApiAssistInvocation,
    *,
    all_messages: list[MessageLog] | None = None,
) -> None:
    messages = all_messages if all_messages is not None else _load_session_messages(db, item.session_id)
    actual_sales_replies = _collect_actual_sales_replies(messages, int(item.anchor_message_id or 0)) if item.anchor_message_id else []
    reply_hash = _reply_block_hash(actual_sales_replies)
    if (
        item.actual_sales_reply_hash == reply_hash
        and item.quality_similarity
        and item.quality_status
    ):
        return

    similarity_payload, best_score, quality_status = _build_api_similarity_payload(
        result_payload=dict(item.result_payload or {}),
        actual_sales_replies=actual_sales_replies,
        triggered_at=item.triggered_at,
    )
    item.actual_sales_replies = _jsonable(actual_sales_replies)
    item.actual_sales_reply_text = _reply_block_text(actual_sales_replies, strip_noise=True)
    item.actual_sales_reply_hash = reply_hash
    item.quality_similarity = _jsonable(similarity_payload)
    item.quality_score = best_score
    item.quality_status = quality_status
    item.quality_scored_at = datetime.utcnow()

def _store_api_assist_invocation(
    db: Session,
    *,
    result_payload: dict,
    session_id: str,
    requested_session_id: str,
    external_userid: str,
    sales_userid: str,
    messages: list[MessageLog],
    trigger_source: str = "api",
) -> ApiAssistInvocation | None:
    anchor_message = _find_latest_customer_anchor(messages)
    if not anchor_message:
        return None
    visible_messages = _visible_messages_until_anchor(messages, anchor_message.id)
    item = ApiAssistInvocation(
        session_id=session_id,
        requested_session_id=requested_session_id,
        external_userid=external_userid,
        sales_userid=sales_userid,
        anchor_message_id=anchor_message.id,
        anchor_message_time=anchor_message.timestamp,
        anchor_message_text=anchor_message.content,
        visible_message_ids=[msg.id for msg in visible_messages],
        latest_dialog_count=len(visible_messages),
        trigger_source=trigger_source,
        trigger_kind="api_sidebar_assist",
        triggered_at=datetime.utcnow(),
        stage_status=_jsonable(result_payload.get("stage_status") or {}),
        result_payload=_jsonable(result_payload),
    )
    db.add(item)
    db.flush()
    _refresh_api_invocation_quality(db, item, all_messages=messages)
    return item

def _summary_json_from_source(source: Any) -> dict:
    if not source:
        return {}
    return {
        "topic": source.topic,
        "core_demand": source.core_demand,
        "key_facts": source.key_facts,
        "todo_items": source.todo_items,
        "risks": source.risks,
        "to_be_confirmed": source.to_be_confirmed,
        "status": source.status,
    }

def _safe_summary_text(value: Any, default: str = "未明确", max_len: int | None = None) -> str:
    if value is None or value == "":
        text_value = default
    elif isinstance(value, (dict, list)):
        text_value = json.dumps(value, ensure_ascii=False)
    else:
        text_value = str(value)
    return text_value[:max_len] if max_len else text_value

def _apply_summary_fields(target: Any, summary: dict | None):
    payload = summary or {}
    target.topic = _safe_summary_text(payload.get("topic"), max_len=200)
    target.core_demand = _safe_summary_text(payload.get("core_demand"))
    target.key_facts = payload.get("key_facts") or {}
    target.todo_items = payload.get("todo_items") or []
    target.risks = _safe_summary_text(payload.get("risks"))
    target.to_be_confirmed = _safe_summary_text(payload.get("to_be_confirmed"))
    target.status = _safe_summary_text(payload.get("status"), max_len=50)

def _run_llm1_compare_bundle(context: list[dict], user_id: str) -> dict:
    compare_config = _runtime_llm_values("LLM1_COMPARE")
    if not _runtime_llm_configured(compare_config):
        return {
            "configured": False,
            "status": "not_configured",
            "summary": None,
            "prompt_trace": None,
        }
    request_spec = IntentEngine.build_llm1_request(
        context,
        user_id=user_id,
        api_url=compare_config["api_url"],
        api_key=compare_config["api_key"],
        model=compare_config["model"],
        timeout_seconds=compare_config["timeout_seconds"],
        label="LLM1_COMPARE",
    )
    prompt_trace = dict(request_spec.get("prompt_trace") or {})
    prompt_trace["result"] = "running"
    try:
        bundle = IntentEngine.run_llm1_request(request_spec)
        prompt_trace = dict(bundle.get("prompt_trace") or prompt_trace)
        prompt_trace["result"] = prompt_trace.get("result") or "success"
        return {
            "configured": True,
            "status": "done",
            "summary": bundle.get("summary"),
            "prompt_trace": prompt_trace,
        }
    except Exception as exc:
        prompt_trace["result"] = "error"
        prompt_trace["reason"] = sanitize_text(str(exc))
        logger.error("LLM-1 对比模型调用失败: %s", exc)
        return {
            "configured": True,
            "status": "error",
            "summary": None,
            "prompt_trace": prompt_trace,
        }

def _select_candidate(candidates: list[dict] | None, model_slot: str, knowledge_source: str | None = None) -> dict | None:
    for item in candidates or []:
        if (
            item.get("model_slot") == model_slot
            and (knowledge_source is None or item.get("knowledge_source") == knowledge_source)
            and item.get("status") == "done"
            and item.get("content")
        ):
            return item
    return None

def _first_model_candidate_entry(candidates: list[dict] | None, model_slot: str, knowledge_source: str | None = None) -> dict | None:
    for item in candidates or []:
        if item.get("model_slot") == model_slot and (knowledge_source is None or item.get("knowledge_source") == knowledge_source):
            return item
    return None

def _needs_full_reply_analysis(
    candidates: list[dict] | None,
    *,
    compare_content: str | None = None,
    reply_scores: dict | None = None,
    compare_configured: bool | None = None,
) -> bool:
    candidate_items = candidates or []
    done_candidates = [
        item for item in candidate_items
        if item.get("status") == "done" and str(item.get("content") or "").strip()
    ]
    if not done_candidates:
        return True

    expected_style_ids = {
        str(item.get("id") or "").strip()
        for item in IntentEngine.get_reply_style_options(enabled_only=True)
        if str(item.get("id") or "").strip()
    }
    done_primary_style_ids = {
        str(item.get("style_id") or "").strip()
        for item in done_candidates
        if item.get("model_slot") == "llm2" and str(item.get("style_id") or "").strip()
    }
    if expected_style_ids and done_primary_style_ids != expected_style_ids:
        return True

    compare_enabled = _runtime_llm_configured(_runtime_llm_values("LLM2_COMPARE")) if compare_configured is None else bool(compare_configured)
    if compare_enabled:
        has_compare = bool(str(compare_content or "").strip()) or any(
            item.get("model_slot") == "llm2_compare"
            and item.get("status") == "done"
            and str(item.get("content") or "").strip()
            for item in done_candidates
        )
        if not has_compare:
            return True

    if not isinstance(reply_scores, dict):
        return True
    return False

def _reply_scores_need_refresh(
    stored_scores: dict | None,
    *,
    candidates: list[dict] | None,
    actual_sales_replies: list[dict] | None,
) -> bool:
    if not isinstance(stored_scores, dict):
        return True

    stored_ai = stored_scores.get("ai_candidates") or []
    stored_sales = stored_scores.get("actual_sales_replies") or []

    done_candidates = [
        item for item in (_prepare_candidates_for_scoring(candidates or []))
        if item.get("status") == "done" and str(item.get("content") or "").strip()
    ]
    sales_replies = [
        item for item in (actual_sales_replies or [])
        if str(item.get("content") or "").strip()
    ]

    stored_ai_ids = {
        str(item.get("candidate_id") or "").strip()
        for item in stored_ai
        if str(item.get("candidate_id") or "").strip()
    }
    current_ai_ids = {
        str(item.get("candidate_id") or "").strip()
        for item in done_candidates
        if str(item.get("candidate_id") or "").strip()
    }
    if stored_ai_ids != current_ai_ids:
        return True

    stored_sales_ids = {
        str(item.get("reply_id") or "").strip()
        for item in stored_sales
        if str(item.get("reply_id") or "").strip()
    }
    current_sales_ids = {
        str(item.get("id") or "").strip()
        for item in sales_replies
        if str(item.get("id") or "").strip()
    }
    if stored_sales_ids != current_sales_ids:
        return True

    return False

def _load_or_refresh_reply_scores(
    *,
    stored_scores: dict | None,
    summary_json: dict | None,
    knowledge_payload: dict | None,
    crm_context: dict | None,
    candidates: list[dict] | None,
    actual_sales_replies: list[dict] | None,
) -> dict | None:
    prepared_candidates = _prepare_candidates_for_scoring(candidates or [])
    if not prepared_candidates:
        return stored_scores if isinstance(stored_scores, dict) else None
    if not _reply_scores_need_refresh(
        stored_scores,
        candidates=prepared_candidates,
        actual_sales_replies=actual_sales_replies,
    ):
        return stored_scores
    return IntentEngine.score_reply_candidates(
        summary_json=summary_json,
        knowledge_payload=knowledge_payload or {},
        crm_context=crm_context,
        candidates=prepared_candidates,
        actual_sales_replies=actual_sales_replies or [],
    )

def _prepare_candidates_for_scoring(candidates: list[dict] | None) -> list[dict]:
    prepared: list[dict] = []
    for item in candidates or []:
        row = dict(item or {})
        content = str(row.get("content") or "").strip()
        if not content:
            reply_reference = str(row.get("reply_reference") or "").strip()
            followup_rationale = str(row.get("followup_rationale") or "").strip()
            if reply_reference or followup_rationale:
                parts = []
                if reply_reference:
                    parts.append("【企微回复参考】")
                    parts.append(reply_reference)
                if followup_rationale:
                    parts.append("")
                    parts.append("【跟进思路说明】")
                    parts.append(followup_rationale)
                row["content"] = "\n".join(parts).strip()
        prepared.append(row)
    return prepared


def _merge_reply_candidates(existing: list[dict] | None, incoming: list[dict] | None) -> list[dict]:
    merged: list[dict] = []
    index_by_id: dict[str, int] = {}
    for item in existing or []:
        candidate_id = str((item or {}).get("candidate_id") or "").strip()
        row = dict(item or {})
        if candidate_id:
            index_by_id[candidate_id] = len(merged)
        merged.append(row)
    for item in incoming or []:
        candidate_id = str((item or {}).get("candidate_id") or "").strip()
        row = dict(item or {})
        if candidate_id and candidate_id in index_by_id:
            merged[index_by_id[candidate_id]] = row
        else:
            if candidate_id:
                index_by_id[candidate_id] = len(merged)
            merged.append(row)
    return merged

def _generate_reply_style_candidates(
    *,
    summary_json: dict,
    knowledge: dict,
    knowledge_compare: dict | None,
    crm_context: dict | None,
    actual_sales_replies: list[dict] | None,
    runtime_key: str,
    single_model_single_style: bool | None = None,
    enable_scoring: bool | None = None,
) -> dict:
    stage_started = perf_counter()
    single_model_single_style = (
        settings.API_REPLY_SINGLE_MODEL_SINGLE_STYLE
        if single_model_single_style is None
        else bool(single_model_single_style)
    )
    enable_scoring = (
        settings.API_REPLY_ENABLE_SCORING
        if enable_scoring is None
        else bool(enable_scoring)
    )
    styles = IntentEngine.get_reply_style_options(enabled_only=True)
    if single_model_single_style and styles:
        styles = styles[:1]
    primary_config = _runtime_llm_values("LLM2")
    compare_config = _runtime_llm_values("LLM2_COMPARE")
    compare_model_configured = _runtime_llm_configured(compare_config)
    candidates: list[dict] = []
    llm_runtime = _llm_runtime_config()
    model_specs = [
        {
            "slot": "llm2",
            "label": "主模型",
            "display_name": llm_runtime["llm2"]["display_name"],
            "provider": llm_runtime["llm2"]["provider"],
            "config": primary_config,
        }
    ]
    if not single_model_single_style:
        model_specs.append({
            "slot": "llm2_compare",
            "label": "对比模型",
            "display_name": llm_runtime["llm2_compare"]["display_name"] or llm_runtime["llm2_compare"]["model"] or "对比模型",
            "provider": llm_runtime["llm2_compare"]["provider"],
            "config": compare_config,
        })

    knowledge_specs = [
        {
            "key": "knowledge_v2",
            "label": "知识库1",
            "payload": knowledge if isinstance(knowledge, dict) else {},
            "status": str((knowledge or {}).get("status") or "not_started"),
        },
        {
            "key": "knowledge_external_api",
            "label": "知识库2",
            "payload": knowledge_compare if isinstance(knowledge_compare, dict) else {},
            "status": str((knowledge_compare or {}).get("status") or "not_started"),
        },
    ]

    model_stats: dict[str, dict[str, int]] = {}
    stage_parts_ms: dict[str, float] = {}

    def ensure_model_stats(slot: str) -> dict[str, int]:
        return model_stats.setdefault(slot, {"done": 0, "failures": 0, "placeholders": 0})

    def append_candidate(
        *,
        candidate_id: str,
        model_spec: dict,
        knowledge_spec: dict,
        style: dict,
        status: str,
        content: str = "",
        validation: dict | None = None,
        prompt_trace: dict | None = None,
        reason: str = "",
    ) -> None:
        candidates.append({
            "candidate_id": candidate_id,
            "model_slot": model_spec["slot"],
            "model_label": model_spec["label"],
            "model_display_name": model_spec["display_name"],
            "model_provider": model_spec["provider"],
            "knowledge_source": knowledge_spec["key"],
            "knowledge_source_label": knowledge_spec["label"],
            "style_id": style["id"],
            "style_title": style["title"],
            "style_content": style["content"],
            "reply_style": style,
            "status": status,
            "content": content,
            "validation": validation,
            "prompt_trace": prompt_trace,
            "reason": reason,
        })

    def knowledge_skip_reason(spec: dict) -> tuple[str, str] | None:
        return None

    for model_spec in model_specs:
        slot_started = perf_counter()
        slot = model_spec["slot"]
        stats = ensure_model_stats(slot)
        config = model_spec["config"]
        if slot == "llm2_compare" and not compare_model_configured:
            for knowledge_spec in knowledge_specs:
                for style in styles:
                    candidate_id = f"{slot}__{knowledge_spec['key']}__{style['id']}"
                    append_candidate(
                        candidate_id=candidate_id,
                        model_spec=model_spec,
                        knowledge_spec=knowledge_spec,
                        style=style,
                        status="not_configured",
                        reason="LLM2_COMPARE_API_URL / API_KEY / MODEL 未完整配置",
                    )
                    stats["placeholders"] += 1
            stage_parts_ms["llm2_compare_ms"] = round((perf_counter() - slot_started) * 1000, 2)
            continue

        if slot == "llm2_compare":
            _set_llm_compare_status(runtime_key, "running", "对比模型正在生成", config)

        def _run_kb_candidate(ks: dict, sty: dict) -> dict:
            cid = f"{slot}__{ks['key']}__{sty['id']}"
            task_started = perf_counter()
            skip_meta = knowledge_skip_reason(ks)
            if skip_meta:
                return {"candidate_id": cid, "ks": ks, "sty": sty,
                        "skip_meta": skip_meta, "elapsed_ms": 0.0}
            label_suffix = "KB1" if ks["key"] == "knowledge_v2" else "KB2"
            prompt_label = f"{slot.upper()}_{label_suffix}"
            req_spec = None
            try:
                req_spec = IntentEngine.build_sales_assist_request(
                    summary_json, ks["payload"], crm_context,
                    api_url=config["api_url"], api_key=config["api_key"],
                    model=config["model"], timeout_seconds=config["timeout_seconds"],
                    label=prompt_label, reply_style=sty,
                )
                bundle = IntentEngine.generate_sales_assist_bundle(
                    summary_json, ks["payload"], crm_context, prepared_request=req_spec,
                )
                elapsed = round((perf_counter() - task_started) * 1000, 2)
                return {
                    "candidate_id": cid, "ks": ks, "sty": sty, "skip_meta": None,
                    "status": "done" if bundle.get("content") else "failed_no_content",
                    "bundle": bundle, "req_spec": req_spec, "elapsed_ms": elapsed,
                }
            except Exception as exc:
                elapsed = round((perf_counter() - task_started) * 1000, 2)
                return {
                    "candidate_id": cid, "ks": ks, "sty": sty, "skip_meta": None,
                    "status": "error", "exc": exc, "req_spec": req_spec, "elapsed_ms": elapsed,
                }

        all_tasks = [(ks, sty) for ks in knowledge_specs for sty in styles]
        per_kb_elapsed: dict[str, float] = {}

        if len(all_tasks) > 1:
            fs = [REPLY_CHAIN_EXECUTOR.submit(_run_kb_candidate, ks, sty) for ks, sty in all_tasks]
            task_results = [f.result() for f in fs]
        else:
            task_results = [_run_kb_candidate(ks, sty) for ks, sty in all_tasks]

        for res in task_results:
            ks, sty, cid = res["ks"], res["sty"], res["candidate_id"]
            if res["skip_meta"]:
                status, reason = res["skip_meta"]
                append_candidate(
                    candidate_id=cid, model_spec=model_spec, knowledge_spec=ks,
                    style=sty, status=status, reason=reason,
                )
                stats["placeholders"] += 1
                continue
            per_kb_elapsed[f"{slot}_{ks['key']}_ms"] = res["elapsed_ms"]
            if res["status"] == "done":
                bundle = res["bundle"]
                content = bundle.get("content")
                stats["done"] += 1
                append_candidate(
                    candidate_id=cid, model_spec=model_spec, knowledge_spec=ks,
                    style=sty, status="done", content=content,
                    validation=bundle.get("validation") or IntentEngine.validate_sales_assist_output(
                        content, ks["payload"], crm_context=crm_context
                    ),
                    prompt_trace=bundle.get("prompt_trace"),
                )
            elif res["status"] == "failed_no_content":
                stats["failures"] += 1
                req_spec = res.get("req_spec")
                append_candidate(
                    candidate_id=cid, model_spec=model_spec, knowledge_spec=ks,
                    style=sty, status="failed_no_content",
                    prompt_trace=req_spec.get("prompt_trace") if isinstance(req_spec, dict) else None,
                    reason=f"{model_spec['label']}在{ks['label']}上返回空内容",
                )
            else:
                stats["failures"] += 1
                req_spec = res.get("req_spec")
                append_candidate(
                    candidate_id=cid, model_spec=model_spec, knowledge_spec=ks,
                    style=sty, status="error",
                    prompt_trace=req_spec.get("prompt_trace") if isinstance(req_spec, dict) else None,
                    reason=sanitize_text(str(res.get("exc", ""))),
                )

        part_key = "llm2_compare_ms" if slot == "llm2_compare" else "llm2_ms"
        stage_parts_ms[part_key] = round((perf_counter() - slot_started) * 1000, 2)
        stage_parts_ms.update(per_kb_elapsed)

    primary_stats = ensure_model_stats("llm2")
    compare_stats = ensure_model_stats("llm2_compare")
    if compare_model_configured and not single_model_single_style:
        if compare_stats["done"]:
            _set_llm_compare_status(runtime_key, "done", "对比模型已生成", compare_config)
        elif compare_stats["failures"]:
            _set_llm_compare_status(runtime_key, "error", "对比模型全部失败", compare_config)
        else:
            _set_llm_compare_status(runtime_key, "failed_no_content", "对比模型无可用输出", compare_config)
    else:
        _set_llm_compare_status(
            runtime_key,
            "not_started",
            "API 当前仅输出主模型首个风格结果" if single_model_single_style else "对比模型未启用",
            compare_config if compare_model_configured else None,
        )

    reply_scores = None
    if enable_scoring:
        score_started = perf_counter()
        reply_scores = IntentEngine.score_reply_candidates(
            summary_json=summary_json,
            knowledge_payload=knowledge,
            crm_context=crm_context,
            candidates=candidates,
            actual_sales_replies=actual_sales_replies,
        )
        stage_parts_ms["reply_scoring_ms"] = round((perf_counter() - score_started) * 1000, 2)
    primary_candidate = _select_candidate(candidates, "llm2", "knowledge_v2")
    compare_candidate = _select_candidate(candidates, "llm2", "knowledge_external_api")
    if primary_candidate and knowledge.get("log_id"):
        IntentEngine.update_knowledge_hit_log_outcome(knowledge.get("log_id"), final_response=primary_candidate.get("content"))
    return {
        "candidates": candidates,
        "reply_scores": reply_scores,
        "primary_candidate": primary_candidate,
        "compare_candidate": compare_candidate,
        "llm2_status": "done" if primary_stats["done"] else ("error" if primary_stats["failures"] else "failed_no_content"),
        "llm2_compare_status": (
            "not_started" if single_model_single_style else (
                "not_configured" if not compare_model_configured else (
                    "done" if compare_stats["done"] else ("error" if compare_stats["failures"] else "failed_no_content")
                )
            )
        ),
        "stage_timings_ms": {
            "total_ms": round((perf_counter() - stage_started) * 1000, 2),
            "parts_ms": stage_parts_ms,
        },
    }


def _generate_llm2_compare_extension(
    *,
    summary_json: dict,
    knowledge: dict,
    knowledge_compare: dict | None,
    crm_context: dict | None,
    actual_sales_replies: list[dict] | None,
    runtime_key: str,
    existing_candidates: list[dict] | None,
    single_model_single_style: bool | None = None,
    enable_scoring: bool | None = None,
) -> dict:
    stage_started = perf_counter()
    single_model_single_style = bool(single_model_single_style)
    enable_scoring = bool(enable_scoring)
    styles = IntentEngine.get_reply_style_options(enabled_only=True)
    if single_model_single_style and styles:
        styles = styles[:1]
    compare_config = _runtime_llm_values("LLM2_COMPARE")
    compare_model_configured = _runtime_llm_configured(compare_config)
    llm_runtime = _llm_runtime_config()
    model_spec = {
        "slot": "llm2_compare",
        "label": "对比模型",
        "display_name": llm_runtime["llm2_compare"]["display_name"] or llm_runtime["llm2_compare"]["model"] or "对比模型",
        "provider": llm_runtime["llm2_compare"]["provider"],
        "config": compare_config,
    }
    knowledge_specs = [
        {
            "key": "knowledge_v2",
            "label": "知识库1",
            "payload": knowledge if isinstance(knowledge, dict) else {},
            "status": str((knowledge or {}).get("status") or "not_started"),
        },
        {
            "key": "knowledge_external_api",
            "label": "知识库2",
            "payload": knowledge_compare if isinstance(knowledge_compare, dict) else {},
            "status": str((knowledge_compare or {}).get("status") or "not_started"),
        },
    ]
    compare_candidates: list[dict] = []
    stage_parts_ms: dict[str, float] = {}

    def append_candidate(
        *,
        candidate_id: str,
        knowledge_spec: dict,
        style: dict,
        status: str,
        content: str = "",
        validation: dict | None = None,
        prompt_trace: dict | None = None,
        reason: str = "",
    ) -> None:
        compare_candidates.append({
            "candidate_id": candidate_id,
            "model_slot": model_spec["slot"],
            "model_label": model_spec["label"],
            "model_display_name": model_spec["display_name"],
            "model_provider": model_spec["provider"],
            "knowledge_source": knowledge_spec["key"],
            "knowledge_source_label": knowledge_spec["label"],
            "style_id": style["id"],
            "style_title": style["title"],
            "style_content": style["content"],
            "reply_style": style,
            "status": status,
            "content": content,
            "validation": validation,
            "prompt_trace": prompt_trace,
            "reason": reason,
        })

    def knowledge_skip_reason(spec: dict) -> tuple[str, str] | None:
        return None

    if not compare_model_configured:
        _set_llm_compare_status(runtime_key, "not_configured", "LLM2_COMPARE_API_URL / API_KEY / MODEL 未完整配置", compare_config)
        merged_candidates = _merge_reply_candidates(existing_candidates, compare_candidates)
        return {
            "candidates": compare_candidates,
            "merged_candidates": merged_candidates,
            "reply_scores": None,
            "llm2_compare_status": "not_configured",
            "stage_timings_ms": {
                "total_ms": round((perf_counter() - stage_started) * 1000, 2),
                "parts_ms": stage_parts_ms,
            },
        }

    _set_llm_compare_status(runtime_key, "running", "对比模型正在生成", compare_config)
    slot_started = perf_counter()
    done_count = 0
    failure_count = 0
    for knowledge_spec in knowledge_specs:
        skip_meta = knowledge_skip_reason(knowledge_spec)
        for style in styles:
            candidate_id = f"llm2_compare__{knowledge_spec['key']}__{style['id']}"
            if skip_meta:
                status, reason = skip_meta
                append_candidate(
                    candidate_id=candidate_id,
                    knowledge_spec=knowledge_spec,
                    style=style,
                    status=status,
                    reason=reason,
                )
                continue
            label_suffix = "KB1" if knowledge_spec["key"] == "knowledge_v2" else "KB2"
            prompt_label = f"LLM2_COMPARE_{label_suffix}"
            request_spec = None
            try:
                request_spec = IntentEngine.build_sales_assist_request(
                    summary_json,
                    knowledge_spec["payload"],
                    crm_context,
                    api_url=compare_config["api_url"],
                    api_key=compare_config["api_key"],
                    model=compare_config["model"],
                    timeout_seconds=compare_config["timeout_seconds"],
                    label=prompt_label,
                    reply_style=style,
                )
                bundle = IntentEngine.generate_sales_assist_bundle(
                    summary_json,
                    knowledge_spec["payload"],
                    crm_context,
                    prepared_request=request_spec,
                )
                content = bundle.get("content")
                if content:
                    done_count += 1
                    append_candidate(
                        candidate_id=candidate_id,
                        knowledge_spec=knowledge_spec,
                        style=style,
                        status="done",
                        content=content,
                        validation=bundle.get("validation") or IntentEngine.validate_sales_assist_output(content, knowledge_spec["payload"], crm_context=crm_context),
                        prompt_trace=bundle.get("prompt_trace"),
                    )
                else:
                    failure_count += 1
                    append_candidate(
                        candidate_id=candidate_id,
                        knowledge_spec=knowledge_spec,
                        style=style,
                        status="failed_no_content",
                        prompt_trace=request_spec.get("prompt_trace") if isinstance(request_spec, dict) else None,
                        reason=f"对比模型在{knowledge_spec['label']}上返回空内容",
                    )
            except Exception as exc:
                failure_count += 1
                append_candidate(
                    candidate_id=candidate_id,
                    knowledge_spec=knowledge_spec,
                    style=style,
                    status="error",
                    prompt_trace=request_spec.get("prompt_trace") if isinstance(request_spec, dict) else None,
                    reason=sanitize_text(str(exc)),
                )
    stage_parts_ms["llm2_compare_ms"] = round((perf_counter() - slot_started) * 1000, 2)

    merged_candidates = _merge_reply_candidates(existing_candidates, compare_candidates)
    reply_scores = None
    if enable_scoring:
        score_started = perf_counter()
        reply_scores = IntentEngine.score_reply_candidates(
            summary_json=summary_json,
            knowledge_payload=knowledge,
            crm_context=crm_context,
            candidates=merged_candidates,
            actual_sales_replies=actual_sales_replies or [],
        )
        stage_parts_ms["reply_scoring_ms"] = round((perf_counter() - score_started) * 1000, 2)

    compare_status = "done" if done_count else ("error" if failure_count else "failed_no_content")
    compare_reason = {
        "done": "对比模型已生成",
        "error": "对比模型全部失败",
        "failed_no_content": "对比模型无可用输出",
    }.get(compare_status, "")
    _set_llm_compare_status(runtime_key, compare_status, compare_reason, compare_config)
    return {
        "candidates": compare_candidates,
        "merged_candidates": merged_candidates,
        "reply_scores": reply_scores,
        "llm2_compare_status": compare_status,
        "stage_timings_ms": {
            "total_ms": round((perf_counter() - stage_started) * 1000, 2),
            "parts_ms": stage_parts_ms,
        },
    }


def _complete_sidebar_assist_compare_async(
    *,
    summary_id: int | None,
    invocation_id: str | None,
    session_id: str,
    summary_json: dict,
    knowledge_v2: dict,
    knowledge_external_api: dict | None,
    crm_context: dict | None,
    single_model_single_style: bool,
    enable_scoring: bool,
) -> None:
    db = SessionLocal()
    try:
        summary = db.query(IntentSummary).filter(IntentSummary.id == summary_id).first() if summary_id else None
        api_invocation = (
            db.query(ApiAssistInvocation).filter(ApiAssistInvocation.invocation_id == invocation_id).first()
            if invocation_id else None
        )
        existing_candidates = None
        if summary and isinstance(summary.reply_style_results_v2, list):
            existing_candidates = summary.reply_style_results_v2
        elif api_invocation and isinstance((api_invocation.result_payload or {}).get("reply_style_results_v2"), list):
            existing_candidates = api_invocation.result_payload.get("reply_style_results_v2")

        extension = _generate_llm2_compare_extension(
            summary_json=summary_json,
            knowledge=knowledge_v2,
            knowledge_compare=knowledge_external_api,
            crm_context=crm_context,
            actual_sales_replies=[],
            runtime_key=session_id,
            existing_candidates=existing_candidates,
            single_model_single_style=single_model_single_style,
            enable_scoring=enable_scoring,
        )
        compare_runtime_status = LLM_COMPARE_RUNTIME_STATUS.get(session_id) or {}
        llm2_compare_status = extension.get("llm2_compare_status") or "failed_no_content"
        stage_timing = extension.get("stage_timings_ms") or {}
        parts_ms = dict(stage_timing.get("parts_ms") or {})
        total_ms = _round_timing_ms(stage_timing.get("total_ms")) or 0.0

        def apply_stage_updates(stage_status: dict) -> dict:
            updated = dict(stage_status or {})
            updated["llm2_compare"] = llm2_compare_status
            existing_llm2 = dict(((updated.get("node_timings_ms") or {}).get("llm2")) or {})
            existing_total = _round_timing_ms(existing_llm2.get("total_ms")) or 0.0
            merged_parts = {
                **dict(existing_llm2.get("parts_ms") or {}),
                **parts_ms,
            }
            combined_total = round(existing_total + total_ms, 2) if total_ms else existing_total
            _set_node_timing(
                updated,
                "llm2",
                total_ms=combined_total,
                parts_ms=merged_parts,
                status="done" if llm2_compare_status in {"done", "not_configured", "failed_no_content", "error"} else "running",
            )
            return updated

        if summary:
            summary.reply_style_results_v2 = extension.get("merged_candidates")
            if enable_scoring:
                summary.reply_scores_v2 = extension.get("reply_scores")
            summary.stage_status = apply_stage_updates(dict(summary.stage_status or {}))

        if api_invocation:
            result_payload = dict(api_invocation.result_payload or {})
            result_payload["reply_style_results_v2"] = _serialize_reply_candidates_for_output(extension.get("merged_candidates"))
            if enable_scoring:
                result_payload["reply_scores_v2"] = extension.get("reply_scores")
            result_payload["llm2_compare_runtime_status"] = compare_runtime_status
            result_payload["stage_status"] = apply_stage_updates(dict(result_payload.get("stage_status") or {}))
            result_payload["node_timings_ms"] = _extract_node_timings(result_payload["stage_status"])
            api_invocation.stage_status = result_payload["stage_status"]
            api_invocation.result_payload = _jsonable(result_payload)

        db.commit()
        logger.info(
            "API 侧边栏异步补齐对比模型完成 session_id=%s invocation_id=%s compare_status=%s",
            session_id,
            invocation_id,
            llm2_compare_status,
        )
    except Exception as exc:
        db.rollback()
        compare_config = _runtime_llm_values("LLM2_COMPARE")
        _set_llm_compare_status(session_id, "error", f"异步补齐失败: {sanitize_text(str(exc))}", compare_config)
        logger.error("API 侧边栏异步补齐对比模型失败 session_id=%s invocation_id=%s error=%s", session_id, invocation_id, exc)
    finally:
        db.close()

def _complete_api_reply_scoring_async(
    *,
    summary_id: int | None,
    invocation_id: str | None,
    session_id: str,
) -> None:
    db = SessionLocal()
    try:
        summary = db.query(IntentSummary).filter(IntentSummary.id == summary_id).first() if summary_id else None
        api_invocation = (
            db.query(ApiAssistInvocation).filter(ApiAssistInvocation.invocation_id == invocation_id).first()
            if invocation_id else None
        )
        payload = dict((api_invocation.result_payload or {}) if api_invocation else {})
        candidates = None
        if summary and isinstance(summary.reply_style_results_v2, list):
            candidates = summary.reply_style_results_v2
        elif isinstance(payload.get("reply_style_results_v2"), list):
            candidates = payload.get("reply_style_results_v2")
        if not candidates:
            return

        scoring_started = perf_counter()
        summary_json = _summary_json_from_source(summary) if summary else _api_summary_json_from_payload(payload)
        knowledge_payload = summary.knowledge_v2 if (summary and isinstance(summary.knowledge_v2, dict)) else (
            payload.get("knowledge_v2") if isinstance(payload.get("knowledge_v2"), dict) else {}
        )
        crm_context = summary.crm_info if (summary and isinstance(summary.crm_info, dict)) else (
            payload.get("crm_info") if isinstance(payload.get("crm_info"), dict) else None
        )
        actual_sales_replies = (
            api_invocation.actual_sales_replies
            if api_invocation and isinstance(api_invocation.actual_sales_replies, list)
            else (payload.get("actual_sales_replies") if isinstance(payload.get("actual_sales_replies"), list) else [])
        )
        stored_scores = summary.reply_scores_v2 if (summary and isinstance(summary.reply_scores_v2, dict)) else (
            payload.get("reply_scores_v2") if isinstance(payload.get("reply_scores_v2"), dict) else None
        )
        refreshed_scores = _load_or_refresh_reply_scores(
            stored_scores=stored_scores,
            summary_json=summary_json,
            knowledge_payload=knowledge_payload,
            crm_context=crm_context,
            candidates=candidates,
            actual_sales_replies=actual_sales_replies,
        )
        if not refreshed_scores:
            return

        scoring_ms = round((perf_counter() - scoring_started) * 1000, 2)
        if summary:
            summary.reply_scores_v2 = refreshed_scores
            updated_stage_status = dict(summary.stage_status or {})
            existing_llm2 = dict(((updated_stage_status.get("node_timings_ms") or {}).get("llm2")) or {})
            merged_parts = {
                **dict(existing_llm2.get("parts_ms") or {}),
                "reply_scoring_ms": scoring_ms,
            }
            existing_total = _round_timing_ms(existing_llm2.get("total_ms")) or 0.0
            if scoring_ms and merged_parts.get("reply_scoring_ms") != dict(existing_llm2.get("parts_ms") or {}).get("reply_scoring_ms"):
                _set_node_timing(
                    updated_stage_status,
                    "llm2",
                    total_ms=round(existing_total + scoring_ms, 2) if existing_total else scoring_ms,
                    parts_ms=merged_parts,
                    status=updated_stage_status.get("llm2") or "done",
                )
                summary.stage_status = updated_stage_status

        if api_invocation:
            payload = dict(api_invocation.result_payload or {})
            payload["reply_scores_v2"] = refreshed_scores
            llm_runtime = dict(payload.get("llm_runtime") or {})
            current_runtime = _llm_runtime_config()
            for key, value in current_runtime.items():
                if not isinstance(llm_runtime.get(key), dict):
                    llm_runtime[key] = value
            payload["llm_runtime"] = llm_runtime
            updated_stage_status = dict(payload.get("stage_status") or {})
            existing_llm2 = dict(((updated_stage_status.get("node_timings_ms") or {}).get("llm2")) or {})
            merged_parts = {
                **dict(existing_llm2.get("parts_ms") or {}),
                "reply_scoring_ms": scoring_ms,
            }
            existing_total = _round_timing_ms(existing_llm2.get("total_ms")) or 0.0
            _set_node_timing(
                updated_stage_status,
                "llm2",
                total_ms=round(existing_total + scoring_ms, 2) if existing_total else scoring_ms,
                parts_ms=merged_parts,
                status=updated_stage_status.get("llm2") or "done",
            )
            payload["stage_status"] = updated_stage_status
            payload["node_timings_ms"] = _extract_node_timings(updated_stage_status)
            api_invocation.stage_status = updated_stage_status
            api_invocation.result_payload = _jsonable(payload)

        db.commit()
        logger.info(
            "API 侧边栏异步补齐评分完成 session_id=%s invocation_id=%s",
            session_id,
            invocation_id,
        )
    except Exception as exc:
        db.rollback()
        logger.error("API 侧边栏异步补齐评分失败 session_id=%s invocation_id=%s error=%s", session_id, invocation_id, exc)
    finally:
        db.close()

def _parse_sales_userid_from_session(session_id: str) -> str | None:
    text = str(session_id or "").strip()
    if not text.startswith("single_"):
        return None
    body = text.replace("single_", "", 1)
    external_userid = extract_external_userid(session_id)
    if external_userid:
        body = re.sub(rf"_?{re.escape(external_userid)}$", "", body).strip("_")
    return body.split("_")[0] if body else None

def _build_thread_fact_payload_for_context(
    *,
    session_id: str,
    summary_json: dict | None,
    crm_context: dict | None,
    messages: list[MessageLog],
) -> dict | None:
    payload = build_thread_business_fact(
        session_id=session_id,
        summary=summary_json,
        crm_context=crm_context,
        messages=[{"content": item.content, "sender_type": item.sender_type} for item in messages],
        external_userid=extract_external_userid(session_id),
        sales_userid=_parse_sales_userid_from_session(session_id),
    )
    if not payload:
        return None
    return {
        "session_id": payload.get("session_id"),
        "scenario_label": payload.get("scenario_label"),
        "intent_label": payload.get("intent_label"),
        "language_style": payload.get("language_style"),
        "business_state": payload.get("business_state"),
        "stage_signals": payload.get("stage_signals") or {},
        "merged_facts": payload.get("merged_facts") or {},
        "reply_guard_reason": payload.get("reply_guard_reason"),
    }

def _reply_snapshot_label(snapshot: ReplyChainSnapshot) -> str:
    ts = snapshot.anchor_message_time.strftime("%m-%d %H:%M") if snapshot.anchor_message_time else "历史节点"
    text = sanitize_text(snapshot.anchor_message_text or "")
    short = text[:18] + "..." if len(text) > 18 else text
    return f"客户 {ts} {short or '未命名节点'}"

def _reply_snapshot_meta(snapshot: ReplyChainSnapshot) -> dict:
    return {
        "snapshot_id": str(snapshot.snapshot_id),
        "session_id": snapshot.session_id,
        "anchor_message_id": snapshot.anchor_message_id,
        "anchor_sender_type": snapshot.anchor_sender_type,
        "anchor_message_time": snapshot.anchor_message_time,
        "anchor_message_text": snapshot.anchor_message_text,
        "label": _reply_snapshot_label(snapshot),
        "kind": "anchored_customer_message",
        "has_v1": bool(snapshot.topic or snapshot.core_demand or snapshot.status),
        "has_v2": bool(snapshot.sales_advice_v2),
        "updated_at": snapshot.updated_at,
    }

def _api_invocation_result_payload(item: ApiAssistInvocation) -> dict:
    payload = dict(item.result_payload or {})
    payload["analysis_mode"] = "api_sidebar_assist_snapshot"
    payload["analysis_version"] = _api_invocation_meta(item)
    payload["actual_sales_replies"] = item.actual_sales_replies or []
    payload["api_quality_similarity"] = item.quality_similarity or None
    payload["api_quality_score"] = float(item.quality_score) if item.quality_score is not None else None
    payload["api_quality_status"] = item.quality_status
    payload["api_triggered_at"] = item.triggered_at
    payload["user_id"] = item.session_id
    payload["has_api_invocation"] = True
    payload.setdefault("stage_status", item.stage_status or {})
    llm_runtime = dict(payload.get("llm_runtime") or {})
    current_runtime = _llm_runtime_config()
    for key, value in current_runtime.items():
        if not isinstance(llm_runtime.get(key), dict):
            llm_runtime[key] = value
        else:
            merged_runtime = dict(value)
            merged_runtime.update(llm_runtime.get(key) or {})
            llm_runtime[key] = merged_runtime
    payload["llm_runtime"] = llm_runtime
    return _sanitize_api_sidebar_result_payload(payload)


def _api_invocation_analytics_payload(item: ApiAssistInvocation) -> dict:
    payload = dict(item.result_payload or {})
    payload["analysis_mode"] = "api_sidebar_assist_snapshot"
    payload["analysis_version"] = _api_invocation_meta(item)
    payload["actual_sales_replies"] = item.actual_sales_replies or []
    payload["api_quality_similarity"] = item.quality_similarity or None
    payload["api_quality_score"] = float(item.quality_score) if item.quality_score is not None else None
    payload["api_quality_status"] = item.quality_status
    payload["api_triggered_at"] = item.triggered_at
    payload["user_id"] = item.session_id
    payload["has_api_invocation"] = True
    payload.setdefault("stage_status", item.stage_status or {})
    llm_runtime = dict(payload.get("llm_runtime") or {})
    current_runtime = _llm_runtime_config()
    for key, value in current_runtime.items():
        if not isinstance(llm_runtime.get(key), dict):
            llm_runtime[key] = value
        else:
            merged_runtime = dict(value)
            merged_runtime.update(llm_runtime.get(key) or {})
            llm_runtime[key] = merged_runtime
    payload["llm_runtime"] = llm_runtime
    payload["node_timings_ms"] = _extract_node_timings(payload.get("stage_status"))
    return _jsonable(payload)


def _api_summary_json_from_payload(payload: dict | None) -> dict:
    source = payload or {}
    return {
        "topic": source.get("topic"),
        "core_demand": source.get("core_demand"),
        "key_facts": source.get("key_facts") or {},
        "todo_items": source.get("todo_items") or [],
        "risks": source.get("risks"),
        "to_be_confirmed": source.get("to_be_confirmed"),
        "status": source.get("status"),
    }

def _refresh_api_invocation_result_payload(item: ApiAssistInvocation) -> bool:
    payload = dict(item.result_payload or {})
    changed = False
    actual_sales_replies = item.actual_sales_replies if isinstance(item.actual_sales_replies, list) else []
    if payload.get("actual_sales_replies") != actual_sales_replies:
        payload["actual_sales_replies"] = _jsonable(actual_sales_replies)
        changed = True

    llm_runtime = dict(payload.get("llm_runtime") or {})
    current_runtime = _llm_runtime_config()
    for key, value in current_runtime.items():
        if not isinstance(llm_runtime.get(key), dict):
            llm_runtime[key] = value
            changed = True
        else:
            merged_runtime = dict(value)
            merged_runtime.update(llm_runtime.get(key) or {})
            if merged_runtime != llm_runtime.get(key):
                llm_runtime[key] = merged_runtime
                changed = True
    if llm_runtime:
        payload["llm_runtime"] = llm_runtime

    candidates = payload.get("reply_style_results_v2") if isinstance(payload.get("reply_style_results_v2"), list) else []
    stored_scores = payload.get("reply_scores_v2") if isinstance(payload.get("reply_scores_v2"), dict) else None
    if candidates:
        refreshed_scores = _load_or_refresh_reply_scores(
            stored_scores=stored_scores,
            summary_json=_api_summary_json_from_payload(payload),
            knowledge_payload=payload.get("knowledge_v2") if isinstance(payload.get("knowledge_v2"), dict) else {},
            crm_context=payload.get("crm_info") if isinstance(payload.get("crm_info"), dict) else None,
            candidates=candidates,
            actual_sales_replies=actual_sales_replies,
        )
        if refreshed_scores and refreshed_scores != stored_scores:
            payload["reply_scores_v2"] = refreshed_scores
            changed = True

    if changed:
        item.result_payload = _jsonable(payload)
    return changed

def _sanitize_api_sidebar_result_payload(payload: dict | None) -> dict:
    result = dict(_jsonable(payload) or {})
    stage_status = dict(result.get("stage_status") or {})
    stage_status["llm2_compare"] = "skipped_api_isolated"
    # API 路径不运行对比模型，清除缓存摘要带入的 llm2_compare_ms，防止污染分析面板
    node_timings = dict(stage_status.get("node_timings_ms") or {})
    llm2_node = dict(node_timings.get("llm2") or {})
    if isinstance(llm2_node.get("parts_ms"), dict):
        llm2_node["parts_ms"] = {k: v for k, v in llm2_node["parts_ms"].items() if k != "llm2_compare_ms"}
    node_timings["llm2"] = llm2_node
    stage_status["node_timings_ms"] = node_timings
    result["stage_status"] = stage_status
    result["node_timings_ms"] = _extract_node_timings(stage_status)
    result["llm2_compare_configured"] = False
    result["llm2_compare_runtime_status"] = {
        "status": "skipped_api_isolated",
        "reason": "API 调用链路已禁用 LLM-2 对比模型",
    }
    llm_runtime = dict(result.get("llm_runtime") or {})
    llm2_compare_runtime = dict(llm_runtime.get("llm2_compare") or {})
    llm2_compare_runtime.update({
        "model": "",
        "provider": "",
        "url": "",
        "display_name": "API 调用已禁用对比模型",
    })
    llm_runtime["llm2_compare"] = llm2_compare_runtime
    result["llm_runtime"] = llm_runtime
    candidates = result.get("reply_style_results_v2")
    if isinstance(candidates, list):
        result["reply_style_results_v2"] = [
            item for item in candidates
            if str((item or {}).get("model_slot") or "").strip() != "llm2_compare"
        ]
    return result


def _sales_advice_output_sections(raw_content: str | None) -> dict:
    sections = IntentEngine.split_sales_assist_output(raw_content)
    return {
        "reply_reference": sections.get("reply_reference") or None,
        "followup_rationale": sections.get("followup_rationale") or None,
    }


def _apply_sales_advice_output_fields(target: dict, raw_content: str | None, *, compare: bool = False) -> None:
    sections = _sales_advice_output_sections(raw_content)
    if compare:
        target["reply_reference_compare"] = sections.get("reply_reference")
        target["followup_rationale_compare"] = sections.get("followup_rationale")
        return
    target["reply_reference"] = sections.get("reply_reference")
    target["followup_rationale"] = sections.get("followup_rationale")


def _stage_is_in_progress(stage: str | None) -> bool:
    return str(stage or "").strip() in {"queued", "running"}


def _utc_now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _round_timing_ms(value: Any) -> float | None:
    try:
        if value is None:
            return None
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def _clean_timing_parts(parts_ms: dict[str, Any] | None) -> dict[str, float]:
    cleaned: dict[str, float] = {}
    for key, raw_value in (parts_ms or {}).items():
        rounded = _round_timing_ms(raw_value)
        if rounded is not None:
            cleaned[str(key)] = rounded
    return cleaned


def _set_node_timing(
    stage_status: dict,
    node_key: str,
    *,
    total_ms: Any = None,
    parts_ms: dict[str, Any] | None = None,
    status: str | None = None,
) -> dict:
    node_timings = dict(stage_status.get("node_timings_ms") or {})
    payload = dict(node_timings.get(node_key) or {})
    rounded_total = _round_timing_ms(total_ms)
    if rounded_total is not None:
        payload["total_ms"] = rounded_total
    if parts_ms is not None:
        merged_parts = dict(payload.get("parts_ms") or {})
        merged_parts.update(_clean_timing_parts(parts_ms))
        payload["parts_ms"] = merged_parts
    if status is not None:
        payload["status"] = str(status)
    node_timings[node_key] = payload
    stage_status["node_timings_ms"] = node_timings
    return payload


def _extract_node_timings(stage_status: dict | None) -> dict:
    if not isinstance(stage_status, dict):
        return {}
    return _jsonable(stage_status.get("node_timings_ms") or {})


def _serialize_reply_candidates_for_output(candidates: list[dict] | None) -> list[dict]:
    serialized: list[dict] = []
    for item in candidates or []:
        row = dict(item or {})
        row.update(_sales_advice_output_sections(row.get("content")))
        row.pop("content", None)
        serialized.append(row)
    return serialized


def _attach_external_sales_kb_result(result: dict, query_text: str | None) -> None:
    external_result = _search_sales_kb_api(query_text, top_k=5)
    result["knowledge_external_api"] = external_result
    stage_status = result.setdefault("stage_status", {})
    stage_status["knowledge_external_api"] = external_result.get("status") or "unknown"


def _read_only_current_tail_result(
    *,
    db: Session,
    user_id: str,
    summary: IntentSummary | None,
    recent_logs: list[MessageLog],
    fast_track_signals: list[str],
) -> dict:
    compare_runtime_status = LLM_COMPARE_RUNTIME_STATUS.get(user_id) or {}
    stored_stage_status = dict((summary.stage_status or {}) if summary else {})
    result = {
        "analysis_mode": "current_tail",
        "analysis_version": {
            "snapshot_id": "current_tail",
            "kind": "current_tail",
            "label": "当前末尾",
            "anchor_message_id": None,
        },
        "fast_track": fast_track_signals,
        "has_v1": False,
        "has_v2": False,
        "has_v2_compare": False,
        "llm_runtime": _llm_runtime_config(),
        "llm2_compare_configured": True,
        "llm2_compare_runtime_status": compare_runtime_status,
        "latest_dialog_count": len(recent_logs),
        "input_messages": [_message_ui_dict(item) for item in reversed(recent_logs)],
        "stage_status": {
            "conversation_input": "done" if recent_logs else "empty",
            "fast_track": "done",
            "llm1": "not_started",
            "crm_profile": "not_started",
            "knowledge_v2": "not_started",
            "knowledge_external_api": "not_started",
            "llm2": "not_started",
            "llm2_compare": "not_started",
            **stored_stage_status,
        },
    }
    result["node_timings_ms"] = _extract_node_timings(result["stage_status"])

    crm_context = (summary.crm_info if summary and summary.crm_info else None) or IntentEngine.get_crm_context(extract_external_userid(user_id))
    if crm_context:
        result["crm_info"] = crm_context
        result["crm_status"] = (summary.crm_status if summary and summary.crm_status else crm_context.get("crm_profile_status"))
        result["stage_status"]["crm_profile"] = result["crm_status"] or result["stage_status"].get("crm_profile") or "unknown"

    if not summary:
        return result

    if not _stage_is_in_progress(result["stage_status"].get("llm1")):
        if not _stage_is_in_progress(result["stage_status"].get("llm1")):
            result["stage_status"]["llm1"] = "done"
    result.update({
        "has_v1": True,
        "topic": summary.topic,
        "core_demand": summary.core_demand,
        "key_facts": summary.key_facts,
        "todo_items": summary.todo_items,
        "risks": summary.risks,
        "to_be_confirmed": summary.to_be_confirmed,
        "status": summary.status,
        "at": summary.summarized_at,
    })

    if summary.thread_business_fact:
        result["thread_business_fact"] = summary.thread_business_fact
    else:
        thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == user_id).first()
        if thread_fact:
            result["thread_business_fact"] = _thread_fact_to_dict(thread_fact)

    if summary.knowledge_v2:
        result["knowledge_v2"] = summary.knowledge_v2
        result["knowledge_status"] = summary.knowledge_status
        result["knowledge_confidence_score"] = float(summary.knowledge_confidence_score) if summary.knowledge_confidence_score is not None else None
        result["knowledge_manual_review_required"] = bool(summary.knowledge_manual_review_required)
        result["knowledge_evidence_context"] = (summary.knowledge_v2 or {}).get("evidence_context")
        result["evidence_refs"] = (summary.knowledge_v2 or {}).get("evidence_refs") or []
        result["stage_status"]["knowledge_v2"] = summary.knowledge_status or result["stage_status"].get("knowledge_v2") or "done"
    elif summary.knowledge_status:
        result["knowledge_status"] = summary.knowledge_status
        result["stage_status"]["knowledge_v2"] = summary.knowledge_status

    if summary.knowledge_external_api:
        result["knowledge_external_api"] = summary.knowledge_external_api
        result["stage_status"]["knowledge_external_api"] = (summary.knowledge_external_api or {}).get("status") or result["stage_status"].get("knowledge_external_api") or "unknown"

    if summary.sales_advice_v2:
        if not _stage_is_in_progress(result["stage_status"].get("llm2")):
            result["stage_status"]["llm2"] = "done"
        result["has_v2"] = True
        _apply_sales_advice_output_fields(result, summary.sales_advice_v2)
    if summary.sales_advice_compare_v2:
        if not _stage_is_in_progress(result["stage_status"].get("llm2_compare")):
            result["stage_status"]["llm2_compare"] = "done"
        result["has_v2_compare"] = True
        _apply_sales_advice_output_fields(result, summary.sales_advice_compare_v2, compare=True)
    if summary.sales_advice_compare_prompt_trace_v2:
        result["sales_advice_compare_prompt_trace_v2"] = summary.sales_advice_compare_prompt_trace_v2
        if not summary.sales_advice_compare_v2:
            trace_result = str(summary.sales_advice_compare_prompt_trace_v2.get("result") or "")
            trace_stage = {
                "running": "running",
                "failed_no_content": "failed_no_content",
                "error": "error",
            }.get(trace_result)
            if trace_stage:
                result["stage_status"]["llm2_compare"] = trace_stage
    elif compare_runtime_status.get("status"):
        result["stage_status"]["llm2_compare"] = compare_runtime_status.get("status")
    if summary.reply_style_results_v2:
        result["reply_style_results_v2"] = _serialize_reply_candidates_for_output(summary.reply_style_results_v2)
    if summary.reply_scores_v2:
        result["reply_scores_v2"] = summary.reply_scores_v2
    if summary.assist_validation:
        result["assist_validation"] = summary.assist_validation
    if summary.assist_compare_validation:
        result["assist_compare_validation"] = summary.assist_compare_validation

    return result

class TriggerReq(BaseModel):
    step: int
    snapshot_id: str | None = None
    trigger_source: str | None = None

class KnowledgeRefreshReq(BaseModel):
    snapshot_id: str | None = None
    trigger_source: str | None = None


class ReplyChainSnapshotCreateReq(BaseModel):
    anchor_message_id: int


def _upsert_reply_chain_snapshot(
    db: Session,
    *,
    session_id: str,
    anchor_message: MessageLog,
    all_messages: list[MessageLog],
) -> ReplyChainSnapshot:
    visible_messages = _visible_messages_until_anchor(all_messages, anchor_message.id)
    item = db.query(ReplyChainSnapshot).filter(
        ReplyChainSnapshot.session_id == session_id,
        ReplyChainSnapshot.anchor_message_id == anchor_message.id,
    ).first()
    if not item:
        item = ReplyChainSnapshot(
            session_id=session_id,
            anchor_message_id=anchor_message.id,
        )
        db.add(item)
        db.flush()
    item.anchor_sender_type = anchor_message.sender_type or "customer"
    item.anchor_message_time = anchor_message.timestamp
    item.anchor_message_text = anchor_message.content
    item.visible_message_ids = [msg.id for msg in visible_messages]
    item.latest_dialog_count = len(visible_messages)
    item.actual_sales_replies = _collect_actual_sales_replies(all_messages, anchor_message.id)
    return item

def _snapshot_result_payload(
    *,
    snapshot: ReplyChainSnapshot,
    visible_messages: list[MessageLog],
) -> dict:
    compare_runtime_status = LLM_COMPARE_RUNTIME_STATUS.get(_reply_chain_runtime_key(snapshot.session_id, str(snapshot.snapshot_id))) or {}
    stage_status = dict(snapshot.stage_status or {})
    result = {
        "analysis_mode": "snapshot",
        "analysis_version": _reply_snapshot_meta(snapshot),
        "fast_track": snapshot.fast_track or [],
        "has_v1": False,
        "has_v2": False,
        "has_v2_compare": False,
        "llm_runtime": _llm_runtime_config(),
        "llm2_compare_configured": True,
        "llm2_compare_runtime_status": compare_runtime_status,
        "latest_dialog_count": snapshot.latest_dialog_count or len(visible_messages),
        "input_messages": [_message_ui_dict(item) for item in visible_messages],
        "actual_sales_replies": snapshot.actual_sales_replies or [],
        "stage_status": {
            "conversation_input": "done" if visible_messages else "empty",
            "fast_track": "done",
            "llm1": "not_started",
            "crm_profile": snapshot.crm_status or "not_started",
            "knowledge_v2": "not_started",
            "knowledge_external_api": "not_started",
            "llm2": "not_started",
            "llm2_compare": "not_started",
            **stage_status,
        },
    }
    result["node_timings_ms"] = _extract_node_timings(result["stage_status"])
    if snapshot.crm_info:
        result["crm_info"] = snapshot.crm_info
        result["crm_status"] = snapshot.crm_status
    if snapshot.topic or snapshot.core_demand or snapshot.status:
        result["has_v1"] = True
        result.update({
            "topic": snapshot.topic,
            "core_demand": snapshot.core_demand,
            "key_facts": snapshot.key_facts,
            "todo_items": snapshot.todo_items,
            "risks": snapshot.risks,
            "to_be_confirmed": snapshot.to_be_confirmed,
            "status": snapshot.status,
            "at": snapshot.summarized_at,
        })
        result["stage_status"]["llm1"] = "done"
    if snapshot.thread_business_fact:
        result["thread_business_fact"] = snapshot.thread_business_fact
    if snapshot.knowledge_v2:
        result["knowledge_v2"] = snapshot.knowledge_v2
        result["knowledge_status"] = snapshot.knowledge_status
        result["knowledge_confidence_score"] = float(snapshot.knowledge_confidence_score) if snapshot.knowledge_confidence_score is not None else None
        result["knowledge_manual_review_required"] = bool(snapshot.knowledge_manual_review_required)
        result["evidence_refs"] = (snapshot.knowledge_v2 or {}).get("evidence_refs") or []
        result["knowledge_evidence_context"] = (snapshot.knowledge_v2 or {}).get("evidence_context")
        result["stage_status"]["knowledge_v2"] = snapshot.knowledge_status or "done"
    if snapshot.knowledge_external_api:
        result["knowledge_external_api"] = snapshot.knowledge_external_api
        result["stage_status"]["knowledge_external_api"] = (snapshot.knowledge_external_api or {}).get("status") or "unknown"
    if snapshot.sales_advice_v2:
        result["has_v2"] = True
        _apply_sales_advice_output_fields(result, snapshot.sales_advice_v2)
        if not _stage_is_in_progress(result["stage_status"].get("llm2")):
            result["stage_status"]["llm2"] = "done"
    if snapshot.sales_advice_compare_v2:
        result["has_v2_compare"] = True
        _apply_sales_advice_output_fields(result, snapshot.sales_advice_compare_v2, compare=True)
        if not _stage_is_in_progress(result["stage_status"].get("llm2_compare")):
            result["stage_status"]["llm2_compare"] = "done"
    if snapshot.sales_advice_compare_prompt_trace_v2:
        result["sales_advice_compare_prompt_trace_v2"] = snapshot.sales_advice_compare_prompt_trace_v2
        if not snapshot.sales_advice_compare_v2:
            trace_result = str((snapshot.sales_advice_compare_prompt_trace_v2 or {}).get("result") or "")
            trace_stage = {
                "running": "running",
                "failed_no_content": "failed_no_content",
                "error": "error",
            }.get(trace_result)
            if trace_stage:
                result["stage_status"]["llm2_compare"] = trace_stage
    if snapshot.assist_validation:
        result["assist_validation"] = snapshot.assist_validation
    if snapshot.assist_compare_validation:
        result["assist_compare_validation"] = snapshot.assist_compare_validation
    if snapshot.reply_style_results_v2:
        result["reply_style_results_v2"] = _serialize_reply_candidates_for_output(snapshot.reply_style_results_v2)
    if snapshot.reply_scores_v2:
        result["reply_scores_v2"] = snapshot.reply_scores_v2
    return result

def refresh_snapshot_knowledge_task(session_id: str, snapshot_id: str, channel: str, analytics_record_id: str | None = None) -> None:
    db = SessionLocal()
    final_status = "unknown"
    final_error = None
    try:
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if not snapshot:
            raise RuntimeError(f"未找到历史节点快照: {snapshot_id}")
        started = perf_counter()
        all_messages = _load_session_messages(db, session_id)
        visible_messages = _visible_messages_until_anchor(all_messages, snapshot.anchor_message_id)
        conversation_input_ms = round((perf_counter() - started) * 1000, 2)
        recent_visible_logs = visible_messages[-15:]
        started = perf_counter()
        fast_track_signals = collect_fast_track_signals(recent_visible_logs)
        fast_track_ms = round((perf_counter() - started) * 1000, 2)
        snapshot.fast_track = fast_track_signals
        snapshot.visible_message_ids = [msg.id for msg in visible_messages]
        snapshot.latest_dialog_count = len(visible_messages)
        snapshot.actual_sales_replies = _collect_actual_sales_replies(all_messages, snapshot.anchor_message_id)
        stage_status = dict(snapshot.stage_status or {})
        stage_status["conversation_input"] = "done" if visible_messages else "empty"
        stage_status["fast_track"] = "done"
        _set_node_timing(stage_status, "conversation_input", total_ms=conversation_input_ms, status="done" if visible_messages else "empty")
        _set_node_timing(stage_status, "fast_track", total_ms=fast_track_ms, status="done")
        stage_status["last_requested_action"] = channel
        stage_status[channel] = "running"
        snapshot.stage_status = stage_status
        db.commit()

        if not snapshot.topic and not snapshot.core_demand and not snapshot.status:
            stage_status[channel] = "skipped_no_summary"
            if channel == "knowledge_v2":
                snapshot.knowledge_v2 = {
                    "status": "skipped_no_summary",
                    "query_text": "",
                    "hits": [],
                    "evidence_refs": [],
                    "evidence_context": _empty_knowledge_evidence_context(),
                }
                snapshot.knowledge_status = "skipped_no_summary"
            else:
                snapshot.knowledge_external_api = {
                    "status": "skipped_no_summary",
                    "query_text": "",
                    "hits": [],
                    "replyable_hits": [],
                    "human_only_hits": [],
                    "evidence_refs": [],
                }
            snapshot.stage_status = stage_status
            db.commit()
            final_status = "skipped_no_summary"
            return

        summary_json = {**_summary_json_from_source(snapshot), "fast_track_signals": fast_track_signals}
        crm_started = perf_counter()
        crm_context = IntentEngine.get_crm_context(extract_external_userid(session_id))
        crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
        snapshot.crm_info = crm_context
        snapshot.crm_status = crm_context.get("crm_profile_status")
        stage_status["crm_profile"] = snapshot.crm_status or "unknown"
        _set_node_timing(stage_status, "crm_profile", total_ms=crm_profile_ms, status=snapshot.crm_status or "unknown")
        thread_fact_payload = _build_thread_fact_payload_for_context(
            session_id=session_id,
            summary_json=summary_json,
            crm_context=crm_context,
            messages=visible_messages,
        )
        snapshot.thread_business_fact = thread_fact_payload
        query_features = {}
        retrieval_query = ""
        if snapshot.core_demand:
            query_features = IntentEngine.infer_query_features(summary_json, crm_context, thread_fact_payload)
            retrieval_query = IntentEngine.build_retrieval_query(summary_json, thread_fact_payload)

        knowledge_stage_started = perf_counter()
        knowledge_parts_ms: dict[str, float] = {}
        if channel == "knowledge_v2":
            if snapshot.core_demand:
                try:
                    started = perf_counter()
                    knowledge = IntentEngine.retrieve_knowledge_v2(
                        retrieval_query,
                        query_features=query_features,
                        top_k=5,
                        request_id=f"snapshot_k1_{snapshot_id[:12]}",
                        session_id=session_id,
                    )
                    knowledge_parts_ms["knowledge_v2_ms"] = round((perf_counter() - started) * 1000, 2)
                    knowledge["thread_business_fact"] = thread_fact_payload
                except Exception as exc:
                    knowledge = _knowledge_v2_error_payload(
                        query_text=retrieval_query,
                        query_features=query_features,
                        error=str(exc),
                    )
                snapshot.knowledge_v2 = knowledge
                snapshot.knowledge_log_id = knowledge.get("log_id")
                snapshot.knowledge_status = knowledge.get("status")
                snapshot.knowledge_confidence_score = _decimal_score(knowledge.get("confidence_score"))
                snapshot.knowledge_manual_review_required = bool(knowledge.get("manual_review_required"))
                stage_status["knowledge_v2"] = snapshot.knowledge_status or "unknown"
            else:
                snapshot.knowledge_v2 = {
                    "status": "skipped_no_core_demand",
                    "query_text": "",
                    "query_features": {},
                    "hits": [],
                    "evidence_refs": [],
                    "evidence_context": _empty_knowledge_evidence_context(),
                }
                snapshot.knowledge_status = "skipped_no_core_demand"
                snapshot.knowledge_log_id = None
                snapshot.knowledge_confidence_score = None
                snapshot.knowledge_manual_review_required = False
                stage_status["knowledge_v2"] = "skipped_no_core_demand"
        else:
            started = perf_counter()
            external_knowledge = _search_sales_kb_api(retrieval_query if retrieval_query else None, top_k=5)
            knowledge_parts_ms["knowledge_external_api_ms"] = round((perf_counter() - started) * 1000, 2)
            snapshot.knowledge_external_api = external_knowledge
            stage_status["knowledge_external_api"] = external_knowledge.get("status") or "unknown"

        _set_node_timing(
            stage_status,
            "knowledge",
            total_ms=round((perf_counter() - knowledge_stage_started) * 1000, 2),
            parts_ms=knowledge_parts_ms,
            status=stage_status.get(channel) or "unknown",
        )
        snapshot.stage_status = stage_status
        db.commit()
        knowledge_payload = snapshot.knowledge_v2 if channel == "knowledge_v2" else snapshot.knowledge_external_api
        final_status = stage_status.get(channel) or "unknown"
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": session_id,
            "snapshot_id": snapshot_id,
            "action": channel,
            "result": stage_status.get(channel) or "unknown",
            "hit_count": len((knowledge_payload or {}).get("hits") or []),
        })
    except Exception as exc:
        logger.error("历史节点知识库刷新失败 [%s]: %s", channel, exc)
        final_status = "error"
        final_error = str(exc)
        db.rollback()
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if snapshot:
            stage_status = dict(snapshot.stage_status or {})
            stage_status[channel] = "error"
            if channel == "knowledge_v2":
                snapshot.knowledge_v2 = _knowledge_v2_error_payload(
                    query_text="",
                    query_features={},
                    error=str(exc),
                )
                snapshot.knowledge_status = "error"
                snapshot.knowledge_log_id = None
                snapshot.knowledge_confidence_score = None
                snapshot.knowledge_manual_review_required = False
            else:
                snapshot.knowledge_external_api = {
                    "status": "error",
                    "query_text": "",
                    "hits": [],
                    "replyable_hits": [],
                    "human_only_hits": [],
                    "manual_review_required": False,
                    "confidence_score": None,
                    "filters_used": {},
                    "evidence_refs": [],
                    "source": "sales_kb_api",
                    "error": sanitize_text(str(exc)),
                }
            snapshot.stage_status = stage_status
            db.commit()
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": session_id,
            "snapshot_id": snapshot_id,
            "action": channel,
            "result": "error",
            "reason": str(exc),
        })
    finally:
        db.close()
        _sync_wecom_trigger_record_result(
            record_id=analytics_record_id,
            session_id=session_id,
            snapshot_id=snapshot_id,
            request_status=final_status,
            error_message=final_error,
        )

def refresh_session_knowledge_task(user_id: str, channel: str, analytics_record_id: str | None = None) -> None:
    db = SessionLocal()
    final_status = "unknown"
    final_error = None
    try:
        summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
        if not summary:
            raise RuntimeError(f"未找到会话摘要: {user_id}")
        started = perf_counter()
        recent_logs = db.query(MessageLog).filter(
            MessageLog.user_id == user_id,
            MessageLog.is_mock.is_(False),
        ).order_by(MessageLog.id.desc()).limit(15).all()
        conversation_input_ms = round((perf_counter() - started) * 1000, 2)
        started = perf_counter()
        fast_track_signals = collect_fast_track_signals(recent_logs)
        fast_track_ms = round((perf_counter() - started) * 1000, 2)
        summary_json = {
            "topic": summary.topic,
            "core_demand": summary.core_demand,
            "key_facts": summary.key_facts,
            "todo_items": summary.todo_items,
            "risks": summary.risks,
            "to_be_confirmed": summary.to_be_confirmed,
            "status": summary.status,
            "fast_track_signals": fast_track_signals,
        }
        stage_status = {
            "conversation_input": "done" if recent_logs else "empty",
            "fast_track": "done",
            **dict(summary.stage_status or {}),
        }
        _set_node_timing(stage_status, "conversation_input", total_ms=conversation_input_ms, status="done" if recent_logs else "empty")
        _set_node_timing(stage_status, "fast_track", total_ms=fast_track_ms, status="done")
        stage_status["last_requested_action"] = channel
        stage_status[channel] = "running"
        summary.stage_status = stage_status
        db.commit()

        if not summary.topic and not summary.core_demand and not summary.status:
            stage_status[channel] = "skipped_no_summary"
            if channel == "knowledge_v2":
                summary.knowledge_v2 = {
                    "status": "skipped_no_summary",
                    "query_text": "",
                    "hits": [],
                    "evidence_refs": [],
                    "evidence_context": _empty_knowledge_evidence_context(),
                }
                summary.knowledge_status = "skipped_no_summary"
            else:
                summary.knowledge_external_api = {
                    "status": "skipped_no_summary",
                    "query_text": "",
                    "hits": [],
                    "replyable_hits": [],
                    "human_only_hits": [],
                    "evidence_refs": [],
                }
            summary.stage_status = stage_status
            db.commit()
            final_status = "skipped_no_summary"
            return

        crm_started = perf_counter()
        crm_context = IntentEngine.get_crm_context(extract_external_userid(user_id))
        crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
        summary.crm_info = crm_context
        summary.crm_status = crm_context.get("crm_profile_status")
        stage_status["crm_profile"] = summary.crm_status or "unknown"
        _set_node_timing(stage_status, "crm_profile", total_ms=crm_profile_ms, status=summary.crm_status or "unknown")
        current_thread_fact = _upsert_thread_business_fact(
            db,
            session_id=user_id,
            summary_json=summary_json,
            crm_context=crm_context,
            messages=[{"content": item.content, "sender_type": item.sender_type} for item in recent_logs],
            external_userid=extract_external_userid(user_id),
        )
        summary.thread_business_fact = _thread_fact_to_dict(current_thread_fact)
        thread_fact_payload = _thread_fact_prompt_dict(current_thread_fact)
        query_features = {}
        retrieval_query = ""
        if summary.core_demand:
            query_features = IntentEngine.infer_query_features(summary_json, crm_context, thread_fact_payload)
            retrieval_query = IntentEngine.build_retrieval_query(summary_json, thread_fact_payload)

        knowledge_stage_started = perf_counter()
        knowledge_parts_ms: dict[str, float] = {}
        if channel == "knowledge_v2":
            if summary.core_demand:
                try:
                    started = perf_counter()
                    knowledge = IntentEngine.retrieve_knowledge_v2(
                        retrieval_query,
                        query_features=query_features,
                        top_k=5,
                        request_id=f"manual_k1_{uuid.uuid4().hex[:12]}",
                        session_id=user_id,
                    )
                    knowledge_parts_ms["knowledge_v2_ms"] = round((perf_counter() - started) * 1000, 2)
                    knowledge["thread_business_fact"] = thread_fact_payload
                except Exception as exc:
                    knowledge = _knowledge_v2_error_payload(
                        query_text=retrieval_query,
                        query_features=query_features,
                        error=str(exc),
                    )
                summary.knowledge_v2 = knowledge
                summary.knowledge_log_id = knowledge.get("log_id")
                summary.knowledge_status = knowledge.get("status")
                summary.knowledge_confidence_score = _decimal_score(knowledge.get("confidence_score"))
                summary.knowledge_manual_review_required = bool(knowledge.get("manual_review_required"))
                stage_status["knowledge_v2"] = summary.knowledge_status or "unknown"
            else:
                summary.knowledge_v2 = {
                    "status": "skipped_no_core_demand",
                    "query_text": "",
                    "query_features": {},
                    "hits": [],
                    "evidence_refs": [],
                    "evidence_context": _empty_knowledge_evidence_context(),
                }
                summary.knowledge_status = "skipped_no_core_demand"
                summary.knowledge_log_id = None
                summary.knowledge_confidence_score = None
                summary.knowledge_manual_review_required = False
                stage_status["knowledge_v2"] = "skipped_no_core_demand"
        else:
            started = perf_counter()
            external_knowledge = _search_sales_kb_api(retrieval_query if retrieval_query else None, top_k=5)
            knowledge_parts_ms["knowledge_external_api_ms"] = round((perf_counter() - started) * 1000, 2)
            summary.knowledge_external_api = external_knowledge
            stage_status["knowledge_external_api"] = external_knowledge.get("status") or "unknown"

        _set_node_timing(
            stage_status,
            "knowledge",
            total_ms=round((perf_counter() - knowledge_stage_started) * 1000, 2),
            parts_ms=knowledge_parts_ms,
            status=stage_status.get(channel) or "unknown",
        )
        summary.stage_status = stage_status
        db.commit()
        knowledge_payload = summary.knowledge_v2 if channel == "knowledge_v2" else summary.knowledge_external_api
        final_status = stage_status.get(channel) or "unknown"
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": user_id,
            "snapshot_id": None,
            "action": channel,
            "result": stage_status.get(channel) or "unknown",
            "hit_count": len((knowledge_payload or {}).get("hits") or []),
        })
    except Exception as exc:
        logger.error("会话知识库刷新失败 [%s]: %s", channel, exc)
        final_status = "error"
        final_error = str(exc)
        db.rollback()
        summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
        if summary:
            stage_status = dict(summary.stage_status or {})
            stage_status[channel] = "error"
            if channel == "knowledge_v2":
                summary.knowledge_v2 = _knowledge_v2_error_payload(
                    query_text="",
                    query_features={},
                    error=str(exc),
                )
                summary.knowledge_status = "error"
                summary.knowledge_log_id = None
                summary.knowledge_confidence_score = None
                summary.knowledge_manual_review_required = False
            else:
                summary.knowledge_external_api = {
                    "status": "error",
                    "query_text": "",
                    "hits": [],
                    "replyable_hits": [],
                    "human_only_hits": [],
                    "manual_review_required": False,
                    "confidence_score": None,
                    "filters_used": {},
                    "evidence_refs": [],
                    "source": "sales_kb_api",
                    "error": sanitize_text(str(exc)),
                }
            summary.stage_status = stage_status
            db.commit()
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": user_id,
            "action": channel,
            "result": "error",
            "reason": str(exc),
        })
    finally:
        db.close()
        _sync_wecom_trigger_record_result(
            record_id=analytics_record_id,
            session_id=user_id,
            snapshot_id=None,
            request_status=final_status,
            error_message=final_error,
        )

def reanalyze_snapshot_task(session_id: str, snapshot_id: str, step: int = 1, analytics_record_id: str | None = None):
    db = SessionLocal()
    runtime_key = _reply_chain_runtime_key(session_id, snapshot_id)
    final_status = "unknown"
    final_error = None
    try:
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if not snapshot:
            raise RuntimeError(f"未找到历史节点快照: {snapshot_id}")

        started = perf_counter()
        all_messages = _load_session_messages(db, session_id)
        visible_messages = _visible_messages_until_anchor(all_messages, snapshot.anchor_message_id)
        conversation_input_ms = round((perf_counter() - started) * 1000, 2)
        recent_visible_logs = visible_messages[-15:]
        started = perf_counter()
        fast_track_signals = collect_fast_track_signals(recent_visible_logs)
        fast_track_ms = round((perf_counter() - started) * 1000, 2)
        snapshot.fast_track = fast_track_signals
        snapshot.visible_message_ids = [msg.id for msg in visible_messages]
        snapshot.latest_dialog_count = len(visible_messages)
        snapshot.actual_sales_replies = _collect_actual_sales_replies(all_messages, snapshot.anchor_message_id)
        stage_status = dict(snapshot.stage_status or {})
        stage_status["conversation_input"] = "done" if visible_messages else "empty"
        stage_status["fast_track"] = "done"
        _set_node_timing(stage_status, "conversation_input", total_ms=conversation_input_ms, status="done" if visible_messages else "empty")
        _set_node_timing(stage_status, "fast_track", total_ms=fast_track_ms, status="done")
        if step == 1:
            log_reply_chain_event("STEP_START", {
                "source": "frontend",
                "business_object": session_id,
                "snapshot_id": snapshot_id,
                "step": step,
                "stage": "llm1_snapshot",
            })
            llm1_started_at = _utc_now_iso()
            stage_status["llm1"] = "running"
            stage_status["llm1_started_at"] = llm1_started_at
            stage_status["llm1_display_reset"] = True
            snapshot.stage_status = stage_status
            db.commit()
            if not visible_messages:
                stage_status["llm1"] = "skipped_no_messages"
                stage_status["llm1_completed_at"] = _utc_now_iso()
                stage_status["llm1_display_reset"] = False
                _set_node_timing(stage_status, "llm1", total_ms=0, status="skipped_no_messages")
                snapshot.stage_status = stage_status
                db.commit()
                final_status = "skipped_no_messages"
                return
            context = [{"content": item.content, "sender_type": item.sender_type} for item in visible_messages]
            llm1_started = perf_counter()
            summary = IntentEngine._run_llm1(f"{session_id}#{snapshot.anchor_message_id}", context)
            _set_node_timing(
                stage_status,
                "llm1",
                total_ms=round((perf_counter() - llm1_started) * 1000, 2),
                status="done" if summary else "error",
            )
            if not summary:
                raise RuntimeError("LLM-1 未返回有效结构化摘要")
            _apply_summary_fields(snapshot, summary)
            snapshot.summarized_at = datetime.utcnow()
            snapshot.llm1_compare_summary = None
            snapshot.llm1_compare_prompt_trace = None
            snapshot.knowledge_log_id = None
            snapshot.knowledge_v2 = None
            snapshot.knowledge_external_api = None
            snapshot.knowledge_status = None
            snapshot.knowledge_confidence_score = None
            snapshot.knowledge_manual_review_required = False
            snapshot.reply_style_results_v2 = None
            snapshot.reply_scores_v2 = None
            snapshot.sales_advice_v2 = None
            snapshot.sales_advice_compare_v2 = None
            snapshot.sales_advice_compare_prompt_trace_v2 = None
            snapshot.assist_validation = None
            snapshot.assist_compare_validation = None
            crm_started = perf_counter()
            crm_context = IntentEngine.get_crm_context(extract_external_userid(session_id))
            crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
            snapshot.crm_info = crm_context
            snapshot.crm_status = crm_context.get("crm_profile_status")
            _set_node_timing(stage_status, "crm_profile", total_ms=crm_profile_ms, status=snapshot.crm_status or "unknown")
            snapshot.thread_business_fact = _build_thread_fact_payload_for_context(
                session_id=session_id,
                summary_json={**_summary_json_from_source(snapshot), "fast_track_signals": fast_track_signals},
                crm_context=crm_context,
                messages=visible_messages,
            )
            stage_status["llm1"] = "done"
            stage_status["llm1_completed_at"] = _utc_now_iso()
            stage_status["llm1_display_reset"] = False
            stage_status["crm_profile"] = snapshot.crm_status or "unknown"
            stage_status["knowledge_v2"] = "not_started"
            stage_status["knowledge_external_api"] = "not_started"
            stage_status["llm2"] = "not_started"
            stage_status["llm2_compare"] = "not_started"
            snapshot.stage_status = stage_status
            db.commit()
            log_reply_chain_event("STEP_RESULT", {
                "source": "frontend",
                "business_object": session_id,
                "snapshot_id": snapshot_id,
                "step": step,
                "stage": "llm1_snapshot",
                "result": "success",
                "message_count": len(visible_messages),
            })
            final_status = "done"
            return

        if step != 2:
            raise RuntimeError(f"不支持的阶段: {step}")

        log_reply_chain_event("STEP_START", {
            "source": "frontend",
            "business_object": session_id,
            "snapshot_id": snapshot_id,
            "step": step,
            "stage": "llm2_snapshot",
        })
        if not snapshot.topic and not snapshot.core_demand and not snapshot.status:
            stage_status["llm2"] = "skipped_no_summary"
            snapshot.stage_status = stage_status
            db.commit()
            final_status = "skipped_no_summary"
            return
        stage_status["llm2"] = "running"
        stage_status["llm2_compare"] = "running"
        snapshot.stage_status = stage_status
        db.commit()

        summary_json = {**_summary_json_from_source(snapshot), "fast_track_signals": fast_track_signals}
        crm_started = perf_counter()
        crm_context = IntentEngine.get_crm_context(extract_external_userid(session_id))
        crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
        snapshot.crm_info = crm_context
        snapshot.crm_status = crm_context.get("crm_profile_status")
        stage_status["crm_profile"] = snapshot.crm_status or "unknown"
        _set_node_timing(stage_status, "crm_profile", total_ms=crm_profile_ms, status=snapshot.crm_status or "unknown")
        thread_fact_payload = _build_thread_fact_payload_for_context(
            session_id=session_id,
            summary_json=summary_json,
            crm_context=crm_context,
            messages=visible_messages,
        )
        snapshot.thread_business_fact = thread_fact_payload

        if not _knowledge_payload_ready(snapshot.knowledge_v2):
            stage_status["llm2"] = "skipped_missing_knowledge"
            stage_status["llm2_compare"] = "skipped_missing_knowledge"
            snapshot.stage_status = stage_status
            db.commit()
            log_reply_chain_event("STEP_RESULT", {
                "source": "frontend",
                "business_object": session_id,
                "snapshot_id": snapshot_id,
                "step": step,
                "stage": "llm2_snapshot",
                "result": "skipped_missing_knowledge",
                "reason": "知识库1证据未准备好，请先执行 05 环节重新检索。",
            })
            final_status = "skipped_missing_knowledge"
            return

        knowledge = dict(snapshot.knowledge_v2 or {})
        knowledge["thread_business_fact"] = thread_fact_payload
        external_knowledge = (
            dict(snapshot.knowledge_external_api or {})
            if _knowledge_payload_ready(snapshot.knowledge_external_api)
            else _knowledge_generation_placeholder(
                "knowledge_external_api",
                reason="知识库2证据未准备好，请先执行 05 环节重新检索。",
            )
        )

        generation_bundle = _generate_reply_style_candidates(
            summary_json=summary_json,
            knowledge=knowledge,
            knowledge_compare=external_knowledge,
            crm_context=crm_context,
            actual_sales_replies=snapshot.actual_sales_replies or [],
            runtime_key=runtime_key,
            single_model_single_style=False,
            enable_scoring=True,
        )
        snapshot.reply_style_results_v2 = generation_bundle.get("candidates")
        snapshot.reply_scores_v2 = generation_bundle.get("reply_scores")
        primary_candidate = generation_bundle.get("primary_candidate")
        compare_candidate = generation_bundle.get("compare_candidate")
        first_compare_entry = _first_model_candidate_entry(
            generation_bundle.get("candidates"),
            "llm2",
            "knowledge_external_api",
        )
        snapshot.sales_advice_v2 = primary_candidate.get("content") if primary_candidate else None
        snapshot.assist_validation = primary_candidate.get("validation") if primary_candidate else None
        snapshot.sales_advice_compare_v2 = compare_candidate.get("content") if compare_candidate else None
        snapshot.assist_compare_validation = compare_candidate.get("validation") if compare_candidate else None
        snapshot.sales_advice_compare_prompt_trace_v2 = (compare_candidate or first_compare_entry or {}).get("prompt_trace")
        llm2_stage_timing = generation_bundle.get("stage_timings_ms") or {}
        _set_node_timing(
            stage_status,
            "llm2",
            total_ms=llm2_stage_timing.get("total_ms"),
            parts_ms=llm2_stage_timing.get("parts_ms"),
            status=generation_bundle.get("llm2_status") or "failed_no_content",
        )
        stage_status["llm2"] = generation_bundle.get("llm2_status") or "failed_no_content"
        stage_status["llm2_compare"] = generation_bundle.get("llm2_compare_status") or "not_started"
        compare_result = (
            (compare_candidate or first_compare_entry or {}).get("prompt_trace", {}) or {}
        ).get("result") or generation_bundle.get("llm2_compare_status")

        if not primary_candidate:
            snapshot.stage_status = stage_status
            db.commit()
            log_reply_chain_event("STEP_RESULT", {
                "source": "frontend",
                "business_object": session_id,
                "snapshot_id": snapshot_id,
                "step": step,
                "stage": "llm2_snapshot",
                "result": stage_status["llm2"],
            })
            final_status = stage_status["llm2"]
            return

        snapshot.stage_status = stage_status
        db.commit()
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": session_id,
            "snapshot_id": snapshot_id,
            "step": step,
            "stage": "llm2_snapshot",
            "result": "success",
            "compare_result": compare_result,
            "knowledge_status": snapshot.knowledge_status,
            "knowledge_log_id": snapshot.knowledge_log_id,
        })
        final_status = stage_status.get("llm2") or "done"
    except Exception as e:
        logger.error("历史节点 AI 推理过程崩坏: %s", e)
        final_status = "error"
        final_error = str(e)
        db.rollback()
        snapshot = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.snapshot_id == snapshot_id,
            ReplyChainSnapshot.session_id == session_id,
        ).first()
        if snapshot:
            stage_status = dict(snapshot.stage_status or {})
            stage_status["conversation_input"] = stage_status.get("conversation_input") or "done"
            stage_status["fast_track"] = stage_status.get("fast_track") or "done"
            if step == 1:
                stage_status["llm1"] = "error"
                stage_status["llm1_display_reset"] = False
            elif step == 2:
                stage_status["llm2"] = "error"
                compare_runtime = LLM_COMPARE_RUNTIME_STATUS.get(runtime_key) or {}
                stage_status["llm2_compare"] = compare_runtime.get("status") or stage_status.get("llm2_compare") or "error"
            snapshot.stage_status = stage_status
            db.commit()
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": session_id,
            "snapshot_id": snapshot_id,
            "step": step,
            "result": "error",
            "reason": str(e),
        })
    finally:
        db.close()
        _sync_wecom_trigger_record_result(
            record_id=analytics_record_id,
            session_id=session_id,
            snapshot_id=snapshot_id,
            request_status=final_status,
            error_message=final_error,
        )

def reanalyze_session_task(user_id: str, step: int = 1, analytics_record_id: str | None = None):
    """由界面主动测算触发：抓取特定会话并投递给 LLM 大语言模型进行意图重算"""
    db = SessionLocal()
    final_status = "unknown"
    final_error = None
    try:
        if step == 1:
            log_reply_chain_event("STEP_START", {
                "source": "frontend",
                "business_object": user_id,
                "step": step,
                "stage": "llm1",
            })
            logger.info("正在执行阶段 1: 调用 LLM-1 提取会话 %s 结构化画像", user_id)
            existing_summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
            llm1_requested_at = None
            llm1_started_at = _utc_now_iso()
            if existing_summary:
                existing_stage_status = {
                    "conversation_input": "done",
                    "fast_track": "done",
                    **dict(existing_summary.stage_status or {}),
                }
                llm1_requested_at = existing_stage_status.get("llm1_requested_at")
                existing_stage_status["llm1"] = "running"
                existing_stage_status["llm1_started_at"] = llm1_started_at
                existing_stage_status["llm1_display_reset"] = True
                existing_summary.stage_status = existing_stage_status
                db.commit()
            started = perf_counter()
            context_logs = db.query(MessageLog).filter(
                MessageLog.user_id == user_id,
                MessageLog.is_mock.is_(False),
            ).order_by(MessageLog.id.desc()).limit(15).all()
            conversation_input_ms = round((perf_counter() - started) * 1000, 2)
            context = [{"content": l.content, "sender_type": l.sender_type} for l in reversed(context_logs)]
            started = perf_counter()
            fast_track_signals = collect_fast_track_signals(context_logs)
            fast_track_ms = round((perf_counter() - started) * 1000, 2)
            if context:
                llm1_started = perf_counter()
                summary_payload = IntentEngine._run_llm1(user_id, context)
                llm1_ms = round((perf_counter() - llm1_started) * 1000, 2)
                if not summary_payload:
                    raise RuntimeError("LLM-1 未返回有效结构化摘要")
                summary_record = IntentSummary(user_id=user_id)
                _apply_summary_fields(summary_record, summary_payload)
                summary_record.llm1_compare_summary = None
                summary_record.llm1_compare_prompt_trace = None
                crm_started = perf_counter()
                crm_context = IntentEngine.get_crm_context(extract_external_userid(user_id))
                crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
                summary_record.crm_info = crm_context
                summary_record.crm_status = crm_context.get("crm_profile_status")
                thread_fact = _upsert_thread_business_fact(
                    db,
                    session_id=user_id,
                    summary_json={
                        "topic": summary_record.topic,
                        "core_demand": summary_record.core_demand,
                        "key_facts": summary_record.key_facts,
                        "todo_items": summary_record.todo_items,
                        "risks": summary_record.risks,
                        "to_be_confirmed": summary_record.to_be_confirmed,
                        "status": summary_record.status,
                        "fast_track_signals": fast_track_signals,
                    },
                    crm_context=crm_context,
                    messages=context,
                    external_userid=extract_external_userid(user_id),
                )
                summary_record.thread_business_fact = _thread_fact_to_dict(thread_fact)
                llm1_completed_at = _utc_now_iso()
                stage_status_payload = {
                    "conversation_input": "done" if context else "empty",
                    "fast_track": "done",
                    "llm1": "done",
                    "llm1_requested_at": llm1_requested_at or llm1_started_at,
                    "llm1_started_at": llm1_started_at,
                    "llm1_completed_at": llm1_completed_at,
                    "llm1_display_reset": False,
                    "crm_profile": summary_record.crm_status or "unknown",
                    "knowledge_v2": "not_started",
                    "knowledge_external_api": "not_started",
                    "llm2": "not_started",
                    "llm2_compare": "not_started",
                }
                _set_node_timing(stage_status_payload, "conversation_input", total_ms=conversation_input_ms, status="done" if context else "empty")
                _set_node_timing(stage_status_payload, "fast_track", total_ms=fast_track_ms, status="done")
                _set_node_timing(stage_status_payload, "llm1", total_ms=llm1_ms, status="done")
                _set_node_timing(stage_status_payload, "crm_profile", total_ms=crm_profile_ms, status=summary_record.crm_status or "unknown")
                summary_record.stage_status = stage_status_payload
                db.add(summary_record)
                db.commit()
                logger.info("会话 %s 的 AI 结构化特征(V1) 已生成完毕", user_id)
                log_reply_chain_event("STEP_RESULT", {
                    "source": "frontend",
                    "business_object": user_id,
                    "step": step,
                    "stage": "llm1",
                    "result": "success",
                    "message_count": len(context),
                })
                final_status = "done"
            else:
                summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
                if summary:
                    stage_status_payload = {
                        "conversation_input": "empty",
                        "fast_track": "done",
                        **dict(summary.stage_status or {}),
                    }
                    stage_status_payload["llm1_completed_at"] = _utc_now_iso()
                    stage_status_payload["llm1_display_reset"] = False
                    _set_node_timing(stage_status_payload, "conversation_input", total_ms=conversation_input_ms, status="empty")
                    _set_node_timing(stage_status_payload, "fast_track", total_ms=fast_track_ms, status="done")
                    _set_node_timing(stage_status_payload, "llm1", total_ms=0, status="skipped_no_messages")
                    summary.stage_status = stage_status_payload
                    db.commit()
                log_reply_chain_event("STEP_RESULT", {
                    "source": "frontend",
                    "business_object": user_id,
                    "step": step,
                    "stage": "llm1",
                    "result": "skipped_no_messages",
                })
                final_status = "skipped_no_messages"
        
        elif step == 2:
            log_reply_chain_event("STEP_START", {
                "source": "frontend",
                "business_object": user_id,
                "step": step,
                "stage": "llm2",
            })
            logger.info("正在执行阶段 2: 调遣 DeepSeek 针对 %s 发放销售实战指令", user_id)
            summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
            if summary:
                # 重新构建 json 体传给 llm2
                started = perf_counter()
                recent_logs = db.query(MessageLog).filter(
                    MessageLog.user_id == user_id,
                    MessageLog.is_mock.is_(False),
                ).order_by(MessageLog.id.desc()).limit(15).all()
                conversation_input_ms = round((perf_counter() - started) * 1000, 2)
                started = perf_counter()
                fast_track_signals = collect_fast_track_signals(recent_logs)
                fast_track_ms = round((perf_counter() - started) * 1000, 2)
                summary_json = {
                    "topic": summary.topic, "core_demand": summary.core_demand, "key_facts": summary.key_facts,
                    "todo_items": summary.todo_items, "risks": summary.risks, "to_be_confirmed": summary.to_be_confirmed,
                    "status": summary.status,
                    "fast_track_signals": fast_track_signals
                }
                stage_status = {
                    "conversation_input": "done" if recent_logs else "empty",
                    "fast_track": "done",
                    **dict(summary.stage_status or {}),
                }
                _set_node_timing(stage_status, "conversation_input", total_ms=conversation_input_ms, status="done" if recent_logs else "empty")
                _set_node_timing(stage_status, "fast_track", total_ms=fast_track_ms, status="done")
                stage_status["llm2"] = "running"
                stage_status["llm2_compare"] = "running"
                summary.stage_status = stage_status
                db.commit()
                # CRM Context
                crm_started = perf_counter()
                crm_context = IntentEngine.get_crm_context(extract_external_userid(user_id))
                crm_profile_ms = round((perf_counter() - crm_started) * 1000, 2)
                summary.crm_info = crm_context
                summary.crm_status = crm_context.get("crm_profile_status")
                stage_status["crm_profile"] = summary.crm_status or "unknown"
                _set_node_timing(stage_status, "crm_profile", total_ms=crm_profile_ms, status=summary.crm_status or "unknown")
                current_thread_fact = _upsert_thread_business_fact(
                    db,
                    session_id=user_id,
                    summary_json=summary_json,
                    crm_context=crm_context,
                    messages=[{"content": item.content, "sender_type": item.sender_type} for item in recent_logs],
                    external_userid=extract_external_userid(user_id),
                )
                summary.thread_business_fact = _thread_fact_to_dict(current_thread_fact)
                db.commit()
                thread_fact_payload = _thread_fact_prompt_dict(current_thread_fact)
                if not _knowledge_payload_ready(summary.knowledge_v2):
                    stage_status["llm2"] = "skipped_missing_knowledge"
                    stage_status["llm2_compare"] = "skipped_missing_knowledge"
                    summary.stage_status = stage_status
                    db.commit()
                    log_reply_chain_event("STEP_RESULT", {
                        "source": "frontend",
                        "business_object": user_id,
                        "step": step,
                        "stage": "llm2",
                        "result": "skipped_missing_knowledge",
                        "reason": "知识库1证据未准备好，请先执行 05 环节重新检索。",
                    })
                    final_status = "skipped_missing_knowledge"
                    return
                knowledge = dict(summary.knowledge_v2 or {})
                knowledge["thread_business_fact"] = thread_fact_payload
                external_knowledge = (
                    dict(summary.knowledge_external_api or {})
                    if _knowledge_payload_ready(summary.knowledge_external_api)
                    else _knowledge_generation_placeholder(
                        "knowledge_external_api",
                        reason="知识库2证据未准备好，请先执行 05 环节重新检索。",
                    )
                )
                generation_bundle = _generate_reply_style_candidates(
                    summary_json=summary_json,
                    knowledge=knowledge,
                    knowledge_compare=external_knowledge,
                    crm_context=crm_context,
                    actual_sales_replies=[],
                    runtime_key=user_id,
                    single_model_single_style=False,
                    enable_scoring=True,
                )
                primary_candidate = generation_bundle.get("primary_candidate")
                compare_candidate = generation_bundle.get("compare_candidate")
                first_compare_entry = _first_model_candidate_entry(
                    generation_bundle.get("candidates"),
                    "llm2",
                    "knowledge_external_api",
                )
                summary.reply_style_results_v2 = generation_bundle.get("candidates")
                summary.reply_scores_v2 = generation_bundle.get("reply_scores")
                summary.sales_advice_v2 = primary_candidate.get("content") if primary_candidate else None
                summary.sales_advice_compare_v2 = compare_candidate.get("content") if compare_candidate else None
                summary.sales_advice_compare_prompt_trace_v2 = (compare_candidate or first_compare_entry or {}).get("prompt_trace")
                summary.assist_validation = primary_candidate.get("validation") if primary_candidate else None
                summary.assist_compare_validation = compare_candidate.get("validation") if compare_candidate else None
                stage_status["llm1"] = "done"
                llm2_stage_timing = generation_bundle.get("stage_timings_ms") or {}
                _set_node_timing(
                    stage_status,
                    "llm2",
                    total_ms=llm2_stage_timing.get("total_ms"),
                    parts_ms=llm2_stage_timing.get("parts_ms"),
                    status=generation_bundle.get("llm2_status") or "failed_no_content",
                )
                stage_status["llm2"] = generation_bundle.get("llm2_status") or "failed_no_content"
                stage_status["llm2_compare"] = generation_bundle.get("llm2_compare_status") or "not_started"
                summary.stage_status = stage_status
                db.commit()
                if primary_candidate:
                    title = f"💡 [AI销售辅助更新] - {summary.topic}"
                    QYWXUtils.send_text_card(user_id, title, primary_candidate.get("content"))
                    logger.info("实战建议(V2)指令下发完成")
                    log_reply_chain_event("STEP_RESULT", {
                        "source": "frontend",
                        "business_object": user_id,
                        "step": step,
                        "stage": "llm2",
                        "result": "success",
                        "compare_result": ((compare_candidate or first_compare_entry or {}).get("prompt_trace", {}) or {}).get("result") or generation_bundle.get("llm2_compare_status"),
                        "knowledge_status": knowledge.get("status"),
                        "knowledge_log_id": knowledge.get("log_id"),
                    })
                    final_status = stage_status.get("llm2") or "done"
                else:
                    log_reply_chain_event("STEP_RESULT", {
                        "source": "frontend",
                        "business_object": user_id,
                        "step": step,
                        "stage": "llm2",
                        "result": generation_bundle.get("llm2_status") or "failed_no_content",
                    })
                    final_status = generation_bundle.get("llm2_status") or "failed_no_content"
            else:
                log_reply_chain_event("STEP_RESULT", {
                    "source": "frontend",
                    "business_object": user_id,
                    "step": step,
                    "stage": "llm2",
                    "result": "skipped_no_summary",
                })
                final_status = "skipped_no_summary"
            
    except Exception as e:
        logger.error(f"AI 推理过程崩坏: {e}")
        final_status = "error"
        final_error = str(e)
        if step == 1:
            db.rollback()
            latest_summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
            if latest_summary:
                stage_status = dict(latest_summary.stage_status or {})
                stage_status["llm1"] = "error"
                stage_status["llm1_display_reset"] = False
                latest_summary.stage_status = stage_status
                db.commit()
        log_reply_chain_event("STEP_RESULT", {
            "source": "frontend",
            "business_object": user_id,
            "step": step,
            "result": "error",
            "reason": str(e),
        })
    finally:
        db.close()
        _sync_wecom_trigger_record_result(
            record_id=analytics_record_id,
            session_id=user_id,
            snapshot_id=None,
            request_status=final_status,
            error_message=final_error,
        )

@app.get("/api/sessions/{user_id}/analysis_versions")
async def list_reply_chain_versions(user_id: str):
    from time import perf_counter

    total_started = perf_counter()
    db = SessionLocal()
    try:
        snapshots = db.query(ReplyChainSnapshot).filter(
            ReplyChainSnapshot.session_id == user_id
        ).order_by(ReplyChainSnapshot.anchor_message_time.desc(), ReplyChainSnapshot.updated_at.desc()).all()
        api_invocations = db.query(ApiAssistInvocation).filter(
            ApiAssistInvocation.session_id == user_id
        ).order_by(ApiAssistInvocation.anchor_message_time.desc(), ApiAssistInvocation.triggered_at.desc()).all()
        deduped_api_invocations: list[ApiAssistInvocation] = []
        seen_api_invocation_keys: set[str] = set()
        for item in api_invocations:
            identity_key = _api_invocation_identity_key(item)
            if identity_key in seen_api_invocation_keys:
                continue
            seen_api_invocation_keys.add(identity_key)
            deduped_api_invocations.append(item)
        mixed_versions = [
            *[_reply_snapshot_meta(item) for item in snapshots],
            *[_api_invocation_meta(item) for item in deduped_api_invocations],
        ]
        mixed_versions.sort(
            key=lambda item: (
                item.get("anchor_message_time") or datetime.min,
                item.get("triggered_at") or datetime.min,
            ),
            reverse=True,
        )
        result = {
            "session_id": user_id,
            "versions": [
                {
                    "snapshot_id": "current_tail",
                    "kind": "current_tail",
                    "label": "当前末尾",
                    "anchor_message_id": None,
                    "anchor_message_time": None,
                    "anchor_message_text": None,
                    "has_v1": None,
                    "has_v2": None,
                },
                *mixed_versions,
            ],
        }
        log_session_view_event("ANALYSIS_VERSIONS", {
            "user_id": user_id,
            "version_count": len(result["versions"]),
            "api_version_count_raw": len(api_invocations),
            "api_version_count_deduped": len(deduped_api_invocations),
            "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
        })
        return result
    finally:
        db.close()

@app.post("/api/sessions/{user_id}/analysis_versions")
async def create_reply_chain_version(user_id: str, payload: ReplyChainSnapshotCreateReq):
    db = SessionLocal()
    try:
        all_messages = _load_session_messages(db, user_id)
        anchor_message = next((item for item in all_messages if item.id == payload.anchor_message_id), None)
        if not anchor_message:
            raise HTTPException(status_code=404, detail="未找到指定消息节点")
        if anchor_message.sender_type != "customer":
            raise HTTPException(status_code=400, detail="只能基于客户气泡创建历史节点")
        snapshot = _upsert_reply_chain_snapshot(
            db,
            session_id=user_id,
            anchor_message=anchor_message,
            all_messages=all_messages,
        )
        db.commit()
        db.refresh(snapshot)
        return {
            "status": "success",
            "snapshot": _reply_snapshot_meta(snapshot),
        }
    finally:
        db.close()

@app.get("/api/sessions/{user_id}/analysis")
async def get_latest_analysis(user_id: str, snapshot_id: str | None = Query(default=None)):
    from time import perf_counter

    total_started = perf_counter()
    db = SessionLocal()
    try:
        if snapshot_id and snapshot_id != "current_tail":
            snapshot = db.query(ReplyChainSnapshot).filter(
                ReplyChainSnapshot.snapshot_id == snapshot_id,
                ReplyChainSnapshot.session_id == user_id,
            ).first()
            if snapshot:
                all_messages = _load_session_messages(db, user_id)
                visible_messages = _visible_messages_until_anchor(all_messages, snapshot.anchor_message_id)
                result = _snapshot_result_payload(snapshot=snapshot, visible_messages=visible_messages)
                if not result.get("fast_track"):
                    result["fast_track"] = collect_fast_track_signals(visible_messages[-15:])
                if not result.get("actual_sales_replies"):
                    result["actual_sales_replies"] = _collect_actual_sales_replies(all_messages, snapshot.anchor_message_id)
                if (snapshot.topic or snapshot.core_demand or snapshot.status) and not result.get("crm_info"):
                    crm_context = IntentEngine.get_crm_context(extract_external_userid(user_id))
                    result["crm_info"] = crm_context
                    result["crm_status"] = crm_context.get("crm_profile_status")
                    result.setdefault("stage_status", {})["crm_profile"] = result["crm_status"] or "unknown"
                log_session_view_event("ANALYSIS", {
                    "user_id": user_id,
                    "snapshot_id": snapshot_id,
                    "analysis_mode": result.get("analysis_mode"),
                    "has_v1": result.get("has_v1"),
                    "has_v2": result.get("has_v2"),
                    "has_v2_compare": result.get("has_v2_compare"),
                    "stage_status": result.get("stage_status"),
                    "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
                })
                return result

            api_invocation = db.query(ApiAssistInvocation).filter(
                ApiAssistInvocation.invocation_id == snapshot_id,
                ApiAssistInvocation.session_id == user_id,
            ).first()
            if not api_invocation:
                raise HTTPException(status_code=404, detail="历史节点链路不存在")
            all_messages = _load_session_messages(db, user_id)
            _refresh_api_invocation_quality(db, api_invocation, all_messages=all_messages)
            db.commit()
            result = _api_invocation_result_payload(api_invocation)
            log_session_view_event("ANALYSIS", {
                "user_id": user_id,
                "snapshot_id": snapshot_id,
                "analysis_mode": result.get("analysis_mode"),
                "has_v1": result.get("has_v1"),
                "has_v2": result.get("has_v2"),
                "has_v2_compare": result.get("has_v2_compare"),
                "stage_status": result.get("stage_status"),
                "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
            })
            return result

        # 1. 实时扫描 Fast-Track 信号供前台色块展示
        recent_logs = db.query(MessageLog).filter(
            MessageLog.user_id == user_id,
            MessageLog.is_mock.is_(False),
        ).order_by(MessageLog.id.desc()).limit(15).all()
        fast_track_signals = collect_fast_track_signals(recent_logs)

        # 2. 获取 V1 及 V2 结果
        summary = db.query(IntentSummary).filter(IntentSummary.user_id == user_id).order_by(IntentSummary.id.desc()).first()
        result = _read_only_current_tail_result(
            db=db,
            user_id=user_id,
            summary=summary,
            recent_logs=recent_logs,
            fast_track_signals=fast_track_signals,
        )
        log_session_view_event("ANALYSIS", {
            "user_id": user_id,
            "snapshot_id": snapshot_id or "current_tail",
            "analysis_mode": result.get("analysis_mode"),
            "has_v1": result.get("has_v1"),
            "has_v2": result.get("has_v2"),
            "has_v2_compare": result.get("has_v2_compare"),
            "stage_status": result.get("stage_status"),
            "timings_ms": {"total_ms": round((perf_counter() - total_started) * 1000, 2)},
        })
        return result
    finally:
        db.close()

def _analytics_source_label(source: str | None) -> str:
    return {
        "api": "API触发",
        "web_manual": "网页人工触发",
        "test": "测试触发",
    }.get(_normalize_trigger_source(source, default="web_manual"), source or "-")


ANALYTICS_QUALITY_LABEL_OPTIONS = [
    "这条回复不好",
    "AI 太长",
    "没有正面报价",
    "追问太多",
    "不够像销售",
    "好样本",
]

ANALYTICS_QUALITY_SCORE_KEYS = {
    "kb1_eval_score",
    "kb2_eval_score",
    "step_6_main_kb1_score",
    "step_6_main_kb2_score",
    "step_6_compare_kb1_score",
    "step_6_compare_kb2_score",
    "manual_business_score",
    "manual_quality_score",
    "step_7_kb1_score",
    "step_7_kb2_score",
}

def _analytics_status_label(status: str | None) -> str:
    value = sanitize_text(str(status or "").strip())
    if not value:
        return "-"
    return {
        "success": "成功",
        "done": "已完成",
        "queued": "排队中",
        "running": "执行中",
        "error": "失败",
        "empty": "无可用数据",
        "ok": "正常",
        "scored": "已生成质量分",
        "pending_no_sales_reply": "待销售后续回复",
        "ignored_cross_day": "跨天回复不计统计",
        "reused_inflight": "复用进行中任务",
        "manual_review_required": "需人工复核",
        "low_confidence": "低置信度",
        "failed_no_content": "生成为空",
        "not_configured": "未配置",
        "not_started": "未开始",
        "skipped_no_query": "缺少检索问题，已跳过",
        "skipped_missing_knowledge": "缺少知识证据，已跳过",
        "connection_failed": "连接失败",
    }.get(value, value)

def _analytics_status_scope_label(scope: str | None) -> str:
    value = sanitize_text(str(scope or "").strip())
    return {
        "ongoing": "进行中",
        "future": "待推进",
        "done": "已完成",
    }.get(value, value or "-")

def _analytics_sender_label(sender_type: str | None) -> str:
    value = sanitize_text(str(sender_type or "").strip())
    return {
        "customer": "客户",
        "sales": "销售",
        "assistant": "系统",
        "email": "邮件",
    }.get(value, value or "未知角色")

def _analytics_format_time_text(value: Any) -> str:
    text_value = str(value or "").strip()
    if not text_value:
        return "-"
    try:
        return datetime.fromisoformat(text_value.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text_value.replace("T", " ")[:19] or text_value

def _analytics_trigger_label(kind: str | None, step: int | None = None, channel: str | None = None) -> str:
    if kind == "api_sidebar_assist":
        return "侧边栏API"
    if kind == "refresh_knowledge":
        return "知识库双通道刷新"
    if kind == "trigger_llm":
        return f"人工重跑步骤{step or '-'}"
    if channel:
        return channel
    return kind or "-"

def _analytics_join_lines(items: list[str], fallback: str = "-") -> str:
    values = [sanitize_text(str(item or "").strip()) for item in items if str(item or "").strip()]
    return "\n".join(values) if values else fallback

def _analytics_recent_customer_bundle(message_dicts: list[dict] | None) -> tuple[str, str, int]:
    customer_items = [item for item in (message_dicts or []) if str(item.get("sender_type") or "").strip() == "customer"]
    full_lines = [
        " / ".join(part for part in [str(item.get("time") or "").strip(), sanitize_text(str(item.get("content") or "").strip())] if part)
        for item in customer_items
    ]
    latest_text = sanitize_text(str((customer_items[-1] or {}).get("content") or "").strip()) if customer_items else "-"
    return latest_text or "-", _analytics_join_lines(full_lines), len(customer_items)

def _analytics_recent_dialog_bundle(message_dicts: list[dict] | None) -> tuple[str, str, int]:
    dialog_items = [item for item in (message_dicts or []) if sanitize_text(str(item.get("content") or "").strip())]
    full_lines = [
        f"{_analytics_format_time_text(item.get('time'))} {_analytics_sender_label(item.get('sender_type'))}：{sanitize_text(str(item.get('content') or '').strip())}"
        for item in dialog_items
    ]
    latest_item = dialog_items[-1] if dialog_items else {}
    latest_text = sanitize_text(str(latest_item.get("content") or "").strip()) or "-"
    return latest_text, _analytics_join_lines(full_lines), len(dialog_items)

def _analytics_candidate_text(candidate: dict | None) -> str:
    if not isinstance(candidate, dict):
        return "-"
    reply_reference = sanitize_text(str(candidate.get("reply_reference") or "").strip())
    followup = sanitize_text(str(candidate.get("followup_rationale") or "").strip())
    blocks = []
    if reply_reference:
        blocks.append(reply_reference)
    if followup:
        blocks.append(f"思路：{followup}")
    reason = sanitize_text(str(candidate.get("reason") or "").strip())
    if not blocks and reason:
        blocks.append(reason)
    return "\n".join(blocks) if blocks else "-"

def _analytics_find_candidate(candidates: list[dict] | None, model_slot: str, knowledge_source: str) -> dict | None:
    for item in candidates or []:
        if str(item.get("model_slot") or "").strip() != model_slot:
            continue
        if str(item.get("knowledge_source") or "").strip() != knowledge_source:
            continue
        return item
    return None

def _analytics_score_map(reply_scores: dict | None) -> dict[str, dict]:
    items = (reply_scores or {}).get("ai_candidates") or []
    return {
        str(item.get("candidate_id") or "").strip(): item
        for item in items
        if str(item.get("candidate_id") or "").strip()
    }


def _normalize_quality_annotation_map(payload: dict | None) -> dict[str, dict]:
    source = payload if isinstance(payload, dict) else {}
    normalized: dict[str, dict] = {}
    for key, value in source.items():
        score_key = str(key or "").strip()
        if score_key not in ANALYTICS_QUALITY_SCORE_KEYS or not isinstance(value, dict):
            continue
        labels = []
        for item in (value.get("labels") or []):
            label = str(item or "").strip()
            if label and label not in labels:
                labels.append(label)
        custom_text = sanitize_text(str(value.get("custom_text") or "").strip())
        normalized[score_key] = {
            "labels": labels,
            "custom_text": custom_text,
            "updated_at": sanitize_text(str(value.get("updated_at") or "").strip()) or None,
        }
    return normalized


def _quality_annotation_payload(score_key: str, value: dict | None) -> dict:
    normalized = _normalize_quality_annotation_map({score_key: value}).get(score_key) or {}
    return {
        "score_key": score_key,
        "labels": normalized.get("labels") or [],
        "custom_text": normalized.get("custom_text") or "",
        "updated_at": normalized.get("updated_at"),
    }

def _analytics_knowledge_summary(payload: dict | None) -> str:
    if not isinstance(payload, dict):
        return "-"
    hits = payload.get("hits") or []
    top_titles = []
    for item in hits[:3]:
        title = sanitize_text(str((item or {}).get("title") or (item or {}).get("content") or "").strip())
        if title:
            top_titles.append(title[:120])
    query_text = sanitize_text(str(payload.get("query_text") or "").strip())
    status = _analytics_status_label(payload.get("status"))
    parts = [f"状态：{status}"]
    if query_text:
        parts.append(f"检索：{query_text}")
    if top_titles:
        parts.append(f"命中：{'；'.join(top_titles)}")
    error_text = sanitize_text(str(payload.get("error") or "").strip())
    if error_text:
        parts.append(f"原因：{error_text}")
    return "\n".join(parts)

def _analytics_llm1_summary(result: dict | None) -> str:
    payload = result or {}
    facts = payload.get("key_facts") if isinstance(payload.get("key_facts"), dict) else {}
    timeline_items = facts.get("timeline_items") if isinstance(facts, dict) else []
    timeline_text = []
    for item in timeline_items[:3] if isinstance(timeline_items, list) else []:
        text_value = " · ".join(
            part for part in [
                str((item or {}).get("time") or "").strip(),
                str((item or {}).get("topic") or "").strip(),
                _analytics_status_scope_label((item or {}).get("status_scope")),
            ] if part
        )
        if text_value:
            timeline_text.append(text_value)
    parts = [
        f"主题：{sanitize_text(str(payload.get('topic') or '').strip()) or '-'}",
        f"诉求：{sanitize_text(str(payload.get('core_demand') or '').strip()) or '-'}",
        f"状态：{_analytics_status_label(payload.get('status'))}",
    ]
    if timeline_text:
        parts.append(f"时间线：{'；'.join(timeline_text)}")
    return "\n".join(parts)

def _analytics_crm_summary(result: dict | None) -> str:
    payload = result or {}
    crm = payload.get("crm_info") if isinstance(payload.get("crm_info"), dict) else {}
    thread_fact = payload.get("thread_business_fact") if isinstance(payload.get("thread_business_fact"), dict) else {}
    merged_facts = thread_fact.get("merged_facts") if isinstance(thread_fact.get("merged_facts"), dict) else {}
    timeline_items = payload.get("key_facts", {}).get("timeline_items") if isinstance(payload.get("key_facts"), dict) else []
    ongoing = [
        " · ".join(part for part in [str((item or {}).get("time") or "").strip(), str((item or {}).get("topic") or "").strip()] if part)
        for item in timeline_items or []
        if str((item or {}).get("status_scope") or "").strip() == "ongoing"
    ]
    future = [
        " · ".join(part for part in [str((item or {}).get("time") or "").strip(), str((item or {}).get("topic") or "").strip()] if part)
        for item in timeline_items or []
        if str((item or {}).get("status_scope") or "").strip() == "future"
    ]
    parts = [
        f"客户：{sanitize_text(str(crm.get('crm_contact_name') or '').strip()) or '-'}",
        f"公司：{sanitize_text(str(crm.get('company_name') or '').strip()) or '-'}",
        f"线程焦点：{sanitize_text(str(payload.get('topic') or payload.get('core_demand') or '').strip()) or '-'}",
        f"最新客户回复类型：{sanitize_text(str(merged_facts.get('latest_customer_reply_type') or '').strip()) or '-'}",
    ]
    if ongoing:
        parts.append(f"进行中：{'；'.join(ongoing[:3])}")
    if future:
        parts.append(f"未来：{'；'.join(future[:3])}")
    recent_followup = sanitize_text(str(crm.get("contact_recent_followup") or "").strip())
    if recent_followup:
        parts.append(f"最近跟进：{recent_followup}")
    return "\n".join(parts)

def _manual_business_validation_payload(result: dict | None) -> dict:
    payload = result or {}
    kb1 = payload.get("knowledge_v2") if isinstance(payload.get("knowledge_v2"), dict) else {}
    kb2 = payload.get("knowledge_external_api") if isinstance(payload.get("knowledge_external_api"), dict) else {}
    return {
        "hits": [
            *(kb1.get("hits") or []),
            *(kb2.get("hits") or []),
        ],
        "thread_business_fact": payload.get("thread_business_fact") if isinstance(payload.get("thread_business_fact"), dict) else None,
        "manual_review_required": False,
    }

def _manual_business_score_text_tokens(result: dict | None) -> tuple[str, str]:
    payload = result or {}
    thread_fact = payload.get("thread_business_fact") if isinstance(payload.get("thread_business_fact"), dict) else {}
    merged_facts = thread_fact.get("merged_facts") if isinstance(thread_fact.get("merged_facts"), dict) else {}
    topic = sanitize_text(str(payload.get("topic") or "").strip())
    core_demand = sanitize_text(str(payload.get("core_demand") or "").strip())
    latest_customer = sanitize_text(str(merged_facts.get("latest_customer_message") or "").strip())
    focus_text = " ".join(part for part in [topic, core_demand, latest_customer] if part).strip()
    return focus_text, latest_customer

def _score_manual_business_reply(reply_text: str | None, result: dict | None) -> tuple[int | None, str]:
    clean_text = _clean_reply_block_text(reply_text)
    if not clean_text:
        return None, "-"
    payload = result or {}
    crm_context = payload.get("crm_info") if isinstance(payload.get("crm_info"), dict) else {}
    validation = IntentEngine.validate_sales_assist_output(
        clean_text,
        _manual_business_validation_payload(payload),
        crm_context=crm_context,
    )
    warnings = [str(item or "").strip() for item in (validation.get("warnings") or []) if str(item or "").strip()]
    blocking = [str(item or "").strip() for item in (validation.get("blocking_issues") or []) if str(item or "").strip()]
    soft_warning_patterns = [
        "知识检索本身已标记人工复核",
    ]
    material_warnings = [
        item for item in warnings
        if not any(pattern in item for pattern in soft_warning_patterns)
    ]
    focus_text, latest_customer = _manual_business_score_text_tokens(payload)
    overlap_score = IntentEngine._text_overlap_ratio(clean_text, latest_customer or focus_text)
    domain_tokens = ["发", "安排", "确认", "报价", "付款", "开票", "交稿", "同步", "修改", "跟进", "看", "交付", "邮件", "资料"]
    explicit_tokens = ["下班前", "今天", "明天", "周一", "稍后", "马上", "一会", "收到后", "可以", "没问题", "按"]
    focus_tokens = [token for token in ["报价", "交付", "资料", "付款", "开票", "人数", "预算", "老师", "档期", "原文", "PPT", "翻译", "logo", "文件"] if token in focus_text and token in clean_text]

    intent_score = 20
    if overlap_score >= 0.08 or focus_tokens:
        intent_score += 5
    if overlap_score >= 0.18 or len(focus_tokens) >= 2:
        intent_score += 5
    if any(token in clean_text for token in domain_tokens + explicit_tokens):
        intent_score += 5
    intent_score = max(0, min(35, intent_score))

    next_step_score = 8
    if any(token in clean_text for token in domain_tokens):
        next_step_score += 6
    if any(token in clean_text for token in ["下班前", "今天", "明天", "周一", "稍后", "马上", "一会", "收到后", "后续"]):
        next_step_score += 4
    if "？" in clean_text or "?" in clean_text or any(token in clean_text for token in ["发我", "提供", "确认", "告诉我", "看需不需要"]):
        next_step_score += 2
    next_step_score = max(0, min(20, next_step_score))

    factual_score = 18
    for item in blocking:
        if "没有命中结构化 pricing_rule" in item:
            factual_score -= 3
        elif "能力/流程承诺" in item:
            factual_score -= 4
        else:
            factual_score -= 8
    factual_score -= min(6, len(material_warnings) * 2)
    factual_score = max(0, min(20, factual_score))

    risk_score = 12
    if not blocking and not material_warnings:
        risk_score += 3
    for item in blocking:
        if "没有命中结构化 pricing_rule" in item:
            risk_score -= 2
        elif "能力/流程承诺" in item:
            risk_score -= 3
        else:
            risk_score -= 5
    risk_score -= min(4, len(material_warnings) * 2)
    risk_score = max(0, min(15, risk_score))

    compact_len = len(re.sub(r"\s+", "", clean_text))
    line_count = len(_reply_block_lines_from_text(clean_text))
    wechat_score = 8
    if 6 <= compact_len <= 48:
        wechat_score += 2
    elif compact_len > 90 or compact_len < 3:
        wechat_score -= 2
    if line_count >= 4:
        wechat_score -= 1
    wechat_score = max(0, min(10, wechat_score))

    total_score = intent_score + next_step_score + factual_score + risk_score + wechat_score
    reason = _analytics_join_lines([
        f"意图命中：{intent_score}/35",
        f"下一步推进：{next_step_score}/20",
        f"事实正确：{factual_score}/20",
        f"风险控制：{risk_score}/15",
        f"微信表达：{wechat_score}/10",
        f"风险提示：{sanitize_text('；'.join((blocking + material_warnings)[:3])) or '无明显结构化风险'}",
    ])
    return total_score, reason

def _analytics_step7_summary(result: dict | None) -> str:
    payload = result or {}
    similarity = payload.get("api_quality_similarity") if isinstance(payload.get("api_quality_similarity"), dict) else {}
    if similarity:
        return "\n".join([
            f"状态：{_analytics_status_label(similarity.get('status'))}",
            f"贴合规则：{sanitize_text(str(similarity.get('quality_score_rule') or '').strip()) or '-'}",
            f"最佳贴合代理分：{sanitize_text(str(similarity.get('best_score') or '-').strip())}",
        ])
    validation = payload.get("assist_validation") if isinstance(payload.get("assist_validation"), dict) else {}
    if validation:
        warnings = "；".join([sanitize_text(str(item or "").strip()) for item in (validation.get("warnings") or [])[:3]])
        blocking = "；".join([sanitize_text(str(item or "").strip()) for item in (validation.get("blocking_issues") or [])[:3]])
        parts = [
            f"状态：{_analytics_status_label(validation.get('status'))}",
            f"人工复核：{'是' if validation.get('manual_review_required') else '否'}",
        ]
        if blocking:
            parts.append(f"阻断：{blocking}")
        if warnings:
            parts.append(f"提醒：{warnings}")
        return "\n".join(parts)
    return "-"

def _analytics_message_dicts_from_api(db: Session, item: ApiAssistInvocation) -> list[dict]:
    visible_ids = [int(x) for x in (item.visible_message_ids or []) if str(x).isdigit()]
    if visible_ids:
        logs = db.query(MessageLog).filter(MessageLog.id.in_(visible_ids)).order_by(MessageLog.id.asc()).all()
        return [_message_analytics_dict(log) for log in logs[-15:]]
    logs = db.query(MessageLog).filter(
        MessageLog.user_id == item.session_id,
        MessageLog.is_mock.is_(False),
    ).order_by(MessageLog.id.desc()).limit(15).all()
    return [_message_analytics_dict(log) for log in reversed(logs)]

def _build_trigger_analytics_row(
    *,
    row_id: str,
    triggered_at: datetime | None,
    trigger_source: str,
    trigger_kind: str,
    request_status: str,
    session_id: str,
    snapshot_id: str | None,
    run_id: str | None,
    anchor_message_id: int | None,
    anchor_message_time: datetime | None,
    anchor_message_text: str | None,
    input_messages: list[dict] | None,
    recent_customer_messages: list[dict] | None,
    manual_reply_text: str | None,
    result_payload: dict | None,
    requested_step: int | None = None,
    requested_channel: str | None = None,
    kb1_eval_score: float | None = None,
    kb1_eval_reason: str | None = None,
    kb2_eval_score: float | None = None,
    kb2_eval_reason: str | None = None,
    quality_annotations: dict | None = None,
) -> dict:
    result = _jsonable(result_payload or {})
    stage_status = result.get("stage_status") if isinstance(result.get("stage_status"), dict) else {}
    node_timings = result.get("node_timings_ms") if isinstance(result.get("node_timings_ms"), dict) else _extract_node_timings(stage_status)
    all_input_messages = input_messages or result.get("input_messages") or []
    llm_runtime = result.get("llm_runtime") if isinstance(result.get("llm_runtime"), dict) else {}
    customer_bundle = recent_customer_messages or [item for item in all_input_messages if str(item.get("sender_type") or "").strip() == "customer"]
    recent_customer_last, recent_customer_full, recent_customer_count = _analytics_recent_customer_bundle(customer_bundle)
    recent_dialog_last, recent_dialog_full, recent_dialog_count = _analytics_recent_dialog_bundle(all_input_messages)
    manual_reply_full = _clean_reply_block_text(str(
        ((result.get("api_quality_similarity") or {}).get("actual_sales_reply_text"))
        or manual_reply_text
        or _reply_block_text(result.get("actual_sales_replies") or [], strip_noise=True)
        or ""
    ).strip()) or "-"
    candidates = result.get("reply_style_results_v2") if isinstance(result.get("reply_style_results_v2"), list) else []
    score_map = _analytics_score_map(result.get("reply_scores_v2") if isinstance(result.get("reply_scores_v2"), dict) else {})
    main_kb1 = _analytics_find_candidate(candidates, "llm2", "knowledge_v2")
    main_kb2 = _analytics_find_candidate(candidates, "llm2", "knowledge_external_api")
    compare_kb1 = _analytics_find_candidate(candidates, "llm2_compare", "knowledge_v2")
    compare_kb2 = _analytics_find_candidate(candidates, "llm2_compare", "knowledge_external_api")
    similarity_scores = {
        str(item.get("key") or "").strip(): item
        for item in ((result.get("api_quality_similarity") or {}).get("scores") or [])
        if str(item.get("key") or "").strip()
    }

    def candidate_score_fields(candidate: dict | None) -> tuple[int | None, str]:
        if not isinstance(candidate, dict):
            return None, "-"
        score_entry = score_map.get(str(candidate.get("candidate_id") or "").strip()) or {}
        score_value = score_entry.get("overall_score")
        try:
            score_value = int(score_value) if score_value is not None else None
        except (TypeError, ValueError):
            score_value = None
        return score_value, sanitize_text(str(score_entry.get("score_reason") or "").strip()) or "-"

    main_kb1_score, main_kb1_reason = candidate_score_fields(main_kb1)
    main_kb2_score, main_kb2_reason = candidate_score_fields(main_kb2)
    compare_kb1_score, compare_kb1_reason = candidate_score_fields(compare_kb1)
    compare_kb2_score, compare_kb2_reason = candidate_score_fields(compare_kb2)
    sim_kb1 = similarity_scores.get("knowledge_v2") or {}
    sim_kb2 = similarity_scores.get("knowledge_external_api") or {}
    main_model_name = sanitize_text(str(
        (main_kb1 or {}).get("model_display_name")
        or (main_kb2 or {}).get("model_display_name")
        or ((llm_runtime.get("llm2") or {}).get("display_name"))
        or ((llm_runtime.get("llm2") or {}).get("model"))
        or "主模型"
    ).strip()) or "主模型"
    compare_model_name = sanitize_text(str(
        (compare_kb1 or {}).get("model_display_name")
        or (compare_kb2 or {}).get("model_display_name")
        or ((llm_runtime.get("llm2_compare") or {}).get("display_name"))
        or ((llm_runtime.get("llm2_compare") or {}).get("model"))
        or "对比模型"
    ).strip()) or "对比模型"
    api_quality_score = result.get("api_quality_score")
    try:
        api_quality_score = float(api_quality_score) if api_quality_score is not None else None
    except (TypeError, ValueError):
        api_quality_score = None
    api_quality_similarity = result.get("api_quality_similarity") if isinstance(result.get("api_quality_similarity"), dict) else {}
    api_quality_reason = _analytics_join_lines([
        f"状态：{_analytics_status_label(api_quality_similarity.get('status'))}",
        f"贴合规则：{sanitize_text(str(api_quality_similarity.get('quality_score_rule') or '').strip()) or '-'}",
        f"最佳贴合代理分：{sanitize_text(str(api_quality_similarity.get('best_score') or '-').strip())}",
    ], fallback="-")
    manual_business_score, manual_business_reason = _score_manual_business_reply(manual_reply_full, result)
    annotation_map = _normalize_quality_annotation_map(quality_annotations)
    return {
        "row_id": row_id,
        "triggered_at": triggered_at.isoformat() if triggered_at else None,
        "source": trigger_source,
        "source_label": _analytics_source_label(trigger_source),
        "trigger_kind": trigger_kind,
        "trigger_label": _analytics_trigger_label(trigger_kind, requested_step, requested_channel),
        "request_status": request_status or "-",
        "request_status_label": _analytics_status_label(request_status),
        "session_id": session_id,
        "snapshot_id": snapshot_id,
        "run_id": run_id,
        "anchor_message_id": anchor_message_id,
        "anchor_message_time": anchor_message_time.isoformat() if anchor_message_time else None,
        "anchor_message_text": sanitize_text(str(anchor_message_text or "").strip()) or "-",
        "recent_customer_last": recent_customer_last,
        "recent_customer_full": recent_customer_full,
        "recent_customer_count": recent_customer_count,
        "recent_dialog_last": recent_dialog_last,
        "recent_dialog_full": recent_dialog_full,
        "recent_dialog_count": recent_dialog_count,
        "manual_reply_text": manual_reply_full,
        "step_1_content": _analytics_join_lines([
            f"输入消息数：{len(all_input_messages)}",
            f"当前锚点：{sanitize_text(str(anchor_message_text or '').strip()) or recent_customer_last}",
        ]),
        "step_1_time_ms": (((node_timings or {}).get("conversation_input") or {}).get("total_ms")),
        "step_2_content": _analytics_join_lines([str(item or "").strip() for item in (result.get("fast_track") or [])], fallback="无强信号"),
        "step_2_time_ms": (((node_timings or {}).get("fast_track") or {}).get("total_ms")),
        "step_3_content": _analytics_llm1_summary(result),
        "step_3_time_ms": (((node_timings or {}).get("llm1") or {}).get("total_ms")),
        "step_4_content": _analytics_crm_summary(result),
        "step_4_time_ms": (((node_timings or {}).get("crm_profile") or {}).get("total_ms")),
        "step_5_kb1_content": _analytics_knowledge_summary(result.get("knowledge_v2") if isinstance(result.get("knowledge_v2"), dict) else {}),
        "step_5_kb1_time_ms": ((((node_timings or {}).get("knowledge") or {}).get("parts_ms") or {}).get("knowledge_v2_ms")),
        "step_5_kb2_content": _analytics_knowledge_summary(result.get("knowledge_external_api") if isinstance(result.get("knowledge_external_api"), dict) else {}),
        "step_5_kb2_time_ms": ((((node_timings or {}).get("knowledge") or {}).get("parts_ms") or {}).get("knowledge_external_api_ms")),
        "step_6_main_model_name": main_model_name,
        "step_6_compare_model_name": compare_model_name,
        "step_6_main_kb1_content": _analytics_candidate_text(main_kb1),
        "step_6_main_kb1_score": main_kb1_score,
        "step_6_main_kb1_score_reason": main_kb1_reason,
        "step_6_main_time_ms": ((((node_timings or {}).get("llm2") or {}).get("parts_ms") or {}).get("llm2_ms")),
        "step_6_main_kb2_content": _analytics_candidate_text(main_kb2),
        "step_6_main_kb2_score": main_kb2_score,
        "step_6_main_kb2_score_reason": main_kb2_reason,
        "step_6_compare_kb1_content": _analytics_candidate_text(compare_kb1),
        "step_6_compare_kb1_score": compare_kb1_score,
        "step_6_compare_kb1_score_reason": compare_kb1_reason,
        "step_6_compare_time_ms": ((((node_timings or {}).get("llm2") or {}).get("parts_ms") or {}).get("llm2_compare_ms")),
        "step_6_compare_kb2_content": _analytics_candidate_text(compare_kb2),
        "step_6_compare_kb2_score": compare_kb2_score,
        "step_6_compare_kb2_score_reason": compare_kb2_reason,
        "step_6_time_ms": (((node_timings or {}).get("llm2") or {}).get("total_ms")),
        "step_7_content": _analytics_step7_summary(result),
        "manual_business_score": manual_business_score,
        "manual_business_reason": manual_business_reason,
        "manual_quality_score": api_quality_score,
        "manual_quality_reason": api_quality_reason,
        "step_7_kb1_score": sim_kb1.get("score"),
        "step_7_kb1_reason": sanitize_text(str(sim_kb1.get("reason") or "").strip()) or "-",
        "step_7_kb2_score": sim_kb2.get("score"),
        "step_7_kb2_reason": sanitize_text(str(sim_kb2.get("reason") or "").strip()) or "-",
        "step_7_time_ms": (((node_timings or {}).get("validation") or {}).get("total_ms")),
        "kb1_eval_score": kb1_eval_score,
        "kb1_eval_reason": kb1_eval_reason or "-",
        "kb2_eval_score": kb2_eval_score,
        "kb2_eval_reason": kb2_eval_reason or "-",
        "quality_annotations": annotation_map,
    }

def _analytics_score_averages(rows: list[dict]) -> dict[str, float | None]:
    score_keys = [
        "step_6_main_kb1_score",
        "step_6_main_kb2_score",
        "step_6_compare_kb1_score",
        "step_6_compare_kb2_score",
        "manual_business_score",
        "manual_quality_score",
        "step_7_kb1_score",
        "step_7_kb2_score",
        "kb1_eval_score",
        "kb2_eval_score",
    ]
    averages: dict[str, float | None] = {}
    for key in score_keys:
        values = []
        for row in rows:
            try:
                raw_value = row.get(key)
                if raw_value is None or raw_value == "":
                    continue
                values.append(float(raw_value))
            except (TypeError, ValueError):
                continue
        averages[key] = round(sum(values) / len(values), 2) if values else None
    return averages

@app.get("/api/wecom/trigger_analytics")
async def get_wecom_trigger_analytics(
    start_date: str | None = Query(default=None),
    end_date: str | None = Query(default=None),
    sources: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
):
    db = SessionLocal()
    try:
        today = datetime.now().date()
        parsed_start = datetime.fromisoformat(str(start_date or today.isoformat())[:10])
        parsed_end = datetime.fromisoformat(str(end_date or str(start_date or today.isoformat()))[:10])
        if parsed_end < parsed_start:
            raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
        day_start = datetime(parsed_start.year, parsed_start.month, parsed_start.day)
        day_end = datetime(parsed_end.year, parsed_end.month, parsed_end.day) + timedelta(days=1)
        source_values = {
            _normalize_trigger_source(item.strip(), default="")
            for item in str(sources or "api,web_manual,test").split(",")
            if item.strip()
        } or {"api", "web_manual", "test"}

        rows: list[dict] = []
        api_items = db.query(ApiAssistInvocation).filter(
            ApiAssistInvocation.triggered_at >= day_start,
            ApiAssistInvocation.triggered_at < day_end,
        ).order_by(ApiAssistInvocation.triggered_at.desc()).all()
        for item in api_items:
            normalized_source = _normalize_trigger_source(item.trigger_source, default="api")
            if normalized_source not in source_values:
                continue
            result_payload = _api_invocation_analytics_payload(item)
            message_dicts = _analytics_message_dicts_from_api(db, item)
            rows.append(_build_trigger_analytics_row(
                row_id=str(item.invocation_id),
                triggered_at=item.triggered_at,
                trigger_source=normalized_source,
                trigger_kind=str(item.trigger_kind or "api_sidebar_assist"),
                request_status=str(item.quality_status or "success"),
                session_id=item.session_id,
                snapshot_id=str(item.invocation_id),
                run_id=None,
                anchor_message_id=item.anchor_message_id,
                anchor_message_time=item.anchor_message_time,
                anchor_message_text=item.anchor_message_text,
                input_messages=message_dicts,
                recent_customer_messages=[m for m in message_dicts if m.get("sender_type") == "customer"],
                manual_reply_text=item.actual_sales_reply_text,
                result_payload=result_payload,
                kb1_eval_score=item.kb1_eval_score,
                kb1_eval_reason=item.kb1_eval_reason,
                kb2_eval_score=item.kb2_eval_score,
                kb2_eval_reason=item.kb2_eval_reason,
                quality_annotations=item.quality_annotations,
            ))

        trigger_items = db.query(WecomTriggerRecord).filter(
            WecomTriggerRecord.triggered_at >= day_start,
            WecomTriggerRecord.triggered_at < day_end,
        ).order_by(WecomTriggerRecord.triggered_at.desc()).all()
        for item in trigger_items:
            normalized_source = _normalize_trigger_source(item.trigger_source, default="web_manual")
            if normalized_source not in source_values:
                continue
            rows.append(_build_trigger_analytics_row(
                row_id=str(item.record_id),
                triggered_at=item.triggered_at,
                trigger_source=normalized_source,
                trigger_kind=item.trigger_kind,
                request_status=item.request_status,
                session_id=item.session_id,
                snapshot_id=item.snapshot_id,
                run_id=item.run_id,
                anchor_message_id=item.anchor_message_id,
                anchor_message_time=item.anchor_message_time,
                anchor_message_text=item.anchor_message_text,
                input_messages=item.input_messages or [],
                recent_customer_messages=item.recent_customer_messages or [],
                manual_reply_text=item.actual_sales_reply_text,
                result_payload=item.result_payload or {},
                requested_step=item.requested_step,
                requested_channel=item.requested_channel,
                kb1_eval_score=item.kb1_eval_score,
                kb1_eval_reason=item.kb1_eval_reason,
                kb2_eval_score=item.kb2_eval_score,
                kb2_eval_reason=item.kb2_eval_reason,
                quality_annotations=item.quality_annotations,
            ))

        rows.sort(key=lambda item: str(item.get("triggered_at") or ""), reverse=True)
        rows = rows[:limit]
        return {
            "status": "success",
            "filters": {
                "start_date": parsed_start.date().isoformat(),
                "end_date": parsed_end.date().isoformat(),
                "sources": sorted(source_values),
                "limit": limit,
            },
            "summary": {
                "row_count": len(rows),
                "score_averages": _analytics_score_averages(rows),
            },
            "quality_annotation_options": ANALYTICS_QUALITY_LABEL_OPTIONS,
            "rows": rows,
        }
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式必须为 YYYY-MM-DD")
    finally:
        db.close()


@app.post("/api/wecom/trigger_analytics/quality_annotation")
async def save_wecom_trigger_quality_annotation(payload: TriggerAnalyticsQualityAnnotationRequest, db: Session = Depends(get_db)):
    score_key = str(payload.score_key or "").strip()
    if score_key not in ANALYTICS_QUALITY_SCORE_KEYS:
        raise HTTPException(status_code=400, detail="不支持的质量分字段")
    row_source = _normalize_trigger_source(payload.row_source, default="web_manual")
    row_id = str(payload.row_id or "").strip()
    if not row_id:
        raise HTTPException(status_code=400, detail="缺少记录 ID")

    labels = []
    for item in (payload.labels or []):
        label = str(item or "").strip()
        if label and label in ANALYTICS_QUALITY_LABEL_OPTIONS and label not in labels:
            labels.append(label)
    custom_text = sanitize_text(str(payload.custom_text or "").strip())
    target_row = (
        db.query(ApiAssistInvocation).filter(ApiAssistInvocation.invocation_id == row_id).first()
        if row_source == "api"
        else db.query(WecomTriggerRecord).filter(WecomTriggerRecord.record_id == row_id).first()
    )
    if not target_row:
        raise HTTPException(status_code=404, detail="未找到对应触发记录")

    annotation_map = _normalize_quality_annotation_map(getattr(target_row, "quality_annotations", None))
    annotation_map[score_key] = {
        "labels": labels,
        "custom_text": custom_text,
        "updated_at": datetime.utcnow().isoformat(),
    }
    target_row.quality_annotations = annotation_map
    db.add(target_row)
    db.commit()
    db.refresh(target_row)
    return {
        "status": "success",
        "row_source": row_source,
        "row_id": row_id,
        "annotation": _quality_annotation_payload(score_key, annotation_map.get(score_key)),
    }

@app.get("/api/thread_facts/{session_id}")
async def get_thread_business_fact(session_id: str, db: Session = Depends(get_db)):
    item = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Thread business fact not found")
    return {"status": "success", "thread_business_fact": _thread_fact_to_dict(item)}

@app.post("/api/sessions/{user_id}/trigger_llm")
async def trigger_llm_post(user_id: str, req: TriggerReq, background_tasks: BackgroundTasks, request: Request):
    """前端手动拔枪按钮请求"""
    run_id = uuid.uuid4().hex[:12]
    trigger_source = _resolve_trigger_source(request=request, explicit=req.trigger_source, default="web_manual")
    runtime_key = _reply_chain_runtime_key(user_id, req.snapshot_id)
    executor_task_key = _reply_chain_executor_task_key(user_id, req.snapshot_id)
    analytics_record_id = _create_wecom_trigger_record(
        session_id=user_id,
        snapshot_id=req.snapshot_id,
        run_id=run_id,
        trigger_source=trigger_source,
        trigger_kind="trigger_llm",
        requested_step=req.step,
        request_status="queued",
        request_payload={"step": req.step, "snapshot_id": req.snapshot_id},
    )
    payload = {
        "source": trigger_source,
        "client_host": request.client.host if request.client else "",
        "user_agent": request.headers.get("user-agent", "")[:160],
        "business_object": user_id,
        "step": req.step,
        "snapshot_id": req.snapshot_id,
        "run_id": run_id,
        "result": "accepted",
    }
    log_reply_chain_event("TRIGGER_REQUEST", payload)
    logger.info(f"收到前台发起的 LLM 手动指令！正在执行阶段 {req.step} ...")
    _mark_reply_chain_requested(user_id, req.snapshot_id, req.step, run_id)
    _future, accepted = _submit_reply_chain_task(user_id, req.snapshot_id, req.step, analytics_record_id)
    if not accepted:
        _update_wecom_trigger_record(
            analytics_record_id,
            request_status="reused_inflight",
            stage_status={"dedupe_status": "reused_inflight", "requested_step": req.step},
            error_message="同一会话已有相同后台任务执行中，本次触发复用已有任务",
        )
    return {
        "status": "success",
        "msg": f"已将步骤 {req.step} 投入后台大模型运算池" if accepted else f"当前节点已有任务执行中，继续复用已有后台任务（步骤 {req.step}）",
        "run_id": run_id,
        "runtime_key": runtime_key,
        "executor_task_key": executor_task_key,
        "accepted": accepted,
        "analytics_record_id": analytics_record_id,
        "llm_runtime": _llm_runtime_config(),
        "llm2_compare_configured": True,
    }

@app.post("/api/sessions/{user_id}/refresh_knowledge")
async def refresh_knowledge_post(user_id: str, req: KnowledgeRefreshReq, request: Request):
    run_id = uuid.uuid4().hex[:12]
    trigger_source = _resolve_trigger_source(request=request, explicit=req.trigger_source, default="web_manual")
    snapshot_id = req.snapshot_id
    channels = ["knowledge_v2", "knowledge_external_api"]
    accepted_channels: list[str] = []
    reused_channels: list[str] = []
    executor_task_keys: dict[str, str] = {}
    analytics_record_id = _create_wecom_trigger_record(
        session_id=user_id,
        snapshot_id=snapshot_id,
        run_id=run_id,
        trigger_source=trigger_source,
        trigger_kind="refresh_knowledge",
        request_status="queued",
        request_payload={"snapshot_id": snapshot_id, "channels": channels},
    )
    for channel in channels:
        task_key = _knowledge_refresh_executor_task_key(user_id, snapshot_id, channel)
        executor_task_keys[channel] = task_key
        with REPLY_CHAIN_ACTIVE_LOCK:
            active = REPLY_CHAIN_ACTIVE_FUTURES.get(task_key)
            is_running = bool(active and not active.done())
        if not is_running:
            _mark_knowledge_refresh_requested(user_id, snapshot_id, channel, run_id)
        _future, accepted = _submit_knowledge_refresh_task(user_id, snapshot_id, channel, analytics_record_id)
        if accepted:
            accepted_channels.append(channel)
        else:
            reused_channels.append(channel)
            _update_wecom_trigger_record(
            analytics_record_id,
            request_status="reused_inflight",
            stage_status={"dedupe_status": "reused_inflight", "requested_channel": channel},
            error_message=f"{channel} 已有后台刷新任务执行中，本次触发复用已有任务",
        )
    payload = {
        "source": trigger_source,
        "client_host": request.client.host if request.client else "",
        "user_agent": request.headers.get("user-agent", "")[:160],
        "business_object": user_id,
        "snapshot_id": snapshot_id,
        "run_id": run_id,
        "action": "knowledge_refresh",
        "accepted_channels": accepted_channels,
        "reused_channels": reused_channels,
        "result": "accepted" if accepted_channels else "reused",
    }
    log_reply_chain_event("KNOWLEDGE_REFRESH_REQUEST", payload)
    return {
        "status": "success",
        "msg": "已将知识库双通道刷新投入后台任务池" if accepted_channels else "当前节点已有知识库刷新任务执行中，继续复用已有后台任务",
        "run_id": run_id,
        "runtime_key": _reply_chain_runtime_key(user_id, snapshot_id),
        "executor_task_keys": executor_task_keys,
        "analytics_record_id": analytics_record_id,
        "accepted_channels": accepted_channels,
        "reused_channels": reused_channels,
        "accepted": bool(accepted_channels),
    }

class SidebarAssistRequest(BaseModel):
    external_userid: str
    userid: str
    force_refresh: bool = False
    sync_archive_before_read: bool = settings.SIDEBAR_ASSIST_SYNC_ARCHIVE_BEFORE_READ_DEFAULT
    trigger_source: str | None = None

class SessionArchiveSyncRequest(BaseModel):
    session_id: str | None = None
    userid: str | None = None
    external_userid: str | None = None
    limit: int = 100

@app.post("/api/wecom/sidebar_assist")
async def sidebar_assist(request: SidebarAssistRequest, http_request: Request):
    """供企微原生右侧边栏应用调用的同步直出接口"""
    from time import perf_counter

    total_started = perf_counter()
    timings_ms = {}
    stage_status = {}
    knowledge_status = "not_started"
    crm_context = None
    knowledge_base = None
    knowledge_v2 = None
    external_knowledge = None
    assist_bundle = None
    compare_bundle = None
    all_messages: list[MessageLog] = []
    dedupe_request_key: str | None = None
    dedupe_future: Future | None = None
    dedupe_future_owner = False

    def mark_timing(name: str, started_at: float):
        timings_ms[name] = round((perf_counter() - started_at) * 1000, 2)

    db = SessionLocal()
    try:
        stage_started = perf_counter()
        external_userid = request.external_userid.strip()
        sales_userid = request.userid.strip()
        trigger_source = _resolve_trigger_source(request=http_request, explicit=request.trigger_source, default="api")
        mark_timing("normalize_request_ms", stage_started)

        if not external_userid or not sales_userid:
            timings_ms["total_ms"] = round((perf_counter() - total_started) * 1000, 2)
            log_sidebar_result("ERROR", {
                "reason": "empty_required_params",
                "external_userid": external_userid,
                "sales_userid": sales_userid,
                "timings_ms": timings_ms,
            })
            return {
                "status": "error",
                "msg": "external_userid 和 userid 不能为空",
                "timings_ms": timings_ms
            }

        stage_started = perf_counter()
        requested_session_id = build_single_session_id(sales_userid, external_userid)
        mark_timing("build_session_id_ms", stage_started)
        stage_status["requested_session_id"] = requested_session_id

        if request.sync_archive_before_read:
            stage_started = perf_counter()
            archive_sync = await _sync_archive_for_session(requested_session_id)
            mark_timing("archive_sync_ms", stage_started)
            stage_status["archive_sync"] = archive_sync.get("status")
            if archive_sync.get("reason"):
                stage_status["archive_sync_reason"] = archive_sync.get("reason")
            stage_status["archive_sync_timeout_seconds"] = archive_sync.get("timeout_seconds")
            db.close()
            db = SessionLocal()
        else:
            stage_status["archive_sync"] = "scheduled_async" if _schedule_archive_sync_for_session(requested_session_id) else "skipped_by_request"

        stage_started = perf_counter()
        session_id, recent_logs, session_lookup = find_existing_single_session_id(db, sales_userid, external_userid, limit=15)
        mark_timing("load_recent_messages_ms", stage_started)
        stage_status["session_lookup"] = session_lookup
        logger.info(
            "收到侧边栏辅助请求：客服[%s] -> 客户[%s] -> 请求会话[%s] -> 命中会话[%s]",
            sales_userid,
            external_userid,
            requested_session_id,
            session_id,
        )
        stage_started = perf_counter()
        all_messages = _load_session_messages(db, session_id) if session_id else []
        mark_timing("load_all_messages_ms", stage_started)
        _set_node_timing(
            stage_status,
            "conversation_input",
            total_ms=(timings_ms.get("load_recent_messages_ms") or 0) + (timings_ms.get("load_all_messages_ms") or 0),
            parts_ms={
                "session_lookup_ms": timings_ms.get("load_recent_messages_ms"),
                "full_session_load_ms": timings_ms.get("load_all_messages_ms"),
            },
            status="done" if recent_logs else "empty",
        )
        stage_started = perf_counter()
        dedupe_request_key = _sidebar_assist_request_key(
            session_id=session_id or requested_session_id,
            messages=all_messages,
            force_refresh=bool(request.force_refresh),
        )
        if not request.force_refresh:
            cached_result = _get_cached_sidebar_assist_result(dedupe_request_key)
            if cached_result:
                mark_timing("dedupe_lookup_ms", stage_started)
                cached_result = _decorate_sidebar_assist_result(
                    cached_result,
                    status="reused_recent_cache",
                    reason="短时间内消息内容未变化，直接复用最近一次已完成结果",
                    request_key=dedupe_request_key,
                )
                log_sidebar_result("REUSED", {
                    "external_userid": external_userid,
                    "sales_userid": sales_userid,
                    "session_id": session_id,
                    "requested_session_id": requested_session_id,
                    "dedupe_status": "reused_recent_cache",
                    "dedupe_reason": "same_content_recent_cache",
                    "latest_dialog_count": len(recent_logs),
                    "request_key": dedupe_request_key,
                    "timings_ms": {
                        **timings_ms,
                        "total_ms": round((perf_counter() - total_started) * 1000, 2),
                    },
                })
                return cached_result
        dedupe_future, dedupe_future_owner = _acquire_sidebar_assist_future(dedupe_request_key)
        mark_timing("dedupe_lookup_ms", stage_started)
        if not dedupe_future_owner:
            reused_result = await asyncio.wrap_future(dedupe_future)
            reused_result = _decorate_sidebar_assist_result(
                reused_result,
                status="reused_inflight",
                reason="同一会话、同一消息内容的请求仍在处理中，当前请求直接复用进行中的结果",
                request_key=dedupe_request_key,
            )
            log_sidebar_result("REUSED", {
                "external_userid": external_userid,
                "sales_userid": sales_userid,
                "session_id": session_id,
                "requested_session_id": requested_session_id,
                "dedupe_status": "reused_inflight",
                "dedupe_reason": "same_content_inflight",
                "latest_dialog_count": len(recent_logs),
                "request_key": dedupe_request_key,
                "timings_ms": {
                    **timings_ms,
                    "total_ms": round((perf_counter() - total_started) * 1000, 2),
                },
            })
            return reused_result

        # Fast-Track 前置扫描：既返回给侧边栏展示，也作为 LLM-2 的输入信号
        stage_started = perf_counter()
        fast_track_signals = collect_fast_track_signals(recent_logs)
        mark_timing("fast_track_scan_ms", stage_started)
        _set_node_timing(
            stage_status,
            "fast_track",
            total_ms=timings_ms.get("fast_track_scan_ms"),
            status="done",
        )
        
        # 1. 触发同步测算 (阶段1 和 阶段2)
        stage_started = perf_counter()
        summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
        mark_timing("load_summary_ms", stage_started)
        if summary and isinstance(summary.stage_status, dict) and summary.stage_status.get("node_timings_ms"):
            stored_timings = _jsonable(summary.stage_status.get("node_timings_ms") or {})
            current_timings = _jsonable(stage_status.get("node_timings_ms") or {})
            stage_status["node_timings_ms"] = {
                **stored_timings,
                **current_timings,
            }
        
        need_v1 = request.force_refresh or summary is None
        if need_v1:
            stage_status["analysis_mode"] = "force_refresh" if request.force_refresh else "first_analyze"
            if recent_logs:
                context = [{"content": l.content, "sender_type": l.sender_type} for l in reversed(recent_logs)]
                stage_started = perf_counter()
                summary_payload = IntentEngine._run_llm1(session_id, context)
                mark_timing("llm1_analyze_ms", stage_started)
                _set_node_timing(
                    stage_status,
                    "llm1",
                    total_ms=timings_ms.get("llm1_analyze_ms"),
                    status="done" if summary_payload else "error",
                )
                if not summary_payload:
                    raise RuntimeError("LLM-1 未返回有效结构化摘要")
                summary_record = IntentSummary(user_id=session_id)
                _apply_summary_fields(summary_record, summary_payload)
                summary_record.llm1_compare_summary = None
                summary_record.llm1_compare_prompt_trace = None
                summary_record.stage_status = {
                    "conversation_input": "done" if context else "empty",
                    "fast_track": "done",
                    "llm1": "done",
                    "crm_profile": "not_started",
                    "knowledge_v2": "not_started",
                    "knowledge_external_api": "not_started",
                    "llm2": "not_started",
                    "llm2_compare": "not_started",
                }
                db.add(summary_record)
                db.commit()
                stage_status["llm1"] = "done"
            else:
                stage_status["llm1"] = "skipped_no_recent_logs"
                _set_node_timing(stage_status, "llm1", total_ms=0, status="skipped_no_recent_logs")
                
            stage_started = perf_counter()
            summary = db.query(IntentSummary).filter(IntentSummary.user_id == session_id).order_by(IntentSummary.id.desc()).first()
            mark_timing("reload_summary_ms", stage_started)
        else:
            stage_status["analysis_mode"] = "complete_missing_v2" if not summary.sales_advice_v2 else "cache"
            stage_status["llm1"] = "skipped_summary_exists"

        if summary and (
            request.force_refresh
            or not summary.sales_advice_v2
            or not summary.sales_advice_compare_v2
        ):
            # 同步执行 V2
            summary_json = {
                "topic": summary.topic, "core_demand": summary.core_demand, "key_facts": summary.key_facts,
                "todo_items": summary.todo_items, "risks": summary.risks, "to_be_confirmed": summary.to_be_confirmed,
                "fast_track_signals": fast_track_signals
            }
            # CRM 画像按外部联系人 ID 查询，不使用组合会话 ID
            stage_started = perf_counter()
            crm_context = IntentEngine.get_crm_context(external_userid)
            mark_timing("crm_profile_ms", stage_started)
            _set_node_timing(
                stage_status,
                "crm_profile",
                total_ms=timings_ms.get("crm_profile_ms"),
                status=(crm_context or {}).get("crm_profile_status") or "unknown",
            )
            thread_fact = _upsert_thread_business_fact(
                db,
                session_id=session_id,
                summary_json={**summary_json, "status": summary.status},
                crm_context=crm_context,
                messages=[{"content": item.content, "sender_type": item.sender_type} for item in recent_logs],
                external_userid=external_userid,
                sales_userid=sales_userid,
            )
            summary.crm_info = crm_context
            summary.crm_status = crm_context.get("crm_profile_status")
            summary.thread_business_fact = _thread_fact_to_dict(thread_fact)
            db.commit()

            knowledge_stage_started = perf_counter()
            knowledge_parts_ms: dict[str, float] = {}
            external_knowledge = None
            if summary.core_demand:
                stage_started = perf_counter()
                thread_fact_payload = _thread_fact_to_dict(thread_fact)
                query_features = IntentEngine.infer_query_features(summary_json, crm_context, thread_fact_payload)
                retrieval_query = IntentEngine.build_retrieval_query(summary_json, thread_fact_payload)
                knowledge_v2 = IntentEngine.retrieve_knowledge_v2(
                    retrieval_query,
                    query_features=query_features,
                    top_k=5,
                    request_id=f"sidebar_{uuid.uuid4().hex[:12]}",
                    session_id=session_id,
                )
                knowledge_v2["thread_business_fact"] = thread_fact_payload
                knowledge_base = knowledge_v2
                knowledge_parts_ms["knowledge_v2_ms"] = round((perf_counter() - stage_started) * 1000, 2)
                stage_started = perf_counter()
                external_knowledge = _search_sales_kb_api(retrieval_query, top_k=5)
                knowledge_parts_ms["knowledge_external_api_ms"] = round((perf_counter() - stage_started) * 1000, 2)
                mark_timing("knowledge_retrieve_ms", knowledge_stage_started)
                knowledge_status = knowledge_v2.get("status", "error")
                stage_status["knowledge_v2"] = knowledge_status
            else:
                knowledge_v2 = {"status": "skipped_no_core_demand", "hits": [], "evidence_context": {"rules": [], "faqs": [], "cases": []}, "evidence_refs": []}
                knowledge_base = knowledge_v2
                knowledge_status = "skipped_no_core_demand"
                stage_started = perf_counter()
                external_knowledge = _search_sales_kb_api(None, top_k=5)
                knowledge_parts_ms["knowledge_external_api_ms"] = round((perf_counter() - stage_started) * 1000, 2)
                mark_timing("knowledge_retrieve_ms", knowledge_stage_started)
            stage_status["knowledge_external_api"] = external_knowledge.get("status") or "unknown"
            _set_node_timing(
                stage_status,
                "knowledge",
                total_ms=timings_ms.get("knowledge_retrieve_ms"),
                parts_ms=knowledge_parts_ms,
                status=stage_status.get("knowledge_v2") or stage_status.get("knowledge_external_api") or "unknown",
            )
            summary.knowledge_log_id = knowledge_v2.get("log_id")
            summary.knowledge_v2 = knowledge_v2
            summary.knowledge_external_api = external_knowledge
            summary.knowledge_status = knowledge_v2.get("status")
            summary.knowledge_confidence_score = _decimal_score(knowledge_v2.get("confidence_score"))
            summary.knowledge_manual_review_required = bool(knowledge_v2.get("manual_review_required"))

            stage_started = perf_counter()
            generation_bundle = _generate_reply_style_candidates(
                summary_json=summary_json,
                knowledge=knowledge_v2,
                knowledge_compare=external_knowledge,
                crm_context=crm_context,
                actual_sales_replies=[],
                runtime_key=session_id,
                single_model_single_style=True,
                enable_scoring=False,
            )
            mark_timing("llm2_generate_ms", stage_started)
            primary_candidate = generation_bundle.get("primary_candidate")
            compare_candidate = generation_bundle.get("compare_candidate")
            first_compare_entry = _first_model_candidate_entry(
                generation_bundle.get("candidates"),
                "llm2",
                "knowledge_external_api",
            )
            summary.reply_style_results_v2 = generation_bundle.get("candidates")
            summary.reply_scores_v2 = generation_bundle.get("reply_scores")
            summary.sales_advice_v2 = primary_candidate.get("content") if primary_candidate else None
            summary.sales_advice_compare_v2 = compare_candidate.get("content") if compare_candidate else None
            summary.sales_advice_compare_prompt_trace_v2 = (compare_candidate or first_compare_entry or {}).get("prompt_trace")
            summary.assist_validation = primary_candidate.get("validation") if primary_candidate else None
            summary.assist_compare_validation = compare_candidate.get("validation") if compare_candidate else None
            llm2_stage_timing = generation_bundle.get("stage_timings_ms") or {}
            llm2_compare_status = "skipped_api_isolated"
            _set_node_timing(
                stage_status,
                "llm2",
                total_ms=llm2_stage_timing.get("total_ms"),
                parts_ms=llm2_stage_timing.get("parts_ms"),
                status=generation_bundle.get("llm2_status") or "failed_no_content",
            )
            summary.stage_status = {
                **dict(summary.stage_status or {}),
                **stage_status,
                "llm1": stage_status.get("llm1") or "done",
                "crm_profile": summary.crm_status or stage_status.get("crm_profile") or "unknown",
                "knowledge_v2": summary.knowledge_status or stage_status.get("knowledge_v2") or "not_started",
                "knowledge_external_api": external_knowledge.get("status") or stage_status.get("knowledge_external_api") or "unknown",
                "llm2": generation_bundle.get("llm2_status"),
                "llm2_compare": llm2_compare_status,
            }
            db.commit()
            stage_status["llm2"] = generation_bundle.get("llm2_status")
            stage_status["llm2_compare"] = llm2_compare_status
            if primary_candidate:
                assist_bundle = {"validation": primary_candidate.get("validation")}
            if compare_candidate:
                compare_bundle = {"validation": compare_candidate.get("validation")}
        else:
            stage_status["llm2"] = "skipped_advice_exists" if summary and summary.sales_advice_v2 else "skipped_no_summary"
            stage_status["llm2_compare"] = "skipped_api_isolated"

        # 3. 封装全量字段给侧边栏前端
        compare_runtime_status = {
            "status": "skipped_api_isolated",
            "reason": "API 调用链路已禁用 LLM-2 对比模型",
        }
        result = {
            "status": "success",
            "external_userid": external_userid,
            "sales_userid": sales_userid,
            "session_id": session_id,
            "requested_session_id": requested_session_id,
            "latest_dialog_count": len(recent_logs),
            "fast_track": fast_track_signals,
            "archive_sync": archive_sync if request.sync_archive_before_read else {"status": "skipped_by_request"},
            "has_v1": False,
            "has_v2": False,
            "has_v2_compare": False,
            "llm_runtime": _llm_runtime_config(),
            "llm2_compare_configured": False,
            "llm2_compare_runtime_status": compare_runtime_status,
            "stage_status": stage_status
        }
        
        if summary:
            result.update({
                "has_v1": True,
                "topic": summary.topic,
                "core_demand": summary.core_demand,
                "key_facts": summary.key_facts,
                "risks": summary.risks,
                "todo_items": summary.todo_items,
                "to_be_confirmed": summary.to_be_confirmed
            })
            if summary.sales_advice_v2:
                result["has_v2"] = True
                _apply_sales_advice_output_fields(result, summary.sales_advice_v2)
                if assist_bundle:
                    result["assist_validation"] = assist_bundle.get("validation")
                    result["evidence_refs"] = assist_bundle.get("evidence_refs") or []
                # 回传检索到的原始知识块参考文档
                if knowledge_base is not None:
                    result["knowledge_v2"] = knowledge_base
                    result["knowledge_base"] = [hit.get("content") for hit in knowledge_base.get("hits", [])] if isinstance(knowledge_base, dict) else knowledge_base
                    if isinstance(knowledge_base, dict):
                        result["knowledge_evidence_context"] = knowledge_base.get("evidence_context")
                        result["evidence_refs"] = result.get("evidence_refs") or knowledge_base.get("evidence_refs") or []
                        result["knowledge_confidence_score"] = knowledge_base.get("confidence_score")
                        result["knowledge_manual_review_required"] = knowledge_base.get("manual_review_required")
                elif summary.core_demand:
                    stage_started = perf_counter()
                    current_thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
                    thread_fact_payload = _thread_fact_prompt_dict(current_thread_fact)
                    summary_payload = {
                        "topic": summary.topic,
                        "core_demand": summary.core_demand,
                        "key_facts": summary.key_facts,
                        "risks": summary.risks,
                        "status": summary.status,
                    }
                    query_features = IntentEngine.infer_query_features(summary_payload, crm_context, thread_fact_payload)
                    retrieval_query = IntentEngine.build_retrieval_query(summary_payload, thread_fact_payload)
                    knowledge_v2 = IntentEngine.retrieve_knowledge_v2(
                        retrieval_query,
                        query_features=query_features,
                        top_k=5,
                        request_id=f"sidebar_cache_{uuid.uuid4().hex[:12]}",
                        session_id=session_id,
                    )
                    knowledge_v2["thread_business_fact"] = thread_fact_payload
                    result["knowledge_v2"] = knowledge_v2
                    result["knowledge_base"] = [hit.get("content") for hit in knowledge_v2.get("hits", [])]
                    result["knowledge_evidence_context"] = knowledge_v2.get("evidence_context")
                    result["evidence_refs"] = knowledge_v2.get("evidence_refs") or []
                    validation = IntentEngine.validate_sales_assist_output(summary.sales_advice_v2, knowledge_v2, crm_context=crm_context)
                    result["assist_validation"] = validation
                    result["knowledge_confidence_score"] = knowledge_v2.get("confidence_score")
                    result["knowledge_manual_review_required"] = knowledge_v2.get("manual_review_required")
                    mark_timing("knowledge_retrieve_ms", stage_started)
                    knowledge_status = knowledge_v2.get("status", "error")
                else:
                    result["knowledge_base"] = []
                    result["knowledge_v2"] = {"status": "skipped_no_core_demand", "hits": [], "evidence_refs": []}
                    result["evidence_refs"] = []
                    knowledge_status = "skipped_no_core_demand"
                # KB2 外部知识库结果写入 result_payload
                if external_knowledge is not None:
                    result["knowledge_external_api"] = external_knowledge
                elif summary.knowledge_external_api is not None:
                    result["knowledge_external_api"] = summary.knowledge_external_api
            if summary.sales_advice_compare_v2:
                result["has_v2_compare"] = True
                _apply_sales_advice_output_fields(result, summary.sales_advice_compare_v2, compare=True)
                if compare_bundle:
                    result["assist_compare_validation"] = compare_bundle.get("validation")
            if summary.reply_style_results_v2:
                result["reply_style_results_v2"] = _serialize_reply_candidates_for_output(summary.reply_style_results_v2)
            if summary.reply_scores_v2:
                result["reply_scores_v2"] = summary.reply_scores_v2
            if summary.sales_advice_compare_prompt_trace_v2:
                result["sales_advice_compare_prompt_trace_v2"] = summary.sales_advice_compare_prompt_trace_v2
            
            # 始终回传 CRM 画像供前端显示
            if crm_context is None:
                stage_started = perf_counter()
                crm_context = IntentEngine.get_crm_context(external_userid)
                mark_timing("crm_profile_ms", stage_started)
                _set_node_timing(
                    stage_status,
                    "crm_profile",
                    total_ms=timings_ms.get("crm_profile_ms"),
                    status=(crm_context or {}).get("crm_profile_status") or "unknown",
                )
            result["crm_info"] = crm_context
            current_thread_fact = db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
            if not current_thread_fact:
                current_thread_fact = _upsert_thread_business_fact(
                    db,
                    session_id=session_id,
                    summary_json={
                        "topic": summary.topic,
                        "core_demand": summary.core_demand,
                        "key_facts": summary.key_facts,
                        "todo_items": summary.todo_items,
                        "risks": summary.risks,
                        "to_be_confirmed": summary.to_be_confirmed,
                        "status": summary.status,
                    },
                    crm_context=crm_context,
                    messages=[{"content": item.content, "sender_type": item.sender_type} for item in recent_logs],
                    external_userid=external_userid,
                    sales_userid=sales_userid,
                )
                db.commit()
            result["thread_business_fact"] = _thread_fact_to_dict(current_thread_fact)
        else:
            if crm_context is None:
                stage_started = perf_counter()
                crm_context = IntentEngine.get_crm_context(external_userid)
                mark_timing("crm_profile_ms", stage_started)
                _set_node_timing(
                    stage_status,
                    "crm_profile",
                    total_ms=timings_ms.get("crm_profile_ms"),
                    status=(crm_context or {}).get("crm_profile_status") or "unknown",
                )
            result["crm_info"] = crm_context
            result["thread_business_fact"] = _thread_fact_to_dict(
                db.query(ThreadBusinessFact).filter(ThreadBusinessFact.session_id == session_id).first()
            )

        result["crm_status"] = (crm_context or {}).get("crm_profile_status", "unknown")
        result["knowledge_status"] = knowledge_status
        timings_ms["total_ms"] = round((perf_counter() - total_started) * 1000, 2)
        result["timings_ms"] = timings_ms
        result["node_timings_ms"] = _extract_node_timings(stage_status)
        result = _sanitize_api_sidebar_result_payload(_jsonable(result))
        api_invocation = _store_api_assist_invocation(
            db,
            result_payload=result,
            session_id=session_id,
            requested_session_id=requested_session_id,
            external_userid=external_userid,
            sales_userid=sales_userid,
            messages=all_messages,
            trigger_source=trigger_source,
        )
        db.commit()
        log_sidebar_result("SUCCESS", {
            "external_userid": external_userid,
            "sales_userid": sales_userid,
            "session_id": session_id,
            "requested_session_id": requested_session_id,
            "session_lookup": session_lookup,
            "latest_dialog_count": len(recent_logs),
            "has_v1": result.get("has_v1"),
            "has_v2": result.get("has_v2"),
            "has_v2_compare": result.get("has_v2_compare"),
            "crm_status": result.get("crm_status"),
            "knowledge_status": result.get("knowledge_status"),
            "api_invocation_id": str(api_invocation.invocation_id) if api_invocation else None,
            "dedupe_status": "fresh",
            "request_key": dedupe_request_key,
            "stage_status": stage_status,
            "timings_ms": timings_ms,
        })
        final_result = _api_invocation_result_payload(api_invocation) if api_invocation else result
        final_result = _decorate_sidebar_assist_result(
            final_result,
            status="fresh",
            reason="本次请求触发了新的侧边栏分析流程",
            request_key=dedupe_request_key,
        )
        if summary and api_invocation:
            REPLY_CHAIN_EXECUTOR.submit(
                _complete_api_reply_scoring_async,
                summary_id=summary.id,
                invocation_id=str(api_invocation.invocation_id),
                session_id=session_id,
            )
        if dedupe_request_key:
            _remember_sidebar_assist_result(dedupe_request_key, final_result)
        if dedupe_request_key and dedupe_future_owner and dedupe_future:
            _resolve_sidebar_assist_future(dedupe_request_key, dedupe_future, result=final_result)
        return final_result
    except Exception as e:
        if dedupe_request_key and dedupe_future_owner and dedupe_future:
            _resolve_sidebar_assist_future(dedupe_request_key, dedupe_future, error=e)
        logger.error(f"侧边栏接口异常退出: {e}")
        timings_ms["total_ms"] = round((perf_counter() - total_started) * 1000, 2)
        log_sidebar_result("ERROR", {
            "error": str(e),
            "request_key": dedupe_request_key,
            "stage_status": stage_status,
            "timings_ms": timings_ms,
        })
        return {
            "status": "error",
            "msg": str(e),
            "stage_status": stage_status,
            "timings_ms": timings_ms
        }
    finally:
        db.close()

@app.post("/api/sync/session")
async def sync_single_session_messages(payload: SessionArchiveSyncRequest, request: Request):
    requested_session_id = str(payload.session_id or "").strip()
    if not requested_session_id:
        userid = str(payload.userid or "").strip()
        external_userid = str(payload.external_userid or "").strip()
        if not userid or not external_userid:
            raise HTTPException(status_code=400, detail="session_id 或 userid + external_userid 至少提供一组")
        requested_session_id = build_single_session_id(userid, external_userid)

    log_reply_chain_event("SYNC_REQUEST", {
        "source": "frontend",
        "client_host": request.client.host if request.client else "",
        "user_agent": request.headers.get("user-agent", "")[:160],
        "business_object": requested_session_id,
        "step": "sync_session",
        "result": "started",
    })

    sync_result = await _sync_archive_for_session(requested_session_id)
    if sync_result.get("status") == "timeout":
        raise HTTPException(status_code=504, detail=sync_result.get("msg"))
    if sync_result.get("status") == "error":
        raise HTTPException(status_code=500, detail=sync_result.get("msg"))

    db = SessionLocal()
    try:
        resolved_session_id = requested_session_id
        if not payload.session_id and payload.userid and payload.external_userid:
            resolved_session_id, _, _ = find_existing_single_session_id(
                db,
                str(payload.userid).strip(),
                str(payload.external_userid).strip(),
                limit=max(1, min(payload.limit, 500)),
            )
        messages = db.query(MessageLog).filter(
            MessageLog.user_id == resolved_session_id,
            MessageLog.is_mock.is_(False),
        ).order_by(MessageLog.id.desc()).limit(max(1, min(payload.limit, 500))).all()
        messages = list(reversed(messages))
        return {
            "status": "success",
            "requested_session_id": requested_session_id,
            "resolved_session_id": resolved_session_id,
            "sync_result": sync_result,
            "message_count": len(messages),
            "messages": [
                {"id": m.id, "sender": m.sender_type, "content": m.content, "time": m.timestamp}
                for m in messages
            ],
        }
    finally:
        db.close()

@app.post("/api/sync/today")
async def sync_today_messages(request: Request):
    """触发真实会话存档同步"""
    from archive_service import ArchiveService
    log_reply_chain_event("SYNC_REQUEST", {
        "source": "frontend",
        "client_host": request.client.host if request.client else "",
        "user_agent": request.headers.get("user-agent", "")[:160],
        "business_object": "today_archive_sync",
        "step": "sync",
        "result": "started",
    })
    try:
        timeout_seconds = max(30, int(settings.ARCHIVE_SYNC_TIMEOUT_SECONDS))
        res = await asyncio.wait_for(asyncio.to_thread(ArchiveService.sync_today_data), timeout=timeout_seconds)
    except asyncio.TimeoutError:
        reason = f"企微会话存档同步超过 {timeout_seconds} 秒未返回，可能是外部 SDK、网络或企微服务器阻塞"
        diagnostics = ArchiveService.get_runtime_diagnostics()
        log_reply_chain_event("SYNC_RESULT", {
            "source": "frontend",
            "business_object": "today_archive_sync",
            "step": "sync",
            "result": "timeout",
            "reason": reason,
            "timeout_seconds": timeout_seconds,
        })
        raise HTTPException(status_code=504, detail={
            "message": reason,
            "timeout_seconds": timeout_seconds,
            "diagnostics": diagnostics,
        })
    if res.get("status") == "error":
        diagnostics = ArchiveService.get_runtime_diagnostics()
        log_reply_chain_event("SYNC_RESULT", {
            "source": "frontend",
            "business_object": "today_archive_sync",
            "step": "sync",
            "result": "error",
            "reason": res.get("msg"),
        })
        raise HTTPException(status_code=500, detail={
            "message": res.get("msg"),
            "sdk_error_code": res.get("sdk_error_code"),
            "current_seq": res.get("current_seq"),
            "timings_ms": res.get("timings_ms"),
            "diagnostics": diagnostics,
        })
    log_reply_chain_event("SYNC_RESULT", {
        "source": "frontend",
        "business_object": "today_archive_sync",
        "step": "sync",
        "result": res.get("status"),
        "message": res.get("msg"),
    })
    return res

def archive_sdk_status():
    from archive_service import ArchiveService
    return {
        "status": "running", 
        "archive_sdk": ArchiveService.get_token_status()
    }
