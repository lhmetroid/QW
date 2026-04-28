from __future__ import annotations

import csv
from io import BytesIO, StringIO


class EmailImportService:
    SUBJECT_ALIASES = ["主题", "邮件主题", "subject", "title"]
    CONTENT_ALIASES = ["正文", "内容", "邮件内容", "摘要", "body", "content", "本次邮件内容提纯后", "PureEmailText", "原始邮件内容", "OriThisEmail"]
    SENDER_ALIASES = ["发件人", "sender", "from", "from_email"]
    RECEIVER_ALIASES = ["收件人", "receiver", "to", "to_email"]
    SENT_AT_ALIASES = ["发送时间", "日期", "时间", "sent_at", "date", "created_at"]
    THREAD_ID_ALIASES = ["thread_id", "会话ID", "线程ID", "邮件线程ID", "conversation_id"]
    SESSION_ID_ALIASES = ["session_id", "企微会话ID", "会话标识", "session"]
    EXTERNAL_USERID_ALIASES = ["external_userid", "客户ID", "外部联系人ID", "external_user_id"]
    SALES_USERID_ALIASES = ["sales_userid", "销售ID", "跟进人ID", "userid", "user_id"]
    ATTACHMENT_NAMES_ALIASES = ["attachment_names", "attachment_name", "attachments", "files", "附件", "附件名", "文件名"]
    ATTACHMENT_SUMMARY_ALIASES = ["attachment_summary", "attachment_desc", "attachment_note", "附件说明", "附件摘要", "文件说明"]
    ATTACHMENT_CONTENT_ALIASES = ["attachment_content", "attachment_text", "附件内容", "文件内容", "附件文本"]
    ATTACHMENT_TIME_ALIASES = ["attachment_time", "file_time", "附件时间", "文件时间"]
    USE_RANGE_ALIASES = ["CorrectUseRange", "人工最终确认本次邮件的用途", "用途范围", "use_range"]
    INTENT_TYPE_ALIASES = ["IntentType", "最终用途类别", "意图类别", "intent_type"]
    EMAIL_STAGE_ALIASES = ["EmailStage", "邮件阶段", "阶段", "email_stage"]

    @classmethod
    def _normalize_header(cls, value) -> str:
        return str(value or "").strip().replace(" ", "").replace("\n", "").lower()

    @classmethod
    def _find_column_index(cls, headers: list, aliases: list[str]) -> int | None:
        normalized = [cls._normalize_header(item) for item in headers]
        for alias in aliases:
            target = cls._normalize_header(alias)
            if target in normalized:
                return normalized.index(target)
        return None

    @staticmethod
    def _cell(row: tuple, index: int | None):
        if index is None or index >= len(row):
            return None
        value = row[index]
        if value is None:
            return None
        return str(value).strip() if isinstance(value, str) else value

    @staticmethod
    def _normalize_attachment_names(value) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        normalized: list[str] = []
        for block in text.replace("；", ";").replace("，", ",").split(";"):
            for token in [item.strip() for item in block.split(",") if item.strip()]:
                if token and token not in normalized:
                    normalized.append(token[:120])
        return normalized

    @classmethod
    def _rows_to_records(cls, rows: list[tuple]) -> list[dict]:
        if not rows:
            return []
        headers = list(rows[0])
        subject_idx = cls._find_column_index(headers, cls.SUBJECT_ALIASES)
        content_idx = cls._find_column_index(headers, cls.CONTENT_ALIASES)
        sender_idx = cls._find_column_index(headers, cls.SENDER_ALIASES)
        receiver_idx = cls._find_column_index(headers, cls.RECEIVER_ALIASES)
        sent_at_idx = cls._find_column_index(headers, cls.SENT_AT_ALIASES)
        thread_id_idx = cls._find_column_index(headers, cls.THREAD_ID_ALIASES)
        session_id_idx = cls._find_column_index(headers, cls.SESSION_ID_ALIASES)
        external_userid_idx = cls._find_column_index(headers, cls.EXTERNAL_USERID_ALIASES)
        sales_userid_idx = cls._find_column_index(headers, cls.SALES_USERID_ALIASES)
        attachment_names_idx = cls._find_column_index(headers, cls.ATTACHMENT_NAMES_ALIASES)
        attachment_summary_idx = cls._find_column_index(headers, cls.ATTACHMENT_SUMMARY_ALIASES)
        attachment_content_idx = cls._find_column_index(headers, cls.ATTACHMENT_CONTENT_ALIASES)
        attachment_time_idx = cls._find_column_index(headers, cls.ATTACHMENT_TIME_ALIASES)
        use_range_idx = cls._find_column_index(headers, cls.USE_RANGE_ALIASES)
        intent_type_idx = cls._find_column_index(headers, cls.INTENT_TYPE_ALIASES)
        email_stage_idx = cls._find_column_index(headers, cls.EMAIL_STAGE_ALIASES)
        if subject_idx is None and content_idx is None:
            raise ValueError("未找到邮件主题或正文列")

        records = []
        for row_number, row in enumerate(rows[1:], start=2):
            subject = cls._cell(row, subject_idx)
            content = cls._cell(row, content_idx)
            if str(subject or "").strip() in {"OriThisEmail", "PureEmailText"} or str(content or "").strip() in {"OriThisEmail", "PureEmailText"}:
                continue
            if not subject and not content:
                continue
            records.append(
                {
                    "row": row_number,
                    "subject": str(subject or "").strip(),
                    "content": str(content or "").strip(),
                    "sender": cls._cell(row, sender_idx),
                    "receiver": cls._cell(row, receiver_idx),
                    "sent_at": cls._cell(row, sent_at_idx),
                    "thread_id": cls._cell(row, thread_id_idx),
                    "session_id": cls._cell(row, session_id_idx),
                    "external_userid": cls._cell(row, external_userid_idx),
                    "sales_userid": cls._cell(row, sales_userid_idx),
                    "attachment_names": cls._normalize_attachment_names(cls._cell(row, attachment_names_idx)),
                    "attachment_summary": cls._cell(row, attachment_summary_idx),
                    "attachment_content": cls._cell(row, attachment_content_idx),
                    "attachment_time": cls._cell(row, attachment_time_idx),
                    "use_range_label": cls._cell(row, use_range_idx),
                    "intent_type_label": cls._cell(row, intent_type_idx),
                    "email_stage_label": cls._cell(row, email_stage_idx),
                }
            )
        return records

    @classmethod
    def parse_file(cls, raw: bytes, filename: str) -> list[dict]:
        lower = (filename or "").lower()
        if lower.endswith((".xlsx", ".xlsm")):
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                raise ValueError(f"openpyxl 不可用: {exc}") from exc
            try:
                workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            except Exception as exc:
                raise ValueError(f"Excel 文件无法解析: {exc}") from exc
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            return cls._rows_to_records(rows)

        if lower.endswith(".csv"):
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("gbk", errors="ignore")
            reader = csv.reader(StringIO(text))
            rows = [tuple(row) for row in reader]
            return cls._rows_to_records(rows)

        raise ValueError("仅支持 .xlsx/.xlsm/.csv 邮件整理文件")
